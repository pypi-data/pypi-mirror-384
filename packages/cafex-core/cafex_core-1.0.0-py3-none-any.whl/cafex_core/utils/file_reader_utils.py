"""
Description: This module contains the FileReaderUtil class, which provides methods to
read various types of files. The supported file types include PDF, Word, CSV, and Excel files.
"""
import csv
import os
import platform
import warnings
import zipfile
from typing import Union


import pandas as pd
import openpyxl
from bs4 import BeautifulSoup
from docx import Document
from fitz import fitz

from cafex_core.utils.exceptions import CoreExceptions
from cafex_core.logging.logger_ import CoreLogger

if platform.system().upper() == "WINDOWS":
    import py7zr


class FileReaderUtil:
    """
    Description:
        |  This class contains different methods that can be used to read various types of files,
        |  such as Word files, CSV files, and PDF files.
    """

    def __init__(self):
        self.__exceptions_generic = CoreExceptions()
        self.logger = CoreLogger(name=__name__).get_logger()

    def read_pdf_file(self, filepath: str) -> list:
        """
        Reads a PDF file and returns its content as a list of strings.

        Args:
            filepath (str): The complete path of the PDF file to be read.

        Returns:
            list: A list of strings representing the text content of the PDF.

        Examples:
            >> pdf_content = FileReaderUtil().read_pdf_file("path/to/file.pdf")
        """
        if not filepath.endswith('.pdf'):
            self.logger.error("The provided file is not a PDF.")
            raise ValueError("The provided file is not a PDF.")
        content_list = []
        try:
            with fitz.open(filepath) as doc:
                for page in doc:
                    text = page.get_text("text")
                    paragraph = text.replace("\n", " ").replace("\r", "")
                    content_list.append(paragraph)
            self.logger.info("Successfully read PDF file: %s", filepath)
            return content_list

        except Exception as e:
            error_description = f"An error occurred while reading the PDF file: {str(e)}"
            self.__exceptions_generic.raise_generic_exception(
                message=error_description,
                insert_report=True,
                trim_log=True,
                log_local=True,
                fail_test=True,
            )
            raise e

    def read_excel(self, filepath: str, sheet_reference: Union[str, int, None] = None) -> list:
        """
        Reads an Excel file and returns the specified sheet's content.

        Args:
            filepath (str): The complete path of the Excel file to be read.
            sheet_reference (str or int, optional): The name or index of the sheet to read.
                If None, the active sheet will be read.

        Returns:
            list: The content of the specified sheet.

        Examples:
            >> values = FileReaderUtil().read_excel("path/to/file.xlsx", "Sheet1")
            >> values = FileReaderUtil().read_excel("path/to/file.xlsx", 0)  # By index
        """
        warnings.filterwarnings("ignore", category=PendingDeprecationWarning)
        warnings.filterwarnings("ignore", category=DeprecationWarning)

        try:
            workbook = openpyxl.load_workbook(filepath, data_only=True)
            if isinstance(sheet_reference, str):
                if sheet_reference in workbook.sheetnames:
                    sheet = workbook[sheet_reference]
                else:
                    self.logger.error("Sheet name %s does not exist.", sheet_reference)
                    raise ValueError(f"Sheet name '{sheet_reference}' does not exist.")
            elif isinstance(sheet_reference, int):
                if 0 <= sheet_reference < len(workbook.worksheets):
                    sheet = workbook.worksheets[sheet_reference]
                else:
                    self.logger.error("Sheet index %s is out of range.", sheet_reference)
                    raise ValueError(f"Sheet index '{sheet_reference}' is out of range.")
            else:
                sheet = workbook.active
            content_list = [list(row) for row in sheet.iter_rows(values_only=True)]
            self.logger.info("Successfully read Excel file: %s, sheet: %s", filepath, sheet.title)
            return content_list
        except Exception as e:
            error_description = f"An error occurred while reading the Excel file: {str(e)}"
            self.__exceptions_generic.raise_generic_exception(
                message=error_description,
                insert_report=True,
                trim_log=True,
                log_local=True,
                fail_test=True,
            )
            raise e

    def read_word_file(self, filepath: str) -> list:
        """
        Reads a Word file and returns its content as a list of strings.

        Args:
            filepath (str): The complete path of the Word file to be read.

        Returns:
            list: A list of strings representing the text content of the Word file.

        Examples:
            >> word_content = FileReaderUtil().read_word_file("path/to/file.docx")
        """
        try:
            with zipfile.ZipFile(filepath) as document:
                xml_content = document.read("word/document.xml")
            string_content = xml_content.decode("utf-8", errors="ignore")
            soup = BeautifulSoup(string_content, features="lxml")
            text_elements = soup.find_all("w:t")
            content_list = [text.text for text in text_elements]
            self.logger.info("Successfully read Word file: %s", filepath)
            return content_list
        except Exception as e:
            error_description = f"An error occurred while reading the Word file: {str(e)}"
            self.__exceptions_generic.raise_generic_exception(
                message=error_description,
                insert_report=True,
                trim_log=True,
                log_local=True,
                fail_test=True,
            )
            raise e

    def read_csv_file(self, filepath: str) -> list:
        """
        Reads a CSV file and returns its content as a list of rows.

        Args:
            filepath (str): The complete path of the CSV file to be read.

        Returns:
            list: A list of rows from the CSV file.

        Examples:
            >> csv_content = FileReaderUtil().read_csv_file("path/to/file.csv")
        """
        try:
            with open(filepath, "r", newline='', encoding='utf-8') as csv_file:
                reader = csv.reader(csv_file)
                content_list = list(reader)
            self.logger.info("Successfully read CSV file: %s", filepath)
            return content_list
        except Exception as e:
            error_description = f"An error occurred while reading the CSV file: {str(e)}"
            self.__exceptions_generic.raise_generic_exception(
                message=error_description,
                insert_report=True,
                trim_log=True,
                log_local=True,
                fail_test=True,
            )
            raise e

    def extract_7z_file(self, zip_file_name: str, target_dir: str) -> str:
        """
        Extracts a 7Z file to the specified target directory.

        Args:
            zip_file_name (str): The complete path of the 7Z file to be extracted.
            target_dir (str): The target directory where the files will be extracted.

        Returns:
            str: The path of the first extracted file.

        Examples:
            >> extracted_file = FileReaderUtil().
            extract_7z_file("path/to/file.7z", "path/to/extract")
        """
        if platform.system().upper() != "WINDOWS":
            error_message = "7Z extraction is only supported on Windows."
            self.logger.error(error_message)
            raise EnvironmentError(error_message)

        try:
            with py7zr.SevenZipFile(zip_file_name, mode="r") as archive:
                archive.extractall(path=target_dir)
                extracted_file_names = archive.getnames()
                if not extracted_file_names:
                    raise ValueError("No files were extracted from the archive.")
                extracted_file_path = os.path.join(target_dir, extracted_file_names[0])
                self.logger.info("Successfully extracted 7Z file: %s to %s",
                                 zip_file_name, target_dir)
                return extracted_file_path
        except Exception as e:
            error_description = f"The provided file is not supported: {str(e)}"
            self.__exceptions_generic.raise_generic_exception(
                message=error_description,
                insert_report=True,
                trim_log=True,
                log_local=True,
                fail_test=True,
            )
            raise e

    def unzip_zip_file(self, source_path: str, destination_path: str) -> None:
        """
        Unzips a ZIP file to the specified destination path.

        Args:
            source_path (str): The complete path of the ZIP file to be unzipped.
            destination_path (str): The destination path where the files will be extracted.

        Returns:
            None

        Examples:
            >> FileReaderUtil().unzip_zip_file("path/to/file.zip", "path/to/extract")
        """
        try:
            if not os.path.exists(destination_path):
                os.makedirs(destination_path)
            with zipfile.ZipFile(source_path, 'r') as zip_file:
                zip_file.extractall(destination_path)
            self.logger.info("Successfully unzipped ZIP file: %s to %s",
                             source_path, destination_path)
        except Exception as e:
            error_description = f"The provided file is not supported: {str(e)}"
            self.__exceptions_generic.raise_generic_exception(
                message=error_description,
                insert_report=True,
                trim_log=True,
                log_local=True,
                fail_test=True,
            )
            raise e

    def read_txt_file(self, filepath: str) -> str:
        """
        Reads the contents of a text file.

        Args:
            filepath (str): The complete path of the text file to be read.

        Returns:
            str: The contents of the text file.

        Raises:
            Exception: If an error occurs while reading the text file.

        Examples:
            >> content = FileReaderUtil().read_txt_file("path/to/file.txt")
        """
        try:
            with open(filepath, 'r', encoding='utf-8') as txt_file:
                content = txt_file.read()
            self.logger.info("Successfully read text file: %s", filepath)
            return content
        except Exception as e:
            error_description = f"An error occurred while reading the text file: {str(e)}"
            self.__exceptions_generic.raise_generic_exception(
                message=error_description,
                insert_report=True,
                trim_log=True,
                log_local=True,
                fail_test=True,
            )
            raise e

    def write_txt_file(self, filepath: str, content: str) -> None:
        """
        Writes content to a text file.

        Args:
            filepath (str): The complete path of the text file to be written.
            content (str): The content to write to the text file.

        Returns:
            None

        Raises:
            Exception: If an error occurs while writing to the text file.

        Examples:
            >> FileReaderUtil().write_txt_file("path/to/file.txt", "Hello, World!")
        """
        try:
            with open(filepath, 'w', encoding='utf-8') as txt_file:
                txt_file.write(content)
            self.logger.info("Successfully wrote to text file: %s", filepath)
        except Exception as e:
            error_description = f"An error occurred while writing to the text file: {str(e)}"
            self.__exceptions_generic.raise_generic_exception(
                message=error_description,
                insert_report=True,
                trim_log=True,
                log_local=True,
                fail_test=True,
            )
            raise e

    def write_excel_file(self, filepath: str, data: list[dict]) -> None:
        """
        Writes data to an Excel file.

        Args:
            filepath (str): The complete path of the Excel file to be written.
            data (list of dict): The data to write to the Excel file.

        Returns:
            None

        Raises:
            Exception: If an error occurs while writing to the Excel file.

        Examples:
            >> data = [{'Name': 'John', 'Age': 30}, {'Name': 'Jane', 'Age': 25}]
            >> FileReaderUtil().write_excel_file("path/to/file.xlsx", data)
        """
        try:
            df = pd.DataFrame(data)
            df.to_excel(filepath, index=False)
            self.logger.info("Successfully wrote to Excel file: %s", filepath)
        except Exception as e:
            error_description = f"An error occurred while writing to the Excel file: {str(e)}"
            self.__exceptions_generic.raise_generic_exception(
                message=error_description,
                insert_report=True,
                trim_log=True,
                log_local=True,
                fail_test=True,
            )
            raise e

    def write_word_file(self, filepath: str, content: list[str]) -> None:
        """
        Writes content to a Word file.

        Args:
            filepath (str): The complete path of the Word file to be written.
            content (list of str): The content to write to the Word file, where each
            string is a paragraph.

        Returns:
            None

        Raises:
            Exception: If an error occurs while writing to the Word file.

        Examples:
            >> content = ["First paragraph.", "Second paragraph."]
            >> FileReaderUtil().write_word_file("path/to/file.docx", content)
        """
        try:
            doc = Document()
            for paragraph in content:
                doc.add_paragraph(paragraph)
            doc.save(filepath)
            self.logger.info("Successfully wrote to Word file: %s", filepath)
        except Exception as e:
            error_description = f"An error occurred while writing to the Word file: {str(e)}"
            self.__exceptions_generic.raise_generic_exception(
                message=error_description,
                insert_report=True,
                trim_log=True,
                log_local=True,
                fail_test=True,
            )
            raise e

    def write_csv_file(self, filepath: str, data: list[list]) -> None:
        """
        Writes data to a CSV file.

        Args:
            filepath (str): The complete path of the CSV file to be written.
            data (list of list): The data to write to the CSV file, where each inner list
            represents a row.

        Returns:
            None

        Raises:
            Exception: If an error occurs while writing to the CSV file.

r4[p'        Examples:
            >> data = [['Name', 'Age'], ['John', 30], ['Jane', 25]]
            >> FileReaderUtil().write_csv_file("path/to/file.csv", data)
        """
        try:
            with open(filepath, 'w', newline='', encoding='utf-8') as csv_file:
                writer = csv.writer(csv_file)
                writer.writerows(data)
            self.logger.info("Successfully wrote to CSV file: %s", filepath)
        except Exception as e:
            error_description = f"An error occurred while writing to the CSV file: {str(e)}"
            self.__exceptions_generic.raise_generic_exception(
                message=error_description,
                insert_report=True,
                trim_log=True,
                log_local=True,
                fail_test=True,
            )
            raise e

    def delete_file(self, filepath: str) -> None:
        """
        Deletes a file from the filesystem.

        Args:
            filepath (str): The complete path of the file to be deleted.

        Returns:
            None

        Raises:
            Exception: If an error occurs while deleting the file, or if the file does not exist.

        Examples:
            >> FileReaderUtil().delete_file("path/to/file.txt")
        """
        try:
            if os.path.exists(filepath):
                os.remove(filepath)
                self.logger.info("Successfully deleted file: %s", filepath)
            error_message = f"The file '{filepath}' does not exist."
            self.logger.error(error_message)
            raise error_message
        except Exception as e:
            error_description = f"An error occurred while deleting the file: {str(e)}"
            self.__exceptions_generic.raise_generic_exception(
                message=error_description,
                insert_report=True,
                trim_log=True,
                log_local=True,
                fail_test=True,
            )
            raise e
