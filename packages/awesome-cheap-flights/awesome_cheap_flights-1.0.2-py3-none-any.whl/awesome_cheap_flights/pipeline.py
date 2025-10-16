from __future__ import annotations

import csv
import json
import re
import time
from collections import defaultdict
from concurrent.futures import Future, ThreadPoolExecutor, as_completed
from dataclasses import asdict, dataclass, fields
from datetime import datetime, timedelta
from itertools import product
from pathlib import Path
from threading import Lock
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

from fast_flights import FlightData, Passengers, Result
from fast_flights.core import parse_response
from fast_flights.fallback_playwright import CODE as PLAYWRIGHT_FALLBACK_CODE
from fast_flights.flights_impl import TFSData
from fast_flights.primp import Client
from selectolax.lexbor import LexborHTMLParser

from rich import box
from rich.console import Console
from rich.markup import escape
from rich.progress import (
    BarColumn,
    Progress,
    SpinnerColumn,
    TaskProgressColumn,
    TextColumn,
    TimeElapsedColumn,
    TimeRemainingColumn,
)
from rich.table import Table

TIME_PATTERN = re.compile(
    r"(?P<hour>\d{1,2}):(?P<minute>\d{2})\s*(?P<ampm>AM|PM)\s*on\s*"
    r"(?P<weekday>[A-Za-z]{3}),\s*(?P<month>[A-Za-z]{3})\s*(?P<day>\d{1,2})"
    r"(?:\s*\+(?P<plus>\d+)\s*day[s]?)?"
)


DURATION_PATTERN = re.compile(r'(?:(?P<hours>\d+)\s*h(?:ours?)?)?(?:\s*(?P<minutes>\d+)\s*m(?:in)?)?', re.IGNORECASE)
TIME_LABEL_PATTERN = re.compile(r"(\d{1,2}:\d{2})\s*(AM|PM)?", re.IGNORECASE)

DEFAULT_ITINERARY_LEG_LIMIT = 0
DEFAULT_ITINERARY_MAX_COMBINATIONS = 0
RECOMMENDED_ITINERARY_LEG_LIMIT = 10


console = Console(stderr=True, highlight=False)


class ProgressReporter:
    def __init__(
        self,
        total_steps: int,
        *,
        config: Optional["SearchConfig"] = None,
        plan: Optional[PlanConfig] = None,
        output_path: Optional[Path] = None,
    ) -> None:
        self.total_steps = total_steps
        self.rows_collected = 0
        self.processed = 0
        self.skipped = 0
        self._task_id: Optional[int] = None
        self._progress: Optional[Progress] = None
        self._start = time.perf_counter()
        self._config = config
        self._plan = plan
        self._output_path = output_path
        self._lock = Lock()

    def __enter__(self) -> "ProgressReporter":
        if self.total_steps > 0 and console.is_terminal:
            self._progress = Progress(
                SpinnerColumn(),
                TextColumn("{task.description}"),
                BarColumn(bar_width=None),
                TaskProgressColumn(),
                TimeElapsedColumn(),
                TimeRemainingColumn(),
                console=console,
                redirect_stdout=False,
                redirect_stderr=False,
            )
            self._progress.__enter__()
            self._task_id = self._progress.add_task("Preparing", total=self.total_steps)
        self._render_overview()
        tip = "Press Ctrl+C anytime to pause and save a draft CSV."
        if self._progress is not None:
            self._progress.log(f"[cyan]{tip}")
        else:
            console.log(f"[cyan]{tip}")
        return self

    def __exit__(self, exc_type, exc, exc_tb) -> None:
        if self._progress is not None:
            self._progress.__exit__(exc_type, exc, exc_tb)
        self._render_summary()

    def record_success(self, label: str, rows_added: int, details: str | None = None) -> None:
        with self._lock:
            self.processed += 1
            self.rows_collected += rows_added
            message = f"+{rows_added} rows"
            total_note = f"total {self.rows_collected}"
            entry = label if not details else f"{label} · {details}"
            log_line = f"[green]{entry}[/] {message} {total_note}"
            if self._progress is not None and self._task_id is not None:
                self._progress.update(self._task_id, advance=1, description=label)
                self._progress.log(log_line)
            else:
                console.log(log_line)

    def record_skip(self, label: str, reason: str, details: str | None = None) -> None:
        with self._lock:
            self.processed += 1
            self.skipped += 1
            note = reason if reason.endswith(".") else f"{reason}."
            total_note = f"total {self.rows_collected}"
            entry = label if not details else f"{label} · {details}"
            log_line = f"[yellow]{entry}[/] {note} {total_note}"
            if self._progress is not None and self._task_id is not None:
                self._progress.update(self._task_id, advance=1, description=label)
                self._progress.log(log_line)
            else:
                console.log(log_line)

    def log_warning(self, message: str) -> None:
        text = message if message.endswith(".") else f"{message}."
        with self._lock:
            if self._progress is not None:
                self._progress.log(f"[yellow]{text}")
            else:
                console.log(f"[yellow]{text}")

    def log_error(self, message: str) -> None:
        text = message if message.endswith(".") else f"{message}."
        with self._lock:
            if self._progress is not None:
                self._progress.log(f"[red]{text}")
            else:
                console.log(f"[red]{text}")

    def _render_summary(self) -> None:
        elapsed_minutes = (time.perf_counter() - self._start) / 60 or 0.0
        table = Table(title="Flight Capture Summary", box=box.SIMPLE_HEAD)
        table.add_column("Metric", justify="left")
        table.add_column("Value", justify="right")
        table.add_row("Itineraries processed", str(self.processed))
        table.add_row("Itineraries skipped", str(self.skipped))
        table.add_row("Rows collected", str(self.rows_collected))
        table.add_row("Elapsed minutes", f"{elapsed_minutes:.2f}")
        console.print(table)

    def _render_overview(self) -> None:
        if self._config is None or self._plan is None:
            return
        table = Table(title="Search Overview", box=box.SIMPLE_HEAD)
        table.add_column("Field", justify="left")
        table.add_column("Value", justify="right")
        table.add_row("Plan", self._plan.name)
        table.add_row("Path", " → ".join(self._plan.path))
        table.add_row("Journeys", str(self.total_steps))
        table.add_row("Passengers", str(self._config.passenger_count))
        table.add_row("Currency", self._config.currency_code)
        default_max = self._config.filters.max_stops
        max_label = "All" if default_max is None else str(default_max)
        table.add_row("Default max stops", max_label)
        table.add_row("Include hidden", "yes" if self._plan.options.include_hidden else "no")
        table.add_row("Hidden hop limit", str(self._plan.options.max_hidden_hops))
        table.add_row("Max leg results", str(self._config.request.max_leg_results))
        table.add_row("Request delay", f"{self._config.request.delay:.2f}s")
        table.add_row("Retry limit", str(self._config.request.retries))
        table.add_row("Concurrency", str(self._config.concurrency))
        table.add_row("HTTP proxy", self._config.http_proxy or "-")
        if self._output_path is not None:
            table.add_row("Output", escape(str(self._output_path)))
        console.print(table)


class RunInterrupted(Exception):
    def __init__(
        self,
        *,
        plan: PlanConfig,
        output_path: Path,
        rows: List[SegmentRow],
        processed: int,
        skipped: int,
        collected: int,
        total: int,
    ) -> None:
        super().__init__("Search interrupted by user")
        self.plan = plan
        self.output_path = output_path
        self.rows = rows
        self.processed = processed
        self.skipped = skipped
        self.collected = collected
        self.total = total

    @property
    def remaining(self) -> int:
        value = self.total - self.processed
        return value if value > 0 else 0


def _warn(message: str, reporter: Optional[ProgressReporter] = None) -> None:
    if reporter is not None:
        reporter.log_warning(message)
    else:
        text = message if message.endswith(".") else f"{message}."
        console.log(f"[yellow]{text}")


def _error(message: str, reporter: Optional[ProgressReporter] = None) -> None:
    if reporter is not None:
        reporter.log_error(message)
    else:
        text = message if message.endswith(".") else f"{message}."
        console.log(f"[red]{text}")


MAX_EXCEPTION_PREVIEW = 160


def _exception_summary(exc: Exception) -> str:
    detail = str(exc).strip()
    if not detail:
        return exc.__class__.__name__
    first_line = detail.splitlines()[0]
    if len(first_line) > MAX_EXCEPTION_PREVIEW:
        return first_line[: MAX_EXCEPTION_PREVIEW - 3] + "..."
    return first_line


def _friendly_exception(exc: Exception, *, debug: bool) -> str:
    if debug:
        detail = str(exc).strip()
        return detail or exc.__class__.__name__
    summary = _exception_summary(exc)
    if "no flights found" in summary.lower():
        return "No flights found"
    return "Request failed"


def _debug_hint(debug: bool) -> str:
    return "" if debug else " (re-run with --debug to view full error)"


def _core_fetch(params: Dict[str, str], proxy: Optional[str]) -> object:
    client = Client(impersonate="chrome_126", verify=False, proxy=proxy)
    res = client.get("https://www.google.com/travel/flights", params=params)
    assert res.status_code == 200, f"{res.status_code} Result: {res.text_markdown}"
    return res


def _fallback_fetch(params: Dict[str, str], proxy: Optional[str]) -> object:
    client = Client(impersonate="chrome_100", verify=False, proxy=proxy)
    res = client.post(
        "https://try.playwright.tech/service/control/run",
        json={
            "code": PLAYWRIGHT_FALLBACK_CODE
            % (
                "https://www.google.com/travel/flights"
                + "?"
                + "&".join(f"{k}={v}" for k, v in params.items())
            ),
            "language": "python",
        },
    )
    assert res.status_code == 200, f"{res.status_code} Result: {res.text_markdown}"

    class DummyResponse:
        status_code = 200
        text = json.loads(res.text)["output"]
        text_markdown = text

    return DummyResponse


def _get_flights_from_filter(
    filter_payload: TFSData,
    *,
    currency: str,
    mode: str = "common",
    proxy: Optional[str] = None,
) -> Result:
    data = filter_payload.as_b64()
    params = {
        "tfs": data.decode("utf-8"),
        "hl": "en",
        "tfu": "EgQIABABIgA",
        "curr": currency,
    }

    def resolve(mode_value: str) -> object:
        if mode_value in {"common", "fallback"}:
            try:
                return _core_fetch(params, proxy)
            except AssertionError as exc:
                if mode_value == "fallback":
                    return _fallback_fetch(params, proxy)
                raise exc
        if mode_value == "local":
            from fast_flights.local_playwright import local_playwright_fetch

            return local_playwright_fetch(params)
        return _fallback_fetch(params, proxy)

    response = resolve(mode)
    try:
        return parse_response(response)
    except RuntimeError as exc:
        if mode == "fallback":
            return _get_flights_from_filter(
                filter_payload,
                currency=currency,
                mode="force-fallback",
                proxy=proxy,
            )
        raise exc
ALLOWED_SEAT_CLASSES = ("economy", "premium-economy", "business", "first")


@dataclass
class RequestSettings:
    delay: float = 1.0
    retries: int = 2
    max_leg_results: int = 10
    seat_class: str = "economy"

    def __post_init__(self) -> None:
        self.delay = float(self.delay)
        try:
            self.retries = int(self.retries)
        except (TypeError, ValueError) as exc:
            raise ValueError(f"Invalid retry count: {self.retries}") from exc
        if self.retries < 1:
            raise ValueError("Retry count must be at least 1")
        try:
            self.max_leg_results = int(self.max_leg_results)
        except (TypeError, ValueError) as exc:
            raise ValueError(f"Invalid max leg results: {self.max_leg_results}") from exc
        if self.max_leg_results < 1:
            raise ValueError("Max leg results must be at least 1")
        seat = str(self.seat_class or "").strip().lower()
        if not seat:
            raise ValueError("Seat class must be provided")
        if seat not in ALLOWED_SEAT_CLASSES:
            allowed = ", ".join(ALLOWED_SEAT_CLASSES)
            raise ValueError(f"Seat class must be one of: {allowed}")
        self.seat_class = seat


@dataclass
class ItinerarySettings:
    leg_limit: int = DEFAULT_ITINERARY_LEG_LIMIT
    max_combinations: int = DEFAULT_ITINERARY_MAX_COMBINATIONS

    def __post_init__(self) -> None:
        try:
            self.leg_limit = int(self.leg_limit)
        except (TypeError, ValueError) as exc:
            raise ValueError(f"Invalid itinerary leg limit: {self.leg_limit}") from exc
        if self.leg_limit < 0:
            raise ValueError("Itinerary leg limit must be >= 0")
        try:
            self.max_combinations = int(self.max_combinations)
        except (TypeError, ValueError) as exc:
            raise ValueError(f"Invalid itinerary max combinations: {self.max_combinations}") from exc
        if self.max_combinations < 0:
            raise ValueError("Itinerary max combinations must be >= 0")


@dataclass
class FilterSettings:
    max_stops: Optional[int] = None
    include_hidden: bool = True
    max_hidden_hops: int = 1

    def __post_init__(self) -> None:
        if self.max_stops is not None:
            try:
                self.max_stops = int(self.max_stops)
            except (TypeError, ValueError) as exc:
                raise ValueError(f"Invalid max stops: {self.max_stops}") from exc
            if self.max_stops < 0 or self.max_stops > 2:
                raise ValueError("Max stops must be between 0 and 2")
        self.include_hidden = bool(self.include_hidden)
        try:
            self.max_hidden_hops = int(self.max_hidden_hops)
        except (TypeError, ValueError) as exc:
            raise ValueError(f"Invalid hidden hop limit: {self.max_hidden_hops}") from exc
        if self.max_hidden_hops < 0 or self.max_hidden_hops > 2:
            raise ValueError("Hidden hop limit must be between 0 and 2")


@dataclass
class OutputSettings:
    directory: Path = Path("output")
    filename_pattern: str = "{timestamp}_{plan}.csv"
    filename: Optional[str] = None

    def __post_init__(self) -> None:
        self.directory = Path(self.directory)
        if self.filename is not None:
            name = str(self.filename).strip()
            if not name:
                raise ValueError("Output filename must be non-empty")
            if not name.endswith(".csv"):
                name = f"{name}.csv"
            self.filename = name
            return
        pattern = str(self.filename_pattern).strip() or "{timestamp}_{plan}"
        if "{plan}" not in pattern:
            pattern = f"{{plan}}_{pattern}"
        if "{timestamp}" not in pattern:
            pattern = f"{pattern}_{{timestamp}}"
        if not pattern.endswith(".csv"):
            pattern = f"{pattern}.csv"
        self.filename_pattern = pattern


@dataclass
class LegFilter:
    max_stops: Optional[int] = None

    def __post_init__(self) -> None:
        if self.max_stops is None:
            return
        try:
            self.max_stops = int(self.max_stops)
        except (TypeError, ValueError) as exc:
            raise ValueError(f"Invalid max stops: {self.max_stops}") from exc
        if self.max_stops < 0 or self.max_stops > 2:
            raise ValueError("Max stops must be between 0 and 2")


@dataclass
class PlanOptions:
    include_hidden: bool
    max_hidden_hops: int

    def __post_init__(self) -> None:
        self.include_hidden = bool(self.include_hidden)
        try:
            self.max_hidden_hops = int(self.max_hidden_hops)
        except (TypeError, ValueError) as exc:
            raise ValueError(f"Invalid hidden hop limit: {self.max_hidden_hops}") from exc
        if self.max_hidden_hops < 0 or self.max_hidden_hops > 2:
            raise ValueError("Hidden hop limit must be between 0 and 2")


@dataclass
class PlanOutput:
    directory: Optional[Path] = None
    filename_pattern: Optional[str] = None
    filename: Optional[str] = None

    def __post_init__(self) -> None:
        if self.directory is not None:
            self.directory = Path(self.directory)
        if self.filename is not None and not str(self.filename).endswith(".csv"):
            self.filename = f"{self.filename}.csv"


@dataclass
class LegDeparture:
    dates: List[str]
    max_stops: Optional[int] = None

    def __post_init__(self) -> None:
        if not self.dates:
            raise ValueError("Leg departure requires at least one date")
        normalized: List[str] = []
        for token in self.dates:
            value = str(token).strip()
            if not value:
                continue
            normalized.append(value)
        if not normalized:
            raise ValueError("Leg departure requires at least one non-empty date")
        self.dates = normalized
        if self.max_stops is None:
            return
        try:
            self.max_stops = int(self.max_stops)
        except (TypeError, ValueError) as exc:
            raise ValueError(f"Invalid max stops: {self.max_stops}") from exc
        if self.max_stops < 0 or self.max_stops > 2:
            raise ValueError("Max stops must be between 0 and 2")


@dataclass
class PlanConfig:
    name: str
    places: Dict[str, List[str]]
    path: List[str]
    departures: Dict[Tuple[str, str], LegDeparture]
    options: PlanOptions
    filters: Dict[Tuple[str, str], LegFilter]
    output: Optional[PlanOutput] = None


@dataclass
class PlanExecution:
    assignment: Dict[str, str]
    departures: Dict[Tuple[str, str], str]


@dataclass
class SearchConfig:
    schema_version: str
    currency_code: str
    passenger_count: int
    request: RequestSettings
    filters: FilterSettings
    output: OutputSettings
    itinerary: ItinerarySettings
    http_proxy: Optional[str]
    concurrency: int
    debug: bool
    plans: Sequence[PlanConfig]

    def __post_init__(self) -> None:
        self.currency_code = str(self.currency_code or "USD").upper()
        try:
            self.passenger_count = int(self.passenger_count)
        except (TypeError, ValueError) as exc:
            raise ValueError(f"Invalid passenger count: {self.passenger_count}") from exc
        if self.passenger_count < 1:
            raise ValueError("Passenger count must be at least 1")
        try:
            self.concurrency = int(self.concurrency)
        except (TypeError, ValueError) as exc:
            raise ValueError(f"Invalid concurrency: {self.concurrency}") from exc
        if self.concurrency < 1:
            raise ValueError("Concurrency must be at least 1")
        self.debug = bool(self.debug)
        if not self.plans:
            raise ValueError("At least one plan must be provided")


@dataclass
class LegFlight:
    airline_name: str
    departure_at: str
    stops: str
    stop_notes: str
    duration_hours: Optional[float]
    price: Optional[int]
    is_best: bool
    seat_class: str
    hidden_departure_at: str


@dataclass
class LayoverDetails:
    stop_text: str
    layover_codes: str
    hidden_departure_at: str


@dataclass
class SegmentRow:
    plan_name: str
    journey_id: str
    journey_label: str
    variant: str
    leg_sequence: int
    origin_place: str
    origin_code: str
    destination_place: str
    destination_code: str
    hidden_via_places: str
    hidden_via_codes: str
    departure_date: str
    departure_at: str
    departure_time: str
    duration_hours: Optional[float]
    airline: str
    stops: str
    stop_notes: str
    price: Optional[int]
    is_best: bool
    currency: str
    seat_class: str
    hidden_departure_at: str


@dataclass
class PlanRunResult:
    plan: PlanConfig
    output_path: Path
    rows: List[SegmentRow]


def standardize_time(raw: str, year_hint: int) -> str:
    if not raw:
        return ""
    cleaned = raw.replace("\u202f", " ").replace("\xa0", " ").replace("\u2009", " ")
    match = TIME_PATTERN.search(cleaned)
    if not match:
        return ""
    hour = int(match.group("hour"))
    minute = int(match.group("minute"))
    ampm = match.group("ampm")
    month = match.group("month")
    day = int(match.group("day"))
    plus = int(match.group("plus")) if match.group("plus") else 0
    dt = datetime.strptime(
        f"{year_hint} {month} {day} {hour}:{minute:02d} {ampm}",
        "%Y %b %d %I:%M %p",
    )
    if plus:
        dt += timedelta(days=plus)
    return dt.strftime("%Y-%m-%d %H:%M:%S")


def parse_price_to_int(price: str) -> Optional[int]:
    digits = "".join(ch for ch in price if ch.isdigit())
    return int(digits) if digits else None


def parse_duration_to_hours(raw: str) -> Optional[float]:
    if not raw:
        return None
    match = DURATION_PATTERN.search(raw)
    if not match:
        return None
    hours = int(match.group("hours")) if match.group("hours") else 0
    minutes = int(match.group("minutes")) if match.group("minutes") else 0
    total_minutes = hours * 60 + minutes
    if total_minutes == 0:
        return None
    value = total_minutes / 60
    return round(value, 2)


def _normalize_time_label(value: str) -> str:
    token = value.strip().upper().replace(".", "")
    for pattern in ("%I:%M %p", "%H:%M"):
        try:
            dt = datetime.strptime(token, pattern)
            return dt.strftime("%H:%M")
        except ValueError:
            continue
    return token


def _build_hidden_departure_labels(values: Sequence[str]) -> str:
    labels: List[str] = []
    for entry in values:
        if not entry:
            continue
        code_match = re.search(r"[A-Z]{3}", entry)
        code = code_match.group(0) if code_match else ""
        for match in TIME_LABEL_PATTERN.finditer(entry):
            label = _normalize_time_label(match.group(0))
            if code:
                label = f"{code} {label}"
            if label and label not in labels:
                labels.append(label)
    return " · ".join(labels)


def split_datetime_components(timestamp: str) -> Tuple[str, str]:
    if not timestamp:
        return "", ""
    try:
        dt = datetime.strptime(timestamp, "%Y-%m-%d %H:%M:%S")
        return dt.strftime("%Y-%m-%d"), dt.strftime("%H:%M")
    except ValueError:
        if " " in timestamp:
            date_part, time_part = timestamp.split(" ", 1)
            return date_part.strip(), time_part.strip()[:5]
        return timestamp, ""


def extract_stop_codes(raw: str) -> str:
    if not raw:
        return ""
    codes = []
    for token in re.findall(r"[A-Z]{3}", raw):
        if token not in codes:
            codes.append(token)
    return " ".join(codes)


def describe_stops(stop_value: object, stop_text: str) -> str:
    if stop_text:
        return stop_text
    if isinstance(stop_value, int):
        if stop_value == 0:
            return "Nonstop"
        suffix = "stop" if stop_value == 1 else "stops"
        return f"{stop_value} {suffix}"
    return str(stop_value)


def build_flight_data(origin_code: str, destination_code: str, departure_date: str) -> List[FlightData]:
    return [
        FlightData(
            date=departure_date,
            from_airport=origin_code,
            to_airport=destination_code,
        )
    ]


def safe_text(node) -> str:
    return node.text(strip=True) if node is not None else ""


def parse_layover_details(html: str) -> Dict[Tuple[str, str, str], LayoverDetails]:
    parser = LexborHTMLParser(html)
    details: Dict[Tuple[str, str, str], LayoverDetails] = {}

    for idx, container in enumerate(parser.css('div[jsname="IWWDBc"], div[jsname="YdtKid"]')):
        items = container.css("ul.Rk10dc li")
        if idx != 0:
            items = items[:-1]
        for item in items:
            name = safe_text(item.css_first("div.sSHqwe.tPgKwe.ogfYpf span"))
            dp_ar_node = item.css("span.mv1WYe div")
            try:
                departure_time = dp_ar_node[0].text(strip=True)
            except IndexError:
                departure_time = ""
            price_text = safe_text(item.css_first(".YMlIz.FpEdX")) or "0"
            price_clean = price_text.replace(",", "")
            stop_text = safe_text(item.css_first(".BbR8Ec .ogfYpf"))
            layover_values: List[str] = []
            for span in item.css("span.rGRiKd"):
                val = span.text(strip=True)
                if val and val not in layover_values:
                    layover_values.append(val)
            layover_codes = extract_stop_codes(" ".join(layover_values))
            hidden_departure_at = _build_hidden_departure_labels(layover_values)
            key = (name, " ".join(departure_time.split()), price_clean)
            details.setdefault(
                key,
                LayoverDetails(
                    stop_text=stop_text,
                    layover_codes=layover_codes,
                    hidden_departure_at=hidden_departure_at,
                ),
            )

    return details


def fetch_leg_html(
    *,
    origin_code: str,
    destination_code: str,
    departure_date: str,
    max_stops: Optional[int],
    passenger_count: int,
    currency_code: str,
    proxy: Optional[str],
    seat_class: str,
) -> str:
    flight_data = build_flight_data(origin_code, destination_code, departure_date)
    filter_payload = TFSData.from_interface(
        flight_data=flight_data,
        trip="one-way",
        passengers=Passengers(adults=passenger_count),
        seat=seat_class,
        max_stops=max_stops,
    )
    params = {
        "tfs": filter_payload.as_b64().decode("utf-8"),
        "hl": "en",
        "tfu": "EgQIABABIgA",
        "curr": currency_code,
    }
    try:
        response = _core_fetch(params, proxy)
        return response.text
    except AssertionError:
        try:
            response = _fallback_fetch(params, proxy)
            return response.text
        except Exception:  # noqa: BLE001
            return ""


def fetch_leg_flights(
    *,
    config: SearchConfig,
    origin_code: str,
    destination_code: str,
    departure_date: str,
    max_stops: Optional[int],
    reporter: Optional[ProgressReporter] = None,
) -> List[LegFlight]:
    last_exc: Optional[Exception] = None
    result: Optional[Result] = None
    layover_lookup: Dict[Tuple[str, str, str], LayoverDetails] = {}

    for attempt in range(1, config.request.retries + 1):
        try:
            flight_data = build_flight_data(origin_code, destination_code, departure_date)
            filter_payload = TFSData.from_interface(
                flight_data=flight_data,
                trip="one-way",
                passengers=Passengers(adults=config.passenger_count),
                seat=config.request.seat_class,
                max_stops=max_stops,
            )
            result = _get_flights_from_filter(
                filter_payload,
                currency=config.currency_code,
                mode="common",
                proxy=config.http_proxy,
            )
            html = fetch_leg_html(
                origin_code=origin_code,
                destination_code=destination_code,
                departure_date=departure_date,
                max_stops=max_stops,
                passenger_count=config.passenger_count,
                currency_code=config.currency_code,
                proxy=config.http_proxy,
                seat_class=config.request.seat_class,
            )
            if html:
                layover_lookup = parse_layover_details(html)
            break
        except Exception as exc:  # noqa: BLE001
            last_exc = exc
            wait_time = config.request.delay * attempt
            detail = _friendly_exception(exc, debug=config.debug)
            hint = _debug_hint(config.debug)
            _warn(
                (
                    f"Leg fail {origin_code}->{destination_code} "
                    f"{departure_date} try {attempt}: {detail}{hint}"
                ),
                reporter,
            )
            if attempt < config.request.retries:
                time.sleep(wait_time)

    if result is None:
        if last_exc:
            _warn(
                (
                    f"Leg skip {origin_code}->{destination_code} "
                    f"{departure_date} after {config.request.retries} tries"
                ),
                reporter,
            )
            detail = _friendly_exception(last_exc, debug=config.debug)
            hint = _debug_hint(config.debug)
            _error(f"Leg last error: {detail}{hint}", reporter)
        return []

    flights: List[LegFlight] = []
    seen: set[Tuple[str, str, str]] = set()
    base_year = int(departure_date.split("-", 1)[0])

    for flight in result.flights:
        if len(flights) >= config.request.max_leg_results:
            break
        key = (flight.name, flight.departure, flight.price)
        if key in seen:
            continue
        seen.add(key)
        departure_std = standardize_time(flight.departure, base_year)
        if not departure_std:
            continue
        layover = layover_lookup.get(key)
        stop_text = layover.stop_text if layover else ""
        layover_codes = layover.layover_codes if layover else ""
        hidden_departure = layover.hidden_departure_at if layover else ""
        notes = layover_codes or extract_stop_codes(stop_text)
        flights.append(
            LegFlight(
                airline_name=flight.name,
                departure_at=departure_std,
                stops=describe_stops(flight.stops, stop_text),
                stop_notes=notes,
                duration_hours=parse_duration_to_hours(flight.duration),
                price=parse_price_to_int(flight.price),
                is_best=flight.is_best,
                seat_class=config.request.seat_class,
                hidden_departure_at=hidden_departure,
            )
        )

    return flights


def _iter_edges(path: Sequence[str]) -> List[Tuple[str, str]]:
    return [(path[idx], path[idx + 1]) for idx in range(len(path) - 1)]


def _iter_hidden_pairs(path: Sequence[str]) -> Iterable[Tuple[int, int]]:
    for start in range(len(path) - 2):
        end = start + 2
        if path[start] == path[end]:
            continue
        yield start, end


def _slugify_plan_name(value: str) -> str:
    if not value:
        return "plan"
    slug = re.sub(r"[^A-Za-z0-9]+", "-", value.strip().lower()).strip("-")
    return slug or "plan"


def _resolve_plan_output_path(config: SearchConfig, plan: PlanConfig) -> Path:
    directory = config.output.directory
    pattern = config.output.filename_pattern
    filename_override: Optional[str] = config.output.filename
    if plan.output is not None:
        if plan.output.directory is not None:
            directory = plan.output.directory
        if plan.output.filename is not None:
            filename_override = plan.output.filename
        elif plan.output.filename_pattern is not None:
            pattern = plan.output.filename_pattern
    timestamp = datetime.now().astimezone().strftime("%Y%m%d_%H%M%S")
    plan_token = _slugify_plan_name(plan.name)
    if filename_override is not None:
        filename = filename_override.format(plan=plan_token, timestamp=timestamp)
    else:
        filename = pattern.format(plan=plan_token, timestamp=timestamp)
    return Path(directory) / filename


def _build_journey_label(
    plan: PlanConfig,
    execution: PlanExecution,
    *,
    journey_index: int,
) -> str:
    edges = _iter_edges(plan.path)
    first_edge = edges[0] if edges else None
    first_date = execution.departures.get(first_edge, "?") if first_edge else "?"
    origin_place = plan.path[0] if plan.path else "?"
    destination_place = plan.path[-1] if plan.path else "?"
    origin_code = execution.assignment.get(origin_place, "?")
    destination_code = execution.assignment.get(destination_place, "?")
    return (
        f"{plan.name}-{journey_index:04d} "
        f"{origin_code}->{destination_code} {first_date}"
    )


def _segment_row_signature(row: SegmentRow) -> Tuple[object, ...]:
    return (
        row.journey_id,
        row.plan_name,
        row.variant,
        row.leg_sequence,
        row.origin_place,
        row.origin_code,
        row.destination_place,
        row.destination_code,
        row.hidden_via_places,
        row.hidden_via_codes,
        row.departure_date,
        row.departure_at,
        row.departure_time,
        row.duration_hours,
        row.airline,
        row.stops,
        row.stop_notes,
        row.price,
        row.currency,
    )


def _prefer_segment_row(existing: SegmentRow, candidate: SegmentRow) -> SegmentRow:
    if candidate.is_best and not existing.is_best:
        return candidate
    if existing.is_best and not candidate.is_best:
        return existing
    if candidate.price is not None and existing.price is not None:
        if candidate.price < existing.price:
            return candidate
        if candidate.price > existing.price:
            return existing
    return existing


def _integrate_segment_rows(
    segment_rows: Sequence[SegmentRow],
    rows: List[SegmentRow],
    index: Dict[Tuple[object, ...], int],
) -> Tuple[List[SegmentRow], bool]:
    new_rows: List[SegmentRow] = []
    updated = False
    for row in segment_rows:
        key = _segment_row_signature(row)
        existing_idx = index.get(key)
        if existing_idx is None:
            index[key] = len(rows)
            rows.append(row)
            new_rows.append(row)
            continue
        existing_row = rows[existing_idx]
        preferred = _prefer_segment_row(existing_row, row)
        if preferred is not existing_row:
            rows[existing_idx] = preferred
            updated = True
    return new_rows, updated


def _segment_row_sort_key(row: SegmentRow) -> Tuple[float, str, str, str]:
    price = float(row.price) if isinstance(row.price, (int, float)) else float("inf")
    departure = row.departure_at or row.departure_date or ""
    return price, departure, row.origin_code, row.destination_code


def _summarize_segment_rows(rows: Sequence[SegmentRow]) -> str:
    if not rows:
        return ""
    leg_stats: Dict[str, Dict[str, object]] = {}
    for row in rows:
        leg = _build_leg_key(row)
        data = leg_stats.setdefault(
            leg,
            {
                "count": 0,
                "min_price": None,
                "currency": row.currency,
                "airline": row.airline,
                "departure": row.departure_at,
            },
        )
        data["count"] = int(data["count"]) + 1
        if row.price is not None:
            if data["min_price"] is None or row.price < data["min_price"]:
                data["min_price"] = row.price
                data["currency"] = row.currency
                data["airline"] = row.airline
                data["departure"] = row.departure_at

    summary_entries: List[str] = []
    sorted_legs = sorted(
        leg_stats.items(),
        key=lambda item: (-int(item[1]["count"]), item[0]),
    )
    for leg, stats in sorted_legs[:3]:
        entry = f"{leg} x{int(stats['count'])}"
        min_price = stats.get("min_price")
        currency = stats.get("currency")
        if isinstance(min_price, int):
            price_label = f"{currency} {min_price:,}" if currency else f"{min_price:,}"
            entry += f" ({price_label})"
        summary_entries.append(entry)
    remaining = len(sorted_legs) - len(summary_entries)
    if remaining > 0:
        summary_entries.append(f"+{remaining} more legs")
    return "; ".join(summary_entries)


def _build_leg_key(row: SegmentRow) -> str:
    return f"{row.origin_place}->{row.destination_place}"


LEG_EXPORT_FIELDS = (
    "price",
    "currency",
    "seat_class",
    "departure_at",
    "departure_time",
    "airline",
    "stops",
    "stop_notes",
    "duration_hours",
    "hidden_departure_at",
    "variant",
)


def _build_plan_itinerary_combinations(
    plan: PlanConfig,
    rows: Sequence[SegmentRow],
    itinerary: ItinerarySettings,
) -> Tuple[List[Dict[str, object]], bool, bool, int, int]:
    journeys: Dict[str, List[SegmentRow]] = defaultdict(list)
    for row in rows:
        if row.variant != "scheduled":
            continue
        journeys[row.journey_id].append(row)

    combinations: List[Dict[str, object]] = []
    expected_leg_count = max(len(plan.path) - 1, 0)
    truncated = False
    clamped = False

    leg_limit = itinerary.leg_limit
    combo_limit = itinerary.max_combinations

    for journey_id, journey_rows in journeys.items():
        legs_map: Dict[int, List[SegmentRow]] = defaultdict(list)
        for row in journey_rows:
            legs_map[row.leg_sequence].append(row)
        if len(legs_map) != expected_leg_count or not legs_map:
            continue
        ordered_sequences = sorted(legs_map.keys())
        leg_rows: List[List[SegmentRow]] = []
        for idx in ordered_sequences:
            flights = sorted(legs_map[idx], key=_segment_row_sort_key)
            if leg_limit > 0 and len(flights) > leg_limit:
                flights = flights[:leg_limit]
                clamped = True
            leg_rows.append(flights)
        for combo in product(*leg_rows):
            record: Dict[str, object] = {
                "plan_name": plan.name,
                "journey_id": combo[0].journey_id,
                "journey_label": combo[0].journey_label,
            }
            total_price = 0
            total_duration = 0.0
            price_complete = True
            currency = None
            for row in combo:
                leg_key = _build_leg_key(row)
                for field in LEG_EXPORT_FIELDS:
                    record[f"{leg_key}_{field}"] = getattr(row, field)
                if row.price is not None:
                    total_price += row.price
                else:
                    price_complete = False
                if row.duration_hours:
                    total_duration += row.duration_hours
                if row.currency:
                    currency = row.currency
            record["total_price"] = total_price if price_complete else None
            record["total_currency"] = currency
            record["total_duration_hours"] = round(total_duration, 2) if total_duration else None
            combinations.append(record)
            if combo_limit and len(combinations) >= combo_limit:
                truncated = True
                break
        if truncated:
            break

    combinations.sort(
        key=lambda entry: (
            entry.get("total_price") is None,
            entry.get("total_price") if entry.get("total_price") is not None else float("inf"),
            entry.get("journey_id"),
        )
    )
    return combinations, truncated, clamped, leg_limit, combo_limit


def _build_summary_headers(
    plan: PlanConfig,
    combinations: Sequence[Dict[str, object]],
) -> List[str]:
    base_headers = ["plan_name", "journey_id", "journey_label"]
    leg_headers: List[str] = []
    leg_keys = [f"{plan.path[idx]}->{plan.path[idx + 1]}" for idx in range(max(len(plan.path) - 1, 0))]
    for leg_key in leg_keys:
        for field in LEG_EXPORT_FIELDS:
            leg_headers.append(f"{leg_key}_{field}")
    trailing_headers = ["total_price", "total_currency", "total_duration_hours"]
    # Include any additional keys that might not be covered (fallback for hidden variants in future)
    known = set(base_headers + leg_headers + trailing_headers)
    extra_headers: List[str] = []
    for record in combinations:
        for key in record.keys():
            if key not in known and key not in extra_headers:
                extra_headers.append(key)
    return base_headers + leg_headers + trailing_headers + extra_headers


def _write_plan_summary_excel(
    config: SearchConfig,
    plan: PlanConfig,
    rows: Sequence[SegmentRow],
    csv_path: Path,
) -> Tuple[Optional[Path], int, bool, bool, int, int]:
    combinations, truncated, clamped, leg_limit, combo_limit = _build_plan_itinerary_combinations(
        plan,
        rows,
        config.itinerary,
    )
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except ImportError:  # pragma: no cover - runtime guard
        _warn("openpyxl not installed; skipping Excel summary export")
        return None, 0, truncated, clamped, leg_limit, combo_limit

    headers = _build_summary_headers(plan, combinations)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Itineraries"
    worksheet.append(headers)

    for record in combinations:
        worksheet.append([record.get(header, "") for header in headers])

    for idx, header in enumerate(headers, start=1):
        values = [header]
        values.extend(record.get(header, "") for record in combinations)
        max_length = max(len(str(value)) if value is not None else 0 for value in values)
        worksheet.column_dimensions[get_column_letter(idx)].width = min(60, max(12, max_length + 2))

    summary_name = f"{csv_path.stem}_itineraries.xlsx"
    summary_path = csv_path.with_name(summary_name)
    workbook.save(summary_path)
    return summary_path, len(combinations), truncated, clamped, leg_limit, combo_limit


def _segment_row_from_mapping(row: Dict[str, str]) -> SegmentRow:
    def get_value(key: str, default: str = "") -> str:
        if key not in row:
            raise ValueError(f"CSV is missing required column '{key}'")
        return row.get(key, default)

    def parse_int(value: str) -> Optional[int]:
        value = value.strip()
        if not value:
            return None
        try:
            return int(float(value))
        except ValueError as exc:
            raise ValueError(f"Invalid integer value '{value}'") from exc

    def parse_float(value: str) -> Optional[float]:
        value = value.strip()
        if not value:
            return None
        try:
            return float(value)
        except ValueError as exc:
            raise ValueError(f"Invalid float value '{value}'") from exc

    def parse_bool(value: str) -> bool:
        return value.strip().lower() in {"1", "true", "yes", "y"}

    def get_optional_value(key: str, default: str = "") -> str:
        value = row.get(key, default)
        if value is None:
            return ""
        return value

    plan_name = get_value("plan_name")
    journey_id = get_value("journey_id")
    journey_label = get_value("journey_label")
    variant = get_value("variant")
    leg_sequence_raw = get_value("leg_sequence")
    try:
        leg_sequence = int(leg_sequence_raw.strip())
    except ValueError as exc:
        raise ValueError(f"Invalid leg_sequence '{leg_sequence_raw}'") from exc
    origin_place = get_value("origin_place")
    origin_code = get_value("origin_code")
    destination_place = get_value("destination_place")
    destination_code = get_value("destination_code")
    hidden_via_places = get_value("hidden_via_places")
    hidden_via_codes = get_value("hidden_via_codes")
    departure_date = get_value("departure_date")
    departure_at = get_value("departure_at")
    departure_time = get_value("departure_time")
    duration_hours = parse_float(get_value("duration_hours"))
    airline = get_value("airline")
    stops = get_value("stops")
    stop_notes = get_value("stop_notes")
    price = parse_int(get_value("price"))
    is_best = parse_bool(get_value("is_best"))
    currency = get_value("currency")
    seat_class = get_optional_value("seat_class", "")
    hidden_departure_at = get_optional_value("hidden_departure_at", "")

    return SegmentRow(
        plan_name=plan_name,
        journey_id=journey_id,
        journey_label=journey_label,
        variant=variant,
        leg_sequence=leg_sequence,
        origin_place=origin_place,
        origin_code=origin_code,
        destination_place=destination_place,
        destination_code=destination_code,
        hidden_via_places=hidden_via_places,
        hidden_via_codes=hidden_via_codes,
        departure_date=departure_date,
        departure_at=departure_at,
        departure_time=departure_time,
        duration_hours=duration_hours,
        airline=airline,
        stops=stops,
        stop_notes=stop_notes,
        price=price,
        is_best=is_best,
        currency=currency,
        seat_class=seat_class,
        hidden_departure_at=hidden_departure_at,
    )


def _collect_segments_from_csv(csv_path: Path) -> Dict[str, List[SegmentRow]]:
    if not csv_path.exists():
        raise FileNotFoundError(f"CSV file not found: {csv_path}")
    with csv_path.open("r", encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle)
        if reader.fieldnames is None:
            raise ValueError("CSV file has no header row")
        rows_by_plan: Dict[str, List[SegmentRow]] = defaultdict(list)
        for index, raw in enumerate(reader, start=1):
            try:
                segment = _segment_row_from_mapping(raw)
            except ValueError as exc:
                raise ValueError(f"Row {index} invalid: {exc}") from exc
            rows_by_plan[segment.plan_name].append(segment)
    return rows_by_plan


def _infer_plan_config_from_segments(plan_name: str, rows: Sequence[SegmentRow]) -> PlanConfig:
    if not rows:
        raise ValueError("No rows available to infer plan configuration")
    scheduled_map: Dict[int, List[SegmentRow]] = defaultdict(list)
    for row in rows:
        if row.variant == "scheduled":
            scheduled_map[row.leg_sequence].append(row)
    if not scheduled_map:
        for row in rows:
            scheduled_map[row.leg_sequence].append(row)
    ordered_sequences = sorted(scheduled_map.keys())
    if not ordered_sequences:
        raise ValueError("Unable to infer leg order from CSV")

    raw_path: List[str] = []
    for seq in ordered_sequences:
        sample = scheduled_map[seq][0]
        if not raw_path:
            raw_path.append(sample.origin_place)
        if raw_path[-1] != sample.origin_place:
            raw_path.append(sample.origin_place)
        raw_path.append(sample.destination_place)
    dedup_path: List[str] = []
    for place in raw_path:
        if not dedup_path or dedup_path[-1] != place:
            dedup_path.append(place)
    if len(dedup_path) < 2:
        raise ValueError("Inferred path must contain at least two points")

    place_codes: Dict[str, List[str]] = {}
    codes_map: Dict[str, set[str]] = defaultdict(set)
    for row in rows:
        codes_map[row.origin_place].add(row.origin_code)
        codes_map[row.destination_place].add(row.destination_code)
    for place, codes in codes_map.items():
        place_codes[place] = sorted(code for code in codes if code)
        if not place_codes[place]:
            place_codes[place] = [""]

    departures: Dict[Tuple[str, str], LegDeparture] = {}
    for seq in ordered_sequences:
        samples = scheduled_map[seq]
        sample = samples[0]
        key = (sample.origin_place, sample.destination_place)
        dates = sorted({row.departure_date for row in samples if row.departure_date})
        if not dates:
            dates = ["1970-01-01"]
        departures[key] = LegDeparture(dates=dates)

    options = PlanOptions(include_hidden=any(row.variant == "hidden" for row in rows), max_hidden_hops=1)
    plan_filters: Dict[Tuple[str, str], LegFilter] = {}
    return PlanConfig(
        name=plan_name,
        places=place_codes,
        path=dedup_path,
        departures=departures,
        options=options,
        filters=plan_filters,
        output=None,
    )


def _build_config_from_rows(rows_by_plan: Dict[str, List[SegmentRow]]) -> SearchConfig:
    plans: List[PlanConfig] = []
    for plan_name, segments in rows_by_plan.items():
        plan = _infer_plan_config_from_segments(plan_name, segments)
        plans.append(plan)
    if not plans:
        raise ValueError("CSV does not contain any plan data")

    currency = next(
        (row.currency for segments in rows_by_plan.values() for row in segments if row.currency),
        "USD",
    )
    return SearchConfig(
        schema_version="v2",
        currency_code=currency,
        passenger_count=1,
        request=RequestSettings(),
        filters=FilterSettings(),
        output=OutputSettings(),
        itinerary=ItinerarySettings(),
        http_proxy=None,
        concurrency=1,
        debug=False,
        plans=plans,
    )


def build_config_from_csv(csv_path: Path) -> SearchConfig:
    rows_by_plan = _collect_segments_from_csv(csv_path)
    return _build_config_from_rows(rows_by_plan)


def convert_csv_to_workbooks(config: Optional[SearchConfig], csv_path: Path) -> None:
    rows_by_plan = _collect_segments_from_csv(csv_path)
    if config is None:
        config = _build_config_from_rows(rows_by_plan)
    plan_lookup: Dict[str, PlanConfig] = {plan.name: plan for plan in config.plans}

    processed = False
    for plan_name, plan_rows in rows_by_plan.items():
        plan = plan_lookup.get(plan_name)
        if plan is None:
            try:
                plan = _infer_plan_config_from_segments(plan_name, plan_rows)
            except ValueError as exc:
                _warn(f"Cannot infer plan '{plan_name}': {exc}")
                continue
            config.plans = list(config.plans) + [plan]
            plan_lookup[plan_name] = plan
        fake_csv_path = csv_path.with_name(f"{csv_path.stem}_{plan_name}.csv")
        (
            workbook_path,
            itinerary_count,
            truncated,
            clamped,
            leg_limit,
            combo_limit,
        ) = _write_plan_summary_excel(
            config,
            plan,
            plan_rows,
            fake_csv_path,
        )
        processed = True
        if workbook_path is not None:
            console.log(
                f"[green]{plan_name}: itineraries workbook saved with {itinerary_count} rows. Path: {workbook_path}.[/]"
            )
            limit_label = "∞" if leg_limit == 0 else str(leg_limit)
            recommended_limit_label = (
                "∞" if RECOMMENDED_ITINERARY_LEG_LIMIT == 0 else str(RECOMMENDED_ITINERARY_LEG_LIMIT)
            )
            combo_label = "∞" if combo_limit == 0 else str(combo_limit)
            console.log(
                (
                    f"[cyan]{plan_name}: itinerary leg limit {limit_label} (recommended <= {recommended_limit_label}), "
                    f"max combinations {combo_label}.[/]"
                )
            )
            if clamped:
                _warn(
                    (
                        "Itinerary workbook limited to the first "
                        f"{leg_limit} flights per leg."
                        " Adjust via defaults.itinerary.leg_limit or --itinerary-leg-limit."
                    )
                )
            if truncated:
                _warn(
                    (
                        "Itinerary workbook truncated after "
                        f"{combo_limit} combinations."
                        " Adjust via defaults.itinerary.max_combinations or --itinerary-max-combos."
                    )
                )

    if not processed:
        _warn(f"No matching plans found in {csv_path}")


def _build_progress_label(
    plan: PlanConfig,
    execution: PlanExecution,
    *,
    journey_index: int,
    total: int,
    config: SearchConfig,
) -> str:
    origin_code = execution.assignment.get(plan.path[0], "?") if plan.path else "?"
    destination_code = execution.assignment.get(plan.path[-1], "?") if plan.path else "?"
    edges = _iter_edges(plan.path)
    first_edge = edges[0] if edges else None
    first_date = execution.departures.get(first_edge, "?") if first_edge else "?"
    return (
        f"{plan.name} #{journey_index}/{total} "
        f"{origin_code}->{destination_code} {first_date} · "
        f"{config.passenger_count} pax · {config.currency_code}"
    )


def _resolve_leg_max_stops(
    config: SearchConfig,
    plan: PlanConfig,
    origin_place: str,
    destination_place: str,
    *,
    hidden: bool,
) -> Optional[int]:
    if hidden:
        limit = plan.options.max_hidden_hops
        global_limit = config.filters.max_hidden_hops
        if global_limit is not None:
            limit = min(limit, global_limit)
        return limit
    leg_departure = plan.departures.get((origin_place, destination_place))
    if leg_departure and leg_departure.max_stops is not None:
        return leg_departure.max_stops
    override = plan.filters.get((origin_place, destination_place))
    if override and override.max_stops is not None:
        return override.max_stops
    return config.filters.max_stops


def _flight_contains_codes(flight: LegFlight, codes: Sequence[str]) -> bool:
    if not codes:
        return True
    haystack = f"{flight.stop_notes} {flight.stops}"
    found = {token.upper() for token in extract_stop_codes(haystack).split() if token}
    return all(code.upper() in found for code in codes)


def _build_segment_rows(
    *,
    config: SearchConfig,
    plan: PlanConfig,
    journey_id: str,
    journey_label: str,
    variant: str,
    leg_sequence: int,
    origin_place: str,
    origin_code: str,
    destination_place: str,
    destination_code: str,
    departure_date_hint: str,
    flights: Sequence[LegFlight],
    via_places: Sequence[str],
    via_codes: Sequence[str],
) -> List[SegmentRow]:
    rows: List[SegmentRow] = []
    via_place_label = " → ".join(via_places)
    via_code_label = " → ".join(via_codes)
    for flight in flights:
        departure_date, departure_time = split_datetime_components(flight.departure_at)
        rows.append(
            SegmentRow(
                plan_name=plan.name,
                journey_id=journey_id,
                journey_label=journey_label,
                variant=variant,
                leg_sequence=leg_sequence,
                origin_place=origin_place,
                origin_code=origin_code,
                destination_place=destination_place,
                destination_code=destination_code,
                hidden_via_places=via_place_label,
                hidden_via_codes=via_code_label,
                departure_date=departure_date or departure_date_hint,
                departure_at=flight.departure_at,
                departure_time=departure_time,
                duration_hours=flight.duration_hours,
                airline=flight.airline_name,
                stops=flight.stops,
                stop_notes=flight.stop_notes,
                price=flight.price,
                is_best=flight.is_best,
                currency=config.currency_code,
                seat_class=flight.seat_class,
                hidden_departure_at=flight.hidden_departure_at,
            )
        )
    return rows


def collect_segments_for_execution(
    *,
    config: SearchConfig,
    plan: PlanConfig,
    execution: PlanExecution,
    journey_index: int,
    journey_label: str,
    reporter: Optional[ProgressReporter],
) -> List[SegmentRow]:
    rows: List[SegmentRow] = []
    edges = _iter_edges(plan.path)
    journey_id = f"{_slugify_plan_name(plan.name)}-{journey_index:04d}"

    for seq_index, (origin_place, destination_place) in enumerate(edges):
        origin_code = execution.assignment.get(origin_place)
        destination_code = execution.assignment.get(destination_place)
        if not origin_code or not destination_code:
            _warn(f"Missing airport code for {origin_place}->{destination_place}", reporter)
            continue
        departure_date = execution.departures.get((origin_place, destination_place))
        if not departure_date:
            _warn(
                f"Missing departure date for {origin_place}->{destination_place}",
                reporter,
            )
            continue
        max_stops = _resolve_leg_max_stops(
            config,
            plan,
            origin_place,
            destination_place,
            hidden=False,
        )
        flights = fetch_leg_flights(
            config=config,
            origin_code=origin_code,
            destination_code=destination_code,
            departure_date=departure_date,
            max_stops=max_stops,
            reporter=reporter,
        )
        rows.extend(
            _build_segment_rows(
                config=config,
                plan=plan,
                journey_id=journey_id,
                journey_label=journey_label,
                variant="scheduled",
                leg_sequence=seq_index,
                origin_place=origin_place,
                origin_code=origin_code,
                destination_place=destination_place,
                destination_code=destination_code,
                departure_date_hint=departure_date,
                flights=flights,
                via_places=(),
                via_codes=(),
            )
        )
        if config.request.delay:
            time.sleep(config.request.delay)

    if plan.options.include_hidden:
        hidden_offset = len(edges)
        for hidden_index, (start_idx, end_idx) in enumerate(_iter_hidden_pairs(plan.path)):
            origin_place = plan.path[start_idx]
            destination_place = plan.path[end_idx]
            via_places = plan.path[start_idx + 1 : end_idx]
            if not via_places:
                continue
            origin_code = execution.assignment.get(origin_place)
            destination_code = execution.assignment.get(destination_place)
            via_codes = [execution.assignment.get(place, "") for place in via_places]
            if not origin_code or not destination_code or any(not code for code in via_codes):
                _warn(
                    f"Missing code for hidden path {origin_place}->{destination_place}",
                    reporter,
                )
                continue
            first_edge = (plan.path[start_idx], plan.path[start_idx + 1])
            departure_date = execution.departures.get(first_edge)
            if not departure_date:
                _warn(
                    f"Missing departure date for hidden path {origin_place}->{destination_place}",
                    reporter,
                )
                continue
            hidden_limit = _resolve_leg_max_stops(
                config,
                plan,
                origin_place,
                destination_place,
                hidden=True,
            )
            if hidden_limit <= 0:
                continue
            flights = fetch_leg_flights(
                config=config,
                origin_code=origin_code,
                destination_code=destination_code,
                departure_date=departure_date,
                max_stops=hidden_limit,
                reporter=reporter,
            )
            filtered = [
                flight
                for flight in flights
                if _flight_contains_codes(flight, via_codes)
            ]
            rows.extend(
                _build_segment_rows(
                    config=config,
                    plan=plan,
                    journey_id=journey_id,
                    journey_label=journey_label,
                    variant="hidden",
                    leg_sequence=hidden_offset + hidden_index,
                    origin_place=origin_place,
                    origin_code=origin_code,
                    destination_place=destination_place,
                    destination_code=destination_code,
                    departure_date_hint=departure_date,
                    flights=filtered,
                    via_places=via_places,
                    via_codes=via_codes,
                )
            )
            if config.request.delay:
                time.sleep(config.request.delay)

    return rows


def run_plan(config: SearchConfig, plan: PlanConfig) -> PlanRunResult:
    if len(plan.path) < 2:
        raise ValueError(f"Plan '{plan.name}' requires at least two path entries")
    for place in plan.path:
        if place not in plan.places:
            raise ValueError(f"Plan '{plan.name}' is missing place '{place}' definition")
    for key, codes in plan.places.items():
        if not codes:
            raise ValueError(f"Plan '{plan.name}' place '{key}' has no airport codes")

    output_path = _resolve_plan_output_path(config, plan)
    edges = _iter_edges(plan.path)
    departure_entries = [plan.departures.get(edge) for edge in edges]
    for edge, entry in zip(edges, departure_entries):
        if entry is None or not entry.dates:
            raise ValueError(
                f"Plan '{plan.name}' missing departure dates for {edge[0]}->{edge[1]}"
            )

    place_keys = list(plan.places.keys())
    code_options = [plan.places[key] for key in place_keys]
    executions: List[PlanExecution] = []
    for codes in product(*code_options):
        assignment = dict(zip(place_keys, codes))
        for date_combo in product(*(entry.dates for entry in departure_entries)):
            departures = {edge: date for edge, date in zip(edges, date_combo)}
            executions.append(PlanExecution(assignment=assignment, departures=departures))

    total_steps = len(executions)
    rows: List[SegmentRow] = []
    with ProgressReporter(total_steps, config=config, plan=plan, output_path=output_path) as reporter:
        executor_ref: Optional[ThreadPoolExecutor] = None
        futures_map: Optional[Dict[Future, Tuple[int, PlanExecution, str, str]]] = None
        tasks: List[Tuple[int, PlanExecution, str, str]] = []
        signature_index: Dict[Tuple[object, ...], int] = {}
        for idx, execution in enumerate(executions, start=1):
            journey_label = _build_journey_label(
                plan,
                execution,
                journey_index=idx,
            )
            progress_label = _build_progress_label(
                plan,
                execution,
                journey_index=idx,
                total=total_steps,
                config=config,
            )
            tasks.append((idx, execution, journey_label, progress_label))
        try:
            if config.concurrency <= 1:
                for idx, execution, journey_label, progress_label in tasks:
                    segment_rows = collect_segments_for_execution(
                        config=config,
                        plan=plan,
                        execution=execution,
                        journey_index=idx,
                        journey_label=journey_label,
                        reporter=reporter,
                    )
                    if segment_rows:
                        new_rows, updated = _integrate_segment_rows(
                            segment_rows,
                            rows,
                            signature_index,
                        )
                        summary_source = new_rows or segment_rows
                        summary = _summarize_segment_rows(summary_source)
                        if new_rows:
                            reporter.record_success(
                                progress_label,
                                len(new_rows),
                                summary or journey_label,
                            )
                        elif updated:
                            reporter.record_skip(
                                progress_label,
                                "Duplicate flights",
                                summary or journey_label,
                            )
                        else:
                            reporter.record_skip(
                                progress_label,
                                "Duplicate flights",
                                summary or journey_label,
                            )
                    else:
                        reporter.record_skip(progress_label, "No flights", journey_label)
                    if config.request.delay:
                        time.sleep(config.request.delay)
            else:
                def process_task(
                    task: Tuple[int, PlanExecution, str, str]
                ) -> Tuple[str, List[SegmentRow], str]:
                    idx, execution, journey_label, progress_label = task
                    segment_rows = collect_segments_for_execution(
                        config=config,
                        plan=plan,
                        execution=execution,
                        journey_index=idx,
                        journey_label=journey_label,
                        reporter=None,
                    )
                    return progress_label, segment_rows, journey_label

                with ThreadPoolExecutor(max_workers=config.concurrency) as executor:
                    executor_ref = executor
                    futures_map = {executor.submit(process_task, task): task for task in tasks}
                    for future in as_completed(futures_map):
                        progress_label, segment_rows, journey_label = future.result()
                        if segment_rows:
                            new_rows, updated = _integrate_segment_rows(
                                segment_rows,
                                rows,
                                signature_index,
                            )
                            summary_source = new_rows or segment_rows
                            summary = _summarize_segment_rows(summary_source)
                            if new_rows:
                                reporter.record_success(
                                    progress_label,
                                    len(new_rows),
                                    summary or journey_label,
                                )
                            elif updated:
                                reporter.record_skip(
                                    progress_label,
                                    "Duplicate flights",
                                    summary or journey_label,
                                )
                            else:
                                reporter.record_skip(
                                    progress_label,
                                    "Duplicate flights",
                                    summary or journey_label,
                                )
                        else:
                            reporter.record_skip(progress_label, "No flights", journey_label)
        except KeyboardInterrupt:  # noqa: PERF203
            if futures_map:
                for future in futures_map:
                    try:
                        future.cancel()
                    except Exception:  # pragma: no cover
                        pass
            if executor_ref is not None:
                executor_ref.shutdown(wait=False, cancel_futures=True)
            raise RunInterrupted(
                plan=plan,
                output_path=output_path,
                rows=list(rows),
                processed=reporter.processed,
                skipped=reporter.skipped,
                collected=reporter.rows_collected,
                total=total_steps,
            ) from None

    return PlanRunResult(plan=plan, output_path=output_path, rows=rows)


def write_csv(
    rows: Sequence[SegmentRow],
    output_path: Path,
    *,
    include_header_only: bool = False,
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    if rows:
        fieldnames = list(asdict(rows[0]).keys())
        with output_path.open("w", newline="", encoding="utf-8") as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            writer.writeheader()
            for row in rows:
                writer.writerow(asdict(row))
        return
    header_fields = [field.name for field in fields(SegmentRow)]
    if not include_header_only:
        _warn("No flight data available to write")
        return
    with output_path.open("w", newline="", encoding="utf-8") as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=header_fields)
        writer.writeheader()


def run_search(config: SearchConfig) -> List[PlanRunResult]:
    return [run_plan(config, plan) for plan in config.plans]


def execute_search(config: SearchConfig) -> List[SegmentRow]:
    try:
        plan_results = run_search(config)
    except RunInterrupted as interrupted:
        draft_path = interrupted.output_path.with_name(
            f"draft-{interrupted.output_path.name}"
        )
        write_csv(interrupted.rows, draft_path, include_header_only=True)
        console.log(
            "[yellow]Search interrupted by Ctrl+C. Partial results saved successfully.[/]"
        )
        draft_display = escape(str(draft_path))
        console.log(
            (
                f"[yellow]Plan {interrupted.plan.name}. Processed {interrupted.processed}/"
                f"{interrupted.total} journeys; remaining {interrupted.remaining}."
                f" Rows collected {interrupted.collected}, skipped {interrupted.skipped}."
                f" Draft path: {draft_display}.[/]"
            )
        )
        return interrupted.rows

    all_rows: List[SegmentRow] = []
    if not plan_results:
        _warn("No flights captured")
        return all_rows

    for result in plan_results:
        write_csv(result.rows, result.output_path, include_header_only=True)
        (
            workbook_path,
            itinerary_count,
            truncated,
            clamped,
            leg_limit,
            combo_limit,
        ) = _write_plan_summary_excel(
            config,
            result.plan,
            result.rows,
            result.output_path,
        )
        all_rows.extend(result.rows)
        console.log(
            f"[green]{result.plan.name}: saved {len(result.rows)} rows. Output path: {result.output_path}.[/]"
        )
        if workbook_path is not None:
            console.log(
                f"[green]{result.plan.name}: itineraries workbook saved with {itinerary_count} rows. Path: {workbook_path}.[/]"
            )
            limit_label = "∞" if leg_limit == 0 else str(leg_limit)
            recommended_limit_label = (
                "∞" if RECOMMENDED_ITINERARY_LEG_LIMIT == 0 else str(RECOMMENDED_ITINERARY_LEG_LIMIT)
            )
            combo_label = "∞" if combo_limit == 0 else str(combo_limit)
            console.log(
                (
                    f"[cyan]{result.plan.name}: itinerary leg limit {limit_label}"
                    f" (recommended <= {recommended_limit_label}), max combinations {combo_label}.[/]"
                )
            )
            if clamped:
                _warn(
                    (
                        "Itinerary workbook limited to the first "
                        f"{leg_limit} flights per leg."
                        " Adjust via defaults.itinerary.leg_limit or --itinerary-leg-limit."
                    )
                )
            if truncated:
                _warn(
                    (
                        "Itinerary workbook truncated after "
                        f"{combo_limit} combinations."
                        " Adjust via defaults.itinerary.max_combinations or --itinerary-max-combos."
                    )
                )

    if not all_rows:
        _warn("No flights captured")

    return all_rows






__all__ = [
    "DEFAULT_ITINERARY_LEG_LIMIT",
    "DEFAULT_ITINERARY_MAX_COMBINATIONS",
    "RECOMMENDED_ITINERARY_LEG_LIMIT",
    "SearchConfig",
    "LegFlight",
    "SegmentRow",
    "PlanRunResult",
    "ItinerarySettings",
    "run_search",
    "execute_search",
    "convert_csv_to_workbooks",
    "build_config_from_csv",
    "RunInterrupted",
]
