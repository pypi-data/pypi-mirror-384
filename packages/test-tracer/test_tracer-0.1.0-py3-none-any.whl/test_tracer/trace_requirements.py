"""
Software Test <-> Requirements Tracing Script

Run this script to extract test cases from unit and integration tests and export to Excel.
Requirement IDs are extracted from test decorators, and test case descriptions are extracted from docstrings.
Exceptions are raised if the docstring is malformed.

Expected docstring format:
    <One line description of the test case>
    Test Case Description:
    Precondition:
    <Test case precondition>
    Steps:
    <Test case steps>
    Expected Result:
    <Test case expected result>
    Tests Requirements:
    <List of requirement IDs>

Currently, this script only extracts test cases from pytest unit tests. Integration tests have been temporarily formatted
into pytest unit tests with the same docstring format.
"""

import ast
import os
from pathlib import Path
from collections import defaultdict
from openpyxl import Workbook

from test_tracer.track_soup import add_tracked_soup
from test_tracer.test_case import TestCase, TestDefinition



def extract_requirement_ids(requirements_file: Path) -> list[str]:
    """Get IDs from the requirements file.

    Strips whitespace and empty lines.
    """
    with open(requirements_file, "r") as f:
        lines = f.readlines()

    return [l.strip() for l in lines if l.strip()]


def get_test_case_definition_from_docstring(test_docstring: str) -> TestDefinition:
    """Split docstrings into TestDefinition objects."""
    docstring_lines = test_docstring.split("\n")

    # TODO make this more robust for different docstring formatting.
    idx_pre_con = docstring_lines.index("Precondition:")
    idx_steps = docstring_lines.index("Steps:")
    idx_result = docstring_lines.index("Expected result:")
    idx_req_ids = docstring_lines.index("Requirements:")

    return TestDefinition(
        description=docstring_lines[0].strip(),
        precondition="\n".join(docstring_lines[idx_pre_con + 1 : idx_steps]).strip(),
        steps="\n".join(docstring_lines[idx_steps + 1 : idx_result]).strip(),
        expected_result="\n".join(
            docstring_lines[idx_result + 1 : idx_req_ids]
        ).strip(),
        req_ids=[
            _l.replace("*", "").strip() for _l in docstring_lines[idx_req_ids + 1 :]
        ],
    )


def extract_tests_from_python(file_path: Path) -> list[TestCase]:
    """Walk Python test files and extract test cases."""
    with open(file_path, "r") as f:
        tree = ast.parse(f.read(), filename=file_path)

    # Analyse non-recursively to get the test cases
    tests = []
    for node in ast.walk(tree):
        for child in ast.iter_child_nodes(node):
            child.parent = node
        if isinstance(node, ast.FunctionDef) and node.name.split("_")[0] == "test":


            test_path = (
                f"{file_path}.{node.parent.name}.{node.name}"
                if isinstance(node.parent, ast.ClassDef)
                else f"{file_path}::{node.name}"
            )

            test_docstring = ast.get_docstring(node)
            if test_docstring is None:
                raise ValueError(
                    f"Test {node.name} in file {file_path} is missing docstring!"
                )

            try:
                definition = get_test_case_definition_from_docstring(test_docstring)
            except IndexError:
                raise ValueError(
                    f"Test {node.name} in file {file_path} docstring is malformed!"
                    f"Please update the test docstring to match the format."
                )

            tests.append(
                TestCase(
                    id=None,
                    name=node.name,
                    file_path=str(file_path),
                    test_path=test_path,
                    definition=definition,
                    soup_components=None
                )
            )

    return tests


def extract_tests_from_c_sharp(file_path: Path) -> list[TestCase]:
    tests = []
    with open(file_path, "r") as f:
        file = f.read()
        sections = file.split("/// <summary>")[1:]
        for section in sections:
            lines = section.split("///")
            lines = [line.strip() for line in lines if line.strip()]
            _test_definition = TestDefinition(
                description=lines[0].strip(),
                precondition=lines[1].split(":")[1],
                steps=lines[2].split(":")[1],
                expected_result=lines[3].split(":")[1],
                req_ids=lines[4].split(": ")[1].split(", "),
            )
            _function_name = lines[5].split("\n")[2].split()[2]
            _t = TestCase(
                id=None,
                name=_function_name,
                file_path=str(file_path),
                test_path=f"{file_path}::{_function_name}",
                definition=_test_definition,
                soup_components=None
            )
            tests.append(_t)

    return tests


def scan_tests_for_requirements(test_dir: Path) -> list[TestCase]:
    test_data = []
    for dirpath, _, filenames in os.walk(test_dir):
        for filename in filenames:

            if not filename.lower().startswith("test"):
                continue

            file_path = Path(os.path.join(dirpath, filename))

            if filename.endswith(".py"):
                print(f"Python file : {filename}")
                test_data += extract_tests_from_python(file_path)

            elif filename.endswith(".cs"):
                print(f"CS file : {filename}")
                test_data += extract_tests_from_c_sharp(file_path)

    return test_data


def export_to_excel(
    tests: list[TestCase], requirement_ids: list[str], bidirectional: bool
) -> Workbook:
    wb = Workbook()

    # Test Case Definition tab
    ws = wb.active
    ws.title = "Test Case Definition"
    ws.append(
        [
            "Test Case ID",
            "Test Name",
            "Test Path",
            "File Path",
            "Description",
            "Precondition",
            "Steps",
            "Expected Result",
            "Requirement IDs",
            "SOUP Components",
        ]
    )

    for i, _t in enumerate(tests):
        _t.id = f"SWT-{i+1:03}"  # TODO: Persistent test case IDs
        ws.append(
            [
                _t.id,
                _t.name,
                str(_t.test_path),
                str(_t.file_path),
                _t.definition.description,
                _t.definition.precondition,
                _t.definition.steps,
                _t.definition.expected_result,
                ", ".join(_t.definition.req_ids),
                _t.soup_components,
            ]
        )

    # Requirement to Test Case Mapping tab
    ws = wb.create_sheet(title="Requirement to Test Mapping")
    ws.append(["Requirement ID", "Test Case ID"])

    req_id_to_test_cases = defaultdict(list)
    for test in tests:
        for req_id in test.definition.req_ids:
            req_id_to_test_cases[req_id].append(test.id)

    for req_id in requirement_ids:
        ws.append([req_id] + req_id_to_test_cases[req_id])

    # Bidirectional mapping tab
    if bidirectional:
        ws = wb.create_sheet(title="Bi-Directional Mapping")
        ws.append(["Requirement ID/Test Case ID"] + sorted(_t.id for _t in tests))
        # Leetcode would be so disappointed
        for req_id in requirement_ids:
            mapping = []
            for test in tests:
                mapping.append("x" if test.id in req_id_to_test_cases[req_id] else "")
            ws.append([req_id] + mapping)

    return wb


def trace_requirements(
    requirements_file: Path, test_dir: Path, output_excel: str, bidirectional: bool, soup: Path|None
) -> None:
    """Extract test cases from unit and integration tests and export to Excel."""
    requirement_ids = extract_requirement_ids(requirements_file)
    tests = scan_tests_for_requirements(test_dir=test_dir)

    if soup:
        tests = add_tracked_soup(tests, test_dir, soup_file_path=soup)

    wb = export_to_excel(tests, requirement_ids, bidirectional=bidirectional)
    wb.save(output_excel)

    print(f"✅ Traceability matrix saved to {output_excel}")

