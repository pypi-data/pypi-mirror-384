"""
Screen Export & Enrichment Module

Provides enhanced export capabilities for screening results:
- Sector/industry enrichment
- Excel export with formatting
- CSV export
- Rich data presentation
"""

import pandas as pd
from pathlib import Path
from typing import List, Optional
from datetime import datetime

from ..utils.logger import setup_logger
from ..data.database import DatabaseManager
from ..data.lookup_tables import LookupTableManager

logger = setup_logger(__name__)


class ScreenExporter:
    """
    Enhanced export functionality for screening results

    Features:
    - Enrich results with sector/industry data
    - Export to Excel with formatting
    - Export to CSV
    - Add calculated fields (% from SMA, etc.)
    """

    def __init__(self, db: DatabaseManager, lookup_tables: LookupTableManager):
        """
        Initialize screen exporter

        Args:
            db: Database manager
            lookup_tables: Lookup table manager for company info
        """
        self.db = db
        self.lookup = lookup_tables

    def enrich_results(self, results: pd.DataFrame) -> pd.DataFrame:
        """
        Enrich screening results with additional data

        Args:
            results: DataFrame with screening results

        Returns:
            Enriched DataFrame with sector, industry, and calculated fields
        """
        if results.empty:
            return results

        enriched = results.copy()

        # Add sector/industry data
        enriched = self._add_company_info(enriched)

        # Add calculated fields
        enriched = self._add_calculated_fields(enriched)

        return enriched

    def _add_company_info(self, df: pd.DataFrame) -> pd.DataFrame:
        """Add sector, industry, and company name"""
        if 'ticker' not in df.columns:
            return df

        # Get company info for all tickers
        tickers = df['ticker'].tolist()

        sector_map = {}
        industry_map = {}
        name_map = {}

        for ticker in tickers:
            info = self.lookup.get_company_info(ticker)
            if info:
                sector_map[ticker] = info.get('sector', 'Unknown')
                industry_map[ticker] = info.get('industry', 'Unknown')
                name_map[ticker] = info.get('company_name', ticker)

        # Add columns
        df['company_name'] = df['ticker'].map(name_map).fillna(df['ticker'])
        df['sector'] = df['ticker'].map(sector_map).fillna('Unknown')
        df['industry'] = df['ticker'].map(industry_map).fillna('Unknown')

        # Reorder columns (company_name, sector, industry after ticker)
        cols = df.columns.tolist()
        if 'ticker' in cols:
            ticker_idx = cols.index('ticker')
            new_cols = cols[:ticker_idx+1] + ['company_name', 'sector', 'industry'] + \
                      [c for c in cols[ticker_idx+1:] if c not in ['company_name', 'sector', 'industry']]
            df = df[new_cols]

        logger.info(f"✓ Enriched {len(df)} results with sector/industry data")

        return df

    def _add_calculated_fields(self, df: pd.DataFrame) -> pd.DataFrame:
        """Add useful calculated fields"""

        # % distance from SMA20
        if 'price' in df.columns and 'sma_20' in df.columns:
            df['pct_from_sma20'] = ((df['price'] - df['sma_20']) / df['sma_20'] * 100).round(2)

        # % distance from SMA50
        if 'price' in df.columns and 'sma_50' in df.columns:
            df['pct_from_sma50'] = ((df['price'] - df['sma_50']) / df['sma_50'] * 100).round(2)

        # Bollinger Band position (%)
        if all(c in df.columns for c in ['price', 'bb_upper', 'bb_lower']):
            bb_range = df['bb_upper'] - df['bb_lower']
            df['bb_position_pct'] = ((df['price'] - df['bb_lower']) / bb_range * 100).round(2)

        # P/E vs sector average (requires sector data)
        # TODO: Calculate sector averages

        return df

    def export_to_excel(
        self,
        results: pd.DataFrame,
        output_path: str,
        sheet_name: str = "Screen Results",
        enrich: bool = True
    ):
        """
        Export results to Excel with formatting

        Args:
            results: DataFrame with screening results
            output_path: Path to save Excel file
            sheet_name: Name for the worksheet
            enrich: Whether to enrich with sector/industry data
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils.dataframe import dataframe_to_rows

        except ImportError:
            logger.error("openpyxl not installed. Install with: pip install openpyxl")
            raise

        if results.empty:
            logger.warning("No results to export")
            return

        # Enrich if requested
        if enrich:
            results = self.enrich_results(results)

        # Create output directory
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name[:31]  # Excel limit is 31 chars

        # Write data
        for r_idx, row in enumerate(dataframe_to_rows(results, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)

                # Header formatting
                if r_idx == 1:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center")

        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50
            ws.column_dimensions[column_letter].width = adjusted_width

        # Freeze header row
        ws.freeze_panes = "A2"

        # Save
        wb.save(output_path)
        logger.info(f"✓ Exported {len(results)} results to Excel: {output_path}")

    def export_to_csv(
        self,
        results: pd.DataFrame,
        output_path: str,
        enrich: bool = True
    ):
        """
        Export results to CSV

        Args:
            results: DataFrame with screening results
            output_path: Path to save CSV file
            enrich: Whether to enrich with sector/industry data
        """
        if results.empty:
            logger.warning("No results to export")
            return

        # Enrich if requested
        if enrich:
            results = self.enrich_results(results)

        # Create output directory
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        # Export
        results.to_csv(output_path, index=False)
        logger.info(f"✓ Exported {len(results)} results to CSV: {output_path}")

    def create_comparison_report(
        self,
        results: pd.DataFrame,
        output_path: str
    ):
        """
        Create multi-sheet Excel report with analysis

        Args:
            results: DataFrame with screening results
            output_path: Path to save Excel file
        """
        try:
            import openpyxl
            from openpyxl.chart import BarChart, Reference
            from openpyxl.utils.dataframe import dataframe_to_rows
        except ImportError:
            logger.error("openpyxl not installed")
            raise

        if results.empty:
            logger.warning("No results to create report")
            return

        # Enrich data
        enriched = self.enrich_results(results)

        # Create workbook
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Sheet 1: All Results
        ws_all = wb.create_sheet("All Results")
        for r_idx, row in enumerate(dataframe_to_rows(enriched, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws_all.cell(row=r_idx, column=c_idx, value=value)

        # Sheet 2: By Sector
        if 'sector' in enriched.columns:
            # Build aggregation dictionary dynamically
            agg_dict = {'ticker': 'count'}
            col_names = ['Count']

            if 'price' in enriched.columns:
                agg_dict['price'] = 'mean'
                col_names.append('Avg Price')

            if 'rsi' in enriched.columns:
                agg_dict['rsi'] = 'mean'
                col_names.append('Avg RSI')

            if 'pe_ratio' in enriched.columns:
                agg_dict['pe_ratio'] = 'mean'
                col_names.append('Avg P/E')

            sector_summary = enriched.groupby('sector').agg(agg_dict).round(2)
            sector_summary.columns = col_names

            ws_sector = wb.create_sheet("By Sector")
            for r_idx, row in enumerate(dataframe_to_rows(sector_summary, index=True, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    ws_sector.cell(row=r_idx, column=c_idx, value=value)

        # Sheet 3: By Industry (top 10)
        if 'industry' in enriched.columns:
            industry_counts = enriched['industry'].value_counts().head(10)

            ws_industry = wb.create_sheet("Top Industries")
            ws_industry.cell(1, 1, "Industry")
            ws_industry.cell(1, 2, "Count")

            for idx, (industry, count) in enumerate(industry_counts.items(), 2):
                ws_industry.cell(idx, 1, industry)
                ws_industry.cell(idx, 2, count)

        # Save
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)

        logger.info(f"✓ Created comparison report: {output_path}")
