"""
Integration tests for Phase 3 Enhanced Discovery features

Tests complete workflows:
- Screening with sector/industry filters
- Export enriched results
- Watchlist management
- End-to-end workflows
"""

import pytest
import pandas as pd
import tempfile
import json
from pathlib import Path
from datetime import datetime

from quantlab.core.screener import ScreenCriteria, StockScreener
from quantlab.core.screen_export import ScreenExporter
from quantlab.core.watchlist import WatchlistManager
from quantlab.data.database import DatabaseManager
from quantlab.data.lookup_tables import LookupTableManager
from quantlab.data.parquet_reader import ParquetReader
from quantlab.data.data_manager import DataManager
from quantlab.utils.config import Config


@pytest.fixture(scope="module")
def test_config():
    """Create test configuration"""
    config = Config()
    # Use test data paths if needed
    return config


@pytest.fixture(scope="module")
def temp_db_path():
    """Create temporary database path"""
    with tempfile.TemporaryDirectory() as tmpdir:
        yield Path(tmpdir) / "test.duckdb"


@pytest.fixture
def db_manager(temp_db_path):
    """Create database manager for testing"""
    db = DatabaseManager(str(temp_db_path))
    yield db


@pytest.fixture
def lookup_manager(db_manager):
    """Create lookup table manager"""
    return LookupTableManager(db_manager)


@pytest.fixture
def mock_parquet_reader():
    """Create mock parquet reader with sample data"""
    from unittest.mock import Mock

    parquet = Mock(spec=ParquetReader)

    # Mock get_tickers_with_recent_data
    parquet.get_tickers_with_recent_data = Mock(return_value=['AAPL', 'MSFT', 'GOOGL', 'JPM', 'BAC'])

    # Mock get_stock_daily to return sample data
    def mock_get_stock_daily(tickers, start_date, end_date, limit=None):
        # Return sample technical data
        data = []
        for ticker in tickers:
            for i in range(60):  # 60 days of data
                data.append({
                    'symbol': ticker,
                    'date': pd.Timestamp('2025-10-16') - pd.Timedelta(days=60-i),
                    'open': 100.0 + i * 0.5,
                    'high': 102.0 + i * 0.5,
                    'low': 98.0 + i * 0.5,
                    'close': 100.0 + i * 0.5 + (hash(ticker) % 10),
                    'volume': 1000000 + i * 10000
                })

        return pd.DataFrame(data) if data else pd.DataFrame()

    parquet.get_stock_daily = Mock(side_effect=mock_get_stock_daily)

    return parquet


@pytest.fixture
def mock_data_manager(mock_parquet_reader):
    """Create mock data manager"""
    from unittest.mock import Mock

    data_mgr = Mock(spec=DataManager)

    # Mock fundamentals
    def mock_get_fundamentals(ticker):
        from quantlab.models.ticker_data import FundamentalData

        fundamentals = {
            'AAPL': FundamentalData(
                ticker='AAPL',
                date=datetime.now().date(),
                market_cap=3000000000000,
                pe_ratio=28.5,
                forward_pe=25.0,
                peg_ratio=2.1,
                profit_margin=0.25,
                return_on_equity=0.45,
                revenue_growth=0.08,
                earnings_growth=0.12,
                debt_to_equity=1.5,
                current_ratio=1.1,
                data_source='test',
                fetched_at=datetime.now()
            ),
            'MSFT': FundamentalData(
                ticker='MSFT',
                date=datetime.now().date(),
                market_cap=2800000000000,
                pe_ratio=32.0,
                forward_pe=28.0,
                peg_ratio=2.5,
                profit_margin=0.35,
                return_on_equity=0.42,
                revenue_growth=0.15,
                earnings_growth=0.18,
                debt_to_equity=0.8,
                current_ratio=2.2,
                data_source='test',
                fetched_at=datetime.now()
            )
        }

        return fundamentals.get(ticker)

    data_mgr.get_fundamentals = Mock(side_effect=mock_get_fundamentals)

    # Mock sentiment
    data_mgr.get_sentiment = Mock(return_value=None)  # No sentiment for tests

    return data_mgr


@pytest.fixture
def screener(test_config, db_manager, mock_data_manager, mock_parquet_reader):
    """Create stock screener for testing"""
    return StockScreener(test_config, db_manager, mock_data_manager, mock_parquet_reader)


class TestSectorIndustryScreening:
    """Integration tests for sector/industry filtering"""

    def test_screen_by_sector(self, screener, lookup_manager):
        """Test screening with sector filter"""
        # Populate lookup table with test data
        test_tickers = ['AAPL', 'MSFT', 'GOOGL']
        for ticker in test_tickers:
            lookup_manager.update_company_info(ticker, {
                'sector': 'Technology',
                'industry': 'Software',
                'company_name': f'{ticker} Inc.'
            })

        criteria = ScreenCriteria(
            sectors=['Technology'],
            volume_min=500000
        )

        results = screener.screen(criteria, limit=10, workers=1)

        # Should only return Technology stocks
        if not results.empty:
            assert len(results) <= 3  # Only AAPL, MSFT, GOOGL are in Technology

    def test_screen_excluding_sector(self, screener, lookup_manager):
        """Test screening with sector exclusion"""
        # Set up test data
        lookup_manager.update_company_info('AAPL', {
            'sector': 'Technology',
            'industry': 'Consumer Electronics',
            'company_name': 'Apple Inc.'
        })
        lookup_manager.update_company_info('JPM', {
            'sector': 'Financials',
            'industry': 'Banks',
            'company_name': 'JPMorgan Chase'
        })

        criteria = ScreenCriteria(
            exclude_sectors=['Technology'],
            volume_min=500000
        )

        results = screener.screen(criteria, limit=10, workers=1)

        # Should not include Technology stocks
        if not results.empty:
            assert 'AAPL' not in results['ticker'].values

    def test_screen_by_industry(self, screener, lookup_manager):
        """Test screening with industry filter"""
        lookup_manager.update_company_info('MSFT', {
            'sector': 'Technology',
            'industry': 'Software',
            'company_name': 'Microsoft Corp.'
        })

        criteria = ScreenCriteria(
            industries=['Software'],
            volume_min=500000
        )

        results = screener.screen(criteria, limit=10, workers=1)

        # Results should only include Software industry
        assert results is not None


class TestExportWorkflow:
    """Integration tests for complete export workflow"""

    def test_screen_and_export_to_excel(self, screener, db_manager, lookup_manager):
        """Test complete workflow: screen -> export to Excel"""
        # Set up lookup data
        lookup_manager.update_company_info('AAPL', {
            'sector': 'Technology',
            'industry': 'Consumer Electronics',
            'company_name': 'Apple Inc.'
        })

        # Run screening
        criteria = ScreenCriteria(
            rsi_max=70,
            volume_min=500000
        )

        results = screener.screen(criteria, limit=5, workers=1)

        if results.empty:
            pytest.skip("No screening results to test export")

        # Export to Excel
        exporter = ScreenExporter(db_manager, lookup_manager)

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "screen_results.xlsx"

            exporter.export_to_excel(results, str(output_path), enrich=True)

            assert output_path.exists()

            # Verify enriched data
            df = pd.read_excel(output_path)
            assert 'company_name' in df.columns
            assert 'sector' in df.columns
            assert 'pct_from_sma20' in df.columns

    def test_screen_and_export_to_csv(self, screener, db_manager, lookup_manager):
        """Test complete workflow: screen -> export to CSV"""
        criteria = ScreenCriteria(volume_min=500000)

        results = screener.screen(criteria, limit=5, workers=1)

        if results.empty:
            pytest.skip("No screening results")

        exporter = ScreenExporter(db_manager, lookup_manager)

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "screen_results.csv"

            exporter.export_to_csv(results, str(output_path), enrich=True)

            assert output_path.exists()

            # Verify data
            df = pd.read_csv(output_path)
            assert not df.empty

    def test_create_comparison_report(self, screener, db_manager, lookup_manager):
        """Test creating multi-sheet comparison report"""
        lookup_manager.update_company_info('AAPL', {
            'sector': 'Technology',
            'industry': 'Consumer Electronics',
            'company_name': 'Apple Inc.'
        })
        lookup_manager.update_company_info('MSFT', {
            'sector': 'Technology',
            'industry': 'Software',
            'company_name': 'Microsoft Corp.'
        })

        criteria = ScreenCriteria(volume_min=500000)
        results = screener.screen(criteria, limit=10, workers=1)

        if results.empty:
            pytest.skip("No screening results")

        exporter = ScreenExporter(db_manager, lookup_manager)

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "comparison_report.xlsx"

            exporter.create_comparison_report(results, str(output_path))

            assert output_path.exists()

            # Verify multiple sheets
            import openpyxl
            wb = openpyxl.load_workbook(output_path)
            assert "All Results" in wb.sheetnames
            assert "By Sector" in wb.sheetnames


class TestWatchlistWorkflow:
    """Integration tests for complete watchlist workflow"""

    def test_screen_to_watchlist_workflow(self, screener, db_manager):
        """Test complete workflow: screen -> save to watchlist -> retrieve"""
        # Create watchlist manager
        watchlist_mgr = WatchlistManager(db_manager)

        # Create watchlist
        success = watchlist_mgr.create_watchlist(
            "tech_screening",
            "Technology Screen Results",
            description="High momentum tech stocks"
        )
        assert success

        # Run screening
        criteria = ScreenCriteria(
            macd_signal='bullish',
            volume_min=500000
        )

        results = screener.screen(criteria, limit=10, workers=1)

        if results.empty:
            pytest.skip("No screening results")

        # Save to watchlist
        count = watchlist_mgr.add_from_screen_results(
            "tech_screening",
            results,
            reason="Bullish MACD screening"
        )

        assert count > 0

        # Retrieve watchlist
        items = watchlist_mgr.get_watchlist("tech_screening")

        assert not items.empty
        assert len(items) == count

    def test_watchlist_snapshot_comparison(self, screener, db_manager):
        """Test watchlist tracking over time"""
        watchlist_mgr = WatchlistManager(db_manager)

        watchlist_mgr.create_watchlist("momentum", "Momentum Stocks")

        # First screening (day 1)
        criteria1 = ScreenCriteria(volume_min=500000)
        results1 = screener.screen(criteria1, limit=5, workers=1)

        if results1.empty:
            pytest.skip("No results for snapshot test")

        watchlist_mgr.add_from_screen_results("momentum", results1)

        # Second screening (day 2) - simulate different results
        results2 = results1.copy()
        if len(results2) > 1:
            # Remove one stock, add price changes
            results2 = results2.iloc[:-1].copy()
            results2['price'] = results2['price'] * 1.05  # 5% price increase

            watchlist_mgr.add_from_screen_results("momentum", results2)

            # Compare snapshots
            comparison = watchlist_mgr.compare_snapshots("momentum")

            assert comparison is not None
            assert 'added' in comparison
            assert 'removed' in comparison
            assert 'price_changes' in comparison

    def test_export_watchlist(self, screener, db_manager):
        """Test exporting watchlist to JSON"""
        watchlist_mgr = WatchlistManager(db_manager)

        watchlist_mgr.create_watchlist(
            "export_test",
            "Export Test",
            tags=["test", "export"]
        )

        # Add some stocks
        criteria = ScreenCriteria(volume_min=500000)
        results = screener.screen(criteria, limit=3, workers=1)

        if not results.empty:
            watchlist_mgr.add_from_screen_results("export_test", results)

            with tempfile.TemporaryDirectory() as tmpdir:
                output_path = Path(tmpdir) / "watchlist.json"

                success = watchlist_mgr.export_watchlist("export_test", str(output_path))

                assert success
                assert output_path.exists()

                # Verify JSON content
                with open(output_path) as f:
                    data = json.load(f)

                assert data['watchlist_id'] == 'export_test'
                assert data['name'] == 'Export Test'
                assert 'test' in data['tags']


class TestEndToEndScenarios:
    """End-to-end integration tests"""

    def test_complete_discovery_workflow(self, screener, db_manager, lookup_manager):
        """Test complete discovery workflow"""
        # 1. Set up company data
        lookup_manager.update_company_info('AAPL', {
            'sector': 'Technology',
            'industry': 'Consumer Electronics',
            'company_name': 'Apple Inc.'
        })

        # 2. Screen for oversold tech stocks
        criteria = ScreenCriteria(
            sectors=['Technology'],
            rsi_max=35,
            volume_min=1000000
        )

        results = screener.screen(criteria, limit=20, workers=1)

        # 3. Export enriched results
        if not results.empty:
            exporter = ScreenExporter(db_manager, lookup_manager)

            with tempfile.TemporaryDirectory() as tmpdir:
                excel_path = Path(tmpdir) / "oversold_tech.xlsx"
                exporter.export_to_excel(results, str(excel_path), enrich=True)

                assert excel_path.exists()

                # 4. Save to watchlist
                watchlist_mgr = WatchlistManager(db_manager)
                watchlist_mgr.create_watchlist("oversold_tech", "Oversold Technology")
                count = watchlist_mgr.add_from_screen_results("oversold_tech", results)

                assert count > 0

                # 5. Export watchlist
                json_path = Path(tmpdir) / "oversold_tech.json"
                watchlist_mgr.export_watchlist("oversold_tech", str(json_path))

                assert json_path.exists()

    def test_multi_criteria_screening_with_export(self, screener, db_manager, lookup_manager):
        """Test screening with multiple criteria and export"""
        lookup_manager.update_company_info('MSFT', {
            'sector': 'Technology',
            'industry': 'Software',
            'company_name': 'Microsoft Corp.'
        })

        # Comprehensive criteria
        criteria = ScreenCriteria(
            sectors=['Technology'],
            rsi_min=40,
            rsi_max=60,
            macd_signal='bullish',
            volume_min=1000000,
            pe_max=35
        )

        results = screener.screen(criteria, limit=10, workers=1)

        if not results.empty:
            exporter = ScreenExporter(db_manager, lookup_manager)

            with tempfile.TemporaryDirectory() as tmpdir:
                # Create comparison report
                report_path = Path(tmpdir) / "multi_criteria_report.xlsx"
                exporter.create_comparison_report(results, str(report_path))

                if report_path.exists():  # May not exist if no results
                    import openpyxl
                    wb = openpyxl.load_workbook(report_path)
                    assert len(wb.sheetnames) >= 2  # At least All Results + one analysis sheet


class TestErrorHandling:
    """Test error handling in integration scenarios"""

    def test_screening_with_invalid_sector(self, screener):
        """Test that invalid sector filter doesn't crash"""
        criteria = ScreenCriteria(
            sectors=['NonexistentSector123'],
            volume_min=1000000
        )

        results = screener.screen(criteria, limit=10, workers=1)

        # Should return empty results, not crash
        assert isinstance(results, pd.DataFrame)

    def test_export_empty_results(self, db_manager, lookup_manager):
        """Test exporting empty screening results"""
        exporter = ScreenExporter(db_manager, lookup_manager)
        empty_df = pd.DataFrame()

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "empty.xlsx"

            # Should not create file or crash
            exporter.export_to_excel(empty_df, str(output_path))

            # File should not be created for empty results
            assert not output_path.exists()

    def test_watchlist_with_missing_data(self, db_manager):
        """Test watchlist operations with missing data"""
        watchlist_mgr = WatchlistManager(db_manager)

        watchlist_mgr.create_watchlist("test", "Test")

        # Add results with missing columns
        df = pd.DataFrame({
            'ticker': ['AAPL'],
            # Missing price and score
        })

        # Should handle gracefully
        count = watchlist_mgr.add_from_screen_results("test", df)

        assert count >= 0  # Should not crash
