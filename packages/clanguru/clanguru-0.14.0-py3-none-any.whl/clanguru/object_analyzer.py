import fnmatch
import os
import re
import subprocess
from concurrent.futures import ThreadPoolExecutor
from dataclasses import dataclass, field
from enum import Enum, auto
from functools import cached_property
from pathlib import Path
from typing import Any, Optional

from jinja2 import Environment, FileSystemLoader
from mashumaro import DataClassDictMixin
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from py_app_dev.core.exceptions import UserNotificationException
from py_app_dev.core.logging import time_it
from py_app_dev.core.subprocess import SubprocessExecutor

from clanguru.compilation_options_manager import CompilationDatabase, CompileCommand


class SymbolLinkage(Enum):
    EXTERN = auto()
    LOCAL = auto()


@dataclass
class Symbol:
    name: str
    linkage: SymbolLinkage


@dataclass
class ObjectDependencies:
    path: Path
    symbols: list[Symbol] = field(default_factory=list)

    @property
    def name(self) -> str:
        return self.path.name

    @cached_property
    def required_symbols(self) -> set[str]:
        """Collects all EXTERN symbols into a set."""
        return {symbol.name for symbol in self.symbols if symbol.linkage == SymbolLinkage.EXTERN}

    @cached_property
    def provided_symbols(self) -> set[str]:
        """Collects all LOCAL symbols into a set."""
        return {symbol.name for symbol in self.symbols if symbol.linkage == SymbolLinkage.LOCAL}


@dataclass
class ObjectReportData:
    object_dependencies: ObjectDependencies
    compile_command: CompileCommand

    @property
    def name(self) -> str:
        return self.object_dependencies.name

    @property
    def source_file(self) -> Path:
        return self.compile_command.get_file_path()

    @property
    def object_file(self) -> Path:
        return self.object_dependencies.path

    @property
    def symbols(self) -> list[Symbol]:
        return self.object_dependencies.symbols

    @property
    def required_symbols(self) -> set[str]:
        return self.object_dependencies.required_symbols

    @property
    def provided_symbols(self) -> set[str]:
        return self.object_dependencies.provided_symbols


class ObjectReportDataPath:
    def __init__(self, root_path: Path, rel_path: Path | None = None):
        self.root_path = root_path
        self.rel_path = rel_path

    @property
    def full_path(self) -> Path:
        return self.root_path / self.rel_path if self.rel_path else self.root_path


@dataclass
class ObjectReportDataTree:
    """Represents the object report data node in the directory tree where the objects are the leafs."""

    name: str | None
    path: ObjectReportDataPath
    children: list["ObjectReportDataTree"] = field(default_factory=list)
    objects: list[ObjectReportData] = field(default_factory=list)

    @property
    def id(self) -> str:
        return self.path.rel_path.as_posix() if self.path.rel_path else (self.name or "root")


def determine_common_path(file_paths: list[Path]) -> Path:
    """Determine common path for files."""
    try:
        common_base_str = os.path.commonpath([str(path) for path in file_paths])
        common_base = Path(common_base_str)

        # If all files are in the same directory, return the parent to ensure we have at least one directory level
        if all(path == common_base for path in file_paths):
            return common_base.parent if common_base.parent != common_base else common_base

        return common_base
    except ValueError:
        raise UserNotificationException("No common base path found. Ensure all files are in the same directory tree.") from None


def build_tree_from_paths(paths: list[list[str]]) -> dict[str, Any]:
    """Builds a tree from a list of path components."""
    root: dict[str, Any] = {}
    for path in paths:
        current_level = root
        for part in path:
            current_level = current_level.setdefault(part, {})
    return root


def collapse_objects_report_data_tree(node: ObjectReportDataTree) -> ObjectReportDataTree:
    """Recursively collapse nodes that have a total of one child or object."""
    # First, recursively collapse all children.
    collapsed_children = [collapse_objects_report_data_tree(child) for child in node.children]
    node.children = collapsed_children

    # Collapse if the node has exactly one child and no objects
    if len(node.children) == 1 and len(node.objects) == 0:
        child = node.children[0]
        combined_name = f"{node.name}/{child.name}" if node.name else child.name
        return ObjectReportDataTree(
            name=combined_name,
            children=child.children,
            objects=child.objects,
            path=ObjectReportDataPath(
                rel_path=child.path.rel_path,
                root_path=node.path.root_path,
            ),
        )

    # After collapsing individual children, check if we should collect objects from leaf children
    # Only collect objects from children that have exactly 1 object and no children (single-object leaf nodes)
    # But only do this for named nodes (not the root node which has name=None)
    if node.name is not None and collapsed_children:  # Don't collapse into the root node
        # Separate children into those we can collapse (single-object leaf nodes) and those we keep
        children_to_keep = []
        collected_objects = list(node.objects)  # Start with current objects

        for child in collapsed_children:
            if len(child.children) == 0 and len(child.objects) == 1:
                # This is a single-object leaf node - collect its object
                collected_objects.extend(child.objects)
            else:
                # Keep this child (either has multiple objects or has children)
                children_to_keep.append(child)

        # Return the node with potentially collected objects and remaining children
        return ObjectReportDataTree(
            name=node.name,
            children=children_to_keep,
            objects=collected_objects,
            path=ObjectReportDataPath(
                rel_path=node.path.rel_path,
                root_path=node.path.root_path,
            ),
        )

    # This node cannot be collapsed, so return it as is.
    return node


def create_objects_report_data_tree_expanded(objects_report_data: list[ObjectReportData]) -> ObjectReportDataTree:
    if not objects_report_data:
        raise ValueError("No object report data available to create tree.")

    files = [obj.source_file for obj in objects_report_data]
    common_path = determine_common_path(files)
    root = ObjectReportDataTree(name=None, path=ObjectReportDataPath(rel_path=None, root_path=common_path))
    node_map: dict[Path, ObjectReportDataTree] = {common_path: root}

    for report_data in sorted(objects_report_data, key=lambda obj: obj.source_file):
        relative_path = report_data.source_file.relative_to(common_path)
        parent_path = common_path
        # Create directory nodes.
        for part in relative_path.parts[:-1]:
            current_path = parent_path / part
            if current_path not in node_map:
                # Calculate relative path for this node
                node_rel_path = current_path.relative_to(common_path)
                new_node = ObjectReportDataTree(name=part, path=ObjectReportDataPath(rel_path=node_rel_path, root_path=common_path))
                node_map[current_path] = new_node
                node_map[parent_path].children.append(new_node)
            parent_path = current_path

        # Add object to its parent node.
        node_map[parent_path].objects.append(report_data)
    return root


@time_it("create_objects_report_data_tree")
def create_objects_report_data_tree(objects_report_data: list[ObjectReportData]) -> ObjectReportDataTree:
    """
    Create the objects report data tree structure based on the source files.

    Example based on this list of files:
        C:/temp/my/project/components/comp_a/src/comp_a.c
        C:/temp/my/project/components/comp_b/src/comp_b.c
        C:/temp/my/project/mcal/src/mcal.c
        C:/temp/my/project/mcal/src/drivers/adc.c

    Common path will be: C:/temp/my/project

    Nodes:

    - components
    - comp_a/src/comp_a.c with parent components
    - comp_b/src/comp_b.c with parent components
    - mcal/src
    - mcal.c with parent mcal/src
    - drivers/src/adc.c with parent mcal/src

    """
    if not objects_report_data:
        raise ValueError("No object report data available to create tree.")

    # Step 1: Build the full tree structure.
    root = create_objects_report_data_tree_expanded(objects_report_data)

    # Step 2: Collapse the tree.
    collapsed_root = collapse_objects_report_data_tree(root)
    return collapsed_root


@dataclass
class ObjectsGraphDataEdgesData(DataClassDictMixin):
    id: str
    source: str
    target: str


@dataclass
class ObjectsGraphDataEdges(DataClassDictMixin):
    data: ObjectsGraphDataEdgesData


@dataclass
class ObjectsGraphDataNodesData(DataClassDictMixin):
    content: str
    font_size: int
    id: str
    label: str
    size: int
    parent: Optional[str] = None


@dataclass
class ObjectsGraphDataNodes(DataClassDictMixin):
    data: ObjectsGraphDataNodesData


@dataclass
class ObjectsGraphData(DataClassDictMixin):
    edges: list[ObjectsGraphDataEdges]
    nodes: list[ObjectsGraphDataNodes]


@time_it("create_objects_graph_data_nodes")
def create_objects_graph_data_nodes(
    object_tree: ObjectReportDataTree, node_connections: dict[str, int], use_parent_deps: bool = True, exclude_isolated_objects: bool = True
) -> list[ObjectsGraphDataNodes]:
    """
    Create graph data nodes from the object report data tree.

    Every node in the tree that has a name becomes a graph node.
    Every object in the leaf nodes becomes a graph node.
    """
    nodes: list[ObjectsGraphDataNodes] = []

    def traverse_tree(node: ObjectReportDataTree, parent_name: str | None = None) -> None:
        """Recursively traverse the tree and create graph nodes."""
        current_parent: Optional[str] = None
        if use_parent_deps:
            # Skip root node (name is None)
            if node.name is not None:
                # Create a directory node
                dir_node_data = ObjectsGraphDataNodesData(
                    content=node.name,
                    font_size=12,
                    id=node.path.rel_path.as_posix() if node.path.rel_path else node.name,
                    label=node.name,
                    parent=parent_name if use_parent_deps else None,
                    size=0,
                )
                nodes.append(ObjectsGraphDataNodes(data=dir_node_data))
                current_parent = node.id
            else:
                current_parent = parent_name

        # Create nodes for individual objects
        for obj in node.objects:
            obj_file_name = obj.compile_command.file.name
            label = obj.source_file.relative_to(node.path.full_path).as_posix() if node.path.full_path else obj_file_name
            # The ID the whole relative path of the object file
            id = obj.source_file.relative_to(node.path.root_path).as_posix() if node.path.root_path else obj_file_name

            node_size = node_connections.get(id, 0)
            # Skip isolated objects if the option is set
            if exclude_isolated_objects and node_size == 0:
                continue
            obj_node_data = ObjectsGraphDataNodesData(
                content=label,
                font_size=10,
                id=id,
                label=label,
                parent=current_parent if use_parent_deps else None,
                size=5 + node_size * 2,  # Scaling based on connections
            )
            nodes.append(ObjectsGraphDataNodes(data=obj_node_data))

        # Recursively process children
        for child in node.children:
            traverse_tree(child, current_parent)

    traverse_tree(object_tree)
    return nodes


@time_it("create_objects_graph_data_edges")
def create_objects_graph_data_edges(objects_report_data: list[ObjectReportData]) -> tuple[list[ObjectsGraphDataEdges], dict[str, int]]:
    if not objects_report_data:
        raise ValueError("No object report data available to create tree.")

    files = [obj.source_file for obj in objects_report_data]
    common_path = determine_common_path(files)

    # Pre-compute object IDs to avoid repeated relative path calculations
    object_ids = [obj.source_file.relative_to(common_path).as_posix() for obj in objects_report_data]
    node_connections = dict.fromkeys(object_ids, 0)

    # Build symbol-to-objects mapping for faster lookups
    symbol_providers: dict[str, list[int]] = {}  # symbol -> list of object indices that provide it

    for obj_idx, obj in enumerate(objects_report_data):
        for symbol in obj.provided_symbols:
            if symbol not in symbol_providers:
                symbol_providers[symbol] = []
            symbol_providers[symbol].append(obj_idx)

    edges: list[ObjectsGraphDataEdges] = []
    edge_set = set()  # To avoid duplicate edges between the same pair

    # For each object, find all objects that provide symbols it requires
    for obj_idx, obj in enumerate(objects_report_data):
        obj_id = object_ids[obj_idx]

        # Find all providers for symbols this object requires
        connected_objects = set()
        for required_symbol in obj.required_symbols:
            if required_symbol in symbol_providers:
                for provider_idx in symbol_providers[required_symbol]:
                    if provider_idx != obj_idx:  # Don't connect to self
                        connected_objects.add(provider_idx)

        # Create edges for all connected objects
        for connected_idx in connected_objects:
            connected_id = object_ids[connected_idx]

            # Ensure edge ID is consistent regardless of order
            source_name, target_name = sorted([obj_id, connected_id])
            edge_id = f"{source_name}.{target_name}"

            if edge_id not in edge_set:
                edges.append(ObjectsGraphDataEdges(data=ObjectsGraphDataEdgesData(id=edge_id, source=obj_id, target=connected_id)))
                edge_set.add(edge_id)
                node_connections[obj_id] += 1
                node_connections[connected_id] += 1

    return edges, node_connections


class NmExecutor:
    @staticmethod
    def run(obj_file: Path) -> ObjectDependencies:
        obj_data = ObjectDependencies(obj_file)
        executor = SubprocessExecutor(command=["nm", obj_file], capture_output=True, print_output=False)
        completed_process = executor.execute(handle_errors=False)
        if completed_process:
            if completed_process.returncode != 0:
                raise subprocess.CalledProcessError(completed_process.returncode, completed_process.args, stderr=completed_process.stderr)

            # Process the output
            for line in completed_process.stdout.splitlines():
                symbol = NmExecutor.get_symbol(line)
                if symbol:
                    obj_data.symbols.append(symbol)
        else:
            raise UnboundLocalError("nm command failed")
        return obj_data

    @staticmethod
    def get_symbol(nm_symbol_output: str) -> Optional[Symbol]:
        # Regex to capture optional address, mandatory uppercase symbol type, and symbol name
        # Group 1: Symbol Type Letter (e.g., 'U', 'T', 'D', 'B', etc.)
        # Group 2: Symbol Name
        pattern: re.Pattern[str] = re.compile(r"^\s*(?:[0-9A-Fa-f]+\s+)?([A-Z])\s+(\S+)")
        match = pattern.match(nm_symbol_output)

        if match:
            symbol_type_letter = match.group(1)
            symbol_name = match.group(2)

            # Determine linkage based on the symbol type letter
            linkage = SymbolLinkage.EXTERN if symbol_type_letter == "U" else SymbolLinkage.LOCAL

            return Symbol(name=symbol_name, linkage=linkage)

        return None


def parse_objects(obj_files: list[Path], max_workers: Optional[int] = None) -> list[ObjectDependencies]:
    """Run the nm executor on each object file in parallel, collecting all the resulting ObjectData in the same order as `obj_files`."""
    with ThreadPoolExecutor(max_workers=max_workers) as pool:
        # executor.map preserves input order in its output sequence
        results = list(pool.map(NmExecutor.run, obj_files))

    return results


def filter_object_data_symbols(object_data: list[ObjectDependencies], exclude_patterns: list[str] | None) -> list[ObjectDependencies]:
    """
    Filter object symbols by glob patterns returning fresh ``ObjectData`` instances.

    The original ``object_data`` collection is left untouched to avoid stale
    ``cached_property`` values (``required_symbols`` / ``provided_symbols``). By
    reconstructing each ``ObjectData`` we ensure those sets are recomputed from
    the filtered symbol list.

    Args:
        object_data: Parsed object data instances.
        exclude_patterns: Glob (``fnmatch``) patterns for symbols to remove. ``None``
            or an empty list -> no filtering.

    Returns:
        New list with symbols whose names do not match any exclusion pattern.

    """
    if not exclude_patterns:
        return object_data

    filtered: list[ObjectDependencies] = []
    for obj in object_data:
        kept_symbols = [s for s in obj.symbols if not any(fnmatch.fnmatch(s.name, pat) for pat in exclude_patterns)]
        # Recreate ObjectData to avoid stale cached_property values
        filtered.append(ObjectDependencies(path=obj.path, symbols=kept_symbols))
    return filtered


def filter_external_symbols_only(object_data: list[ObjectDependencies]) -> list[ObjectDependencies]:
    """
    Filter object data to keep only symbols with external linkage (EXTERN).

    For dependency analysis, we typically only care about external interfaces
    and dependencies, not internal implementation details (LOCAL symbols).

    Args:
        object_data: Parsed object data instances.

    Returns:
        New list with only symbols that have external linkage (SymbolLinkage.EXTERN).

    """
    filtered: list[ObjectDependencies] = []
    for obj in object_data:
        extern_symbols = [s for s in obj.symbols if s.linkage == SymbolLinkage.EXTERN]
        # Recreate ObjectData to avoid stale cached_property values
        filtered.append(ObjectDependencies(path=obj.path, symbols=extern_symbols))
    return filtered


def create_report_data(compilation_database: CompilationDatabase, object_data: list[ObjectDependencies]) -> list[ObjectReportData]:
    """Create a list of ObjectReportData by matching object files from the compilation database with parsed object data."""
    report_data: list[ObjectReportData] = []
    object_data_map = {obj.path: obj for obj in object_data}
    for command in compilation_database.commands:
        output_path = command.get_output_path()
        if output_path and output_path in object_data_map:
            report_data.append(ObjectReportData(object_dependencies=object_data_map[output_path], compile_command=command))
    return report_data


class ObjectsDependenciesReportGenerator:
    def __init__(
        self,
        objects_data: list[ObjectReportData],
        use_parent_deps: bool = False,
        exclude_isolated_objects: bool = True,
    ):
        self.objects_data = objects_data
        self.use_parent_deps = use_parent_deps
        self.exclude_isolated_objects = exclude_isolated_objects

    def generate_report(self, output_file: Path) -> None:
        """Generates the HTML report by rendering the Jinja2 template with the graph data."""
        graph_data = self.generate_graph_data()

        env = Environment(loader=FileSystemLoader(Path(__file__).parent), autoescape=True)
        template = env.get_template("object_analyzer.html.jinja")
        rendered_html = template.render(graph_data=graph_data.to_dict())

        # Write the rendered HTML to the output file
        with open(output_file, "w", encoding="utf-8") as file:
            file.write(rendered_html)

    def generate_graph_data(self) -> ObjectsGraphData:
        """Converts a list of ObjectData into a dictionary suitable for Cytoscape.js containing nodes and edges representing object dependencies."""
        objects = self.objects_data
        objects_report_data_tree = create_objects_report_data_tree(objects)
        edges, nodes_connections = create_objects_graph_data_edges(objects)
        nodes = create_objects_graph_data_nodes(objects_report_data_tree, nodes_connections, self.use_parent_deps, self.exclude_isolated_objects)
        return ObjectsGraphData(edges=edges, nodes=nodes)


class ExcelColumnMapper:
    """Manages Excel column mappings for the Objects sheet."""

    def __init__(self) -> None:
        # Define column positions
        self.OBJECT_NAME = 1
        self.FILE_PATH = 2
        self.PROVIDED_TOTAL = 3
        self.PROVIDED_USED = 4
        self.PROVIDED_DEPENDENCIES = 5
        self.REQUIRED_COUNT = 6
        self.REQUIRED_DEPENDENCIES = 7
        self.TOTAL_SYMBOLS = 8

        # Define column groups for headers
        self.OBJECT_INFO_START = self.OBJECT_NAME
        self.OBJECT_INFO_END = self.FILE_PATH

        self.PROVIDED_INTERFACES_START = self.PROVIDED_TOTAL
        self.PROVIDED_INTERFACES_END = self.PROVIDED_DEPENDENCIES

        self.REQUIRED_INTERFACES_START = self.REQUIRED_COUNT
        self.REQUIRED_INTERFACES_END = self.REQUIRED_DEPENDENCIES

        self.TOTAL_COLUMN = self.TOTAL_SYMBOLS

        # Total number of columns
        self.TOTAL_COLUMNS = self.TOTAL_SYMBOLS

        # Header texts
        self.MAIN_HEADERS = {
            self.OBJECT_INFO_START: "Object Info",
            self.PROVIDED_INTERFACES_START: "Provided Interfaces",
            self.REQUIRED_INTERFACES_START: "Required Interfaces",
            self.TOTAL_COLUMN: "Total",
        }

        self.SUB_HEADERS = [
            "Object Name",  # OBJECT_NAME
            "File Path",  # FILE_PATH
            "Total",  # PROVIDED_TOTAL
            "Used",  # PROVIDED_USED
            "Dependencies",  # PROVIDED_DEPENDENCIES
            "Count",  # REQUIRED_COUNT
            "Dependencies",  # REQUIRED_DEPENDENCIES
            "Symbols",  # TOTAL_SYMBOLS
        ]

        # Column ranges for merging
        self.MERGE_RANGES = [
            "A1:B1",  # Object Info
            "C1:E1",  # Provided Interfaces
            "F1:G1",  # Required Interfaces
        ]

        # Columns that get main headers (for styling)
        self.MAIN_HEADER_COLUMNS = [self.OBJECT_INFO_START, self.PROVIDED_INTERFACES_START, self.REQUIRED_INTERFACES_START, self.TOTAL_COLUMN]


class ObjectsDataExcelReportGenerator:
    """Excel report generator for object data with dependency analysis."""

    def __init__(
        self,
        objects_data: list[ObjectReportData],
        use_parent_deps: bool = False,
        create_traceability_matrix: bool = False,
    ) -> None:
        self.objects_data = objects_data
        self.use_parent_deps = use_parent_deps
        self.create_traceability_matrix = create_traceability_matrix
        self.columns = ExcelColumnMapper()

    def generate_report(self, output_file: Path) -> None:
        wb = Workbook()
        self._create_objects_sheet(wb)
        if self.create_traceability_matrix:
            self._create_dependency_matrix_sheet(wb)
        wb.save(output_file)

    def _create_objects_sheet(self, wb: Workbook) -> None:
        """Create the Objects sheet with object statistics."""
        ws = wb.active
        if ws is None:
            raise RuntimeError("Failed to create worksheet")
        ws.title = "Objects"

        # Create grouped headers
        self._create_grouped_headers(ws, self.columns)

        for row, object_data in enumerate(self.objects_data, 3):  # Start from row 3 due to grouped headers
            # Calculate dependencies
            objects_requiring_from_this = [
                other_obj.name for other_obj in self.objects_data if other_obj != object_data and object_data.provided_symbols.intersection(other_obj.required_symbols)
            ]
            objects_providing_to_this = [
                other_obj.name for other_obj in self.objects_data if other_obj != object_data and other_obj.provided_symbols.intersection(object_data.required_symbols)
            ]

            # Calculate used provided interfaces (symbols from this object that are actually required by other objects)
            used_provided_interfaces = set()
            for other_obj in self.objects_data:
                if other_obj != object_data:
                    used_provided_interfaces.update(object_data.provided_symbols.intersection(other_obj.required_symbols))

            # Choose file path based on use_object_paths setting
            file_path = object_data.source_file

            ws.cell(row=row, column=self.columns.OBJECT_NAME, value=object_data.name)
            ws.cell(row=row, column=self.columns.FILE_PATH, value=str(file_path))
            ws.cell(row=row, column=self.columns.PROVIDED_TOTAL, value=len(object_data.provided_symbols))
            ws.cell(row=row, column=self.columns.PROVIDED_USED, value=len(used_provided_interfaces))
            ws.cell(row=row, column=self.columns.PROVIDED_DEPENDENCIES, value="\n".join(objects_requiring_from_this) if objects_requiring_from_this else "None")
            ws.cell(row=row, column=self.columns.REQUIRED_COUNT, value=len(object_data.required_symbols))
            ws.cell(row=row, column=self.columns.REQUIRED_DEPENDENCIES, value="\n".join(objects_providing_to_this) if objects_providing_to_this else "None")
            ws.cell(row=row, column=self.columns.TOTAL_SYMBOLS, value=len(object_data.symbols))

        self._auto_adjust_columns(ws, self.columns.TOTAL_COLUMNS)

    def _create_dependency_matrix_sheet(self, wb: Workbook) -> None:
        """Create the Dependency Matrix sheet showing interface dependencies."""
        ws = wb.create_sheet("Dependency Matrix")

        interface_usage = self._calculate_interface_usage()
        obj_names = [obj.name for obj in self.objects_data]

        fixed_headers = ["Object", "Interface", "Usage Count"]
        headers = [*fixed_headers, *obj_names]
        self._create_header_row(ws, headers)

        # Freeze the first row and fixed headers
        freeze_column = get_column_letter(len(fixed_headers) + 1)
        ws.freeze_panes = f"{freeze_column}2"

        current_row = 2
        for obj in self.objects_data:
            if not obj.provided_symbols:
                continue

            for interface_name in sorted(obj.provided_symbols):
                self._create_dependency_row(ws, current_row, obj.name, interface_name, interface_usage[interface_name])
                current_row += 1

        self._auto_adjust_columns(ws, len(headers))

    def _create_header_row(self, ws: Any, headers: list[str]) -> None:
        """Create and style header row."""
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

    def _create_dependency_row(self, ws: Any, row: int, obj_name: str, interface_name: str, usage_count: int) -> None:
        """Create a single dependency matrix row."""
        ws.cell(row=row, column=1, value=obj_name).font = Font(bold=True)
        ws.cell(row=row, column=2, value=interface_name)

        usage_cell = ws.cell(row=row, column=3, value=usage_count)
        usage_cell.font = Font(bold=usage_count > 0)
        usage_cell.alignment = Alignment(horizontal="center", vertical="center")

        for col, requiring_obj in enumerate(self.objects_data, 4):
            if interface_name in requiring_obj.required_symbols:
                cell = ws.cell(row=row, column=col, value="X")
                cell.font = Font(bold=True, color="FF0000")
                cell.alignment = Alignment(horizontal="center", vertical="center")

    def _calculate_interface_usage(self) -> dict[str, int]:
        """Calculate usage count for each interface."""
        interface_usage = {}

        for obj in self.objects_data:
            for symbol in obj.provided_symbols:
                usage_count = sum(1 for other_obj in self.objects_data if symbol in other_obj.required_symbols)
                interface_usage[symbol] = usage_count

        return interface_usage

    def _auto_adjust_columns(self, ws: Any, column_count: int) -> None:
        """Auto-adjust column widths."""
        for col in range(1, column_count + 1):
            ws.column_dimensions[get_column_letter(col)].auto_size = True

    def _create_grouped_headers(self, ws: Any, column_mapper: ExcelColumnMapper) -> None:
        """Create grouped headers with main categories and sub-headers."""
        # First row - main category headers
        for start_col in column_mapper.MAIN_HEADER_COLUMNS:
            ws.cell(row=1, column=start_col, value=column_mapper.MAIN_HEADERS[start_col])

        # Merge cells for main headers
        for merge_range in column_mapper.MERGE_RANGES:
            ws.merge_cells(merge_range)

        # Second row - specific column headers
        sub_headers = column_mapper.SUB_HEADERS

        # Style main headers
        for col in column_mapper.MAIN_HEADER_COLUMNS:
            cell = ws.cell(row=1, column=col)
            cell.font = Font(bold=True, size=12)
            cell.fill = PatternFill(start_color="B0B0B0", end_color="B0B0B0", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Style and populate sub-headers
        for col, header in enumerate(sub_headers, 1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Add borders to separate groups
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for row in range(1, 3):
            for col in range(1, column_mapper.TOTAL_COLUMNS + 1):
                ws.cell(row=row, column=col).border = thin_border
