"""
Excel exporter for statistical results.
"""

from typing import Any, Dict, List, Optional, Union
import pandas as pd
from ..utils import format_p_value, format_effect_size, format_confidence_interval


class ExcelExporter:
    """
    Exporter for converting standardized statistical results to Excel format.
    """
    
    def __init__(self, data: Dict[str, Any]):
        """
        Initialize the exporter with standardized data.
        
        Args:
            data: Standardized statistical results dictionary
        """
        self.data = data
    
    def export(self, filename: str, sheet_name: str = "Statistical Results") -> None:
        """
        Export the statistical results to Excel format.
        
        Args:
            filename: Path to the Excel file to create
            sheet_name: Name of the Excel sheet
        """
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            # Create main results sheet
            self._create_main_sheet(writer, sheet_name)
            
            # Create coefficients sheet if available
            if self.data.get('coefficients'):
                self._create_coefficients_sheet(writer, "Coefficients")
            
            # Create model info sheet if available
            if self.data.get('model_info'):
                self._create_model_info_sheet(writer, "Model Information")
            
            # Create additional info sheet if available
            if self.data.get('additional_info'):
                self._create_additional_info_sheet(writer, "Additional Information")
    
    def _create_main_sheet(self, writer: pd.ExcelWriter, sheet_name: str) -> None:
        """Create the main results sheet."""
        # Create summary information
        summary_data = []
        
        # Test information
        test_name = self.data.get('test_name', 'Statistical Test')
        test_type = self.data.get('test_type', 'unknown')
        summary_data.append(['Test Name', test_name])
        summary_data.append(['Test Type', test_type])
        
        # Sample size
        sample_size = self.data.get('sample_size')
        if sample_size:
            if isinstance(sample_size, dict):
                for key, value in sample_size.items():
                    summary_data.append([f'Sample Size ({key})', value])
            else:
                summary_data.append(['Sample Size', sample_size])
        
        # Degrees of freedom
        df = self.data.get('degrees_of_freedom')
        if df is not None:
            if isinstance(df, dict):
                for key, value in df.items():
                    summary_data.append([f'Degrees of Freedom ({key})', value])
            else:
                summary_data.append(['Degrees of Freedom', df])
        
        summary_data.append(['', ''])  # Empty row
        
        # Main results
        summary_data.append(['Statistic', 'Value', 'p-value', 'Significance'])
        
        statistic = self.data.get('statistic')
        p_value = self.data.get('p_value')
        
        if statistic is not None and p_value is not None:
            if isinstance(statistic, dict):
                for key, value in statistic.items():
                    p_val = p_value.get(key.replace('stat_', 'p_'), p_value) if isinstance(p_value, dict) else p_value
                    significance = "***" if p_val < 0.001 else "**" if p_val < 0.01 else "*" if p_val < 0.05 else ""
                    
                    summary_data.append([
                        key.replace('stat_', '').title(),
                        value,
                        p_val,
                        significance
                    ])
            else:
                significance = "***" if p_value < 0.001 else "**" if p_value < 0.01 else "*" if p_value < 0.05 else ""
                summary_data.append([
                    'Test Statistic',
                    statistic,
                    p_value,
                    significance
                ])
        
        # Effect size
        effect_size = self.data.get('effect_size')
        if effect_size is not None:
            if isinstance(effect_size, dict):
                for key, value in effect_size.items():
                    summary_data.append([
                        key.replace('_', ' ').title(),
                        value,
                        '',
                        ''
                    ])
            else:
                summary_data.append([
                    'Effect Size',
                    effect_size,
                    '',
                    ''
                ])
        
        # Confidence interval
        ci = self.data.get('confidence_interval')
        if ci is not None:
            if isinstance(ci, dict) and 'lower' in ci and 'upper' in ci:
                ci_str = f"[{ci['lower']:.3f}, {ci['upper']:.3f}]"
                summary_data.append([
                    '95% CI',
                    ci_str,
                    '',
                    ''
                ])
        
        # Create DataFrame and write to Excel
        df_summary = pd.DataFrame(summary_data)
        df_summary.to_excel(writer, sheet_name=sheet_name, index=False, header=False)
        
        # Format the sheet
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]
        
        # Format headers
        from openpyxl.styles import Font, PatternFill, Alignment
        
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Apply formatting to header row (row 8, 0-indexed)
        for col in range(1, 5):
            cell = worksheet.cell(row=8, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
    
    def _create_coefficients_sheet(self, writer: pd.ExcelWriter, sheet_name: str) -> None:
        """Create the coefficients sheet."""
        coefficients = self.data.get('coefficients', [])
        if not coefficients:
            return
        
        # Prepare data
        data = []
        for coef in coefficients:
            variable = coef.get('variable', 'Unknown')
            coefficient = coef.get('coefficient', 0)
            std_error = coef.get('std_error')
            p_value = coef.get('p_value')
            ci_lower = coef.get('ci_lower')
            ci_upper = coef.get('ci_upper')
            
            # Format confidence interval
            ci_str = f"[{ci_lower:.3f}, {ci_upper:.3f}]" if ci_lower is not None and ci_upper is not None else ""
            
            data.append([
                variable,
                coefficient,
                std_error if std_error is not None else "",
                p_value if p_value is not None else "",
                ci_str
            ])
        
        # Create DataFrame
        df = pd.DataFrame(data, columns=['Variable', 'Coefficient', 'SE', 'p-value', '95% CI'])
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # Format the sheet
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]
        
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # Format headers
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        for col in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
    
    def _create_model_info_sheet(self, writer: pd.ExcelWriter, sheet_name: str) -> None:
        """Create the model information sheet."""
        model_info = self.data.get('model_info', {})
        if not model_info:
            return
        
        # Prepare data
        data = []
        for key, value in model_info.items():
            if isinstance(value, (int, float)):
                data.append([key.replace('_', ' ').title(), value])
        
        # Create DataFrame
        df = pd.DataFrame(data, columns=['Statistic', 'Value'])
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # Format the sheet
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]
        
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # Format headers
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        for col in range(1, 3):
            cell = worksheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
    
    def _create_additional_info_sheet(self, writer: pd.ExcelWriter, sheet_name: str) -> None:
        """Create the additional information sheet."""
        additional_info = self.data.get('additional_info', {})
        if not additional_info:
            return
        
        # Prepare data
        data = []
        for key, value in additional_info.items():
            if isinstance(value, (str, int, float, bool)):
                data.append([key.replace('_', ' ').title(), str(value)])
            elif isinstance(value, list):
                data.append([key.replace('_', ' ').title(), ', '.join(map(str, value))])
        
        # Create DataFrame
        df = pd.DataFrame(data, columns=['Information', 'Value'])
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # Format the sheet
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]
        
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # Format headers
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        for col in range(1, 3):
            cell = worksheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
    
    def to_dataframe(self) -> pd.DataFrame:
        """
        Convert the results to a pandas DataFrame.
        
        Returns:
            DataFrame with the main results
        """
        # Create main results DataFrame
        data = []
        
        statistic = self.data.get('statistic')
        p_value = self.data.get('p_value')
        
        if statistic is not None and p_value is not None:
            if isinstance(statistic, dict):
                for key, value in statistic.items():
                    p_val = p_value.get(key.replace('stat_', 'p_'), p_value) if isinstance(p_value, dict) else p_value
                    significance = "***" if p_val < 0.001 else "**" if p_val < 0.01 else "*" if p_val < 0.05 else ""
                    
                    data.append({
                        'Statistic': key.replace('stat_', '').title(),
                        'Value': value,
                        'p_value': p_val,
                        'Significance': significance
                    })
            else:
                significance = "***" if p_value < 0.001 else "**" if p_value < 0.01 else "*" if p_value < 0.05 else ""
                data.append({
                    'Statistic': 'Test Statistic',
                    'Value': statistic,
                    'p_value': p_value,
                    'Significance': significance
                })
        
        return pd.DataFrame(data)
