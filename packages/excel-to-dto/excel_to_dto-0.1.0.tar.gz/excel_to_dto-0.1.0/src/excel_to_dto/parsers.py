"""Парсеры для Excel файлов."""

from __future__ import annotations

import logging
import warnings
from dataclasses import fields, is_dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Type, TypeVar

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from excel_to_dto.exceptions import (
    EmptyValueError,
    FieldNotFoundError,
    ParsingError,
    SheetNotFoundError,
    ValidationError,
)
from excel_to_dto.models import (
    ExcelFileConfig,
    FieldConfig,
    HorizontalSheetConfig,
    SheetConfig,
    VerticalSheetConfig,
)
from excel_to_dto.utils import TypeConverter, clean_key, is_empty_value

logger = logging.getLogger(__name__)

T = TypeVar("T")


class ExcelParser:
    """Главный класс для парсинга Excel файлов.
    
    Поддерживает парсинг вертикальных (ключ-значение) и горизонтальных
    (табличных) листов с автоматическим преобразованием в dataclass модели.
    
    Example:
        >>> parser = ExcelParser()
        >>> data = parser.parse_vertical_sheet(
        >>>     "data.xlsx",
        >>>     "Info",
        >>>     DataModel
        >>> )
    """

    def __init__(self) -> None:
        """Инициализация парсера."""
        self.converter = TypeConverter()
        logger.debug("ExcelParser инициализирован")

    def register_converter(
        self,
        target_type: Type[T],
        converter_func: Any,
    ) -> None:
        """Регистрирует кастомный конвертер типов.
        
        Args:
            target_type: Целевой тип
            converter_func: Функция преобразования
        """
        self.converter.register(target_type, converter_func)

    def parse_vertical_sheet(
        self,
        file_path: str | Path,
        sheet_name: str,
        model_class: Type[T],
        key_column: int = 0,
        value_column: int = 1,
        start_row: int = 0,
        skip_empty_rows: bool = True,
    ) -> T:
        """Парсит вертикальный лист Excel (ключ-значение).
        
        Структура вертикального листа:
        | Ключ        | Значение   |
        |-------------|------------|
        | Название    | Значение1  |
        | Параметр    | 100.0      |
        
        Args:
            file_path: Путь к Excel файлу
            sheet_name: Имя листа
            model_class: Класс dataclass для маппинга
            key_column: Номер колонки с ключами (0-based)
            value_column: Номер колонки со значениями (0-based)
            start_row: Номер строки начала чтения (0-based)
            skip_empty_rows: Пропускать пустые строки
            
        Returns:
            Экземпляр model_class с данными из листа
            
        Raises:
            FileNotFoundError: Если файл не найден
            SheetNotFoundError: Если лист не найден
            ParsingError: Если ошибка парсинга
            ValidationError: Если ошибка валидации
        """
        config = VerticalSheetConfig(
            sheet_name=sheet_name,
            model_class=model_class,
            key_column=key_column,
            value_column=value_column,
            start_row=start_row,
            skip_empty_rows=skip_empty_rows,
        )
        
        return self._parse_vertical_sheet(file_path, config)

    def parse_horizontal_sheet(
        self,
        file_path: str | Path,
        sheet_name: str,
        model_class: Type[T],
        header_row: int = 0,
        start_row: int = 1,
        end_row: Optional[int] = None,
        skip_empty_rows: bool = True,
    ) -> List[T]:
        """Парсит горизонтальный лист Excel (табличный формат).
        
        Структура горизонтального листа:
        | ID | Название   | Значение |
        |----|------------|----------|
        | 1  | Элемент 1  | 100.0    |
        | 2  | Элемент 2  | 200.0    |
        
        Args:
            file_path: Путь к Excel файлу
            sheet_name: Имя листа
            model_class: Класс dataclass для маппинга
            header_row: Номер строки с заголовками (0-based)
            start_row: Номер строки начала данных (0-based)
            end_row: Номер последней строки (опционально)
            skip_empty_rows: Пропускать пустые строки
            
        Returns:
            Список экземпляров model_class
            
        Raises:
            FileNotFoundError: Если файл не найден
            SheetNotFoundError: Если лист не найден
            ParsingError: Если ошибка парсинга
            ValidationError: Если ошибка валидации
        """
        config = HorizontalSheetConfig(
            sheet_name=sheet_name,
            model_class=model_class,
            header_row=header_row,
            start_row=start_row,
            end_row=end_row,
            skip_empty_rows=skip_empty_rows,
        )
        
        return self._parse_horizontal_sheet(file_path, config)

    def parse_file(
        self,
        file_path: str | Path,
        config: ExcelFileConfig,
        result_class: Type[T],
    ) -> T:
        """Парсит Excel файл с несколькими листами.
        
        Args:
            file_path: Путь к Excel файлу
            config: Конфигурация файла
            result_class: Класс для результата (dataclass)
            
        Returns:
            Экземпляр result_class с данными из всех листов
            
        Raises:
            FileNotFoundError: Если файл не найден
            SheetNotFoundError: Если лист не найден
            ParsingError: Если ошибка парсинга
            ValidationError: Если ошибка валидации
        """
        file_path = Path(file_path)
        logger.info("Парсинг файла: %s", file_path)

        if not file_path.exists():
            raise FileNotFoundError(f"Файл не найден: {file_path}")

        result_data: Dict[str, Any] = {}

        for key, sheet_config in config.sheets.items():
            logger.debug("Парсинг листа '%s' с ключом '%s'", sheet_config.sheet_name, key)
            
            if isinstance(sheet_config, VerticalSheetConfig):
                result_data[key] = self._parse_vertical_sheet(file_path, sheet_config)
            elif isinstance(sheet_config, HorizontalSheetConfig):
                result_data[key] = self._parse_horizontal_sheet(file_path, sheet_config)
            else:
                raise ParsingError(
                    f"Неизвестный тип конфигурации: {type(sheet_config)}",
                    file_path=str(file_path),
                )

        # Создаём экземпляр result_class
        try:
            return result_class(**result_data)
        except TypeError as e:
            raise ValidationError(
                f"Не удалось создать экземпляр {result_class.__name__}: {e}"
            ) from e

    def _parse_vertical_sheet(
        self,
        file_path: str | Path,
        config: VerticalSheetConfig,
    ) -> Any:
        """Внутренний метод парсинга вертикального листа.
        
        Args:
            file_path: Путь к файлу
            config: Конфигурация листа
            
        Returns:
            Экземпляр model_class
        """
        file_path = Path(file_path)
        logger.info(
            "Парсинг вертикального листа '%s' из файла %s",
            config.sheet_name,
            file_path,
        )

        if not file_path.exists():
            raise FileNotFoundError(f"Файл не найден: {file_path}")

        # Загружаем workbook (подавляем предупреждения openpyxl о неподдерживаемых расширениях)
        with warnings.catch_warnings():
            warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
            workbook = load_workbook(filename=file_path, data_only=True)
        
        # Получаем лист
        if config.sheet_name not in workbook.sheetnames:
            raise SheetNotFoundError(config.sheet_name, str(file_path))
        
        worksheet = workbook[config.sheet_name]
        
        # Читаем данные в словарь ключ-значение
        data_dict = self._read_vertical_data(
            worksheet,
            config,
            str(file_path),
        )
        
        # Маппим на dataclass
        return self._map_to_dataclass(
            data_dict,
            config.model_class,
            config.sheet_name,
            str(file_path),
        )

    def _parse_horizontal_sheet(
        self,
        file_path: str | Path,
        config: HorizontalSheetConfig,
    ) -> List[Any]:
        """Внутренний метод парсинга горизонтального листа.
        
        Args:
            file_path: Путь к файлу
            config: Конфигурация листа
            
        Returns:
            Список экземпляров model_class
        """
        file_path = Path(file_path)
        logger.info(
            "Парсинг горизонтального листа '%s' из файла %s",
            config.sheet_name,
            file_path,
        )

        if not file_path.exists():
            raise FileNotFoundError(f"Файл не найден: {file_path}")

        with warnings.catch_warnings():
            warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
            workbook = load_workbook(filename=file_path, data_only=True)
        
        # Получаем лист
        if config.sheet_name not in workbook.sheetnames:
            raise SheetNotFoundError(config.sheet_name, str(file_path))
        
        worksheet = workbook[config.sheet_name]
        
        # Читаем данные
        rows_data = self._read_horizontal_data(
            worksheet,
            config,
            str(file_path),
        )
        
        # Маппим каждую строку на dataclass
        results = []
        for row_index, row_dict in enumerate(rows_data, start=config.start_row):
            try:
                instance = self._map_to_dataclass(
                    row_dict,
                    config.model_class,
                    config.sheet_name,
                    str(file_path),
                )
                results.append(instance)
            except (ParsingError, ValidationError) as e:
                # Добавляем информацию о строке
                if isinstance(e, ParsingError) and e.row is None:
                    e.row = row_index + 1  # +1 для Excel нумерации
                raise

        logger.info("Распарсено %d строк из листа '%s'", len(results), config.sheet_name)
        return results

    def _read_vertical_data(
        self,
        worksheet: Worksheet,
        config: VerticalSheetConfig,
        file_path: str,
    ) -> Dict[str, Any]:
        """Читает данные из вертикального листа в словарь.
        
        Args:
            worksheet: Лист Excel
            config: Конфигурация
            file_path: Путь к файлу
            
        Returns:
            Словарь ключ-значение
        """
        data_dict: Dict[str, Any] = {}
        
        for row_index, row in enumerate(worksheet.iter_rows(min_row=config.start_row + 1)):
            # Получаем ключ и значение
            key = row[config.key_column].value
            value = row[config.value_column].value
            
            # Пропускаем строки где и ключ и значение пустые
            if config.skip_empty_rows:
                if is_empty_value(key) and is_empty_value(value):
                    continue
            
            # Добавляем ключ даже если значение пустое (для корректной обработки EmptyValueError)
            if key is not None:
                key_str = clean_key(str(key))
                data_dict[key_str] = value
                logger.debug("Прочитана пара: '%s' = '%s'", key_str, value)

        return data_dict

    def _read_horizontal_data(
        self,
        worksheet: Worksheet,
        config: HorizontalSheetConfig,
        file_path: str,
    ) -> List[Dict[str, Any]]:
        """Читает данные из горизонтального листа.
        
        Args:
            worksheet: Лист Excel
            config: Конфигурация
            file_path: Путь к файлу
            
        Returns:
            Список словарей (каждая строка = словарь)
        """
        # Читаем заголовки
        header_row = list(worksheet.iter_rows(
            min_row=config.header_row + 1,
            max_row=config.header_row + 1,
        ))[0]
        
        headers = [clean_key(str(cell.value)) if cell.value else f"Column_{i}" 
                   for i, cell in enumerate(header_row)]
        
        logger.debug("Заголовки: %s", headers)
        
        # Читаем данные
        rows_data: List[Dict[str, Any]] = []
        
        max_row = config.end_row + 1 if config.end_row else None
        
        for row in worksheet.iter_rows(min_row=config.start_row + 1, max_row=max_row):
            # Пропускаем пустые строки
            if config.skip_empty_rows:
                if all(is_empty_value(cell.value) for cell in row):
                    continue
            
            # Создаём словарь для строки
            row_dict: Dict[str, Any] = {}
            for i, cell in enumerate(row):
                if i < len(headers):
                    row_dict[headers[i]] = cell.value
            
            rows_data.append(row_dict)
        
        return rows_data

    def _map_to_dataclass(
        self,
        data_dict: Dict[str, Any],
        model_class: Type[T],
        sheet_name: str,
        file_path: str,
    ) -> T:
        """Маппит словарь данных на dataclass.
        
        Args:
            data_dict: Словарь с данными
            model_class: Класс dataclass
            sheet_name: Имя листа
            file_path: Путь к файлу
            
        Returns:
            Экземпляр model_class
            
        Raises:
            ValidationError: Если не удалось создать экземпляр
        """
        if not is_dataclass(model_class):
            raise ValidationError(
                f"{model_class.__name__} не является dataclass",
            )

        field_values: Dict[str, Any] = {}

        # Обрабатываем каждое поле dataclass
        for field in fields(model_class):
            field_name = field.name
            field_type = field.type
            
            # Получаем конфигурацию поля из metadata
            field_metadata = field.metadata.get("excel_config")
            
            if field_metadata and isinstance(field_metadata, FieldConfig):
                # Используем конфигурацию из field_config
                excel_key = field_metadata.key or field_metadata.column or field_name
                
                # Проверяем наличие ключа в data_dict
                if excel_key not in data_dict:
                    # Ключ вообще не найден в листе
                    if field_metadata.required and field_metadata.default is None:
                        raise FieldNotFoundError(
                            excel_key,
                            sheet_name,
                            file_path,
                        )
                    value = field_metadata.default
                else:
                    # Ключ найден, получаем значение
                    value = data_dict.get(excel_key)
                    
                    # Проверяем не пустое ли значение
                    if is_empty_value(value):
                        if field_metadata.required and field_metadata.default is None:
                            # Значение пустое для обязательного поля
                            raise EmptyValueError(
                                field_name,
                                excel_key,
                                sheet_name,
                                file_path,
                            )
                        value = field_metadata.default
                
                # Применяем кастомный конвертер если есть
                if field_metadata.converter and value is not None:
                    value = field_metadata.converter(value)
                # Иначе стандартное преобразование типа
                elif value is not None:
                    value = self.converter.convert(value, field_type, field_name)
                
                # Применяем валидатор если есть
                if field_metadata.validator and value is not None:
                    if not field_metadata.validator(value):
                        raise ValidationError(
                            f"Значение '{value}' не прошло валидацию для поля '{field_name}'",
                            field_name=field_name,
                            value=value,
                        )
                
                field_values[field_name] = value
            else:
                # Без конфигурации - пробуем найти по имени поля
                value = data_dict.get(field_name)
                
                if value is not None:
                    value = self.converter.convert(value, field_type, field_name)
                
                field_values[field_name] = value

        # Создаём экземпляр
        try:
            return model_class(**field_values)
        except TypeError as e:
            raise ValidationError(
                f"Не удалось создать экземпляр {model_class.__name__}: {e}",
            ) from e
