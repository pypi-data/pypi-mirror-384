import sys
import PySide2.QtWidgets as qtw
import PySide2.QtCore as qtc
import PySide2.QtGui as qtg
import PySide2.QtSvg as qsvg
import math
import time
import webbrowser
import colorsys
import weakref
import enum
import re
import openpyxl as pxl
import csv
import os
import json
import traceback
import networkx as nx
import numpy as np
import gc

import matplotlib
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg
from matplotlib.figure import Figure

from Bio import SeqIO, SeqFeature, Seq

import threading

from . import dbdl
from . import dframe
from . import about

DEBUG = False

#

NEI_MOVEMENT_MODIFIER = qtc.Qt.ALT
#NEI_DRAG_MODIFIER = False
NEI_ADD_SELECT_AND_DRAG_MODIFIER = qtc.Qt.CTRL
#NEI_REPLACE_SELECT_MODIFIER = False

NEI_REPLACE_SELECT_AND_DRAG_MODIFIER = False
NEI_REGION_SELECT_MODIFIER = qtc.Qt.SHIFT
NEI_HOLD_ACTIVATE_DELAY = 350

# Import PySide2, the GUI library we're using

# DPI scaling adjustment, otherwise text will be too small on large screens
#   and too big on small screens -- without the rest of the UI scaling to match
if hasattr(qtc.Qt, 'AA_EnableHighDpiScaling'):
    qtw.QApplication.setAttribute(qtc.Qt.AA_EnableHighDpiScaling, True)

if hasattr(qtc.Qt, 'AA_UseHighDpiPixmaps'):
    # Not that there are any :)
    qtw.QApplication.setAttribute(qtc.Qt.AA_UseHighDpiPixmaps, True)

# GUI classes relying on templates
#   Please note that many parts of the UI are generated without
#   a template, or may not even have their own class.
from .uilayouts import resources
from .uilayouts.importframe import Ui_ImportFrame
from .uilayouts.filewidget import Ui_FileWidget
from .uilayouts.graphframe import Ui_GraphFrame
from .uilayouts.constraintsframe import Ui_ConstraintsFrame
from .uilayouts.accsetblock import Ui_AccsetBlock
from .uilayouts.neiview import Ui_NeiView
from .uilayouts.wrapperwindow import Ui_MainWindow
#from .uilayouts.proteininfo import Ui_ProteinInfo
#from .uilayouts.clusterinfo import Ui_ClusterInfo
from .uilayouts.neiview_creator import Ui_NeiviewCreator
#from .uilayouts.taxoninfo import Ui_TaxonInfo
#from .uilayouts.regioninfo import Ui_RegionInfo
from .uilayouts.selector_widget import Ui_SelectorWidget
from .uilayouts.blastconfig import Ui_BlastConfig
from .uilayouts.dbprogressbar import Ui_DBProgressBar
from .uilayouts.excelexportframe import Ui_ExcelExportFrame
from .uilayouts.introwindow import Ui_IntroWindow
from .uilayouts.styledictcreator import Ui_StyledictCreator
from .uilayouts.proteingrouptable import Ui_ProteinGroupTable
from .uilayouts.selection_manager import Ui_SelectionManager
from .uilayouts.info_feature import Ui_InfoFeature
from .uilayouts.info_region import Ui_InfoRegion
from .uilayouts.info_sequence  import Ui_InfoSequence
from .uilayouts.info_taxon  import Ui_InfoTaxon
from .uilayouts.info_protein  import Ui_InfoProtein
from .uilayouts.info_cluster  import Ui_InfoCluster
from .uilayouts.taginfowidget import Ui_TagInfoWidget
from .uilayouts.neiconfig import Ui_NeiConfig
from .uilayouts.neiexport import Ui_NeiExport
from .uilayouts.colocresultsframe import Ui_ColocResultsFrame
from .uilayouts.about_neiview import Ui_AboutNeiview
from .uilayouts.ClusterNetworkViewer2 import Ui_ClusterNetworkViewer2
from .uilayouts.NeiviewLegend import Ui_NeiviewLegend

#Debug only
from collections import Counter
import code
import random
#code.interact(local=locals())

matplotlib.use("Qt5Agg")

class ExceptionHandler(qtc.QObject):
    sig_generic_error = qtc.Signal(object, object, object)
    sig_no_alignments = qtc.Signal()
    def __init__(self):
        super().__init__()
        sys.excepthook = self.exception_hook
        
        self.sig_generic_error.connect(self.on_general_exception)
        self.sig_no_alignments.connect(self.no_alignments_error)
    def exception_hook(self, error_type, value, tb):
        self.sig_generic_error.emit(error_type, value, tb)
    def on_general_exception(self, error_type, value, tb):
        print("".join(traceback.format_exception(error_type, value, tb)))
        msg = (
            f"<html><head/><body><p>The program encountered a {error_type}. Some errors may be harmless, but "
            "regardless, it is recommended to save under a different filename and restart CluSeek. </p>"
            "<p>We'd appreciate it if you'd let us know at <a href='mailto:cluseek@biomed.cas.cz'>cluseek@biomed.cas.cz</a> and include a "
            "description of how you encountered this issue and a copy of the "
            "error details shown when you click <b>Show Details</b> in this error message.</p></body></html>")
        
        
        messagebox = qtw.QMessageBox(
            qtw.QMessageBox.Critical,
            "Exception encountered",
            msg,
            qtw.QMessageBox.Ok)
        messagebox.setTextFormat(qtc.Qt.RichText)
        
        exit_btn = messagebox.addButton("Exit CluSeek", messagebox.DestructiveRole)
        
        messagebox.setDetailedText("".join(traceback.format_exception(error_type, value, tb)))
        
        messagebox.exec()
        
        if messagebox.clickedButton() is exit_btn:
            if AppRoot.introwindow.passed_control_to_main_window:
                AppRoot.mainwindow.close()
            else:
                AppRoot.introwindow.close()
    def on_work_thread_exception(self, exception_msg):
        print("lol")
        print(exception_msg)
        msg = (
            "<html><head/><body><p>The program encountered an error during the data processing pipeline. "
            "As these errors are usually unrecoverable or damage the final results, "
            "<b>it is strongly recommended to exit CluSeek and try again</b>. </p>")
        if "error message from ncbi" in exception_msg.lower():
            msg += (
                "<p><br>"
                "<b>NOTE:</b> It appears this error was caused by an incorrect input "
                "being submitted to the NCBI servers. "
                "<b>This is likely a user error!</b> Check your inputs "
                "AND your BLASTp configuration! </p>")
        msg += (
            "<p>Please contact us at <a href='mailto:cluseek@biomed.cas.cz'>cluseek@biomed.cas.cz</a> and include a "
            "description of how you encountered this issue and a copy of the "
            "error details shown when you click <b>Show Details</b> in this error message.</p></body></html>")
        
        messagebox = qtw.QMessageBox(
            qtw.QMessageBox.Critical,
            "Exception encountered in work thread",
            msg,
            qtw.QMessageBox.Ok)
        messagebox.setTextFormat(qtc.Qt.RichText)
        
        exit_btn = messagebox.addButton("Exit CluSeek", messagebox.DestructiveRole)
        
        messagebox.setDetailedText(exception_msg)
        
        messagebox.exec()
        
        if messagebox.clickedButton() is exit_btn:
            if AppRoot.introwindow.passed_control_to_main_window:
                AppRoot.mainwindow.close()
            else:
                AppRoot.introwindow.close()
    def no_alignments_error(self):
        qtw.QMessageBox.warning(None, "BLAST result contains no alignments",
            "There were no homologs found in a BLAST search!"
                "It is recommended to restart your analysis and "
                "review your inputs and BLAST search conditions.")
class WorkerSignals(qtc.QObject):
    progress_manager_update = qtc.Signal(object, str, int)
    result = qtc.Signal(object)
    finished = qtc.Signal()
    failed = qtc.Signal(object)
class WorkerThread(qtc.QThread):
    def __init__(self, func, *args, **kwargs):
        super().__init__()
        self.func = func
        self.args = args
        self.kwargs = kwargs
        self.signals = WorkerSignals()
        self.setTerminationEnabled(True)
        
        self.finished.connect(self.cbk_finished)
        self.signals.failed.connect(
            AppRoot.exception_handler.on_work_thread_exception)
    @qtc.Slot()
    def run(self):
        try:
            result = self.func(*self.args, **self.kwargs)
        except Exception as exception:
            self.signals.failed.emit(traceback.format_exc())
            #raise exception
        else:
            self.signals.result.emit(result)
            self.signals.finished.emit()
    def safe_start(self):
        assert AppRoot.work_thread is None, "Attempted to launch two simultaneous work threads"
        AppRoot.work_thread = self
        self.start()
    def cbk_finished(self):
        AppRoot.work_thread = None
        print("Work thread finished!")

class ProgressDialog(qtw.QDialog, Ui_DBProgressBar):
    class DotTimer(qtc.QObject):
        def __init__(self, master):
            super().__init__()
            self.timer_id = None
            self.master = master
        def start(self):
            self.stop()
            self.timer_id = self.startTimer(750)
        def stop(self):
            if self.timer_id:
                self.killTimer(self.timer.timer_id)
                self.timer_id = None
        def timerEvent(self, event):
            self.master.cbk_on_tick()
    def __init__(self, title, label_text, minimum=0, maximum=100, on_abort=None,
                 text_only=False):
        super().__init__()
        self.init_ui()
        self.reinitialize(title, label_text, minimum, maximum, on_abort,
                          text_only)
        self.timer = self.DotTimer(self)
        self.timer.start()
        self.trailing_dots = ""
        self.keep_abort_btn_hidden = True
    def init_ui(self):
        self.setupUi(self)
        self.abort_btn.clicked.connect(self.cbk_abort)
    def reinitialize(self, title, label_text, minimum=0, maximum=100, 
                     on_abort=None, text_only=False):
        # Static defaults
        self.canceled = False
        self.progress_pgbr.reset()
        self.setModal(True)
        
        self.trailing_dots = ""
        
        # Parameter-defined values
        self.setWindowTitle(title)
        self.setLabelText(label_text)
        self.progress_pgbr.setMaximum(maximum)
        self.progress_pgbr.setMinimum(minimum)
        self.on_abort = on_abort
        if text_only:
            self.abort_btn.setHidden(True)
            self.progress_pgbr.setHidden(True)
        else:
            self.abort_btn.setHidden(True if self.keep_abort_btn_hidden else False)
            self.progress_pgbr.setHidden(False)
        
        self.show()
        self.activateWindow()
            
    def setLabelText(self, text):
        self.label_text = text
        self.info_lbl.setText(text)
    def cbk_abort(self):
        self.setLabelText("Aborting (Please be patient)")
        self.canceled = True
        self.on_abort()
    def wasCanceled(self):
        return(self.canceled)
    
    def setMaximum(self, val):
        self.progress_pgbr.setMaximum(val)
    def setMinimum(self, val):
        self.progress_pgbr.setMinimum(val)
    def setValue(self, val):
        self.progress_pgbr.setValue(val)
    def closeEvent(self, event):
        # We do not allow the users to close the window
        #   on their own. They have to sit and wait like
        #   good little users, so the app doesn't segfault.
        pass
        #event.ignore()
        #self.show()
    def hideEvent(self, event):
        print("Hiding loading bar...")
    def on_abort(self):
        #To be overidden
        pass
    def cbk_on_tick(self):
        self.trailing_dots += "."
        if len(self.trailing_dots) > 3:
            self.trailing_dots = ""
        self.info_lbl.setText(self.label_text+" "+self.trailing_dots)
#other

def add_question_mark_icon(button):
        button.setIcon(
            button.style().standardIcon(qtw.QStyle.SP_MessageBoxQuestion))

def m_clear_layout(layout):
    #Adapted from qt docs
    #TODO: Can this spawn infinite recursion? God I hope not.
    while layout.count():
        child = layout.takeAt(0)
        if child.widget():
            child.widget().deleteLater()

#
 # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #
 #                                                                   #
 #                         Abstract Classes                          #
 #                                                                   #
 # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #


class AppRoot():
    #This is effectively a container of global variables
    #  in the guise of a class definition.
    #  When you instantiate this class, the constructor modifies
    #  the base class (rather than the instance) to initialize all 
    #  the bells and whistles needed by other parts of the application.
    #
    #  AppRoot is called at will from anywhere within this module, to
    #  provide access to resources which there should never be more
    #  than one of, eg. database managers or Qt's QApplication class
    #  this is also why we use the base class instead of an instance
    #  (though an instance should work just as well for most purposes)
    #
    #  The reason we need to instantiate it first is simply because
    #  we don't want the module to start accessing the hard drive or
    #  building files the moment it is imported -- rather, it has to 
    #  be told to do so explicitly when we are ready for it.
    #
    # TODO: Re-implement this in a way that won't get me locked up
    #       for programming crimes. Please.
    @classmethod
    def __init__(cls):
        #cls.qtapp
        #cls.ui_topclusters = None
        #cls.ui_importmanager = None
        #cls.ui_constraintsframe = None
        #cls.ui_neighborhoodviewer = None
        #cls.ui_infowindow = None
        #cls.introwindow = None
        
        #Colors. Yes really.
        cls.active_color = None
        #TODO: Is this actually being used though?
        #Timer
        cls.timer = TimerHandler()
        cls.work_thread = None
        cls.progress_dialog = None
        
        cls.exception_handler = ExceptionHandler()
        dbdl.ERROR_MANAGER = cls.exception_handler
    @classmethod
    def launch(cls):
        print("Good morning!")
        cls.qtapp = qtw.QApplication(sys.argv)
        cls.mainwindow = MainWindow()
        cls.introwindow = IntroWindow()
        cls.introwindow.show()
        cls.introwindow.activateWindow()
        cls.neiexport = NeiExport()
        
        cls.qtapp.exec_()
    @classmethod
    def on_app_start(cls):
        # Create progress dialog
        cls.progress_dialog = ProgressDialog(title="Launching...", 
                                             label_text="Please wait", 
                                             text_only=True)
        cls.progress_dialog.show()
        cls.progress_dialog.activateWindow()
        
        def on_db_init_complete():
            # * * Run other tasks * *
            # Making a decorator doesn't quite work.
            # It's not as explicit, but it works for now.
            #UIProteinCluster._init_color_manager()
            #ColorableButton._load_icons()
            cls.progress_dialog.setHidden(True)
        worker_db_init = WorkerThread(cls._on_app_start)
        worker_db_init.signals.finished.connect(on_db_init_complete)
        
        worker_db_init.safe_start()
        
    @classmethod
    def _on_app_start(cls):
        cls.dbman = dbdl.DatabaseManager()
        cls.mainwindow.init_dbman_link()
        cls.cluster_manager = dbdl.ManagerClus()
    @classmethod
    def exit_application(cls, dontask=False):
        if cls.introwindow.passed_control_to_main_window:
            if not dontask:
                question = qtw.QMessageBox()
                question.setText("Save project before exiting?")
                question.setInformativeText("Unsaved changes may be lost.")
                question.setStandardButtons(
                    qtw.QMessageBox.Save | qtw.QMessageBox.Discard |
                    qtw.QMessageBox.Cancel)
                question.setDefaultButton(qtw.QMessageBox.Cancel)
                
                answer = question.exec_()
                
                if answer == qtw.QMessageBox.Save:
                    try:
                        AppRoot.mainwindow.save()
                    except:
                        # If the save failed, do not exit.
                        qtw.QMessageBox.critical(
                            None, 
                            "Error saving project on exit",
                            "The application will not exit because an error "
                            "occurred while it was trying to save your data.\n"
                            "Try saving your data again. Who knows, you might get lucky.")
                        return(False)
                elif answer == qtw.QMessageBox.Discard:
                    pass
                elif answer == qtw.QMessageBox.Cancel:
                    return(False)
                else:
                    assert False, "Invalid return dialog code."
        
        time.sleep(2)
        
        print("- - - Good night! - - -")
        if cls.work_thread:
            # Ew, but if the user wants to corrupt their data,
            #   c'est la vie. Better than the work thread staying
            #   alive in the background.
            cls.work_thread.terminate()
        # Can't say I understand how this doesn't trigger a closeEvent
        #   loop, but it seems to work.
        cls.qtapp.closeAllWindows()
        return(True)
    
    @classmethod
    def check_dead(cls, obj):
        weakrefset = weakref.WeakSet([obj])
        del obj
        
        def func():
            gc.collect()
            
        cls.timer.after_delay_do(func, 5000)
    
    @classmethod
    def show_errormessage(cls, message):
        message.exec()
class TimerHandler(qtc.QObject):
    # Q: Why does this exist?
    # A: Sometimes, we need to pass control over to
    #    Qt's loop for a bit to give it time to render --
    #    either because we're working on a time-intensive
    #    task, or because we need the geometry rendered 
    #    before modifying it further.
    def __init__(self):
        super().__init__()
        self.timers = {}
        self.timerlock = threading.Lock
    def after_delay_do(self, callback, delay=10):
        # Start the timer and get its unique ID
        _id = self.startTimer(delay)
        
        assert _id != 0, "Timer failed to start for some reason."
        
        #Store the callback under the event's id
        self.timers[_id] = callback
        
        return(_id)
    def kill_timer(self, _id):
        del self.timers[_id]
        self.killTimer(_id)
    def timerEvent(self, event):
        # Save the event ID and callback function 
        #   in local variables
        _id = event.timerId()
        callback = self.timers[_id]
        
        # Clear the callback and stop the timer
        del self.timers[_id]
        self.killTimer(_id)
        
        # Execute the callback under the timer's ID
        # If this fails, the exception is raised, but we
        #   have already stopped the timer and cleared
        #   the callback, so the TimerHandler doesn't care.
        #   If we hadn't, the timer event would keep
        #   triggering over and over, spamming exceptions.
        callback()

class VariableTypes(enum.Enum):
    # Expect value
    EVALUE = "evalue"
    
    # Num identical aminoacids divided by length of hit sequence
    IDENTITY_PER_HIT = "identity_per_hit"
    
    # Num identical aminoacids divided by length of query sequence
    IDENTITY_PER_QUERY = "identity_per_query"
    
    # Bitscore
    BITSCORE = "bitscore"
    
    # Length of alignment divided by length of query sequence
    #ALIGNMENT_PER_QUERY = "alignment_per_query"
    
    # Length of alignment divided by length of hit sequence
    #ALIGNMENT_PER_HIT = "alignment_per_hit"

class Dataset():
    class Rules():
        def __init__(self, dataset):
            self.dataset = dataset
            self.alignment_rules = {}
            self.sequence_rules = {}
            self.minimum_sequence_score = None
            for accset_name in self.dataset.accsets:
                self.alignment_rules[accset_name] = dict()
        def reset_sequence_rules(self):
            self.sequence_rules = {}
        def wipe_alignment_rule(self, accset_name, variable):
            if variable in self.alignment_rules[accset_name]:
                del self.alignment_rules[accset_name][variable]
        def new_alignment_rule(self, accset_name, variable, minimum=None, maximum=None, *args, **kwargs):
            rule = self.AlignmentRule(variable=variable, 
                                      minimum=minimum, 
                                      maximum=maximum)
            self.alignment_rules[accset_name][variable] = rule
        def get_alignment_rule(self, accset_name, variable):
            return(self.alignment_rules[accset_name].get(variable))
        def new_sequence_rule(self, accset_name, on_true):
            rule = self.SequenceRule(accset_name=accset_name, on_true=on_true)
            self.sequence_rules[accset_name] = rule
        def wipe_sequence_rule(self, accset_name):
            del self.sequence_rules[accset_name]
        def get_sequence_rule(self, accset_name):
            return(self.sequence_rules.get(accset_name))
        def set_minimum_sequence_score(self, min_score):
            self.minimum_sequence_score = min_score
        class Rule():
            # This is just a base class to identify rules of any kind.
            def __init__(self):
                pass
            def evaluate(self):
                pass
        class AlignmentRule(Rule):
            def __init__(self, variable, minimum, maximum, *args, **kwargs):
                super().__init__(*args, **kwargs)
                assert variable in VariableTypes, ("Invalid variable value. "
                                                    "Must be a VariableTypes enum.")
                self.variable = variable
                self.min = minimum # Inclusive
                self.max = maximum # Inclusive
            def evaluate(self, alignment):
                if self.variable is VariableTypes.EVALUE:
                    value = alignment.tophsp.expect
                elif self.variable is VariableTypes.IDENTITY_PER_QUERY:
                    value = alignment.tophsp.identities / alignment.record.query_length
                elif self.variable is VariableTypes.IDENTITY_PER_HIT:
                    value = alignment.tophsp.identities / alignment.length
                
                # We test the min max, and if neither fails, the rule succeeds.
                #   It shouldn't be possible for there to be neither min nor max,
                #   but technically it's not even such an issue.
                if self.min:
                    if value < self.min:
                        return(False)
                if self.max:
                    if value > self.max:
                        return(False)
                return(True)
        class SequenceRule(Rule):
            def __init__(self, accset_name, on_true):
                self.accset_name = accset_name
                self.on_true = on_true
    def __init__(self):
        #General
        self.names   = []
        self.accsets = {}
        self.blasts  = {}
        self.aliases = {}
        
        #Max bp between edges of features
        self.max_feature_distance = 60000
        #Max bp added to the edges
        self.border_size = 60000
        
        self.rules   = None
        self.root    = None
        self.subset  = None
        
        #The global alignment clustering result
        self.homo_clustering = None
        #The local alignment clustering result
        self.heth_clustering = None
    # * Primary Pipeline
    def load_files(self, handles):
        # TODO: STARTHERE: make input/output align
        #ftype = None
        #Load XMLs
        records = dbdl.load_blastp_xmls(handles)
        for name in records:
            record = records[name]
            accset = set(record.als)
            
            self.names.append(name)
            self.accsets[name] = accset
            self.blasts[name] = record
            self.aliases[name] = record.query
        self.rules = self.Rules(self)
        self.build_root()
    def build_root(self):
        self.root = None
        allaccessions = []
        for accset in self.accsets.values():
            allaccessions.extend(accset)
        self.root = dbdl.build_data_tree(
            AppRoot.dbman.obtain_ipgs(allaccessions))
    # The following functions are part of a sequence
    #   sometimes I don't need to run the whole sequence,
    #   but regardless of where you stop, you should start
    #   from the top.
    def proc1_create_subsets(self):
        self.root.scAll
        accsubsets = {}
        for accset_name in self.accsets:
            rules = [x for x in self.rules.alignment_rules[accset_name].values()]
            if not rules:
                accsubsets[accset_name] = set(self.blasts[accset_name].als)
            else:
                accsubsets[accset_name] = set()
            for alignment in self.blasts[accset_name].alignments:
                if all([rule.evaluate(alignment) for rule in rules]):
                    accsubsets[accset_name].add(alignment.accession)
        self.subset = {"accsets": accsubsets}
    def proc2_extend_subset_using_ipg(self):
        not_found = 0
        added = 0
        for accsubset in self.subset["accsets"].values():
            to_add = set()
            for accession in accsubset:
                try:
                    to_add.update(self.root.ptAll[accession].ipt.pts)
                except KeyError:
                    not_found += 1
            accsubset.update(to_add)
            added += len(to_add)
    def proc3_find_target_regions(self, region_filter=None):
        errors = Counter()
        
        # * We tag the marker proteins
        # First purge old marker tags:
        for pt in self.root.ptAll.values():
            if hasattr(pt, "marker"):
                pt.marker = set()
        # Then tag 'em anew:
        for accsubset_name in self.subset["accsets"]:
            accsubset = self.subset["accsets"][accsubset_name]
            for accession in accsubset:
                #TODO:STARTHERE: Account for sequence similarity between markers.
                #NOTE: Not sure if this is still relevant.
                try:
                    if hasattr(self.root.ptAll[accession], "marker"):
                        self.root.ptAll[accession].marker.add(accsubset_name)
                    else:
                        self.root.ptAll[accession].marker = {accsubset_name}
                except KeyError:
                    errors["Marker accession not found in root"] += 1
        
        hits = []
        
        
        for sc in self.root.scAll.values():
            extended_at_least_once = False
            errors["Sequences checked"] += 1
            # * Sort features by position
            features = sorted([ft for ft in sc.fts.values()],
                              reverse=False,
                              key=lambda ft: ft.start)
            
            # Keeps the score of a window. 
            #   If multiple windows merge, keeps best score from among them.
            last_window_best_score = 0
            window = []
            last_window = None
            hits_in_sc = []
            
            # We do a shifting window that is self.max_feature_distance large
            done = False
            while not done:
                # * Update window
                # add newest features
                errors["Windows checked"] += 1
                
                # As long as there's features left in the window,
                #   delete features from the left until we can add at least
                #   the next feature.
                while len(window) > 0 and ((features[0].stop - window[0].start) > self.max_feature_distance):
                    del window[0]
                
                added = 0
                # Fill the window up to capacity by adding features to the right
                while len(features)>0 and (len(window)==0 or ((features[0].stop - window[0].start) <= self.max_feature_distance)):
                    window.append(features[0])
                    del features[0]
                    added += 1
                
                # Flip the done flag at the end of the adding process
                #   if applicable
                if len(features) == 0:
                    done = True
                
                # The scoring variable
                passed = {}
                # Each key corresponds to one accset.
                #   The values are the scores assigned for that given accset.
                #   If multiple homologues from the same accset are present in
                #   a window, the score value is added only once.
                
                # Check rules
                passed_features = set()
                for feature in window:
                    if hasattr(feature.ref, "marker"):
                        for accsubset_name in feature.ref.marker:
                            if accsubset_name in self.rules.sequence_rules:
                                passed_features.add(feature)
                                passed[accsubset_name] = self.rules.sequence_rules[accsubset_name].on_true
                
                # If not all rules passed, just ignore this window and keep going.
                # OLD:
                #if not all(passed.values()): errors["Windows with no hits"] += 1; continue
                # NEW: The final score is the sum of all assigned scores.
                score = sum(passed.values())
                
                score_passed = not (score < self.rules.minimum_sequence_score)
                
                windows_overlap = last_window and window[0].start < last_window[1]
                
                # The following conditionals are order-sensitive 
                #   (don't move them around)
                
                # Note that
                #   last_window is implicit in windows_overlap
                if score_passed and windows_overlap:
                    # NOTE: I'm explicitly duplicating the int values in the
                    #   following section, because some of them appear to have
                    #   been getting kept, causing issues.
                    # Then extend the last_window with current one
                    last_window[1] = int(window[-1].stop)
                    errors["Extended window due to window hit overlap."] += 1
                    extended_at_least_once = True
                    if score > last_window_best_score:
                        last_window_best_score = score
                    
                    window_extended = True
                else:
                    window_extended = False
                
                if last_window and not window_extended:
                    # Add last_window
                    hits_in_sc.append((int(last_window[0]), 
                                       int(last_window[1]), 
                                       int(last_window_best_score),
                                       set(last_window_passed_features)))
                    errors["Non-overlapping hit in same sequence as an existing hit"] += 1
                
                if score_passed and not window_extended:
                    # Replace last_window with current window (if it passed)
                    last_window_passed_features = passed_features
                    last_window_best_score = score
                    last_window = [int(window[0].start), int(window[-1].stop)]
                if not score_passed:
                    # Replace last_window with nothing (if it failed)
                    last_window = None
                
                if score_passed:
                    errors["Windows with hits"] += 1
                else:
                    errors["Windows with no hits"] += 1
                
                    
            # Add the last window
            if last_window:
                hits_in_sc.append((int(last_window[0]), 
                                   int(last_window[1]), 
                                   int(last_window_best_score),
                                   set(passed_features)))
            
            
            errors[f"Sequences with {len(hits_in_sc)} hits"] += 1
            
            
            if extended_at_least_once:
                errors[f"Sequences with {len(hits_in_sc)} hits that had at least one extension"] += 1
            # Generate target regions
            for window_start,window_stop,window_score,passed_features in hits_in_sc:
                # We are ignoring the above values and redefining
                #   the size of region based on actual positions of markers,
                #   not the rolling window that identified it.
                # TODO: Maybe also recompute score to account for
                #       window extensions. Ugh, this needs an overhaul.
                try:
                    start = min(ft.start for ft in passed_features)
                    stop = max(ft.stop for ft in passed_features)
                except ValueError as e:
                    print(window_start, window_stop, window_score, passed_features)
                    raise e
                hits.append(dframe.GeneticRegion(self, sc, start, stop,
                                                 hit_score=max(score, last_window_best_score)))
            
        
        # * Cull identical NZ_ variants
        #culled = []
        #accdict_ = {}
        #for hit in hits:
        #    acc = hit.sc.accession.lstrip("NZ_")
        #    
        #    if acc in accdict_:
        #        hit2 = accdict_[acc]
        #        print("accmatch", hit.start, hit2.start)
        #        if hit.start == hit2.start and hit.length == hit2.length:
        #            #toremove.append(hit)
        #            print("fullmatch")
        #            pass
        #        else:
        #            culled.append(hit)
        #    else:
        #        accdict_[acc] = hit
        #        culled.append(hit)
        #print("NZ_ dereplication done:", len(hits), len(culled))
        #hits = culled
        #del culled
        #del accdict_
        
        # Get contig NZ_ variants
        #nz_variants = set()
        #for hit in hits:
        #    if hit.sc.accession.startswith("NZ_"):
        #        nz_variants.add(hit.sc.accession.lstrip("NZ_"))
        ## Purge non-NZ_ variants
        #purged = 0
        #for hit in hits[::-1]:
        #    if hit.sc.accession in nz_variants:
        #        purged += 1
        #        hits.remove(hit)
        #del nz_variants
        
        # * *  Dereplication culling
        filter_type,use_wgs,ignore_unknown_taxa = region_filter
        
        if filter_type != "nothing":
            # Pseudo generator for empty fields we don't want to match
            #   with anything else.
            _invalid_i = [-1]
            def invalid():
                _invalid_i[0] += 1
                return(_invalid_i[0])
            def get_discriminant(hit):
                genus = hit.sc.tx.genus
                species = hit.sc.tx.species
                strain = hit.sc.tx.strain
                
                if species and species.lower() == "sp.":
                    species = None
                
                if ignore_unknown_taxa:
                    if (filter_type == "strain"
                      and not (genus and species and strain)):
                        return(None)
                    elif (filter_type == "species"
                      and not (genus and species)):
                        return(None)
                    elif (filter_type == "genus"
                      and not genus):
                        return(None)
                
                if genus and species and strain:
                    d = (genus, species, strain)
                elif genus and species and not strain:
                    d = (genus, species, invalid())
                elif genus and not species and strain:
                    # If we don't know the species, we'll match
                    #   by strain at both species and strain level.
                    d = (genus, strain, strain)
                elif genus and not species and not strain:
                    d = (genus, invalid(), invalid())
                else:
                    d = (invalid(), invalid(), invalid())
                
                
                if filter_type == "strain":
                    return(d)
                elif filter_type == "species":
                    return(d[0:2])
                elif filter_type == "genus":
                    return(d[0:1])
                else:
                    raise ValueError("Invalid filter type.")
            def pick_best_hit_from_dereplicated_set(results):
                # We do several rounds of sorting. If there isn't
                #   a clear winner in one, we move on to the next.
                #   Starting with meaningful metrics, and moving on
                #   to the arbitrary to at least preserve determinism.
                
                if len(results) == 1:
                    return(results[0])

                # Results[ Result{'discriminant':"", 'hits':[hit, hit hit]} ]

                sorting_functions = [
                    ("number of clusters", 0, lambda result: len(result["hits"])),
                    #("overall length", 0, lambda result: sum([hit.length() for hit in result["hits"]])),
                    ("reciprocal sum of start positions", 0, lambda result: 1/sum([hit.start for hit in result["hits"]])),
                    # We replace NZ_ with whitespace to deprioritize refseq sequences.
                    ("accession code", "", lambda result: "".join(sorted([hit.sc.accession.replace("NZ_"," ") for hit in result["hits"]])))
                ]
                
                for metric_name, default, sorting_function in sorting_functions:
                    best_score = default
                    best_result = None
                    failed = False
                    for result in results:
                        score = sorting_function(result)
                        if score > best_score:
                            best_result = result
                            best_score = score
                        elif score == best_score:
                            failed = True
                            break
                    if not failed:
                        return(best_result)
                
                print("\tWARNING: Best replicate algorithm failed, choice is arbitrary")
                print(results)
                return(results[0])

            results_to_filter = []
            # Group hits into result dicts
            if use_wgs:
                # Group by wgs run if applicable
                wgs_pattern = re.compile(
                    r"(?P<wgs_run>[A-Z]{4,6}[0-9]{2})(?P<contig_id>[0-9]{6,99})")
                wgs_dict = {}
                
                for hit in hits:
                    wgs_name = re.match(wgs_pattern, hit.sc.accession.split("_")[-1])
                    # If wgs_name is None, there is no wgs run and it
                    #   should be treated as a wgs run unto itself
                    if wgs_name is not None:
                        if wgs_name not in wgs_dict:
                            wgs_dict[wgs_name] = {
                                "discriminant": get_discriminant(hit),
                                "hits": []}
                        wgs_dict[wgs_name]["hits"].append(hit)
                    else:
                        results_to_filter.append({
                            "discriminant": get_discriminant(hit),
                            "hits": [hit]})
                results_to_filter.extend(wgs_dict.values())
                del wgs_dict
            else:
                for hit in hits:
                    results_to_filter.append({
                        "discriminant": get_discriminant(hit),
                        "hits": [hit]})
            
            # Group by discriminant
            results_by_discriminant = {}
            for result in results_to_filter:
                if result["discriminant"] is None:
                    continue
                if result["discriminant"] not in results_by_discriminant:
                    results_by_discriminant[result["discriminant"]] = []
                results_by_discriminant[result["discriminant"]].append(result)
            
            # Pick one result with most hits per discriminant
            dereplicated = []
            for results in results_by_discriminant.values():
                dereplicated.extend(pick_best_hit_from_dereplicated_set(results)["hits"])
            
        else:
            dereplicated = hits
        
        # * Finalize the results and save them.
        self.subset["hits"] = dereplicated
        self.subset["pre-dereplication"] = hits
        errors["Hits dereplicated"] = len(hits) - len(dereplicated)
        errors["Total hits found"] = len(dereplicated)
        for key in errors:
            print(f"\t{key}: {errors[key]}")
        print("")
            
    def update_scs_with_gbdata(self, regions, borders):
        results = AppRoot.dbman.obtain_neighborhood_gbs(regions,
                                                        border_size=borders)
        regdict = {(region.sc.accession, region.start, region.stop): region \
                    for region in regions}
        
        # Read the results and attach them to the contigs
        t0 = time.time()
        times = []
        t1 = time.time()
        for result in results:
            key=(result["scacc"], result["regstart"], result["regstop"])
            # We pass the region as well so it can receive the sequence
            #   of bases along its length and correct its stop position.
            dbdl.update_scs_with_gbdata(result, regdict[key], 
                                        self.root, types_=["cds"])
            t2 = time.time()
            times.append(t2 - t1)
            t1 = t2
pass
#
 # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #
 #                                                                   #
 #                         GUI Classes                               #
 #                                                                   #
 # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #

# Override some classes for file opening
class DatabaseManagerUI(dbdl.DatabaseManager):
    def __init__(self):
        super().__init__()
    def on_error_invalid_path(self, reason=None):
        print("Error: Invalid path")
        qtw.QMessageBox.warning(None,
            "Invalid path",
            "CluSeek could not find the path to the specified file.")
    def on_error_loading_file(self, reason=None):
        print("Error: Failed to load file")
        if reason is None:
            reason = ""
        qtw.QMessageBox.warning(None, "Failed to load file", reason)
    def on_error_different_version(self, reason=None):
        print("Warning: Different version")
        qtw.QMessageBox.information(None,
            "Different version",
            "File was created using different version of CluSeek\nYou may experience issues as a result.")
    def on_error_file_already_exists(self, reason=None):
        qtw.QMessageBox.warning(None, "Could not overwrite file", 
                                "Please check read/write permissions")
    def on_error_is_not_file(self, reason=None):
        qtw.QMessageBox.warning(None, "No file specified", 
                                      "No file was specified.")
    
dbdl.DatabaseManager = DatabaseManagerUI

class IntroWindow(qtw.QWidget):
    def __init__(self):
        super().__init__()
        self.init_ui()
        self.passed_control_to_main_window = False
    def init_ui(self):
        self.lay = qtw.QHBoxLayout()
        self.setLayout(self.lay)
        
        self.intro_frame = IntroFrame(self)
        self.import_frame = ImportManager(self)
        
        self.intro_frame.setHidden(False)
        self.import_frame.setHidden(True)
        
        self.lay.addWidget(self.intro_frame)
        
        self.setWindowTitle(f"CluSeek {about.VERSION}")
        
        #
        
        self.new_analysis_path = None
    
    def switch_to_intro(self):
        self.import_frame.setHidden(True)
        self.lay.removeWidget(self.import_frame)
        self.lay.addWidget(self.intro_frame)
        self.intro_frame.setHidden(False)
        
        rect = self.geometry()
        self.setGeometry(rect.left(), rect.top(), 300, 250)
    def switch_to_import(self):
        self.intro_frame.setHidden(True)
        self.lay.removeWidget(self.intro_frame)
        self.lay.addWidget(self.import_frame)
        self.import_frame.setHidden(False)
        
        rect = self.geometry()
        self.setGeometry(rect.left(), rect.top(), 800, 500)
    def begin_new_analysis(self):
        #path = qtw.QFileDialog.getSaveFileName(
        #    caption="Select where to save the analysis.",
        #    filter="CluSeek Project (*.clp)")
        #path = path[0]
        
        #if path:
        succeeded = AppRoot.dbman.open_new_db()
        print("DB loaded", succeeded)
        if succeeded:
            # Explicitly set db to online mode so we download missing info
            AppRoot.dbman.set_offline(False)
            
            self.import_frame.load_files()
            
            self.proceed_to_main_window()
    
    def open_old_analysis_file(self):
        # Called from IntroFrame
        path = qtw.QFileDialog.getOpenFileName(
            caption="Select a CluSeek analysis file to open.",
            filter="CluSeek Project (*.clp);;All Files (*)")
            # specify suffix?
        path = path[0] # path=(path_to_file, file_type)
        if path:
            succeeded = AppRoot.dbman.open_old_db(path)
            if succeeded:
                # Go offline so we don't redownload stuff that's missing
                #   as we're unlikely to get it on the second go either.
                AppRoot.dbman.set_offline(True)
                
                # Note: It shouldn't happen, but if there are
                #       filewidgets with valid inputs in the
                #       importframe, they WILL be imported
                # For this reason,
                #   we explicitly inform the import_frame that it shouldn't
                #   consider any new inputs that may have been added.
                # The import frame should reset itself automatically,
                #   but I'd rather make very certain no strange behaviour
                #   occurs as it can ruin up peoples' data.
                
                
                def on_data_loaded():
                    # Load neiviews
                    print("Attempting to load saved cluster views ...")
                    for nei_id in AppRoot.dbman.get_available_neiviews():
                        print("Loading neiview:", nei_id)
                        neidata,neiview = AppRoot.dbman.get_neiview(nei_id, AppRoot.ui_constraintsframe.dataset)
                        AppRoot.mainwindow.add_neiview(neiview)
                    
                    # Load colocalization config
                    AppRoot.ui_constraintsframe.load_config_from_dict(
                        AppRoot.dbman.get_coloc_config())
                    
                    AppRoot.dbman.set_offline(False)
                    AppRoot.mainwindow.set_displayed_project_name(
                        os.path.basename(path))
                    self.proceed_to_main_window()
                    self.import_frame.files_loaded.disconnect(on_data_loaded)
                
                self.import_frame.files_loaded.connect(on_data_loaded)
                
                self.import_frame.load_files(no_new_input=True)
                dataset = self.import_frame.dataset
                
    def proceed_to_main_window(self):
        self.passed_control_to_main_window = True
        AppRoot.mainwindow.maintabs.setCurrentIndex(
            AppRoot.mainwindow.maintabs.count()-1)
        AppRoot.mainwindow.show()
        AppRoot.mainwindow.activateWindow()
        AppRoot.mainwindow.showMaximized()
        self.setHidden(True)
    def closeEvent(self, event):
        if not self.passed_control_to_main_window:
            if not AppRoot.exit_application():
                event.ignore()
class IntroFrame(qtw.QWidget, Ui_IntroWindow):
    def __init__(self, intro_window):
        super().__init__(intro_window)
        self.init_ui()
        
        self.intro_window = intro_window
    def init_ui(self):
        self.setupUi(self)
        
        self.open_new_btn.clicked.connect(self.cbk_new_analysis)
        self.open_old_btn.clicked.connect(self.cbk_old_analysis)
        
    def cbk_new_analysis(self):
        self.intro_window.switch_to_import()
        self.conclude()
    def cbk_old_analysis(self): # TODO:
        self.intro_window.open_old_analysis_file()
        self.conclude()
        # TODO: Launch analysis from database. #STARTHERE: Add the option to.
        #    Q: Wtf does this mean?
    def conclude(self):
        # Remove this?
        pass
class ImportManager(qtw.QWidget, Ui_ImportFrame):
    files_loaded = qtc.Signal()
    # ----------- < < <   Subclass Definitions   > > > --------------
    class BlastConfig(qtw.QWidget, Ui_BlastConfig):
        # TODO: https://www.ncbi.nlm.nih.gov/books/NBK53758/
        #   This is a super useful page which defines the [categories]
        #   like [organism]
        default_settings = {
            "db": "nr",
            "entrez_query": "",
            "matrix_name": "BLOSUM62",
            "gapcosts": "11 1",
            "comp_based_statistics": 2,
            "nhits": 5000,
            "word_size": "6",
            "expect": 10**-10,
            "filter": "F"}
        def __init__(self):
            super().__init__()
            self.init_ui()
            
            # Copy defaults over settings.
            self.settings = dict(self.default_settings)
        def init_ui(self):
            self.setupUi(self)
            
            self.apply_btn.clicked.connect(
                self.cbk_apply)
            self.cancel_btn.clicked.connect(
                self.cbk_cancel)
            self.applydefaults_btn.clicked.connect(
                self.cbk_apply_defaults)
            
            # * Set up the initial values
            # First, input all the values into fields
        def cbk_apply(self):
            self.settings = self.read_values()
            self.close()
            print("Blast settings changed.")
            
        # Virtual function reimplementation
        def showEvent(self, event):
            self.sync_UI_and_saved_values()
        
        def cbk_cancel(self):
            self.close()
        def sync_UI_and_saved_values(self):
            self.database_combo.setCurrentIndex(
                self.database_combo.findText(
                    "("+self.settings["db"]+")", 
                    qtc.Qt.MatchFixedString|qtc.Qt.MatchContains))
            
            self.entrez_ledit.setText("")
            
            self.matrix_combo.setCurrentIndex(
                self.matrix_combo.findText(
                    self.settings["matrix_name"], 
                    qtc.Qt.MatchFixedString|qtc.Qt.MatchContains))
            
            gapcosts = self.settings["gapcosts"].split(" ")
            gapcosts = f"Existence: {gapcosts[0]} Extension: {gapcosts[1]}"
            self.gapcost_combo.setCurrentIndex(
                self.gapcost_combo.findText(gapcosts, 
                    qtc.Qt.MatchFixedString|qtc.Qt.MatchContains))
            
            self.compadj_combo.setCurrentIndex(
                self.settings["comp_based_statistics"])
            
            self.maxhits_spin.setValue(self.settings["nhits"])
            
            self.wordsize_combo.setCurrentIndex(
                self.wordsize_combo.findText(
                    self.settings["word_size"], 
                    qtc.Qt.MatchFixedString|qtc.Qt.MatchContains))
            
            self.expect_spin.setValue(
                round(math.log(self.settings["expect"], 10)))
            
            filter = self.settings["filter"]
            self.mask_chk.setChecked("m" in filter)
            self.filter_chk.setChecked("T" in filter or "L" in filter)
        
        def cbk_apply_defaults(self):
            self.database_combo.setCurrentIndex(0)
            self.entrez_ledit.setText("")
            self.matrix_combo.setCurrentIndex(5)
            self.gapcost_combo.setCurrentIndex(2)
            self.compadj_combo.setCurrentIndex(2)
            self.maxhits_spin.setValue(5000)
            self.wordsize_combo.setCurrentIndex(2)
            self.expect_spin.setValue(-10)
            self.filter_chk.setChecked(False)
            self.mask_chk.setChecked(False)
            
        def read_values(self):
            values = {}
            
            # * Database
            values["db"] = self.database_combo.currentText().split("(")[-1]\
                .split(")")[0]
            
            # * Organisms
            values["entrez_query"] = self.entrez_ledit.text()
            
            # * Matrix
            values["matrix_name"] = self.matrix_combo.currentText()
            
            # * Gap Costs
            raw = " ".join(self.gapcost_combo.currentText().\
              lstrip("Existence: ").split(" Extension: "))
            values["gapcosts"] = raw
            
            # * Compositional adjustments
            values["comp_based_statistics"] = self.compadj_combo.currentIndex()
            
            # * Max target sequences
            values["nhits"] = self.maxhits_spin.value()
            
            # * Word size
            values["word_size"] = self.wordsize_combo.currentText()
            
            # * Expect threshold
            values["expect"] = 10 ** self.expect_spin.value()
            
            # ENSURE EXPECT VALUE IS A REAL NUMBER VIA ANOTHER CHECK
            
            
            # * Filter
            #    Low complexity filtering:
            #       F to disable, T or L to enable
            #    Masking:
            #       Prepend m for mask at lookup
            if self.filter_chk.isChecked():
                if self.mask_chk.isChecked():
                    values["filter"] = "mT"
                else:
                    values["filter"] = "T"
            else:
                values["filter"] = "F"
            
            return(values)

    class FileWidget(qtw.QWidget, Ui_FileWidget):
        #Represents a single file to be loaded
        uFASTA_pattern = re.compile(r"^>.+(?:(?:\n|\r|\n\r)[A-Za-z*-]+)+", re.MULTILINE)
        gene_uFASTA_pattern = re.compile(r"^>.+(?:(?:\n|\r|\n\r)[aAcCtTgGnN-]+)+", re.MULTILINE)
        accession_pattern = re.compile(r"[A-Z]{3}[0-9]{5,7}(?:\.[0-9]+){0,1}")
        WARNING_INPUTS = 15
        MAX_INPUTS = 50
        def __init__(self, impman):
            super().__init__()
            self.impman = impman
            self.input_type = None # Need to setup UI first
            self.init_ui()
        def init_ui(self):
            self.setupUi(self)
            self.btn_browse.clicked.connect(self.add_files)
            self.btn_delete.clicked.connect(self.delete)
            
            # Adjust text field size based on text
            self.query_text.document().contentsChanged.connect(self.cbk_text_growth)
            
            # Change functionality when input type changes
            self.querytype_combo.currentIndexChanged.connect(
                self.cbk_input_type_combo_changed)
            
            # Check valid input
            #self.query_text.editingFinished.connect(
            #    self.cbk_text_edited)
            # Now handled through focusOutEvent
            
            def focusOutEvent(event):
                self.cbk_text_edited()
            self.query_text.focusOutEvent = focusOutEvent
            
            # Hide valid input error message until it is needed
            self.error_lbl.setHidden(True)
            
            self.setMinimumHeight(17+20)
            self.setMaximumHeight(17+20)
            self.updateGeometry()
            
            # First refresh to reflect the defaults
            self.cbk_input_type_combo_changed()
            
            # Set text window to resize up to a limit
            self.cbk_text_growth()
        # * Handlers
        def update_minimum_height(self):
            height = (17
                +self.query_text.minimumHeight()
                +(0 if self.error_lbl.isHidden() else 30))
            self.setMinimumHeight(height)
            self.setMaximumHeight(height)
            self.updateGeometry()
        def cbk_text_growth(self):
            text_height = self.query_text.document().size().height()
            size_limit = 24 if self.input_type != "FASTA" else 48
            if text_height < size_limit:
                height = size_limit
            elif text_height < 115:
                height = text_height
            else:
                height = 115
            self.query_text.setMinimumHeight(height)
            self.query_text.setMaximumHeight(height)
            self.update_minimum_height()
        def cbk_input_type_combo_changed(self):
            value = self.querytype_combo.currentText()
            if value == "BLASTP XML":
                self.btn_browse.setEnabled(True)
                self.input_type = "XML"
                self.query_text.setPlaceholderText("C:/path/to/file.xml")
            elif value == "NCBI protein accession":
                self.btn_browse.setEnabled(False)
                self.input_type = "Accession"
                self.query_text.setPlaceholderText("CAA55772.1")
            elif value == "Amino acid FASTA":
                self.btn_browse.setEnabled(True)
                self.input_type = "FASTA"
                self.query_text.setPlaceholderText(
                """>CAA55772.1 lmbY [Streptomyces lincolnensis]
MRHGVVILPEHHWARARELWRYAQELGFDHAWTYDHVKWRWLSDRPWFGAVPTLAAAATATSRIGLGTLV
ANIRLHDPVVFAKEVMTLDDISGGRFLCGVGSGGPDRDILRAGELTKGQWADRYGEFVELMDTLLRQEPV
AFDGTYYSCHETVLHPACVRRPRTPLCVAAAGPARMRLAARHADTWVTMGAPNVFDDAPYADSVPLVKDQ
VAAFERACHDVGRDPATVRRLLVAGPSIGGVLDSAGAFQDAAGLFEDAGINDFVVHWPRPDFPYRGSPAV
LDDIAPILHSAPEEA""")
            self.check_valid_input()
        def set_input_type(self, value):
            if value=="XML":
                self.querytype_combo.setCurrentText("BLASTP XML")
            elif value=="Accession":
                self.querytype_combo.setCurrentText("NCBI protein accession")
            elif value=="FASTA":
                self.querytype_combo.setCurrentText("Amino acid FASTA")
            self.cbk_input_type_combo_changed()
        def set_text(self, value):
            self.query_text.setText(value)
            self.cbk_text_edited()
        def cbk_text_edited(self):
            self.check_valid_input()
        def add_file(self, path):
            # Set the path to the file as input, because this
            #   filetype is too big to display
            self.query_text.setText(path)
        def open_file(self, path):
            # Load the contents of the file as input
            with open(path, mode="r") as file:
                fasta = file.read()
            self.query_text.setText(fasta) 
            self.cbk_text_edited()
        def process_file(self, path):
            if self.input_type == "XML":
                self.add_file(path)
            elif self.input_type == "FASTA":
                self.open_file(path)
            return
        def add_files(self):
            if self.input_type == "XML":
                filter_string = "XML files (*.xml);;Any (*)"
            elif self.input_type == "FASTA":
                filter_string = "FASTA files (*.fsa *.fasta *.txt);;Any (*)"
            else:
                filter_string = "Error"
                print(f"Unknown input_type: {self.input_type}")
            
            paths = qtw.QFileDialog.getOpenFileNames(
                filter=filter_string)
            if not paths or len(paths[0])==0:
                # If the user cancelled out of the dialog, we peace out.
                return
            if isinstance(paths, tuple):
                paths = paths[0]
            else:
                # If the dialog is cancelled,
                #   just exit the function without changes.
                return
            
            # Load file / add path
            self.process_file(paths[0])
            
            # If multiple files are selected, open separate
            #   input boxes for each new one.
            if len(paths) > 1:
                for path in paths[1:]:
                    extrawidget = self.impman.add_filewidget()
                    extrawidget.querytype_combo.setCurrentIndex(self.querytype_combo.currentIndex())
                    extrawidget.process_file(path)
            return
        def delete(self):
            self.impman.filewidgets.remove(self)
            self.deleteLater()
            self.parent().adjustSize()
        def check_valid_input(self):
            text = self.query_text.document().toPlainText()
            
            # Don't scream at people until they actually type something
            if text == "": return
            
            # Red text html for screaming at people their inputs are wrong
            red_text_1 = '''<html><head/><body><p><span style=" color:#ff0004;">'''
            red_text_2 = '''</span></p></body></html>'''
            
            outcome = (True, None)
            if self.input_type == "XML":
                if not os.path.exists(text.replace("\n","")):
                    outcome = (False, red_text_1+"Invalid path"
                           +red_text_2)
            else:
                # Determine what pattern and error message we'll be using
                #   based on input type.
                if self.input_type == "Accession":
                    pattern = self.accession_pattern
                    bad_outcome = (False, red_text_1+
                           "May not be a valid NCBI protein accession code"
                           +red_text_2)
                    found = re.findall(pattern, text)
                elif self.input_type == "FASTA":
                    pattern = self.uFASTA_pattern
                    bad_outcome = (False,
                                   red_text_1
                                   + "Not valid protein FASTA. Proper FASTA has a header that starts with '>' followed by the amino acid sequence on the next line."
                                   + red_text_2)
                    found = re.findall(r">", text)
                
                # If necessary, check with user regarding
                #   1) splitting up multiple joined entries
                #   2) loading up a large number of inputs
                splitup = None
                toomany = None
                if len(found) > 1:
                    splitup = qtw.QMessageBox.question(
                        None, 
                        "Split inputs?",
                        f"It looks like you have entered {len(found)} inputs into the same "
                        "input box. Would you like to split them up into several?")
                    splitup = splitup==qtw.QMessageBox.Yes
                if len(found)>=self.WARNING_INPUTS:
                    # If we find too many, ask the user what the hell.
                    toomany = qtw.QMessageBox.question(
                        None, 
                        "Really load many inputs?",
                        f"You are attempting to load {len(found)} marker proteins into CluSeek.\n"
                        "Are you sure? It is recommended to use less than 10 inputs, and the total should"
                        "not exceed 50.")
                    if toomany==qtw.QMessageBox.Yes:
                        self.set_text("")
                        text=""
                
                if splitup:
                    if self.input_type=="FASTA":
                        elements = [">"+element for element in text.split(">")[1:]]
                    elif self.input_type=="Accession":
                        elements = re.findall(self.accession_pattern, text)
                    
                    self.set_text(elements[0])
                    text=elements[0]
                    for element in elements[1:]:
                        new_widget = self.impman.add_filewidget()
                        new_widget.set_input_type(self.input_type)
                        new_widget.set_text(element)
                
                text=text.strip("\n").replace("\t","").replace(" ","")
                if not re.fullmatch(pattern, text):
                    outcome = bad_outcome
            
            if not outcome[0]:
                self.error_lbl.setHidden(False)
                self.error_lbl.setText(outcome[1])
                self.update_minimum_height()
                return(outcome[0])
            else:
                self.error_lbl.setHidden(True)
                self.update_minimum_height()
                return(outcome[0])
        
        # in_filepath has been replaced with query_text
    # ----------- < < <   Method Definitions   > > > ----------------
    def __init__(self, intro_window):
        super().__init__(intro_window)
        # * Processing vars
        self.intro_window = intro_window
        
        self.filewidgets = [] #Stores the widgets storing path info
        
        self.dataset = None
        #self.conframe = AppRoot.ui_constraintsframe
        self.blastconfig = self.BlastConfig()
        self.xmlpaths = []
        
        #Do this last
        self.init_ui()
    def init_ui(self):
        self.setupUi(self)
        #self.frame_neighborSettings.layout()\
        #    .setAlignment(qtc.Qt.AlignTop)
        self.lay_importFileStack.setAlignment(qtc.Qt.AlignTop)
        
        self.btn_addFile.clicked.connect(self.add_file)
        self.btn_loadFiles.clicked.connect(self.cbk_start)
        self.blastconfig_btn.clicked.connect(self.cbk_show_blastconfig)
        
        self.btn_back.clicked.connect(self.cbk_back_to_intro)
        
        #This should happen last
        self.add_filewidget()        
        
        #Testing:
        #pal = qtg.QPalette(AppRoot.qtapp.palette())
        #pal.setColor(pal.Active, pal.Window, qtg.QColor(256, 96, 96))
        #self.setPalette(pal)
        
        #Testing2
        #self.setStyleSheet("background-color: cyan")
        
        #Testing3
        #Yes! Bingo! Very good. Dopamine kick.
        
        #Testing
        #self.btn_addFile.setStyleSheet(
        #    "QPushButton { "
        ##    "              border: 2px solid green;"
        #    "              border-color: green;"
        #    "              background-color: red;"
        ##    "              color: white;"
        #    "            }")
        #self.btn_addFile.setIcon(qtg.QIcon("./uilayouts/right_to_left.png"))
        #self.btn_loadFiles.setIcon(qtg.QIcon("./uilayouts/right_to_left.png"))
        
        
        #
        pass
    # * * Handlers * *
    def cbk_start(self):
        self.intro_window.begin_new_analysis()
    def cbk_show_blastconfig(self):
        self.blastconfig.show()
        self.blastconfig.activateWindow()
    def load_files(self, no_new_input=False):
        # NOTE: The steps in this function are executed via work threads,
        #   and are declared in reverse order for this reason.
        
        AppRoot.progress_dialog.reinitialize(
          title="Progress", 
          label_text="Processing BLAST data (this may take a while if using remote blast)", 
          on_abort=None,
          text_only=True)
        
        # STEP-0: COLLECT INPUTS
        # * First we scrape the filewidgets for input data
        queries_to_blast = []
        xmlpaths = []
        if not no_new_input:
            for filewidget in self.filewidgets:
                text = filewidget.query_text.document().toPlainText()
                if text == "":
                    continue
                elif filewidget.input_type == "XML":
                    xmlpaths.append(text)
                elif filewidget.input_type == "FASTA":
                    if ">" not in text:
                        text = ">unknown_sequence\n"+text
                    queries_to_blast.append(text)
                elif filewidget.input_type == "Accession":
                    queries_to_blast.append(text)
            
            if len(xmlpaths) == 0 and len(queries_to_blast) == 0:
                print("Loading: No new inputs provided.")
        
            # If we have XML inputs, write them directly into the database.
            for xmlpath in xmlpaths:
                AppRoot.dbman.save_blastpxml_file_into_db(xmlpath)
        
        # STEP-3: DISPLAY DATA
        def on_files_loaded():
            # Load graph in main window.
            AppRoot.ui_constraintsframe.load_dataset(self.dataset)
            AppRoot.progress_dialog.setHidden(True)
            
            # Display main window
            self.intro_window.proceed_to_main_window()
            self.files_loaded.emit()
        # STEP-2: LOAD DATA (And download IPG information)
        def on_remoteblast_complete():
            AppRoot.timer.after_delay_do(worker_loadfiles.safe_start)
        worker_loadfiles = WorkerThread(self._load_files)
        worker_loadfiles.signals.finished.connect(on_files_loaded)
        
        # With this somewhat confusing implementation, the
        #   download manager calls the signal in one thread,
        #   then responds to it in the other.
        worker_loadfiles.signals.progress_manager_update.connect(
            dbdl.DOWNLOAD_MANAGER.cbk_status_update)
        dbdl.DOWNLOAD_MANAGER.set_progress_manager_update_signal(
            worker_loadfiles.signals.progress_manager_update)
        
        # STEP-1: REMOTE BLAST
        worker_remoteblast = WorkerThread(self._run_remote_BLAST, 
                                          queries_to_blast)
        worker_remoteblast.signals.finished.connect(
            on_remoteblast_complete)
        
        # * Then we run the BLAST thread
        AppRoot.progress_dialog.show()
        AppRoot.progress_dialog.activateWindow()
        worker_remoteblast.safe_start()
    def _run_remote_BLAST(self, queries_to_blast):
        if len(queries_to_blast) > 0:
            if len(queries_to_blast) > 50:
                qtw.QMessageBox.warning(None,
                    "Too many queries.",
                    "You are attempting to submit more than 50 queries to the NCBI servers."
                    "This is strongly discouraged by both NCBI and us, as typical CluSeek searches"
                    "should require below 10 marker proteins."
                    "Large searches put strain on NCBI servers and are unlikely to complete due to their size.")
            print("Sending remote BLAST request")
            # * Second, we need to perform remote BLASTs
            
            AppRoot.dbman.remote_blast_inputs(queries_to_blast, 
                                              self.blastconfig.settings)
        else:
            print("Remote BLAST skipped, no queries.")
    def _load_files(self, progress_callback=None):
        self.dataset = Dataset()
        # TODO: load_files must take XML strings, not paths
        self.dataset.load_files(AppRoot.dbman._retrieve_all_blastpxml())
        print("Loaded files.")
    def add_file(self):
        self.add_filewidget()
    # * * Processors * *
    def add_filewidget(self):
        newfilewidget = self.FileWidget(self)
        self.lay_importFileStack.insertWidget(
            self.lay_importFileStack.count()-2, 
            newfilewidget)
        self.filewidgets.append(newfilewidget)
        return(newfilewidget)
    def cbk_back_to_intro(self):
        self.reset()
        self.intro_window.switch_to_intro()
    def reset(self):
        return
        # This was causing problems, it might be easier to just not.
        for filewidget in self.filewidgets:
            filewidget.delete()
        self.blastconfig.cbk_apply_defaults()

#
class MainWindow(qtw.QMainWindow, Ui_MainWindow):
    def __init__(self):
        super().__init__()
        self.init_ui()
        
        # About info dialogs
        self.aboutme = None
        self.approot = AppRoot
        
        AppRoot.on_app_start()
    def init_ui(self):
        self.setupUi(self)
        
        # * License text popup
        self.licensetext = qtw.QTextBrowser()
        self.licensetext.setWindowTitle(f"CluSeek {about.VERSION} - Licenses")
        self.licensetext.setMinimumWidth(550)
        self.licensetext.setMinimumHeight(550)
        
        # * Menubar Callbacks
        self.action_aboutme.triggered.connect(self.cbk_aboutme)
        self.action_aboutqt.triggered.connect(self.cbk_aboutqt)
        
        self.offline_action.triggered.connect(self.cbk_toggle_offlinemode)
        self.file_save_act.triggered.connect(self.save)
        self.file_saveas_act.triggered.connect(self.save_as)
        
        
        # * Other
        # This first
        AppRoot.ui_constraintsframe = ConstraintsFrame()
        self.lay_tab_filter.addWidget(AppRoot.ui_constraintsframe)
        
        # This second
        #AppRoot.ui_importmanager = ImportManager()
        #self.lay_tab_import.addWidget(AppRoot.ui_importmanager)
        
        # Finally these
        #AppRoot.ui_neighborhoodviewer = NeighborhoodViewer()
        #self.lay_tab_neighbors.addWidget(AppRoot.ui_neighborhoodviewer)
        #AppRoot.ui_topclusters = TopClusters()
        #self.lay_tab_test.addWidget(AppRoot.ui_topclusters)
        
        #AppRoot.ui_infowindow = InfoWindow()
        
        self.setWindowTitle(f"CluSeek {about.VERSION}")
    def init_dbman_link(self):
        AppRoot.dbman.on_offline_changed = self.on_offline_changed
    def closeEvent(self, event):
        if not AppRoot.exit_application():
            event.ignore()
    def cbk_aboutme(self):
        def callback_maker(textbox, text):
            def cbk_showlicense():
                textbox.setText(text)
                textbox.show()
                textbox.activateWindow()
            return(cbk_showlicense)
        if not self.aboutme:
            
            #scrollwrap = qtw.QWidget()
            scrollarea = qtw.QScrollArea()
            scrollable = qtw.QWidget()
            self.aboutme = scrollarea
            
            scrollarea.setWidget(scrollable)
            scrollarea.setMinimumWidth(700)
            scrollarea.setMinimumHeight(400)
            scrollarea.setWidgetResizable(True)
            scrollarea.setVerticalScrollBarPolicy(qtc.Qt.ScrollBarAlwaysOn)
            scrollarea.setHorizontalScrollBarPolicy(qtc.Qt.ScrollBarAlwaysOff)
            
            layout = qtw.QGridLayout()
            
            # About Me
            row=0
            label = qtw.QLabel(about.aboutme)
            label.setOpenExternalLinks(True)
            label.setWordWrap(True)
            layout.addWidget(label, row, 0, 1, 3)
            
            # About BLAST
            row+=1
            label = qtw.QLabel(about.aboutblast)
            label.setOpenExternalLinks(True)
            label.setWordWrap(True)
            layout.addWidget(label, row, 0, 1, 3)
            
            # About DIAMOND
            row+=1
            label = qtw.QLabel(about.aboutdiamond)
            label.setOpenExternalLinks(True)
            label.setWordWrap(True)
            layout.addWidget(label, row, 0, 1, 3)
            
            # About Python
            row+=1
            label = qtw.QLabel(about.aboutpython)
            label.setOpenExternalLinks(True)
            label.setWordWrap(True)
            layout.addWidget(label, row, 0, 1, 3)
            #
            cbk_showlicense = callback_maker(self.licensetext, about.pylicense)
            btn = qtw.QPushButton("Python License")
            btn.clicked.connect(cbk_showlicense)
            layout.addWidget(btn, row, 4, 1, 1)
            
            # About PACKAGES
            row+=1
            label = qtw.QLabel(about.aboutpackages)
            label.setOpenExternalLinks(True)
            label.setWordWrap(True)
            layout.addWidget(label, row, 0, 1, 3)
            
            row+=1
            label = qtw.QLabel(about.licensenote)
            label.setOpenExternalLinks(True)
            label.setWordWrap(True)
            layout.addWidget(label, row, 0, 1, 3)
            
            # PACKAGES
            row+=1
            headers = ["Module", "Version", "License Used*", "Copyright Notice/s", ""]
            for col in range(len(headers)):
                label = qtw.QLabel(headers[col])
                layout.addWidget(label, row, col)
            
            row+=1
            line = qtw.QFrame()
            line.setFrameShape(qtw.QFrame.HLine)
            line.setFrameShadow(qtw.QFrame.Sunken)
            layout.addWidget(line, row, 0, 1, 5)
            
            row+=1
            for package_info in about.get_aboutpackages():
                row += 1
                col =  0
                for value in package_info:
                    header = about.packagecsvheaders[col]
                    if header == "LicenseText":
                        cbk_showlicense = callback_maker(self.licensetext, value)
                        btn = qtw.QPushButton("License Text")
                        btn.clicked.connect(cbk_showlicense)
                        layout.addWidget(btn, row, col)
                    else:
                        layout.addWidget(qtw.QLabel(value), row, col)
                    col += 1
            
                        # PyThirdParty
            row+=1
            label = qtw.QLabel(about.aboutpythirdparty)
            label.setOpenExternalLinks(True)
            label.setWordWrap(True)
            layout.addWidget(label, row, 0, 1, 3)
            
            row+=1
            headers = ["Module/Software", "skip", "skip", "Copyright Notice/s", ""]
            for col in range(len(headers)):
                if headers[col] == "skip":
                    continue
                label = qtw.QLabel(headers[col])
                layout.addWidget(label, row, col)
            
            row+=1
            line = qtw.QFrame()
            line.setFrameShape(qtw.QFrame.HLine)
            line.setFrameShadow(qtw.QFrame.Sunken)
            layout.addWidget(line, row, 0, 1, 5)
            
            row+=1
            for package_info in about.get_pythirdparty():
                row += 1
                col =  0
                for value in package_info:
                    header = about.packagecsvheaders[col]
                    if headers[col] == "skip":
                        col += 1
                        continue
                    if header == "LicenseText":
                        cbk_showlicense = callback_maker(self.licensetext, value)
                        btn = qtw.QPushButton("License Text")
                        btn.clicked.connect(cbk_showlicense)
                        layout.addWidget(btn, row, col)
                    else:
                        layout.addWidget(qtw.QLabel(value), row, col)
                    col += 1
            
            row+=1
            label = qtw.QLabel("\n\n")
            label.setOpenExternalLinks(True)
            label.setWordWrap(True)
            layout.addWidget(label, row, 0, 1, 3)
            
            
            
            # About PyInstaller
            row+=1
            label = qtw.QLabel("\n\n")
            label.setOpenExternalLinks(True)
            label.setWordWrap(True)
            layout.addWidget(label, row, 0, 1, 3)
            
            row+=1
            label = qtw.QLabel(about.aboutpyinstaller)
            label.setOpenExternalLinks(True)
            label.setWordWrap(True)
            layout.addWidget(label, row, 0, 1, 3)
            
            row+=1
            label = qtw.QLabel(about.aboutpyinstaller2)
            label.setOpenExternalLinks(True)
            label.setWordWrap(True)
            layout.addWidget(label, row, 0, 1, 3)
            
            row+=1
            label = qtw.QLabel(about.aboutpyinstaller3)
            label.setOpenExternalLinks(True)
            label.setWordWrap(True)
            layout.addWidget(label, row, 0, 1, 3)
            
            # PyInstaller Packages
            row+=1
            headers = ["Module", "Version", "License", "Copyright Notice/s", ""]
            for col in range(len(headers)):
                label = qtw.QLabel(headers[col])
                layout.addWidget(label, row, col)
            
            row+=1
            line = qtw.QFrame()
            line.setFrameShape(qtw.QFrame.HLine)
            line.setFrameShadow(qtw.QFrame.Sunken)
            layout.addWidget(line, row, 0, 1, 5)
            
            row+=1
            for package_info in about.get_aboutpyinstallerpackages():
                row += 1
                col =  0
                for value in package_info:
                    header = about.packagecsvheaders[col]
                    if header == "LicenseText":
                        cbk_showlicense = callback_maker(self.licensetext, value)
                        btn = qtw.QPushButton("License Text")
                        btn.clicked.connect(cbk_showlicense)
                        layout.addWidget(btn, row, col)
                    else:
                        layout.addWidget(qtw.QLabel(value), row, col)
                    col += 1
            
            scrollable.setLayout(layout)
        self.aboutme.setWindowTitle(f"Cluseek {about.VERSION} - About")
        self.aboutme.show()
        self.aboutme.activateWindow()
    def cbk_aboutqt(self):
        meh = qtw.QWidget()
        aboutqt = qtw.QMessageBox.aboutQt(meh, "About Qt")
        meh.show()
        meh.activateWindow()
    def cbk_toggle_offlinemode(self):
        AppRoot.dbman.set_offline(self.offline_action.isChecked())
    # Export callbacks:
    # Debug callbacks
    def cbk_codeinteract(self):
        code.interact(local=locals())
    def set_displayed_project_name(self, filename):
        self.setWindowTitle(f"CluSeek {about.VERSION} - {filename}")
    
    def add_neiview(self, neiview):
        if neiview.displayname:
            displayname = neiview.displayname
        else:
            displayname = f"Cluster view {neiview.id_}"
            neiview.displayname = displayname
        self.maintabs.addTab(neiview, displayname)
    def save_all_neiview_changes(self):
        for index in range(self.maintabs.count()):
            widget = self.maintabs.widget(index)
            if not isinstance(widget, NeiView): continue
            if not widget.is_initialized: continue
            
            widget.neidata.cluster_hierarchy.clear_selection()
            AppRoot.dbman.write_neidata(widget)
    def save(self):
        try:
            if not AppRoot.dbman.opened_db_path:
                self.save_as()
            # Put all changes into db
            self.save_all_neiview_changes()
            AppRoot.dbman.save_coloc_config(
                AppRoot.ui_constraintsframe.get_config_dict())
            
            # Copy db to disk
            AppRoot.dbman.save_active_to(
                AppRoot.dbman.opened_db_path)
        except Exception as e:
            qtw.QMessageBox.critical(
                None, 
                "Error saving project",
                str(traceback.format_exc()))
            raise e
    def save_as(self):
        try:
            # TODO: Add exception error popup
            path = qtw.QFileDialog.getSaveFileName(
                caption="Select where to save the analysis.",
                filter="CluSeek Project (*.clp)")
            path = path[0]
            
            if not path: return
            
            # Put all changes into db
            self.save_all_neiview_changes()
            AppRoot.dbman.save_coloc_config(
                AppRoot.ui_constraintsframe.get_config_dict())
            
            # Copy db to disk
            AppRoot.dbman.save_active_to(path)
            
            # Change active file path
            AppRoot.dbman.opened_db_path = path
            
            self.set_displayed_project_name(os.path.basename(path))
        except Exception as e:
            qtw.QMessageBox.critical(
                None, 
                "Error saving project",
                str(traceback.format_exc()))
            raise e
    def new_project(self):
        qtw.QMessageBox.warning(None, "Feature not yet implemented", "sorry :(")
    def on_offline_changed(self):
        self.offline_action.setChecked(AppRoot.dbman.offline)
#
#[A, B]
class ConstraintsGraph(qtw.QWidget, Ui_GraphFrame):
    variable_texts = {
        "evalue": "E-value",
        "pquery": "% Identity",
        "score": "Bitscore",
    }
    def __init__(self, conframe, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.xvals = []
        self.yvals = []
        
        self.conframe = conframe
        self.dataset = None #Contains all 
        
        #Graph properties
        #self.graph = None #the graph itself
        
        self.accname = None #Active accset name
        self.acci = 0 #For cycling through accessions
        self.data = None #A tuple of values like evalue and %identity
        self.selectionbounds = None
        self.accname_radios = {}
        
        #Event handling (mouse events)
        self.pressedcoords = None
        
        #All possible variables
        self.variables = ["evalue", "pquery", "score"]
        self.variable = None #Active variable
        
        self.yscale = None #Use log or linear
        
        self.init_ui()
    def switch_dataset(self, dataset):
        self.dataset = dataset

        for name in self.dataset.names:
            radio = self.add_marker_radio(name)
        
        radio.setChecked(True)
        
        self.preprocess_graph_data()
        
        self.change_accset(name)
    def init_ui(self):
        self.setupUi(self)
        
        # Initialize Graph
        self.fig = Figure(figsize=(5, 5), dpi=100)
        self.ax = self.fig.add_subplot(111)
        self.graph  = FigureCanvasQTAgg(self.fig)
        
        self.le_min.editingFinished.connect(self.minmax_changed)
        self.le_max.editingFinished.connect(self.minmax_changed)
        
        self.lay_graphcontainer.addWidget(self.graph)
        
        # Hook up handlers
        self.btn_variable_evalue.clicked.connect(self.cbk_variable_changed)
        self.btn_variable_identity.clicked.connect(self.cbk_variable_changed)
        self.btn_variable_bitscore.clicked.connect(self.cbk_variable_changed)
        
        self.btn_loglin_lin.clicked.connect(self.cbk_loglin_changed)
        self.btn_loglin_log.clicked.connect(self.cbk_loglin_changed)
        
        self.btn_wiperule.clicked.connect(self.rule_wipe)
        
        self.info_histogram_btn.clicked.connect(self.cbk_info_histogram_toggle)
        
        #self.save_data_btn.clicked.connect(self.cbk_save_data)
        self.save_graph_btn.clicked.connect(self.cbk_save_graph)
        
        self.conframe.sig_accsetalias_changed.connect(
            self.cbk_accsetalias_changed)
        
        self.btn_variable_bitscore.setHidden(True) # Functionality is not fully there
        self.info_histogram_container.setHidden(True)
        self.cbk_variable_changed()
        self.cbk_loglin_changed()
        
        add_question_mark_icon(self.info_histogram_btn)
    # * Handlers
    def add_marker_radio(self, accname):
        radio = qtw.QRadioButton()
        def callback():
            if radio.isChecked():
                self.change_accset(str(accname))
        radio.clicked.connect(callback)
        radio.setText(self.dataset.aliases[accname])
        self.accname_radios[accname] = radio
        self.markerswitcher_scrollable_lay.addWidget(radio)
        
        return(radio)
    def minmax_changed(self):
        try:
            float(self.le_min.text())
        except ValueError:
            self.le_min.setText("")
        try:
            float(self.le_max.text())
        except ValueError:
            self.le_max.setText("")
        if self.le_min.text() != "" and self.le_max.text() != "":
            self.graph_setlimits(float(self.le_min.text()),
                                 float(self.le_max.text()))
        self.rule_make()
    def change_accset(self, accname):
        self.accname = accname
        #Update the index so it cycles properly
        self.acci = self.dataset.names.index(accname)
        self.redraw()
    def cbk_variable_changed(self):
        if self.btn_variable_evalue.isChecked():
            self.variable = "evalue"
        elif self.btn_variable_identity.isChecked():
            self.variable = "pquery"
        elif self.btn_variable_bitscore.isChecked():
            self.variable = "score"
        self.redraw()
    def cbk_loglin_changed(self):
        if self.btn_loglin_lin.isChecked():
            self.yscale="linear"
        elif self.btn_loglin_log.isChecked():
            self.yscale="log"
        self.redraw()
    def graph_mousepress(self, event):
        if event.button == 1:
            # This only saves the position of the click 
            #    event for later
            self.pressedevent = event
    def graph_mouserelease(self, event):
        if event.button == 1:
            # The moment the button is released, this creates a 
            #   bounding box on the graph between where the mousepress
            #   and mouserelease happened.
            #print(f"MouseRelease-{event.button} at {event.xdata}, {event.ydata}")
            try:
                self.graph_setlimits(min(event.xdata, 
                                         self.pressedevent.xdata),
                                     max(event.xdata, 
                                         self.pressedevent.xdata))
            except TypeError:
                print("Drag and drop likely wasn't within bounds of the graph widget.")
            self.rule_make()
        elif event.button == 3:
            self.graph_wipelimits()
    def graph_addhandlers(self):
        #self.graph.mousePressEvent = self.graph_mousepress
        #self.graph.mouseReleaseEvent = self.graph_mouserelease
        self.graph.mpl_connect("button_press_event", 
                               self.graph_mousepress)
        self.graph.mpl_connect("button_release_event", 
                               self.graph_mouserelease)
    def graph_wipelimits(self):
        if self.selectionbounds is not None:
            self.selectionbounds.remove()
            self.selectionbounds = None
            self.graph.draw()
        #TODO: Wipe current rule limits
        self.le_min.setText("")
        self.le_max.setText("")
    def graph_setlimits(self, xmin, xmax):
        if self.selectionbounds is not None:
            self.selectionbounds.remove()
            self.selectionbounds = None
        bounds = self.ax.get_xbound()
        self.selectionbounds = self.ax.axvspan(xmin, xmax, alpha=0.3, color="red")
        self.ax.set_xbound(bounds)
        self.graph.draw()
        #TODO: Set new rule limits
        self.le_min.setText(str(xmin))
        self.le_max.setText(str(xmax))
        pass
    def rule_make(self):
        try:
            xmin = float(self.le_min.text())
            xmax = float(self.le_max.text())
        except:
            print("Invalid min and max values for rule.")
            return
        rule = (xmin, xmax)
        #TODO: Try converting all the variable names to the new enum.
        #TODO: You know you want to.
        if self.variable == "evalue":
            variable = VariableTypes.EVALUE
        elif self.variable == "pquery":
            variable = VariableTypes.IDENTITY_PER_QUERY
        elif self.variable == "bitscore":
            variable = VariableTypes.BITSCORE
        else:
            assert False, ("Invalid ConstraintsGraph.variable value. "
                           "Must be 'evalue' 'pquery' or 'bitscore'")
        self.dataset.rules.new_alignment_rule(
            accset_name=self.accname, variable=variable, minimum=xmin, maximum=xmax)
        
        # After making the rule, recalc and redraw
        #self.dataset.proc1_create_subsets()
        self.redraw()
    def rule_wipe(self):
        if self.variable == "evalue":
            variable = VariableTypes.EVALUE
        elif self.variable == "pquery":
            variable = VariableTypes.IDENTITY_PER_QUERY
        else:
            assert False, "Invalid ConstraintsGraph.variable value. Must be 'evalue' or 'pquery'"
        self.dataset.rules.wipe_alignment_rule(self.accname, variable)
        self.graph_wipelimits()
        
        # After making the rule, redraw
        #self.dataset.proc1_create_subsets()
        self.redraw()
    def rules_load(self):
        #Updates the displayed limits 
        #  based on current accset and variable        
        if self.variable == "evalue":
            variable = VariableTypes.EVALUE
        elif self.variable == "pquery":
            variable = VariableTypes.IDENTITY_PER_QUERY
        else:
            assert False, "Invalid ConstraintsGraph.variable value. Must be 'evalue' or 'pquery'"
        rule = self.dataset.rules.get_alignment_rule(self.accname, variable)
        if rule is None:
            self.graph_wipelimits()
        else:
            self.graph_setlimits(rule.min, rule.max)
    def preprocess_graph_data(self):
        # This function pre-cooks a lot of data in order to speed up the
        #   graph redraw times to something humans can live with.
        
        self._prep_accession_to_vals = {}
        self._prep_accession_to_ipg = {}
        self._prep_extended_accsets = {}
        self._prep_missing_proteins = set()
        
        for accname in self.dataset.accsets:
            extended_accset = set()
            for accession in self.dataset.accsets[accname]:
                if accession not in self.dataset.root.ptAll:
                    print(" * * Protein not found in root")
                    self._prep_missing_proteins.add(accession)
                    continue
                
                # * Process alignment
                alignment = self.dataset.blasts[accname].als[accession]
                # Get the best hsp
                hsp = max(alignment.hsps, key=lambda hsp: hsp.score)
                
                # Generate values from hsp
                vals = [(hsp.identities / self.dataset.blasts[accname].query_length,
                        hsp.expect,
                        hsp.score,
                        hsp.query_end-hsp.query_start+1,
                        hsp.sbjct_end-hsp.sbjct_start+1)]
                
                # * Get list of identical protein accessions
                ipg_accessions = list(self.dataset.root.ptAll[accession].ipt.pts)
                
                # Map original accession to its ipg accessions
                self._prep_accession_to_ipg[accession] = ipg_accessions
                
                for ipg_accession in ipg_accessions:
                    # Add each accession to extended accset
                    extended_accset.add(ipg_accession)
                    
                    # Map accession to preprocessed values
                    self._prep_accession_to_vals[ipg_accession] = vals
            
            # Add the extended accset to our collection
            self._prep_extended_accsets[accname] = extended_accset
    def redraw_prepare_data(self):
        # Get all accession codes relevant to our graph
        all_accessions = self._prep_extended_accsets[self.accname]
        subset_accessions = set(self.dataset.subset["accsets"][self.accname])
        #for acc in self.dataset.subset["accsets"][self.accname]:
        #    if acc in self._prep_missing_proteins:
        #        continue
        #    subset_accessions.extend(self._prep_accession_to_ipg[acc])
        #subset_accessions = set(subset_accessions)
        
        # Get all colocalized proteins from hits
        pre_dereplication_accessions = set()
        colocalized_accessions = set()
        for hit in self.dataset.subset["pre-dereplication"]:
            pre_dereplication_accessions.update(hit.fts)
        for hit in self.dataset.subset["hits"]:
            colocalized_accessions.update(hit.fts)
        
        # Limit the colocalized proteins to the ones relevant for this accset
        pre_dereplication_accessions = pre_dereplication_accessions & subset_accessions
        colocalized_accessions = colocalized_accessions & subset_accessions
        
        # Generate final dataset
        data_all = []
        data_subset = []
        data_prederep = []
        data_coloc = []
        for accession in all_accessions:
            if accession in self._prep_missing_proteins:
                continue
            
            # Get data
            vals = self._prep_accession_to_vals[accession]
            
            # Add to relevant datasets
            data_all.extend(vals)
            if accession in subset_accessions:
                data_subset.extend(self._prep_accession_to_vals[accession])
            if accession in pre_dereplication_accessions:
                data_prederep.extend(vals)
            if accession in colocalized_accessions:
                data_coloc.extend(vals)
        return(data_all, data_subset, data_prederep, data_coloc)
    def redraw_prepare_data_old(self, count_ipgs=True):
        # These are the three sets of values we'll be outputting.
        data_all = []
        data_subset = []
        data_in_hits = []
        name = self.accname
        errors = Counter()
        
        # Get all accessions in subset, if available
        try:
            subset = self.dataset.subset["accsets"][name]
        except:
            subset = set()
        
        # Get all accessions in hits, if available
        try:
            hit_accessions = set()
            hit_accessions_duplicated = list()
            for hit in self.dataset.subset["hits"]:
                hit_accessions.update(hit.fts)
                hit_accessions_duplicated.extend(list(hit.fts))
            hit_accessions = hit_accessions.intersection(self.dataset.subset["accsets"][name])
        except:
            hit_accessions = set()
            hit_accessions_duplicated = list()
        
        for alignment in self.dataset.blasts[name].alignments:
            hsp = max(alignment.hsps, key=lambda hsp: hsp.score)
            # Get the IPG
            if count_ipgs:
                try:
                    ipg_accs = set(self.dataset.root.ptAll[alignment.accession].ipt.pts)
                except KeyError:
                    errors["Accessions missing in data frame."] += 1
                    ipg_accs = {alignment.accession}
            else:
                ipg_accs = {alignment.accession}
            # This should calculate %query (pquery) as
            #   identities/query length
            # In order: Identities as % of query, evalue, bitscore,
            #   query_coverage, hit_coverage
            vals = [(hsp.identities / self.dataset.blasts[name].query_length,
                    hsp.expect,
                    hsp.score,
                    hsp.query_end-hsp.query_start+1,
                    hsp.sbjct_end-hsp.sbjct_start+1)]
            
            
            # All data
            data_all.extend(vals)
            # Subset after graph restrictions
            if ipg_accs.intersection(subset):
                data_subset.extend(vals)
            # Proteins found in actual potential clusters
            if ipg_accs.intersection(hit_accessions):
                data_in_hits.extend(vals)
        #code.interact(local=locals())
        return(data_all, data_subset, data_in_hits)
    def redraw(self):
        t0 = time.time()
        if not self.dataset:
            print("Histogram has no dataset, no redraw will take place.")
            return
        
        if not self.dataset.subset: return
        
        #Redraws the graph in full
        
        zeroes = 0 # Keeps track of excluded data not shown on graph.
        zeroes_in_subset = 0
        zeroes_in_prederep = 0
        zeroes_in_coloc = 0
        
        data_all,data_subset,data_prederep,data_coloc = self.redraw_prepare_data()
        
        t1 = time.time()
        
        # This next section is incredibly fast, even though it looks massive in code
        if self.variable == "pquery":
            data_all      = [vals[0] for vals in data_all]
            data_subset   = [vals[0] for vals in data_subset]
            data_prederep = [vals[0] for vals in data_prederep]
            data_coloc    = [vals[0] for vals in data_coloc]
            xbounds = {"lower": 0, "upper": 1}
            bins = [x/100 for x in range(0, 101)]
            xscale = "linear"
        elif self.variable == "evalue":
            data_all      = [vals[1] for vals in data_all]
            data_subset   = [vals[1] for vals in data_subset]
            data_prederep = [vals[1] for vals in data_prederep]
            data_coloc    = [vals[1] for vals in data_coloc]
            evalue_threshold = float(self.dataset.blasts[self.accname].expect)
            # Drop all the zeroes. Sorry, no place for you in my graph, at least for now.
            #   Added note saying as much to graph title.
            for dataset in (data_all, data_subset, data_coloc):
                for i in reversed(range(0,len(dataset))):
                    if dataset[i] == 0:
                        if dataset is data_all:
                            zeroes += 1
                        if dataset is data_subset:
                            zeroes_in_subset += 1
                        if dataset is data_coloc:
                            zeroes_in_coloc += 1
                        if dataset is data_prederep:
                            zeroes_in_prederep += 1
                        del dataset[i]
            
            
            # If all the hits are zeroes, we have a problem.
            try:
                bins = [10**x for x in range(math.floor(math.log(min(data_all), 10))-5, 
                                              math.ceil(math.log(max(data_all),10))+5)]
            
                xbounds = {"lower": min(bins+[1]), "upper": max(bins)}
                xscale="log"
            except ValueError:
                bins = 1
                xbounds = {"lower": 1, "upper": 2}
                xscale = "linear"
        elif self.variable == "score": # Unused
            data_all      = [vals[2] for vals in data_all]
            data_subset   = [vals[2] for vals in data_subset]
            data_prederep = [vals[2] for vals in data_prederep]
            data_coloc    = [vals[2] for vals in data_coloc]
            try:
                xbounds = {"lower": min(data_all)-5, "upper": max(data_all)+5}
            except ValueError:
                xbounds = {"lower": None, "upper": None}
            bins = max(data_all) #TODO: Fix this when you care.
            xscale = "linear"
        
        #
        t2 = time.time()
        
        self.graph.deleteLater()
        
        self.fig = Figure(figsize=(5, 5), dpi=100)
        self.ax = self.fig.add_subplot(111)
        self.graph = FigureCanvasQTAgg(self.fig)
        self.graph_addhandlers() #For mouse picking on the graph
        self.rules_load() #To display previously saved rules
        
        self.lay_graphcontainer.addWidget(self.graph)
        
        calc_total = len(data_all)         + zeroes
        calc_coloc = len(data_coloc)       + zeroes_in_coloc
        calc_prederep = len(data_prederep) + zeroes_in_prederep - calc_coloc
        calc_subset = len(data_subset)     + zeroes_in_subset   - calc_coloc - calc_prederep
        calc_ignored = calc_total - calc_subset - calc_prederep - calc_coloc
        
        # Colorblind compatible
        my_colors = [
            (0/255, 114/255, 178/255), # Dark blue
            (86/255, 180/255, 233/255), # Light blue
            (230/255, 159/255, 0/255), # Orange
            (213/255, 94/255, 0/255), # Red
        ]
        
        self.ax.hist(data_all,     bins=bins, color=my_colors[0], alpha=1, 
            label=f"Unselected ({calc_ignored})")
        self.ax.hist(data_subset,  bins=bins, color=my_colors[1], alpha=1,
            label=f"Uncolocalized ({calc_subset})")
        self.ax.hist(data_prederep,  bins=bins, color=my_colors[2], alpha=1,
            label=f"Colocalized but dereplicated ({calc_prederep})")
        self.ax.hist(data_coloc,   bins=bins, color=my_colors[3], alpha=1,
            label=f"Colocalized ({calc_coloc})")
        
        if self.variable == "evalue":
            self.ax.axvline(x=evalue_threshold,
                           ymin=0,
                           ymax=1,
                           color="lightgray",
                           linestyle="--")
        
        self.ax.legend()
        
        titletext = f"{self.dataset.aliases[self.accname]} ({calc_total})"
        if zeroes:
            titletext += f"\nNote: E-values of zero (perfect match) not visible on graph (={zeroes})"
        
        
        self.ax.set_title(titletext)
        self.ax.set_ylabel("n proteins")
        self.ax.set_xlabel(self.variable_texts[self.variable])
        self.ax.set_xscale(xscale)
        self.ax.set_yscale(self.yscale)
        self.ax.set_xbound(**xbounds)
        
        self.ax.grid(True)
        
        t3 = time.time()
        
        self.graph.draw()
        
        t4 = time.time()
        
        #print(" * * GRAPH TIMES")
        total = t4-t0
        #print("Data prep", t1-t0, round(100*(t1-t0)/total))
        #print("Data process", t2-t1, round(100*(t2-t1)/total))
        #print("Graph description", t3-t2, round(100*(t3-t2)/total))
        #print("Graph render", t4-t3, round(100*(t4-t3)/total))
    def export_graph_as_image(self):
        path,suffix = qtw.QFileDialog.getSaveFileName(
                        caption="Save output",
                        filter="PNG(*.png);;JPEG(*.jpg *.jpeg)")
        if not path:
            return
        if suffix == "PNG(*.png)":
            self.graph.print_png(path)
        elif suffix == "JPEG(*.jpg *.jpeg)":
            self.graph.print_jpg(path)
    def cbk_save_graph(self):
        self.export_graph_as_image()
    def cbk_save_data(self):
        self.export_graph_data_as_spreadsheet()
    def cbk_accsetalias_changed(self, accname):
        if self.isVisible() and accname==self.accname:
            self.redraw()
        self.accname_radios[accname].setText(self.dataset.aliases[accname])
    def cbk_info_histogram_toggle(self):
        self.info_histogram_container.setHidden(not self.info_histogram_container.isHidden())
class ColocResultsFrame(qtw.QWidget, Ui_ColocResultsFrame):
    def __init__(self, conframe, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.table_initialized = False
        self.dataset = None
        self.conframe = conframe
        self.init_ui()
    def init_ui(self):
        self.setupUi(self)
        self.coloc_table_basic.horizontalHeader()\
          .setFrameStyle(qtw.QFrame.Panel | qtw.QFrame.Raised)
        self.coloc_table_basic.horizontalHeader().setLineWidth(1)
        
        self.conframe.sig_accsetalias_changed.connect(
            self.cbk_accsetalias_changed)
        self.export_btn.clicked.connect(
            self.export_as_spreadsheet)
        
        self.copy_action = qtw.QAction("Copy selected")
        self.copy_action.triggered.connect(
            self.copy_highlighted_into_clipboard)
        self.copy_action.setShortcut(qtg.QKeySequence.Copy)
        self.copy_action.setShortcutContext(qtc.Qt.WidgetWithChildrenShortcut)
        self.addAction(self.copy_action)

        self.info_lbl.setHidden(True)
        self.info_btn.clicked.connect(
            self.cbk_info_btn_pressed)
        
        add_question_mark_icon(self.info_btn)
    def cbk_info_btn_pressed(self):
        self.info_lbl.setHidden(not self.info_lbl.isHidden())
    def show_results(self, dataset):
        self.dataset = dataset
        headers = ["Score", "Length (bp)", "Internal Length (bp)",
                   "Taxon", "Strain", "Tax ID", "Sequence"]
        
        # Initialize table
        self.table_initialized = True
        self.coloc_table_basic.clear()
        
        self.coloc_table_basic.setRowCount(len(dataset.subset["hits"]))
        self.coloc_table_basic.setColumnCount(len(headers)+len(dataset.accsets))
        
        # Disable sorting so it doesn't interfere during generation
        self.coloc_table_basic.setSortingEnabled(False)
        
        # Headers
        accset_order = list(dataset.accsets)
        for accset_name in accset_order:
            headers.append(dataset.aliases[accset_name])
        self.coloc_table_basic.setHorizontalHeaderLabels(headers)
        
        # Values
        row = -1
        for hit in dataset.subset["hits"]:
            row += 1
            col = 0
            
            # Hitscore
            item = qtw.QTableWidgetItem()
            item.setData(qtc.Qt.DisplayRole, hit.hit_score)
            item.setFlags(qtc.Qt.ItemIsSelectable | qtc.Qt.ItemIsEnabled)
            self.coloc_table_basic.setItem(row, col, item)
            col += 1
            
            # Length
            item = qtw.QTableWidgetItem()
            item.setData(qtc.Qt.DisplayRole, hit.length())
            item.setFlags(qtc.Qt.ItemIsSelectable | qtc.Qt.ItemIsEnabled)
            self.coloc_table_basic.setItem(row, col, item)
            col += 1
            
            # Internal Length
            fts = sorted([ft for ft in hit.fts.values()], key=lambda ft: ft.start)
            
            item = qtw.QTableWidgetItem()
            item.setData(qtc.Qt.DisplayRole, fts[-1].start-fts[0].stop)
            item.setFlags(qtc.Qt.ItemIsSelectable | qtc.Qt.ItemIsEnabled)
            self.coloc_table_basic.setItem(row, col, item)
            col += 1
            
            # Taxon
            item = qtw.QTableWidgetItem()
            item.setData(qtc.Qt.DisplayRole, hit.sc.tx.sciname)
            item.setFlags(qtc.Qt.ItemIsSelectable | qtc.Qt.ItemIsEnabled)
            self.coloc_table_basic.setItem(row, col, item)
            col += 1
            
            # Strain
            item = qtw.QTableWidgetItem()
            item.setData(qtc.Qt.DisplayRole, hit.sc.tx.strain)
            item.setFlags(qtc.Qt.ItemIsSelectable | qtc.Qt.ItemIsEnabled)
            self.coloc_table_basic.setItem(row, col, item)
            col += 1
            
            # TaxID
            item = qtw.QTableWidgetItem()
            item.setData(qtc.Qt.DisplayRole, hit.sc.tx.taxid)
            item.setFlags(qtc.Qt.ItemIsSelectable | qtc.Qt.ItemIsEnabled)
            self.coloc_table_basic.setItem(row, col, item)
            col += 1
            
            # Sequence
            item = qtw.QTableWidgetItem()
            item.setData(qtc.Qt.DisplayRole, hit.sc.accession)
            item.setFlags(qtc.Qt.ItemIsSelectable | qtc.Qt.ItemIsEnabled)
            self.coloc_table_basic.setItem(row, col, item)
            col += 1
            
            # Figure out which markers are present
            markers = {}
            for ft in hit.fts.values():
                if hasattr(ft.ref, "marker"):
                    for marker in ft.ref.marker:
                        if marker not in markers:
                            markers[marker] = []
                        markers[marker].append(ft.ref)
            for i in range(len(accset_order)):
                accset_name = accset_order[i]
                if accset_name in markers:
                    displaytext = ", ".join(
                        [pt.accession for pt in markers[accset_name]])
                else:
                    displaytext = ""
                    
                item = qtw.QTableWidgetItem()
                item.setData(qtc.Qt.DisplayRole, displaytext)
                item.setFlags(qtc.Qt.ItemIsSelectable | qtc.Qt.ItemIsEnabled)
                self.coloc_table_basic.setItem(row, col+i, item)
        
        # Re-size columns
        self.coloc_table_basic.horizontalHeader().resizeSections(
            qtw.QHeaderView.ResizeToContents)
        
        # Re-enable sorting
        self.coloc_table_basic.setSortingEnabled(True)
    def export_as_spreadsheet(self):
        if not self.table_initialized:
            print("Export failed: No spreadsheet initialized.")
            return
        
        path,suffix = qtw.QFileDialog.getSaveFileName(
                        caption="Save output",
                        filter="Excel 2010 Spreadsheet(*.xlsx)")
        
        if not path:
            return
        
        # self proxy
        t = self.coloc_table_basic
        
        # Initialize workbook
        wbk = pxl.Workbook()
        
        sh = wbk.active
        sh.title = "Colocalization Results"
        
        # copy headers
        for col in range(0, t.columnCount()):
            item = t.horizontalHeaderItem(col)
            sh.cell(column=col+1, row=1, value=item.text())
        
        # copy cells
        for row in range(0, t.rowCount()):
            for col in range(0, t.columnCount()):
                item = t.item(row,col)
                sh.cell(column=col+1, row=row+2, value=item.text())
        
        wbk.save(path)
    def cbk_accsetalias_changed(self, accsetname):
        if self.isVisible():
            self.show_results(self.dataset)
    def copy_highlighted_into_clipboard(self):
        output = list()
        ranges_ = self.coloc_table_basic.selectedRanges()
        if not ranges_: return
        for range_ in ranges_:
            for row in range(range_.topRow(), range_.bottomRow()+1):
                output.append(list())
                for col in range(range_.leftColumn(), range_.rightColumn()+1):
                    item = self.coloc_table_basic.item(row, col)
                    
                    value = item.text()
                    
                    output[-1].append(value)
        del col
        del row
        clipboard = AppRoot.qtapp.clipboard()
        clipboard.clear()
        clipboard.setText("\n".join(["\t".join(row) for row in output]))

class ConstraintsFrame(qtw.QWidget, Ui_ConstraintsFrame):
    # ----------- < < <   Subclass Definitions   > > > --------------
    class AccsetBlock(qtw.QWidget, Ui_AccsetBlock):
        def __init__(self, conframe, name):
            super().__init__()
            self.name = name
            
            #List of other AccsetBlocks
            self.conframe = conframe
            #  This widget will add itself to it, and then remove itself
            #  upon deletion, removes itself from the list
            self.conframe.accsetwidgets.append(self)
            
            self.alias = None
            
            #do this last
            self.init_ui()
        def init_ui(self):
            self.setupUi(self)
            self.ch_include.setText(self.name)
            
            #self.btn_view.clicked.connect(self.selected)
            self.ch_include.stateChanged.connect(self.cbk_checked)
            
            # Set the default alias
            if self.name not in self.conframe.dataset.aliases:
                self.in_alias.setText(self.name)
            else:
                self.in_alias.setText(self.conframe.dataset.aliases[self.name])
            self.in_alias.editingFinished.connect(self.cbk_renamed)
            self.in_alias.setCursorPosition(0)
        def easyscore_widgets_on(self):
            self.ch_include.setEnabled(True)
            self.score_spin.setEnabled(False)
            if self.score_spin.value() <= 0:
                self.ch_include.setChecked(qtc.Qt.Unchecked)
            else:
                self.score_spin.setValue(1)
        def easyscore_widgets_off(self):
            self.ch_include.setChecked(qtc.Qt.Checked)
            self.ch_include.setEnabled(False)
            self.score_spin.setEnabled(True)
        def cbk_renamed(self):
            val = self.in_alias.text()
            if val == "":
                val = self.name
            print(f"Alias changed to {val}")
            self.conframe.dataset.aliases[self.name] = val
            self.conframe.sig_accsetalias_changed.emit(self.name)
        def set_alias(self, alias):
            self.in_alias.setText(alias)
            self.cbk_renamed()
        def cbk_checked(self):
            if self.ch_include.checkState():
                if self.conframe.easyscore:
                    self.score_spin.setValue(1)
                    self.conframe.update_minscore()
            else:
                if self.conframe.easyscore:
                    self.score_spin.setValue(0)
                    self.conframe.update_minscore()
        def export_rules(self):
            # This should only be called when the user can
            #   no longer mess with the numbers.
            self.conframe.dataset.rules.new_sequence_rule(
                self.name,
                on_true=self.score_spin.value())
        def remove(self):
            self.conframe.accsetwidgets.remove(self)
            self.deleteLater()
            self.parent().adjustSize()
    class NeiviewCreator(qtw.QWidget, Ui_NeiviewCreator):
        LOCAL_CLUSTERING_SENSITIVITY_OPTIONS = {
            "Fast": "--fast",
            "Somewhat Sensitive": "--mid-sensitive",
            "Sensitive": "--sensitive",
            "More Sensitive": "--more-sensitive",
            "Very Sensitive": "--very-sensitive",
            "Ultra Sensitive": "--ultra-sensitive"}
        LOCAL_CLUSTERING_SENSITIVITY_DEFAULT = "Sensitive"
        COMMUNITY_WEIGHT_VARIABLE_OPTIONS = {
            "None": None,
            "E-value": "evalue",
            "Bitscore": "bitscore",
            "Identity": "identity"}
        COMMUNITY_WEIGHT_VARIABLE_DEFAULT = "Bitscore"
        def __init__(self, cframe):
            super().__init__()
            self.cframe = cframe
            self.init_ui()
        def init_ui(self):
            self.setupUi(self)
            
            # Link callbacks
            self.cancel_btn.clicked.connect(self.cbk_cancel)
            self.create_btn.clicked.connect(self.cbk_create)
                
            self.setWindowTitle("Creating new neighborhood view ...")
            self.setWindowTitle(f"CluSeek {about.VERSION} - Cluster View Creator")
            
            # Set sensitivity combo box options
            for key in self.LOCAL_CLUSTERING_SENSITIVITY_OPTIONS:
                self.clus_local_sensitivity_combo.addItem(key)
            self.clus_local_sensitivity_combo.setCurrentText(
                self.LOCAL_CLUSTERING_SENSITIVITY_DEFAULT)
            
            # Set community detection weight variable options
            for key in self.COMMUNITY_WEIGHT_VARIABLE_OPTIONS:
                self.community_weight_variable_combo.addItem(key)
            self.community_weight_variable_combo.setCurrentText(
                self.COMMUNITY_WEIGHT_VARIABLE_DEFAULT)
            
            # Hide all the info labels
            for helplbl in [
              self.global_alignment_helplbl,
              self.local_alignment_helplbl,
              self.clus_local_pidentity_helplbl,
              self.clus_local_evalue_exponent_helplbl,
              self.clus_local_bitscore_helplbl,
              self.community_detection_helplbl,
              self.community_resolution_helplbl]:
                helplbl.setHidden(True)
            
            # Create info label button callbacks
            self.global_alignment_helpbtn.clicked.connect(
                lambda : self.global_alignment_helplbl.setHidden(
                            not self.global_alignment_helplbl.isHidden()))
            self.local_alignment_helpbtn.clicked.connect(
                lambda : self.local_alignment_helplbl.setHidden(
                            not self.local_alignment_helplbl.isHidden()))
            self.clus_local_pidentity_helpbtn.clicked.connect(
                lambda : self.clus_local_pidentity_helplbl.setHidden(
                            not self.clus_local_pidentity_helplbl.isHidden()))
            self.clus_local_evalue_exponent_helpbtn.clicked.connect(
                lambda : self.clus_local_evalue_exponent_helplbl.setHidden(
                            not self.clus_local_evalue_exponent_helplbl.isHidden()))
            self.clus_local_bitscore_helpbtn.clicked.connect(
                lambda : self.clus_local_bitscore_helplbl.setHidden(
                            not self.clus_local_bitscore_helplbl.isHidden()))
            self.community_detection_helpbtn.clicked.connect(
                lambda : self.community_detection_helplbl.setHidden(
                            not self.community_detection_helplbl.isHidden()))
            self.community_resolution_helpbtn.clicked.connect(
                lambda : self.community_resolution_helplbl.setHidden(
                            not self.community_resolution_helplbl.isHidden()))
            
            add_question_mark_icon(self.global_alignment_helpbtn)
            add_question_mark_icon(self.local_alignment_helpbtn)
            add_question_mark_icon(self.clus_local_pidentity_helpbtn)
            add_question_mark_icon(self.clus_local_evalue_exponent_helpbtn)
            add_question_mark_icon(self.clus_local_bitscore_helpbtn)
            add_question_mark_icon(self.community_detection_helpbtn)
            add_question_mark_icon(self.community_resolution_helpbtn)
        def on_show(self):
            pass
        def cbk_create(self):
            # Rules in dataset only get updated on apply_rules so we should not need
            #   to re-run the search just to cover the possibility of the user
            #   changing the rules in the meantime.
            
            # Borders
            self.cframe.dataset.border_size = \
                self.border_spin.value()
            
            # Search rules --
            rules = self.cframe.dataset.rules
            aboutstrings = []
            variable_types = {
                VariableTypes.EVALUE: "<E-value>",
                VariableTypes.IDENTITY_PER_QUERY: "<Identity per query length>",
                VariableTypes.IDENTITY_PER_HIT: "<Identity per hit length>"}
            for accset_name in self.cframe.dataset.subset["accsets"]:
                aboutstring = None
                if accset_name in rules.sequence_rules:
                    aboutstring = f"{accset_name} ({self.cframe.dataset.aliases[accset_name]})"
                if accset_name in rules.alignment_rules and aboutstring:
                    ruledict = rules.alignment_rules[accset_name]
                    ruled_on_once = False
                    for variable in ruledict:
                        rule = rules.alignment_rules[accset_name][variable]
                        variable_type = variable_types[rule.variable]
                        if not ruled_on_once:
                            aboutstring += f"restricted to {variable_type} of {rule.min} to {rule.max}"
                        else:
                            aboutstring += f"and {variable_type} of {rule.min} to {rule.max}"
                        ruled_on_once = True
                if aboutstring:
                    aboutstrings.append(aboutstring)
            # Settings
            settings = {
                "local_clustering_identity_threshold": 
                    round(self.clus_local_pidentity_spin.value(), 2) \
                    if self.clus_local_pidentity_chk.isChecked() \
                    else None,
                "local_clustering_evalue_threshold": 
                    10**int(self.clus_local_evalue_exponent_spin.value()) \
                    if self.clus_local_evalue_exponent_radio.isChecked() \
                    else None,
                "local_clustering_bitscore_threshold": 
                    self.clus_local_bitscore_spin.value() \
                    if self.clus_local_bitscore_radio.isChecked() \
                    else None,
                "global_clustering_identity_threshold": 
                    self.clus_global_identity_spin.value(),
                "local_clustering_sensitivity_mode":
                    self.LOCAL_CLUSTERING_SENSITIVITY_OPTIONS[
                        self.clus_local_sensitivity_combo.currentText()],
                "community_weight_variable": 
                    self.COMMUNITY_WEIGHT_VARIABLE_OPTIONS[
                        self.community_weight_variable_combo.currentText()],
                "community_resolution": self.community_resolution_spin.value(),
                "borders_size": self.border_spin.value(),
                "loading_existing_session": False,
                "marker_accsets": {x : list(self.cframe.dataset.subset["accsets"][x]) 
                                        for x in self.cframe.dataset.subset["accsets"]},
                "about_rules": aboutstrings, 
            }
            
            
            # Get ID
            ids = [1]
            for index in range(1, AppRoot.mainwindow.maintabs.count(), 1):
                ids.append(AppRoot.mainwindow.maintabs.widget(index).id_+1)
            ids.append(AppRoot.dbman.get_new_neiview_id())
            new_id = max(ids)
            
            neidata = NeiData(list(self.cframe.dataset.subset["hits"]),
                              self.cframe.dataset,
                              settings)
            
            neiview = NeiView(neidata, new_id)
            AppRoot.mainwindow.add_neiview(neiview)
            #AppRoot.mainwindow.lay_tab_neighbors.addWidget(neiview)
            #protein_table = ProteinGroupTable(neiview)
            #AppRoot.mainwindow.lay_tab_proteingroups.addWidget(protein_table)
            
            def on_data_ready():
                neiview.init_displayed_data()
            neidata.on_data_ready = on_data_ready
            def on_selection_manager_ready():
                neiview.selection_manager.load_initial_rules()
                neiview.sig_selection_manager_created.disconnect(
                    on_selection_manager_ready)
            neiview.sig_selection_manager_created.connect(
                on_selection_manager_ready)
            
            neidata.run_pipeline()
            
            #End the dialog
            self.hide()
        def cbk_cancel(self):
            self.hide()
        def cbk_gray_out_rare_changed(self):
            if self.cbk_gray_out_rare_changed.isChecked():
                self.spin_gray_out_rare_threshold.setDisabled(False)
            else:
                self.spin_gray_out_rare_threshold.setDisabled(True)
        def refresh(self):
            # Display the window if hidden
            # and update values based on current dataset state
            
            #n neighbors
            self.nneighborhoods_displaylabel.setText(
                str(len(self.cframe.dataset.subset["hits"])))
            
            self.show()
            self.activateWindow()
    # ----------- < < <   Method Definitions   > > > ----------------
    sig_accsetalias_changed = qtc.Signal(object)
    
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.init_ui()
        
        self.dataset = None
        
        self.currentdisplay = None #None, self.graphframe, (, "group")
        self.graphframe = ConstraintsGraph(self)
        self.colocresultsframe = ColocResultsFrame(self)
        self.neiview_creator = self.NeiviewCreator(self)
        
        self.accsetwidgets = []
        
        self._tabs_added = False
        
        self.intersections = None
    def init_ui(self):
        self.setupUi(self)
        
        self.lay_accsets.setAlignment(qtc.Qt.AlignTop)
        
        self.btn_applyrules.clicked.connect(self.cbk_search)
        
        self.spin_maxdist.valueChanged.connect(\
            self.update_max_feature_distance)
        self.minscore_spin.valueChanged.connect(\
            self.cbk_min_score_changed)
        
        self.easyscore_chk.stateChanged.connect(self.cbk_easyscore_toggle)
        
        self.btn_showneighbors.clicked.connect(self.cbk_show_neiview)
        
        self.btn_viewresults.clicked.connect(self.cbk_show_colocresults)
        
        #self.btn_exportresults.clicked.connect(self.cbk_export_colocresults)
        
        self.btn_show_histogram.clicked.connect(self.show_graph)
        
        # Help
        self.includewgshelp_btn.clicked.connect(
            self.cbk_toggle_includewgshelp)
        self.derepfilterhelp_btn.clicked.connect(
            self.cbk_toggle_derepfilterhelp)
        self.ignoreunknowntxhelp_btn.clicked.connect(
            self.cbk_toggle_ignoreunknowntxhelp)
        self.maxdist_help_btn.clicked.connect(
            self.cbk_toggle_geneclustersizehelp)
        self.cbk_toggle_includewgshelp()
        self.cbk_toggle_derepfilterhelp()
        self.cbk_toggle_ignoreunknowntxhelp()

        self.derepfilter_combo.currentIndexChanged.connect(
            self.cbk_derepfilter_combo_changed)
            
        add_question_mark_icon(self.maxdist_help_btn)
        add_question_mark_icon(self.derepfilterhelp_btn)
        add_question_mark_icon(self.includewgshelp_btn)
        add_question_mark_icon(self.ignoreunknowntxhelp_btn)
    def update_colocresults(self):
        self.colocresultsframe.show_results(self.dataset)
    def add_tab_widget(self):
        if self._tabs_added: return
        self._tabs_added = True
        
        self.coloctabs = qtw.QTabWidget()
        self.constraintsframe_lay.replaceWidget(
            self.tabplaceholder_lbl, self.coloctabs)
        self.tabplaceholder_lbl.setHidden(True)
        self.tabplaceholder_lbl.setParent(None)
        del self.tabplaceholder_lbl
        
        self.graphframe_index = self.coloctabs.addTab(self.graphframe, 
                                                      "Marker homologs")
        self.colocresultsframe_index = self.coloctabs.addTab(self.colocresultsframe, 
                                                             "Colocalization")
        #neiviewmaker = self.coloctabs.addTab(self.neiview_creator, None,
        #                                     "Neighborhood view creator")
        
    def show_colocresults(self):
        self.coloctabs.setCurrentIndex(self.colocresultsframe_index)
    def show_graph(self):
        self.coloctabs.setCurrentIndex(self.graphframe_index)
    def display_none(self):
        pass
    def update_max_feature_distance(self):
        #Updates the maximum spacing between selected features
        self.dataset.max_feature_distance = self.spin_maxdist.value()
    def load_dataset(self, dataset):
        print("Loading dataset...")
        self.dataset = dataset
        AppRoot.active_dataset = dataset
        
        self.graphframe.switch_dataset(self.dataset)
        #Update the maxspacing spinbox
        self.spin_maxdist.setValue(self.dataset.max_feature_distance)
        
        #First remove all the old widgets
        for block in list(self.accsetwidgets):
            #  We use list() to duplicate the list of accsetwidgets,
            #  as the original will change size during iteration
            #  due to things being removed from it.
            block.remove()
        
        #Then add the new ones
        for name in self.dataset.names:
            block = self.AccsetBlock(self, name)
            self.lay_accsets.addWidget(block)
            block.ch_include.setChecked(qtc.Qt.Checked)
        
        # Run the easyscore toggle callback once to switch to default state
        self.cbk_easyscore_toggle()
        AppRoot.mainwindow.maintabs.setCurrentIndex(0)
        
        # Disable the data view button
        self.btn_viewresults.setDisabled(True)
        #self.btn_exportresults.setDisabled(True)
        self.btn_showneighbors.setDisabled(True)
        self.btn_show_histogram.setDisabled(True)
    def apply_rules(self):
        # Get Rules
        self.dataset.rules.reset_sequence_rules()
        for accwidget in self.accsetwidgets:
            accwidget.export_rules()
        # Define region_filter
        includewgs = self.includewgs_chk.isChecked()
        ignore_unknown_taxa = self.ignoreunknowntx_chk.isChecked()
        filterto = None
        index = self.derepfilter_combo.currentIndex()
        if index == 0: filterto = "nothing"
        elif index==1: filterto = "strain"
        elif index==2: filterto = "species"
        elif index==3: filterto = "genus"
        
        region_filter = (filterto, includewgs, ignore_unknown_taxa)
        
        
        self.dataset.proc1_create_subsets()
        self.dataset.proc2_extend_subset_using_ipg()
        self.dataset.proc3_find_target_regions(region_filter=region_filter)
        
        self.update_dataset_stats()
        self.graphframe.redraw()
        self.update_colocresults()
    def update_dataset_stats(self):
        #proteins = []
        #for accset_name in self.dataset.accsets:
        #    proteins.append(f"\t{accset_name}: "
        #                    f"{len(self.dataset.accsets[accset_name])}\n")
        #proteins = "".join(proteins)
        info = (f"Found {len(self.dataset.subset['hits'])} regions "
                f"in {len(set([hit.sc.tx.taxstrid for hit in self.dataset.subset['hits']]))} NCBI Taxa")
        self.out_infolabel.setText(info)
    @property
    def easyscore(self):
        return(True if self.easyscore_chk.checkState() is qtc.Qt.Checked else False)
    def count_intersections(self):
        #To count combinations
        regions = Counter()
        errors = Counter()
        
        #So we can easily convert accessions to accsets
        accession_to_accset = {}
        for accset_name in self.dataset.accsets:
            for accession in self.dataset.accsets[accset_name]:
                accession_to_accset[accession] = accset_name
                
        #Now count the combinations
        region=[]
        for sc in self.dataset.root.scAll.values():
            errors["Sequences processed"]+=1
            last = None
            #Go through features from left to right
            for ft in sorted([feat for feat in sc.fts.values()], key=lambda x: x.start):
                if last:
                    if ft.start - last.stop <= self.dataset.max_feature_distance:
                        region.append(ft)
                    else:
                        #Use a tuple of alphabetically sorted accset names as
                        # the key for the region counter.
                        if len(region) > 1:
                            regions[tuple(sorted(\
                                [accession_to_accset[ft.ref.accession] for ft in region]))] += 1
                            #Refresh region
                        region = []
                last = ft
            #And also save region if the end of features is hit
            if len(region) > 1:
                regions[tuple(sorted(\
                    [accession_to_accset[ft.ref.accession] for ft in region]))] += 1
                #Refresh region
                region = []
        
        #Display
        print(errors)
        print(f"\nTotal regions size: {len(regions)}\n")
        #Then display the bad boys
        for key in sorted(list(regions), key=lambda x: len(x)):
            if regions[key] > 0:
                #Convert the key tuples to alias tuples
                aliases = []
                for accset_name in key:
                    aliases.append(self.dataset.aliases.get(accset_name))
                aliases = tuple(sorted(aliases))
                print(aliases, ":", regions[key], ",")
    def update_minscore(self):
        if not self.easyscore: return
        minscore = 0
        for accwidget in self.accsetwidgets:
            if accwidget.score_spin.value() > 0:
                minscore += 1
        self.minscore_spin.setValue(minscore)
    def cbk_export_colocresults(self):
        self.colocresultsframe.export_as_spreadsheet()
    def cbk_show_colocresults(self):
        #self.update_colocresults()
        self.show_colocresults()
    def cbk_min_score_changed(self):
        self.dataset.rules.set_minimum_sequence_score(self.minscore_spin.value())
    def cbk_show_neiview(self):
        self.neiview_creator.refresh()
    def cbk_code_interact(self):
        code.interact(local=locals())
    def cbk_toggle_includewgshelp(self):
        self.includewgshelp_lbl.setHidden(
            not self.includewgshelp_lbl.isHidden())
    def cbk_toggle_derepfilterhelp(self):
        self.derepfilterhelp_lbl.setHidden(
            not self.derepfilterhelp_lbl.isHidden())
    def cbk_toggle_ignoreunknowntxhelp(self):
        self.ignoreunknowntxhelp_lbl.setHidden(
            not self.ignoreunknowntxhelp_lbl.isHidden())
    def cbk_toggle_geneclustersizehelp(self):
        self.maxdist_help_lbl.setHidden(
            not self.maxdist_help_lbl.isHidden())
    def cbk_easyscore_toggle(self):
        if self.easyscore:
            self.minscore_spin.setEnabled(False)
            for accwidget in self.accsetwidgets:
                accwidget.easyscore_widgets_on()
        else:
            self.minscore_spin.setEnabled(True)
            for accwidget in self.accsetwidgets:
                accwidget.easyscore_widgets_off()
        self.update_minscore()
    def cbk_search(self):
        self.btn_viewresults.setDisabled(False)
        self.btn_showneighbors.setDisabled(False)
        #self.btn_exportresults.setDisabled(False)
        self.btn_show_histogram.setDisabled(False)
        self.add_tab_widget()
        self.apply_rules()
    def contextMenuEvent(self, event):
        event.accept()
        menu = qtw.QMenu(self)
    def reload_redownload(self):
        AppRoot.dbman.set_offline(False)
        AppRoot.introwindow.import_frame.load_files(no_new_input=True)
    def get_config_dict(self):
        aliases = self.dataset.aliases
        
        config = {"aliases": aliases}
        return(config)
    def load_config_from_dict(self, config):
        aliases = config["aliases"] if "aliases" in config else dict()
        
        for accsetwidget in self.accsetwidgets:
            if accsetwidget.name in aliases:
                accsetwidget.set_alias(aliases[accsetwidget.name])
    def cbk_derepfilter_combo_changed(self):
        if self.derepfilter_combo.currentIndex() == 0:
            self.ignoreunknowntx_chk.setDisabled(True)
            self.includewgs_chk.setDisabled(True)
        else:
            self.ignoreunknowntx_chk.setDisabled(False)
            self.includewgs_chk.setDisabled(False)
            
#
    # # # # # # # # # # # # # # # # # # #
    #    Class Overrides                #
    #        - UI functionality.        #
    #          for base dframe types.   #
    #        -> Genetic Region          #
    #        -> Protein Cluster         #
    # # # # # # # # # # # # # # # # # # #

class UIEnabledRegion(dbdl.dframe.GeneticRegion):
    # - - - < Constructor > - - -
    def __init__(self, dataset, *args, **kwargs):
        self.dataset = dataset
        
        #Decides whether to reverse the region's features when rendering
        #self.flipped = False    
        #This is a weak set of live buttons and spacers
        #self.cluster_layout = weakref.WeakSet()
        
        super().__init__(*args, **kwargs)
        
        self.borders = 0
    def export_to_gb(self, output):
        record = SeqIO.SeqRecord(
            id=self.sc.accession,
            name=self.sc.accession,
            seq=Seq.Seq(foobar),
            description="A genetic region/cluster exported from CluSeek",
            features=[foobar],
            annotations={})
dframe.GeneticRegion = UIEnabledRegion
print("Validity check", dbdl.dframe.GeneticRegion is dframe.GeneticRegion)

class UIEnabledProteinCluster(dframe.ProteinCluster, qtc.QObject):
    sig_cluster_selected = qtc.Signal(bool)
    
    def __init__(self, id_, hierarchy, type_):
        super().__init__(id_=id_, hierarchy=hierarchy, type_=type_)
        
        # Additional variables for the UI declared.
        #   The _metadata variable is declared upstream so that basic
        #   clusters are both saving-compatible and their definitions
        #   make sense.
        
        self._metadata.update({
            "name": "", 
            "annotation": "", 
            "styledict_autostyles": {}})
        
        # The display settings unique to this cluster
        #   must only include names that are being changed in this cluster.
        #   Any keys that we want inherited from parent/default 
        #   must not be present in this dict.
        self.styledict_config = {}
        
        # Inherited settings overlaid with cluster-specific settings
        #   A new one is created every time settings change
        self.styledict = None
        self.redraw_id = 0
        
        self._selected = False
        
        self.clusteruis = weakref.WeakSet()
        
        self.update_styledict()
        
    @property
    def name(self):
        return(self._metadata["name"])
    @name.setter
    def name(self, text):
        self._metadata["name"] = text
    @property
    def annotation(self):
        return(self._metadata["annotation"])
    @annotation.setter
    def annotation(self, text):
        self._metadata["annotation"] = text
    @property
    def styledict_autostyles(self):
        return(self._metadata["styledict_autostyles"])
    
    
    def clear_all_styledicts(self):
        self._clear_all_styledicts()
        self.update_styledict()
    def _clear_all_styledicts(self):
        self.styledict_config = {}
        for subcluster in self.subclusters.values():
            subcluster._clear_all_styledicts()
    def get_name(self):
        return(self.name if self.name else self.id_)
    def set_name(self, text):
        self.name = text
        if not isinstance(text, str):
            print(f"Warning: Saved text is of type {type(text)}, value {text}")
    def set_annotation(self, annotation):
        self.annotation = text
    def set_styledict_config(self, new_config):
        self.styledict_config = new_config
        self.update_styledict()
        #self.hierarchy.sig_styledicts_changed.emit()
    def get_full_config(self):
        # Explicit color overrides autocolor,
        #   child customization overrides what is inherited
        #   from the parent.
        
        config,autostyle = dict(),dict()


        # Apply own autostyles
        for styledict_autostyle in self.styledict_autostyles.values():
            autostyle.update(styledict_autostyle)
        
        # Apply own styledict config
        config.update(self.styledict_config)

        if self._parent:
            parent_config,parent_autostyle = self._parent.get_full_config()
            config.update(parent_config)
            autostyle.update(parent_autostyle)
        return(config, autostyle)
    def update_styledict(self):
        # Combine inheritance, autostyles etc.
        self.styledict = dict(StyledictCreator.default_styledict)
        
        config,autostyle = self.get_full_config()
        
        # Config overrides autostyle to limit user frustration,
        #   though these clashes should be kept to a minimum.
        self.styledict.update(autostyle)
        self.styledict.update(config)
        
        # Change redraw_id to force a redraw
        self.redraw_id += 1
        if self.redraw_id > 999999:
            self.redraw_id = 0
        self.styledict["redraw_id"] = self.redraw_id

        self.update_widgets()
        # Propagate changes down the hierarchy
        for subcluster in self.subclusters.values():
            subcluster.update_styledict()
    def on_custom_info_changed(self):
        pass
    def get_annotation(self):
            if self.annotation:
                return(self.custom_info["annotation"])
            else:
                for annotation,count in self._get_composition().most_common():
                    if annotation not in {"hypothetical protein", "unknown"}:
                        return(annotation)
                return("unknown")
    def set_selected(self, value):
        assert isinstance(value, bool), "Selected must be bool."
        if self._selected == value:
            return
        
        self._selected = value
        
        if self._selected:
            self.hierarchy._cluster_selected(self)
        else:
            self.hierarchy._cluster_unselected(self)
        
        self.sig_cluster_selected.emit(value)
        
        # Change line to dotted line or vice versa
        if self._selected:
            self.set_autostyle("_selected", StyledictCreator.styledict_selected)
        else:
            self.unset_autostyle("_selected")
    def toggle_selected(self):
        self.set_selected(not self._selected)
    def get_selected(self):
        return(self._selected)
    def add_clusterui(self, clusterui):
        self.clusteruis.add(clusterui)
    def set_autostyle(self, key, autostyle):
        self.styledict_autostyles[key] = autostyle
        self.update_styledict()
        #if not delay_update:
        #   self.hierarchy.sig_styledicts_changed.emit()
    def unset_autostyle(self, key):
        if key in self.styledict_autostyles:
            del self.styledict_autostyles[key]
            self.update_styledict()
            #if not delay_update:
            #    self.hierarchy.sig_styledicts_changed.emit()
            return(True)
        return(False)
    def update_widgets(self):
        for widget in self.clusteruis:
            widget.update()
dframe.ProteinCluster = UIEnabledProteinCluster

class UIEnabledProteinClusterHierarchy(dframe.ProteinClusterHierarchy, qtc.QObject):
    sig_cluster_selected = qtc.Signal(object, bool)
    sig_cluster_removed = qtc.Signal(object)
    sig_cluster_added = qtc.Signal(object)
    sig_clusters_grouped = qtc.Signal(object, object)
    sig_clusters_ungrouped = qtc.Signal(object, object)
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.selection = set()
        
        self.signal_added_removed = True
    def _cluster_selected(self, cluster):
        self.selection.add(cluster)
        self.sig_cluster_selected.emit(cluster, True)
    def _cluster_unselected(self, cluster):
        self.selection.remove(cluster)
        self.sig_cluster_selected.emit(cluster, False)
    def clear_selection(self):
        for cluster in list(self.selection):
            cluster.set_selected(False)
    
    def remove_cluster(self, cluster):
        if self.signal_added_removed:
            self.sig_cluster_removed.emit(cluster)
        
        self._remove_cluster(cluster)
    def on_new_cluster_registered(self, cluster):
        if self.signal_added_removed:
            self.sig_cluster_added.emit(cluster)
    
    # = The many grouping methods
    def group_to(self, children, new_parent, auto_reparent=True):
        # Can't group to globally aligned clusters
        assert new_parent.type_ != "globally_aligned"
        
        children_with_parents = []
        for child in children:
            if child._parent is not None:
                children_with_parents.append(child)
        
        if len(children_with_parents) > 0:
            if not auto_reparent:
                return
            
            for child in children_with_parents:
                parent = child._parent
                parent.remove_subcluster(child)
                
                if parent is not new_parent:
                    parent.do_remove_check()
        
        # Change cluster type if it is modified
        if new_parent.type_ != "user_defined":
            new_parent.type_ = "user_defined"
        
        for child in children:
            new_parent.add_subcluster(child)
            child.update_styledict()
        
        self.sig_clusters_grouped.emit(list(children), [new_parent])
        return(new_parent)
    def ungroup(self, clusters):
        former_parents = []
        ungrouped = []
        for cluster in clusters:
            if cluster._parent:
                parent = cluster._parent
                parent.remove_subcluster(cluster)
                parent.do_remove_check()
                
                cluster.update_styledict()
                
                ungrouped.append(cluster)
                former_parents.append(parent)
                
            if cluster.subclusters:
                for subcluster in list(cluster.subclusters.values()):
                    cluster.remove_subcluster(subcluster)
                    cluster.update_styledict()
                    
                    ungrouped.append(subcluster)
                cluster.do_remove_check()
                former_parents.append(cluster)
        self.sig_clusters_ungrouped.emit(list(ungrouped), former_parents)
        return(ungrouped)
    def group(self, clusters):
        # Wrapper for group_to that automatically resolves
        #   the "to" part in a semi-intuitive way
        potential_parents = []
        children = []
        
        for cluster in clusters:
            if cluster.type_ == "globally_aligned":
                children.append(cluster)
            else:
                # Implicitly triggers only for 
                #   cluster.type_ != globally_aligned
                
                potential_parents.append(cluster)
                
                bottomlevels = list(cluster.get_bottomlevel_subclusters())
                children.extend(bottomlevels)
        
        #new_parent
        if len(potential_parents) == 0:
            # Make new parent
            new_parent = self.new_cluster(
                type_="user_defined")
        elif len(potential_parents) == 1:
            # Set this cluster as parent
            new_parent = potential_parents[0]
        else:
            # TODO: Alternatively ask the user.
            #   At the moment we just pick one arbitrarily.
            new_parent = max(
                potential_parents,
                key = lambda c: len(c.proteins)+1000*int(bool(c.styledict_config)))
        # TODO: When creating new parent, maybe check children
        #   to steal styledicts from?
        
        # Desired behavior:
        #   Group globally_aligned clusters only.
        #   If the higher group itself is chosen, its styledict 
        #   may be used. Regardless, its children will be taken
        #   away and put under a different parent.
        
        new_parent = self.group_to(children, new_parent)
        return(new_parent)
    def group_selected_clusters(self):
        selection = list(self.selection)
        
        self.clear_selection()
        
        new_parent = self.group(selection)
        
        new_parent.set_selected(True)
    def ungroup_selected_clusters(self):
        selection = list(self.selection)
        
        self.clear_selection()
        
        ungrouped = self.ungroup(selection)
        
        for cluster in ungrouped:
            cluster.set_selected(True)
dframe.ProteinClusterHierarchy = UIEnabledProteinClusterHierarchy

# This class is slotted into dbdl to control
#   download starting/progress.
#   Downloads are initiated by
#   the database managers, but must be approved
#   by the download manager.
#   This manager asks user for permission, and
#   reports download status.
class UIDownloadManager():
    def __init__(self):
        self.download_active = False
        self.size_of_download = None
        self.already_downloaded = None
        self.database = None
        self.progress_manager_update_signal = None
        self.abort = False
        self.abort_dialog = None
        self.abortlock = threading.Lock()
    def set_progress_manager_update_signal(self, signal):
        self.progress_manager_update_signal = signal
    
    def cbk_status_update(self, func, type, number_left):
        func(type, number_left)
    
    # Main thread side
    def ui_confirm_download_start(self):
        # OLD: To be removed
        #dialog = qtw.QMessageBox()
        #dialog.setText("Download starting!")
        #dialog.setInformativeText(
        #    f"Download from NCBI database?\n"
        #    f"There are {self.size_of_download} entries requested, "
        #    f"that are not available locally. If they are not downloaded,"
        #    f"the software will simply use whatever data is available locally.")
        #dialog.setStandardButtons(qtw.QMessageBox.Yes | qtw.QMessageBox.No)
        #response = dialog.exec_()
        #if response == qtw.QMessageBox.Yes:
        #    return True
        #else:
        #    return False
        
        # NEW: We do not ask for permission anymore to keep the
        #   pipeline going. The user can always interrupt.
        
        return True
    def _download_starting(self, type, number_to_download):
        # type = What kind of data is being downloaded. 
        #        Valid so far: "IPG", "GB" adding "BLASTP"
        # number_to_download = number of entries present locally,
        #                       which will need to be downloaded
        # False = Download is OK, proceed
        # True = Do not start download
        assert self.download_active==False, "Only one download can be running at the same time."
        self.download_active = True
        self.size_of_download = number_to_download
        self.already_downloaded = 0
        self.database = type
        self.abort = False
        self.abort_dialog = None
        
        # Ask the user whether he wants the download to happen.
        proceed = self.ui_confirm_download_start()
        abort = not proceed
        
        if not abort:
            print("Download starting...")
            # If this download is happening, create the progress dialog.
            AppRoot.progress_dialog.reinitialize(title = "Download Progress", 
                                              label_text="Downloading", 
                                              minimum=0, 
                                              maximum=self.size_of_download, 
                                              on_abort=self.on_abort)
            AppRoot.progress_dialog.show()
            AppRoot.progress_dialog.activateWindow()
        elif abort:
            print("Download cancelled.")
            # If it's not happening, reset relevant variables.
            self.download_active = False
        return abort
    def on_abort(self):
        self.abortlock.acquire()
        self.abort = True
        self.abortlock.release()
    def is_aborted(self):
        self.abortlock.acquire()
        aborted = self.abort
        self.abortlock.release()
        return(aborted)
    def _download_status(self, type, number_left_to_download):
        self.abortlock.acquire()
        if self.abort:
            print("Aborting...")
        else:
            self.already_downloaded = self.size_of_download - number_left_to_download
            AppRoot.progress_dialog.setValue(self.already_downloaded)
        # Same as before, True interrupts download.
        self.abortlock.release()
    def _download_ended(self, type, number_left):
        print("Download ended.")
        self.download_active = False
        AppRoot.progress_dialog.setValue(self.size_of_download)
        # It is the responsibility of the superior process
        #   to hide the progress_dialog.
        #AppRoot.progress_dialog.setVisible(False)
        self.progress_manager_update_signal = None
    # Work thread side
    #   These functions call their main thread counterparts
    #   in a thread-safe way.
    def download_starting(self, type, number_to_download):
        self.progress_manager_update_signal.emit(
            self._download_starting, type, number_to_download)
    def download_status(self, number_left_to_download):
        self.progress_manager_update_signal.emit(
            self._download_status, "", number_left_to_download)
        return(self.abort)
    def download_ended(self):
        self.progress_manager_update_signal.emit(
            self._download_ended, "", -1)
dbdl.DOWNLOAD_MANAGER = UIDownloadManager()

#    # # # # # # # # # # # # # # # # # # #
#    #           Info Classes            #
#    # # # # # # # # # # # # # # # # # # #

#feature protein cluster region sequence taxon
class InfoFrame(qtw.QWidget):
    def __init__(self, neiview):
        self.neiview = neiview
        super().__init__()
        self.__init_ui()
    def __init_ui(self):
        self.lay = qtw.QVBoxLayout()
        self.lay.setContentsMargins(0,0,0,0)
        self.lay.setSpacing(1)
        self.setLayout(self.lay)
        
        #
        self.content_scroller = qtw.QScrollArea()
        self.content_scroller.setHorizontalScrollBarPolicy(qtc.Qt.ScrollBarAlwaysOff)
        self.content_scroller.setVerticalScrollBarPolicy(qtc.Qt.ScrollBarAsNeeded)
        self.content_scroller.setWidgetResizable(True)
        #self.content_scroller.setMinimumWidth(270)
        #self.content_scroller.setMaximumWidth(270)
        self.lay.addWidget(self.content_scroller)
        #old_scrollable = self.content_scroller.widget()
        
        
        self.content_scrollable = qtw.QWidget()
        
        self.setSizePolicy(
            qtw.QSizePolicy.Minimum,
            qtw.QSizePolicy.Expanding)
        self.content_scroller.setSizePolicy(
            qtw.QSizePolicy.Minimum,
            qtw.QSizePolicy.Expanding)
        self.content_scrollable.setSizePolicy(
            qtw.QSizePolicy.Minimum,
            qtw.QSizePolicy.Expanding)
        
        
        old_resize = self.content_scroller.resizeEvent
        def scroller_resized(ev):
            old_resize(ev)
            self.content_scrollable.setFixedWidth(
                self.neiview.lefttabs_tabs.geometry().width()-25)
            #self.neiview.lefttabs_tabs.setMaximumWidth(
            #    self.content_scrollable.geometry().width()+50)
        self.content_scroller.resizeEvent = scroller_resized
            
        #self.content_scrollable.setMinimumWidth(260)
        #self.content_scrollable.setMaximumWidth(260)

        self.content_lay = qtw.QVBoxLayout()
        self.content_lay.setContentsMargins(2,2,10,2)
        self.content_lay.setSpacing(6)
        self.content_scrollable.setLayout(self.content_lay)
        
        #old_scrollable.setParent(None)
        
        #
        self.info_taxon    = InfoCategory_Taxon()
        self.info_sequence = InfoCategory_Sequence()
        self.info_region   = InfoCategory_Region()
        self.info_feature  = InfoCategory_Feature()
        self.info_protein  = InfoCategory_Protein()
        self.info_cluster  = InfoCategory_Cluster(self.neiview)
        
        self.content_lay.addWidget(self.info_taxon)
        self.content_lay.addWidget(self.info_sequence)
        self.content_lay.addWidget(self.info_region)
        self.content_lay.addWidget(self.info_feature)
        self.content_lay.addWidget(self.info_protein)
        self.content_lay.addWidget(self.info_cluster)
        
        self.content_lay.addStretch()
        #
        #self.content_lay.addStretch()
        
        # This last
        self.content_scroller.setWidget(self.content_scrollable)
        # This lastest
        self.display_info()
    def display_quick(self, item, invert_cluster=False):
        kwargs = {"feature":None,"protein":None,"cluster":None,
                  "region":None,"sequence":None,"taxon":None,
                  "cluster_anchor": None, "partial": False}
        if isinstance(item, ClusterWidget) and item.represents == "bottomlevel" or (invert_cluster and item.represents=="toplevel"):
            kwargs["cluster"] = item.cluster
            kwargs["cluster_anchor"] = item.cluster
            #if item.cluster in self.info_cluster.last_clu_family:
            #    kwargs["partial"]=True
        elif isinstance(item, ClusterWidget) and item.represents == "toplevel" or (invert_cluster and item.represents=="bottomlevel"):
            kwargs["cluster"] = item.cluster.get_toplevel()
            kwargs["cluster_anchor"] = item.cluster
            #if item.cluster in self.info_cluster.last_clu_family:
            #    kwargs["partial"]=True
        elif isinstance(item, TranscriptUI):
            kwargs["feature"] = item.ft
            kwargs["sequence"] = item.ft.sc
            kwargs["region"] = item.tail.regui.reg
            kwargs["taxon"] = item.ft.sc.tx
            kwargs["cluster"] = item.cluster.get_toplevel()
            kwargs["cluster_anchor"] = item.cluster
            if item.ft is not None:
                kwargs["protein"] = item.ft.ref
        elif isinstance(item, FeatureUI) and item.ft is not None:
            kwargs["feature"] = item.ft
            kwargs["sequence"] = item.ft.sc
            kwargs["region"] = item.tail.regui.reg
            kwargs["taxon"] = item.ft.sc.tx
        self.display_info(**kwargs)
    def display_info(self, feature=None, protein=None, cluster=None, cluster_anchor=None, 
                           region=None, sequence=None, taxon=None, partial=False):
        if taxon:
            self.info_taxon.set_unavailable(False)
            self.info_taxon.display(taxon)
        elif partial:
            pass
        else:
            self.info_taxon.set_unavailable(True)
        
        if sequence:
            self.info_sequence.set_unavailable(False)
            self.info_sequence.display(sequence)
        elif partial:
            pass
        else:
            self.info_sequence.set_unavailable(True)
        
        if region:
            self.info_region.set_unavailable(False)
            self.info_region.display(region)
        elif partial:
            pass
        else:
            self.info_region.set_unavailable(True)
        
        if feature:
            self.info_feature.set_unavailable(False)
            self.info_feature.display(feature)
        elif partial:
            pass
        else:
            self.info_feature.set_unavailable(True)
        
        if protein:
            self.info_protein.set_unavailable(False)
            self.info_protein.display(protein)
        elif partial:
            pass
        else:
            self.info_protein.set_unavailable(True)
        
        if cluster:
            self.info_cluster.set_unavailable(False)
            self.info_cluster.display(cluster, cluster_anchor, protein)
        elif partial:
            pass
        else:
            self.info_cluster.set_unavailable(True)
        
class InfoCategory(qtw.QWidget):
    category_name = "N/A"
    def __init__(self):
        super().__init__()
        self.__init_ui()
    def __init_ui(self):
        self.category_lay = qtw.QVBoxLayout()
        self.category_lay.setContentsMargins(0,0,0,0)
        self.category_lay.setSpacing(1)
        self.setLayout(self.category_lay)
        
        self.toggle_btn = self.FoldingButton(self.category_name)
        self.category_lay.addWidget(self.toggle_btn)
        self.toggle_btn.state_changed.connect(
            self.cbk_toggle_contents)
        self.toggle_btn.setMaximumHeight(15)
        
        self.contents = qtw.QWidget()
        self.category_lay.addWidget(self.contents)
        
    def cbk_toggle_contents(self, state):
        self.contents.setHidden(not state)
    def set_unavailable(self, unavailable):
        self.toggle_btn.set_unavailable(unavailable)
        if unavailable:
            self.contents.setHidden(True)
        else:
            self.contents.setHidden(not self.toggle_btn.state)
    class FoldingButton(qtw.QWidget):
        state_changed = qtc.Signal(bool)
        def __init__(self, text):
            super().__init__()
            self.text = text
            self.state = True
            self.unavailable = False
            self.__init_ui()
        def set_unavailable(self, unavailable):
            self.unavailable = unavailable
            self.label.setText(self.get_text())
        def get_text(self):
            if self.unavailable:
                return("  <i>"+self.text+"</i>")
            if self.state:
                return("↓ <b>"+self.text+"</b>")
            else:
                return("● <b>"+self.text+"</b>")
        def __init_ui(self):
            self.lay = qtw.QHBoxLayout()
            self.lay.setContentsMargins(0,0,0,0)
            self.lay.setSpacing(1)
            self.setLayout(self.lay)
            
            def ignore(event):
                event.ignore()
            self.label = qtw.QLabel(self.get_text())
            self.label.mouseReleaseEvent = ignore
            self.lay.addWidget(self.label)
            
            self.line = qtw.QFrame()
            self.line.setFrameShape(qtw.QFrame.HLine)
            self.line.setFrameShadow(qtw.QFrame.Sunken)
            self.line.setSizePolicy(qtw.QSizePolicy.Expanding, 
                                    qtw.QSizePolicy.Preferred)
            self.lay.addWidget(self.line)
        def _on_clicked(self):
            if self.unavailable:
                return
            self.state = not self.state
            self.label.setText(self.get_text())
            self.state_changed.emit(self.state)
        def mouseReleaseEvent(self, event):
            if event.button() == qtc.Qt.LeftButton:
                self._on_clicked()
class InfoCategory_Feature(InfoCategory, Ui_InfoFeature):
    category_name = "Genetic feature"
    def __init__(self):
        super().__init__()
        self.__init_ui()
    def __init_ui(self):
        self.setupUi(self.contents)
        
        self.info_strand_lbl.setHidden(True)
        #
    def display(self, ft):
        self.info_type_lbl.setText(ft.type)
        self.info_length_lbl.setText(str(ft.length()))
        self.info_strand_lbl.setText(ft.strand)
        self.info_start_lbl.setText(str(ft.start))
        self.info_stop_lbl.setText(str(ft.stop))
class InfoCategory_Region(InfoCategory, Ui_InfoRegion):
    category_name = "Genetic region"
    def __init__(self):
        super().__init__()
        self.__init_ui()
    def __init_ui(self):
        self.setupUi(self.contents)
        #
    def display(self, reg):
        self.info_start_lbl.setText(str(reg.start))
        self.info_stop_lbl.setText(str(reg.stop))
        self.info_length_lbl.setText(str(reg.length()))
class InfoCategory_Sequence(InfoCategory, Ui_InfoSequence):
    category_name = "Sequence"
    def __init__(self):
        super().__init__()
        self.__init_ui()
        self.seq = None
    def __init_ui(self):
        self.setupUi(self.contents)
        self.info_genbank_btn.clicked.connect(self.cbk_genbank)
    def cbk_genbank(self):
        webbrowser.open(
            f"https://www.ncbi.nlm.nih.gov/nuccore/{self.seq.accession}")
    def display(self, seq):
        self.info_sequence_acc.setText(seq.accession)
        self.seq = seq
class InfoCategory_Taxon(InfoCategory, Ui_InfoTaxon):
    category_name = "GenBank taxon"
    def __init__(self):
        super().__init__()
        self.__init_ui()
        self.tx = None
    def __init_ui(self):
        self.setupUi(self.contents)
        self.info_genbank_btn.clicked.connect(
            self.cbk_genbank)
    def cbk_genbank(self):
        webbrowser.open(
            f"https://www.ncbi.nlm.nih.gov/Taxonomy/Browser/wwwtax.cgi?id={self.tx.taxid}")
    def display(self, tx):
        self.info_sciname_lbl.setText(tx.sciname)
        self.info_strain_lbl.setText(tx.strain)
        self.info_lineage_lbl.setText("TBD")
        self.info_taxid_lbl.setText(tx.taxid)
        
        self.tx = tx
class InfoCategory_Protein(InfoCategory, Ui_InfoProtein):
    category_name = "Protein"
    def __init__(self):
        super().__init__()
        self.__init_ui()
        self.pt = None
    def __init_ui(self):
        self.setupUi(self.contents)
        self.btn_genbank.clicked.connect(self.cbk_genbank)
    def cbk_genbank(self):
        webbrowser.open(f"https://www.ncbi.nlm.nih.gov/protein/{self.pt.accession}")
    def display(self, pt):
        self.info_accession_lbl.setText(pt.accession)
        self.info_name_lbl.setText(pt.name if pt.name else "N/A")
        self.info_annotation_lbl.setText(pt.type if pt.type else "N/A")
        self.info_AAsequence.setText(pt.seq if pt.seq else "N/A")
        self.info_length_aa.setText(str(len(pt.seq))+" aa" if pt.seq else "N/A")
        
        self.pt = pt
class InfoCategory_Cluster(InfoCategory, Ui_InfoCluster):
    category_name = "Protein group"
    def __init__(self, neiview):
        super().__init__()
        self.neiview = neiview
        self.clu = None
        self.last_clu_toplevel = None
        self.last_clu_localgroup = None
        self.__init_ui()
    def __init_ui(self):
        self.setupUi(self.contents)
        #self.apply_name_btn.clicked.connect(self.cbk_apply_name)
        self.info_name_ledit.editingFinished.connect(self.cbk_apply_name)
        #self.show_related_network_btn.clicked.connect(
        #    self.cbk_show_local_group_network)
        
        
        self.cluster_network = ClusterNetworkViewer2(neiview=self.neiview)
        self.network_view_container.setMinimumHeight(210)
        self.network_view_container.setMinimumWidth(210)
        self.cluster_network.setParent(self) 
        self.network_view_container_lay.addWidget(self.cluster_network)
        
        self.cluster_network.set_mode_mini()
        
        self.cluster_network.closeEvent = self.cbk_cluster_network_closed
        
        self.pop_out_hierarchy_btn.clicked.connect(
            self.cbk_toggle_cluster_network_window)
        
        #
        #self.hierarchy_viewer = ClusterHierarchyViewer(None, neiview=self.neiview)
        #self.hierarchy_scroller.setWidget(self.hierarchy_viewer)
        #self.hierarchy_scroller.closeEvent = self.cbk_hierarchy_closed
        
        self.related_groups_clusterwidget = ClusterContainerWidget(
            self.neiview,
            contents_accept_drops = False,
            row_size = 4)
        self.main_layout.replaceWidget(
            self.cluster_widget_placeholder,
            self.related_groups_clusterwidget)
        self.main_layout.update()
        self.cluster_widget_placeholder.setHidden(True)
        del self.cluster_widget_placeholder
        
    def cbk_apply_name(self):
        text = self.info_name_ledit.text()
        self.clu.set_name(text if text else None)
        self.clu.update_styledict()
        self.neiview.cbk_clunames_changed()
    def cbk_cluster_network_closed(self, event):
        self.cbk_toggle_cluster_network_window()
        event.ignore()
    def cbk_toggle_cluster_network_window(self):
        #"local_group"
        
        if self.cluster_network.isWindow():
            self.cluster_network.setParent(self)
            self.network_view_container_lay.addWidget(
                self.cluster_network)
            self.cluster_network.setHidden(False)
            self.cluster_network.set_mode_mini()
            self.pop_out_hierarchy_btn.setText("Show windowed")
        else:
            self.network_view_container_lay.removeWidget(
                self.cluster_network)
            self.cluster_network.setParent(None)
            #self.cluster_network.setGeometry(150, 150, 900, 900)
            self.cluster_network.set_mode_full()
            self.cluster_network.show()
            self.cluster_network.activateWindow()
            self.cluster_network.showMaximized()
            self.pop_out_hierarchy_btn.setText("Show inline")
        self.main_layout.update()
        #self.neiview.cluster_network.show_clusters([self.clu], mode="standard")
    def display(self, clu, clu_anchor, pt):
        self.info_id_lbl.setText(clu.id_)
        self.info_name_ledit.setText(clu.name if clu.name else "")
        self.info_member_tebr.setText(", ".join(list(clu.proteins)))
        self.info_clustertype.setText(clu.type_)
        
        if self.cluster_network.isWindow():
            self.cluster_network.activateWindow()
        
        lengths = []
        for p in clu.proteins.values():
            if p.seq and len(p.seq)>0:
                lengths.append(len(p.seq))
        range=f"{round(sum(lengths)/len(lengths),1)} ({min(lengths)}-{max(lengths)})"
        self.info_proteinlength.setText(range)
        
        # * Count incidences, maybe this should be a cluster method
        raw_counts = 0
        for protein in clu.proteins.values():
            raw_counts += len(protein.fts)
        self.protein_count_lbl.setText(
            f"{raw_counts} instances")
        
        # * Count annotations
        annotation_count = clu._get_composition()
        annotation_count_list = []
        for key in annotation_count:
            annotation_count_list.append((key, annotation_count[key]))
        annotation_count_list.sort(key=lambda x: x[1], reverse=True)
        
        self.info_annotations_tebr.setText("\n".join(
            [f"{key}: {value}" for key,value in annotation_count_list]))
        
        # No idea what this is
        self.clu = clu
        self.clu_anchor = clu_anchor
        
        # * Set cluster network display
        # First, check if the new cluster is in the graph we already have
        
        local_groups = clu.get_local_groups()
        if ((self.cluster_network.get_subset_mode() == "standard"
                and clu.get_toplevel() != self.last_clu_toplevel)
              or (self.cluster_network.get_subset_mode() == "local_group"
                and not (local_groups & self.last_clu_localgroup))):
            # If not, re-generate
            self.cluster_network.show_clusters([clu], mode="standard")
            
            self.last_clu_toplevel = clu.get_toplevel()
            self.last_clu_localgroup = local_groups
        
        # * Set related groups
        self.related_groups_clusterwidget.clear()
        for related_cluster in {c.get_toplevel() for c in clu.get_related_clusters()}:
            self.related_groups_clusterwidget.add_cluster(related_cluster)
        
        def register_lineages(cluster):
            for subcluster in cluster.subclusters.values():
                self.lineages[subcluster] = subcluster.get_lineage()
                register_lineages(subcluster)
# UI Remake Classes
class ClusterHierarchyViewer(qtw.QWidget):
    cluster_clicked = qtc.Signal(object)
    DEFAULT_CLUSTERWIDGET_SIZE = (50, 20)
    GAP_X = 15
    GAP_Y = 26
    def __init__(self, parent, anchored=True, neiview=None):
        super().__init__(parent)
        
        self.lineages = None
        
        self.anchored = anchored
        self.neiview = neiview
        
        self.clusterwidgets = {}
        
        self.background_lines = []
        self.background_lines_red = []
        self.background_texts = []
    def load_tree(self, cluster_anchor):
        def register_lineages(cluster):
            for subcluster in cluster.subclusters.values():
                self.lineages[subcluster] = subcluster.get_lineage()
                register_lineages(subcluster)
        
        self.cluster_anchor = cluster_anchor
        cluster_toplevel = cluster_anchor.get_toplevel()
        self.lineages = {}
        
        self.lineages[cluster_toplevel] = cluster_toplevel.get_lineage()
        register_lineages(cluster_toplevel)
        
        self._display_clusters(list(self.lineages))
    def cbk_on_cluster_clicked(self, cluster):
        self.cluster_clicked.emit(cluster)
    def _display_clusters(self, clusters_to_display):
        def get_line(clusterA, clusterB):
            A = clusterA.geometry()
            B = clusterB.geometry()
            start = qtc.QPoint(A.left()+A.width()*0.5,
                               A.top()+A.height()*0.5)
            end = qtc.QPoint(B.left()+B.width()*0.5,
                             B.top()+B.height()*0.5)
            middle1 = qtc.QPoint(start.x(), 
                                 A.top()+A.height()+math.ceil(self.GAP_Y/2))
            middle2 = qtc.QPoint(end.x(), middle1.y())
            return(start, middle1, middle2, end)
        def add_line(start, middle1, middle2, end):
            self.background_lines.extend([
                qtc.QLine(start, middle1),
                qtc.QLine(middle1, middle2),
                qtc.QLine(middle2, end)])
        def add_line_red(start, middle1, middle2, end):
            self.background_lines_red.extend([
                qtc.QLine(start, middle1),
                qtc.QLine(middle1, middle2),
                qtc.QLine(middle2, end)])
        def add_text(left,top,width,height,cluster):
            x = left+math.floor(width/2)+3
            y = top+height+math.floor(self.GAP_Y/2)-2
            text = f"n={len(cluster.proteins)}"
            self.background_texts.append((x,y,text))
        # This should discard all existing clusterwidgets
        #   and references to them should only be from this dict
        for clusterwidget in self.clusterwidgets.values():
            clusterwidget.setParent(None)
        
        self.clusterwidgets = {}
        self.background_lines = []
        self.background_lines_red = []
        self.background_texts = []
        
        for cluster in clusters_to_display:
            self.clusterwidgets[cluster] = ClusterWidget(self, cluster, 
                                                         neiview=self.neiview)
            self.clusterwidgets[cluster].cluster_clicked.connect(self.cbk_on_cluster_clicked)
        
        # Which level to display a cluster at.
        #   0 is all the way at the bottom
        cluster_levels = {cluster:0 for cluster in self.lineages}
        for lineage in self.lineages.values():
            for i in range(len(lineage)):
                cluster_levels[lineage[i]] = max(cluster_levels[lineage[i]], i)
        
        levels = {}
        for cluster in cluster_levels:
            i = cluster_levels[cluster]
            if i not in levels:
                levels[i] = []
            levels[i].append(cluster)
        
        # * Base Level (=0)
        # All clusters are sorted by lineage (using .id_ to facilitate > < comparison)
        base_level = sorted(levels[0], key=lambda x: [y.id_ for y in self.lineages[x][::-1]])
        for i in range(len(base_level)):
            cluster = base_level[i]
            left = (self.DEFAULT_CLUSTERWIDGET_SIZE[0]+self.GAP_X) * i
            top = (self.DEFAULT_CLUSTERWIDGET_SIZE[1]+self.GAP_Y)*(len(levels)-1)
            width = self.DEFAULT_CLUSTERWIDGET_SIZE[0]
            height = self.DEFAULT_CLUSTERWIDGET_SIZE[1]
            self.clusterwidgets[cluster].setGeometry(
                left, top, width, height)
            self.clusterwidgets[cluster].setHidden(False)
            add_text(left,top,width,height,cluster)
        
        # subsequent levels
        for level_i in sorted(list(levels)):
            if level_i == 0: continue
            level = levels[level_i]
            for cluster in level:
                x_positions = [self.clusterwidgets[child].geometry().left()\
                                for child in cluster.subclusters.values()]
                left = round(sum(x_positions)/len(x_positions),0)
                top = ((self.DEFAULT_CLUSTERWIDGET_SIZE[1]+self.GAP_Y)
                        *(len(levels)-level_i-1))
                width = self.DEFAULT_CLUSTERWIDGET_SIZE[0]
                height = self.DEFAULT_CLUSTERWIDGET_SIZE[1]
                self.clusterwidgets[cluster].setGeometry(
                    left, top, width, height)
                self.clusterwidgets[cluster].setHidden(False)
                add_text(left,top,width,height,cluster)
                for child in cluster.subclusters.values():
                    add_line(*get_line(self.clusterwidgets[cluster], 
                                       self.clusterwidgets[child]))
        
        # Draw anchor cluster lineage
        if self.anchored:
            lineage = self.cluster_anchor.get_lineage()
            for i in range(len(lineage)-1):
                child = lineage[i]
                parent = lineage[i+1]
                add_line_red(*get_line(self.clusterwidgets[parent],
                                       self.clusterwidgets[child]))
        
        # finally set geometry of self
        self.setFixedSize(
            (self.DEFAULT_CLUSTERWIDGET_SIZE[0]+self.GAP_X)*(len(levels[0])),
            (self.DEFAULT_CLUSTERWIDGET_SIZE[1]+self.GAP_Y)*(len(levels))
        )
        self.redraw_clusters()
    def redraw_clusters(self):
            for clusterwidget in self.clusterwidgets.values():
                clusterwidget.repaint()
            self.repaint()
    def paintEvent(self, event):
        painter = qtg.QPainter()
        painter.begin(self)
        
        painter.drawLines(self.background_lines)
        for x,y,text in self.background_texts:
            painter.drawText(x,y,text)
        
        painter.setPen(qtg.QColor(255,0,0))
        painter.drawLines(self.background_lines_red)
        painter.end()
class StyledictCreator(qtw.QWidget, Ui_StyledictCreator):
    # Preview widget is currently hidden
    line_styles = {"Solid": qtc.Qt.SolidLine,
                   "Dashed": qtc.Qt.DashLine,
                   "Dotted": qtc.Qt.DotLine}
    line_styles_reverse = {}
    for key in line_styles:
        value = line_styles[key]
        line_styles_reverse[value] = key
    
    fill_patterns = {
        "No Pattern": None,
        "Solid": qtc.Qt.SolidPattern,
        "Dense-1": qtc.Qt.Dense1Pattern,
        "Dense-2": qtc.Qt.Dense2Pattern,
        "Dense-3": qtc.Qt.Dense3Pattern,
        "Dense-4": qtc.Qt.Dense4Pattern,
        "Dense-5": qtc.Qt.Dense5Pattern,
        "Dense-6": qtc.Qt.Dense6Pattern,
        "Dense-7": qtc.Qt.Dense7Pattern,
        "Horizontal": qtc.Qt.HorPattern,
        "Vertical": qtc.Qt.VerPattern,
        "Cross": qtc.Qt.CrossPattern,
        "Forward Diagonal": qtc.Qt.FDiagPattern,
        "Backward Diagonal": qtc.Qt.BDiagPattern,
        "Diagonal Cross": qtc.Qt.DiagCrossPattern}
    fill_patterns_reverse = {}
    for key in fill_patterns:
        value = fill_patterns[key]
        fill_patterns_reverse[value] = key

    default_styledict = { 
        # Options (qtc.Qt.)
        #   Pen:
        #       Lines: SolidLine, DashLine, DotLine, DashDotLine, DashDotDotLine
        #       LineCaps: SquareCap, FlatCap, RoundCap
        #       LineJoin: BevelJoin, MiterJoin, RoundJoin
        #   Brush:
        #       Pattern: SolidPattern, Dense1Pattern...Dense7Pattern, NoBrush,
        #                HorPattern, VerPattern, CrossPattern, BDiagPattern,
        #                FDiagPattern, DiagCrossPattern
        #       Maybe don't use gradients or texturepatterns if possible.
        "line_color": qtg.QColor(0,0,0,255),
        "line_pattern": qtc.Qt.SolidLine,
        "line_thickness": 1,
        "background_color": qtg.QColor(255,255,255,255),
        "foreground_color": qtg.QColor(255,255,255,255),
        "foreground_pattern": None,
        "text_color": qtg.QColor(0,0,0,255),
        "font_family": "Arial", #qtg.QFont().family(),
        "font_italic": False,
        "font_bold": False,
        "font_underlined": False,
        "text": None, # Is only ever None
        "redraw_id": 0,
    }
    
    styledict_selected = {
        "line_thickness": 3,}
    
    styledict_rarecluster = {
        "background_color": qtg.QColor(160,160,160,255),}
    
    styledict_singletoncluster = {
        "background_color": qtg.QColor(64,64,64,255),
        "text_color": qtg.QColor(255,255,255,255)}
    
    editing_widgets = [
            ("line_color", "outline_color_inherit", "outline_color_btn"),
            #("line_pattern", "outline_style_inherit", "outline_style_combo"),
            #("line_thickness", "outline_thickness_inherit", "outline_thickness_spinbox"),
            ("background_color", "background_color_inherit", "background_color_btn"),
            ("foreground_color", "foreground_color_inherit", "foreground_color_btn"),
            ("foreground_pattern", "foreground_pattern_inherit", "foreground_pattern_combo"),
            ("text", "text_title_inherit", "text_title_ledit"),
            ("text_color", "text_color_inherit", "text_color_btn"),
            ("font_family", "text_font_inherit", "text_font_combo"),
            ("font_bold", "text_bold_inherit", "text_formatting_bold"),
            ("font_italic", "text_italic_inherit", "text_formatting_italic"),
            ("font_underlined", "text_underlined_inherit", "text_formatting_underlined"),
        ]

    def __init__(self, neiview):
        super().__init__()
        
        # Ignore update calls while the widget is being built
        self.ignore_update_calls = True
        
        # Neiview must be defined
        self.neiview = neiview

        # ... before initing UI
        self.init_ui()
        

        self.neiview.neidata.cluster_hierarchy.sig_cluster_selected.connect(
            self.cbk_cluster_selection_changed)

        self.anchor_cluster = None
        self.displayed_cluster = None
        
        self.current_styledict = None
        self.current_styledict_config = None
        self.preview_styledict = dict(self.default_styledict)
        
        self.last_style = None
        
        self._load_timer = None
        
        # Used to lock update calls when modifying multiple widgets
        #   at once. It is up to the relevant code to then call
        #   self.update_preview() when finished.
        
        # Stop ignoring calls when done
        self.ignore_update_calls = False
    def init_ui(self):
        self.setupUi(self)
        
        # Combo
        self.outline_style_combo.addItems(list(self.line_styles))
        self.foreground_pattern_combo.addItems(list(self.fill_patterns))
        
        # Control button callbacks
        #self.control_reset.clicked.connect(self.cbk_reset)
        #self.control_reset_to_default.clicked.connect(self.cbk_reset_to_default)
        #self.control_cancel.clicked.connect(self.cbk_cancel)
        #self.control_apply.clicked.connect(self.cbk_apply)
        
        # Color setter button callbacks
        self.outline_color_btn.clicked.connect(self.cbk_select_outline_color)
        self.background_color_btn.clicked.connect(self.cbk_select_background_color)
        self.foreground_color_btn.clicked.connect(self.cbk_select_foreground_color)
        self.text_color_btn.clicked.connect(self.cbk_select_font_color)

        # Configuration interface callbacks
        self._initialize_inheritance_callbacks() 
        
        # Color dialog
        self.color_dialog = qtw.QColorDialog()
        
        # Preview
        self.preview = FeatureUI(self, None, "dummy")
        self.preview.setFixedSize(
            NeiView.get_default_display_policy()["size_fixed_width"],
            NeiView.get_default_display_policy()["size_fixed_height"])
        self.container_layout.addWidget(self.preview)
        
        # The preview widget doesn't work correctly, so we just hide it :)
        self.preview.setHidden(True)
        
        # Hierarchy viewer
        self.hierarchy_viewer = ClusterHierarchyViewer(self, neiview=self.neiview)
        self.hierarchy_viewer.cluster_clicked.connect(self.cbk_cluster_clicked)
        self.cluster_hierarchy_lay.addWidget(self.hierarchy_viewer)
        
        # manual overrides, some temporary:
        self.outline_style_combo.setDisabled(True)
        self.outline_thickness_spinbox.setDisabled(True)
        
        self.outline_style_inherit.setChecked(True)
        self.outline_thickness_inherit.setChecked(True)
        self.outline_style_inherit.setDisabled(True)
        self.outline_thickness_inherit.setDisabled(True)
        

    def _initialize_inheritance_callbacks(self):
        # Generates a callback which informs the contents_changed function
        #   which widget changed, so it can use the above dictionary to change
        #   the appropriate inheritance checkmarks and update the UI.
        def create_callback_for(originator_widget):
            def cbk():
                self.contents_changed(originator_widget)
            return(cbk)
        
        all_widgets = []
        for variable, inherit_chk_name, edit_widget_name in self.editing_widgets:
            inherit_chk = getattr(self, inherit_chk_name)
            edit_widget = getattr(self, edit_widget_name)
            all_widgets.append(inherit_chk)
            all_widgets.append(edit_widget)

        for widget in all_widgets:
            if isinstance(widget, qtw.QComboBox):
                widget.currentIndexChanged.connect(
                    create_callback_for(widget))
            elif isinstance(widget, qtw.QAbstractButton):
                widget.clicked.connect(
                    create_callback_for(widget))
            elif isinstance(widget, qtw.QLineEdit):
                widget.editingFinished.connect(
                    create_callback_for(widget))
            elif isinstance(widget, qtw.QSpinBox):
                widget.valueChanged.connect(
                    create_callback_for(widget))
            else:
                raise ValueError("Unknown widget in customization menu.")
        
        
    # Basic operation
    def cbk_cluster_selection_changed(self, cluster, value):
        
        
        def temp():
            selection = self.neiview.neidata.cluster_hierarchy.selection
            
            if self.displayed_cluster in selection:
                pass
            elif len(selection) > 0:
                new = sorted(list(selection), key=lambda x: len(x.styledict_config), reverse=True)[0]
                
                self.enable_all()
                self.load(new)
            else:
                self.load({})
                self.disable_all()
            
            self.update_selected_clusters_label()
            
            self._load_timer = None
        
        if self._load_timer:
            AppRoot.timer.kill_timer(self._load_timer)
        
        self._load_timer = AppRoot.timer.after_delay_do(temp, NEI_HOLD_ACTIVATE_DELAY+10)
    def update_selected_clusters_label(self):
        if not self.displayed_cluster:
            self.selected_cluster_lbl.setText("No protein group selected!")
            return
        
        selection = self.neiview.neidata.cluster_hierarchy.selection
        if len(selection) > 1:
            print(len(selection))
            addition = ", changes will also apply to: " + ", ".join([cluster.get_name() for cluster in (selection-{self.displayed_cluster})])
        else:
            addition = ""

        self.selected_cluster_lbl.setText(
            self.displayed_cluster.get_name() + addition)
    def disable_all(self):
        for variable, inherit_chk_name, edit_widget_name in self.editing_widgets:
            inherit_chk = getattr(self, inherit_chk_name)
            edit_widget = getattr(self, edit_widget_name)

            inherit_chk.setDisabled(True)
            edit_widget.setDisabled(True)
    def enable_all(self):
        for variable, inherit_chk_name, edit_widget_name in self.editing_widgets:
            inherit_chk = getattr(self, inherit_chk_name)
            edit_widget = getattr(self, edit_widget_name)

            inherit_chk.setDisabled(False)
            edit_widget.setDisabled(False)
    def load(self, input_):
        if isinstance(input_, UIEnabledProteinCluster):
            # If we're loading a cluster, we want to display hierarchy
            # TODO: Implement hierarchy display and switching within hierarchy
            self.anchor_cluster = input_
            self.displayed_cluster = input_.get_toplevel()
            self.hierarchy_viewer.load_tree(input_)
            self._load_cluster(input_)
            self._active_hierarchy = input_.hierarchy
        else:
            # Otherwise assume it is a basic styledict
            self.anchor_cluster = None
            self.displayed_cluster = None
            
            self.current_styledict = input_
            self.current_styledict_config = input_

            self._load_styledict(input_, input_)
    def _load_cluster(self, cluster):
        self.displayed_cluster = cluster
        self._load_styledict(cluster.styledict, cluster.styledict_config)
        self.current_styledict = cluster.styledict
        self.current_styledict_config = cluster.styledict_config
        
        self.update_selected_clusters_label()
    def _load_styledict(self, styledict_full, styledict_config={}):
        self.ignore_update_calls = True

        # To avoid issues with an incomplete styledict being loaded
        #   we use a default styledict where possible.
        default = dict(StyledictCreator.default_styledict)
        default.update(styledict_config)
        styledict = default
        del default

        # * * Line
        # * Color
        self.outline_color_btn.color = styledict["line_color"]
        
        # * Style
        self.outline_style_combo.setCurrentIndex(
            self.outline_style_combo.findText(
              self.line_styles_reverse[styledict["line_pattern"]], 
              qtc.Qt.MatchFixedString|qtc.Qt.MatchContains))
        
        # * Thickness
        self.outline_thickness_spinbox.setValue(styledict["line_thickness"])
        
        
        # * * Background
        self.background_color_btn.color = styledict["background_color"]
        
        # * * Foreground
        # * Color
        self.foreground_color_btn.color = styledict["foreground_color"]
        
        # * Pattern
        self.foreground_pattern_combo.setCurrentIndex(
            self.foreground_pattern_combo.findText(
              self.fill_patterns_reverse[styledict["foreground_pattern"]], 
              qtc.Qt.MatchFixedString|qtc.Qt.MatchContains))
        
        # * * Text
        # * Title
        if self.displayed_cluster:
            self.text_title_ledit.setDisabled(False)
            self.text_title_ledit.setText(self.displayed_cluster.get_name())
        else:
            self.text_title_ledit.setText("")
            self.text_title_ledit.setDisabled(True)
        
        # * Color, text_color
        self.text_color_btn.color = styledict["text_color"]
        
        # * Font
        self.text_font_combo.setCurrentFont(styledict["font_family"])
        
        # * Formatting
        self.text_formatting_italic.setChecked(styledict["font_italic"])
        self.text_formatting_bold.setChecked(styledict["font_bold"])
        self.text_formatting_underlined.setChecked(styledict["font_underlined"])
        
        # * * Inheritance
        if self.displayed_cluster and self.displayed_cluster._parent:
            parent_config, parent_autostyle = self.displayed_cluster._parent.get_full_config()
            del parent_autostyle
        else:
            parent_config = None
        
        for variable, inherit_btn_name, setter_widget_name in self.editing_widgets:
            inherit_btn = getattr(self, inherit_btn_name)
            setter_widget = getattr(self, setter_widget_name)

            if parent_config and variable in parent_config:
                inherit_btn.setDisabled(False)
                inherit_btn.setChecked(True)
                inherit_btn.setDisabled(True)
                inherit_btn.setText("Inherited")

                setter_widget.setDisabled(True)
            elif variable in styledict_config:
                inherit_btn.setDisabled(False)
                inherit_btn.setChecked(False)
                inherit_btn.setText("Default")

                setter_widget.setDisabled(False)
            else:
                inherit_btn.setDisabled(False)
                inherit_btn.setChecked(True)
                inherit_btn.setText("Default")

                setter_widget.setDisabled(False)

        self.ignore_update_calls = False
        #self.update_preview()
    def _create_styledict_config(self):
        styledict = {}
        
        if not self.outline_color_inherit.isChecked():
            styledict["line_color"] = self.outline_color_btn.color
        
        if not self.outline_style_inherit.isChecked():
            styledict["line_pattern"] = self.line_styles[
                self.outline_style_combo.currentText()]
        
        if not self.outline_thickness_inherit.isChecked():
            styledict["line_thickness"] = self.outline_thickness_spinbox.value()
        
        if not self.background_color_inherit.isChecked():
            styledict["background_color"] = self.background_color_btn.color
        
        if not self.foreground_color_inherit.isChecked():
            styledict["foreground_color"] = self.foreground_color_btn.color
        
        if not self.foreground_pattern_inherit.isChecked():
            styledict["foreground_pattern"] = self.fill_patterns[
                self.foreground_pattern_combo.currentText()]
        
        if not self.text_title_inherit.isChecked():
            styledict["text"] = None
            if self.displayed_cluster:
                self.displayed_cluster.set_name(self.text_title_ledit.text())
        
        if not self.text_color_inherit.isChecked():
            styledict["text_color"] = self.text_color_btn.color
        
        if not self.text_font_inherit.isChecked():
            styledict["font_family"] = self.text_font_combo.currentFont()
        
        if not self.text_bold_inherit.isChecked():
            styledict["font_bold"]       = self.text_formatting_bold.isChecked()
        if not self.text_italic_inherit.isChecked():
            styledict["font_italic"]     = self.text_formatting_italic.isChecked()
        if not self.text_underlined_inherit.isChecked():
            styledict["font_underlined"] = self.text_formatting_underlined.isChecked()
        
        print(styledict)
        return(styledict)
    
    # Callbacks
    def cbk_cluster_clicked(self, cluster):
        # Useless and abandoned ...
        pass
    def cbk_select_outline_color(self):
        color = self.color_dialog.getColor(self.outline_color_btn.color)
        if not color.isValid(): return
        self.outline_color_btn.color = color
        self.contents_changed(self.outline_color_btn)
    def cbk_select_background_color(self):
        color = self.color_dialog.getColor(self.background_color_btn.color)
        if not color.isValid(): return
        self.background_color_btn.color = color
        self.contents_changed(self.background_color_btn)
    def cbk_select_foreground_color(self):
        color = self.color_dialog.getColor(self.foreground_color_btn.color)
        if not color.isValid(): return
        self.foreground_color_btn.color = color
        self.contents_changed(self.foreground_color_btn)
    def cbk_select_font_color(self):
        color = self.color_dialog.getColor(self.text_color_btn.color)
        if not color.isValid(): return
        self.text_color_btn.color = color
        self.contents_changed(self.text_color_btn)
    def cbk_reset(self):
        self._load_styledict(self.current_styledict, 
                             self.current_styledict_config)
    def cbk_reset_to_default(self):
        self._load_styledict({})
    def cbk_cancel(self):
        self.setHidden(True)
    def cbk_apply(self):
        # TODO: Add variation for FeatureUI, TranscriptUI/Cluster etc.
        for selected_cluster in set(self._active_hierarchy.selection):
            selected_cluster.set_styledict_config(self._create_styledict_config())
        self.last_style = self._create_styledict_config()
        self.hierarchy_viewer.redraw_clusters()
    def cbk_update_preview(self):
        self.update_preview()
    def update_preview(self):
        if self.ignore_update_calls:
            return()
        
        self.cbk_apply()

        # Depricated
        return
        if self.displayed_cluster:
            self.preview_styledict = dict(self.default_styledict)
            config,autostyle = self.displayed_cluster.get_full_config()
            self.preview_styledict.update(autostyle)
            self.preview_styledict.update(config)
            self.preview_styledict.update(self._create_styledict_config())
        else:
            self.preview_styledict = dict(self.default_styledict)
            self.preview_styledict.update(self._create_styledict_config())
        
        self.preview.set_styledict(self.preview_styledict)
        
        if self.displayed_cluster:
            self.preview.text = self.displayed_cluster.get_name()
        
        self.preview.redraw_image()
        self.preview.repaint()
    def contents_changed(self, changed_widget):
        for variable, inherit_chk_name, edit_widget_name in self.editing_widgets:
            edit_widget = getattr(self, edit_widget_name)
            inherit_chk = getattr(self, inherit_chk_name)
            if changed_widget is edit_widget:
                inherit_chk.setChecked(False)
                
                self.update_preview()
                return
            elif changed_widget is inherit_chk:
                self.update_preview()
                return
                    
        
        assert False, "Unknown editing widget triggering a contents_changed() callback in StyledictCreator!"
    
    # Quick
    def apply_last_style(self, cluster):
        if self.last_style:
            if isinstance(self.last_style, dict):
                # If it is a styledict
                
                for selected_cluster in cluster.hierarchy.selection:
                    selected_cluster.set_styledict_config(dict(self.last_style))
            elif isinstance(self.last_style, qtg.QColor):
                # If it's just a color
                config = dict(cluster.styledict_config)
                config["background_color"] = qtg.QColor(self.last_style)
                
                # We don't copy it here since apply_last_style should be
                #   applied immediately after invocation, as there is no
                #   user input required after invocation.
                assert cluster in cluster.hierarchy.selection
                for selected_cluster in cluster.hierarchy.selection:
                    selected_cluster.set_styledict_config(config)
        
    def quick_set_color(self, cluster):
        _selection = set(cluster.hierarchy.selection)
        
        old_color = cluster.styledict["background_color"]
        color = self.color_dialog.getColor(old_color)
        if old_color == color or not color.isValid():
            return
        
        if cluster not in _selection:
            _selection = {cluster}
            print("WARNING: quick_set_color applied to cluster outside of selection")
        for selected_cluster in _selection:
            config = dict(selected_cluster.styledict_config)
            config["background_color"] = color
            
            selected_cluster.set_styledict_config(config)
        
        self.last_style = color
    
    # Static
    @staticmethod
    def get_pen(styledict):
        return(qtg.QPen(qtg.QBrush(styledict["line_color"]),
                                   float(styledict["line_thickness"]),
                                   styledict["line_pattern"]))
    @staticmethod
    def get_background_brush(styledict):
        return(qtg.QBrush(styledict["background_color"], 
                          bs=qtc.Qt.SolidPattern))
    @staticmethod
    def get_foreground_brush(styledict):
        if styledict["foreground_pattern"]:
            return(qtg.QBrush(styledict["foreground_color"],
                              bs=styledict["foreground_pattern"]))
        else:
            return(None)
    @staticmethod
    def get_font(styledict):
        font = qtg.QFont(styledict["font_family"])
        font.setItalic(styledict["font_italic"])
        font.setBold(styledict["font_bold"])
        font.setUnderline(styledict["font_underlined"])
        return(font)

class NeiData():
    def __init__(self, regions_to_display, dataset, settings):
        extended_regions = []
        if not settings["loading_existing_session"]:
            borders_size = settings["borders_size"]
        else:
            borders_size = 0
        
        for region in regions_to_display:
            extended_regions.append(dframe.GeneticRegion(
                dataset=region.dataset, 
                scaffold=region.sc, 
                start=max(region.start-borders_size, 1), 
                stop=region.stop+borders_size))
        self.regions_to_display = extended_regions
        
        self.dataset = dataset
        self.settings = settings
        
        # settings:
        #   local_clustering_identity_threshold: float
        #   local_clustering_evalue_threshold: float # Is this even valid?
        #   global_clustering_identity_threshold: float
        #   borders_size: int
        #   loading_existing_session: bool
        #   marker_accsets: dict of lists of strings
        #   about_rules: list of strings
        
        self.cluster_hierarchy = None
    
    # = = = Creation pipeline
    def run_pipeline(self):
        download_worker = WorkerThread(
            self.p1_download_neighborhoods)
        clustering_worker = WorkerThread(
            self.p2_cluster_proteins)
        
        def on_download_finished():
            AppRoot.timer.after_delay_do(clustering_worker.safe_start)
        def on_clustering_finished():
            AppRoot.progress_dialog.setHidden(True)
            self.on_data_ready()
        
        
        download_worker.signals.finished.connect(on_download_finished)
        clustering_worker.signals.finished.connect(on_clustering_finished)
        
        
        
        AppRoot.progress_dialog.reinitialize(
            title="Neighborhood View progress",
            label_text="Retrieving neighborhoods...",
            on_abort=None,
            text_only=True)
        
        # With this somewhat confusing implementation, the
        #   download manager calls the signal in one thread,
        #   then responds to it in the other.
        download_worker.signals.progress_manager_update.connect(
            dbdl.DOWNLOAD_MANAGER.cbk_status_update)
        dbdl.DOWNLOAD_MANAGER.set_progress_manager_update_signal(
            download_worker.signals.progress_manager_update)
        
        
        AppRoot.progress_dialog.show()
        AppRoot.progress_dialog.activateWindow()
        download_worker.safe_start()
        
        # TODO: This function used to set the TopClusters,
        #       do this from somewhere else.
        pass
    def p1_download_neighborhoods(self):
        # This obtains the features for the requested regions,
        #   and adds them to the root feature tree.
        print(f"Requesting {len(self.regions_to_display)} regions to be displayed.")
        self.dataset.update_scs_with_gbdata(
            self.regions_to_display,
            0)
    def p2_cluster_proteins(self):
        #
        proteins_to_cluster = set()
        for region in self.regions_to_display:
            for feature in region.fts.values():
                if feature.type == "cds":
                    proteins_to_cluster.add(feature.ref.accession)
        
        print(f"Prepped {len(proteins_to_cluster)} proteins for clustering.\n")
        
        # Global alignment-based clustering
        cluster_hierarchy = AppRoot.cluster_manager.run_global_clustering(
            list(proteins_to_cluster), 
            self.dataset.root, 
            self.settings["global_clustering_identity_threshold"])
        
        # Local alignment-based clustering
        cluster_hierarchy = AppRoot.cluster_manager.\
          run_local_clustering_of_global_clusters(
            cluster_hierarchy, 
            evalue=self.settings["local_clustering_evalue_threshold"],
            identity=self.settings["local_clustering_identity_threshold"],
            bitscore=self.settings["local_clustering_bitscore_threshold"],
            sensitivity=self.settings["local_clustering_sensitivity_mode"],
            community_weight_variable=self.settings["community_weight_variable"],
            community_resolution=self.settings["community_resolution"])
        
        self.cluster_hierarchy = cluster_hierarchy
    # This is a virtual function, leave it.
    def on_data_ready(self):
        pass
class NeiBrowser():
    # Comprised of a NeiView and a Protein Group Overview Table
    # Attaches to a NeiDataset and displays the data in it? I guess?
    
    # Depricated as of 2023/08
    pass

class ClusterNetworkViewer2(qtw.QWidget, Ui_ClusterNetworkViewer2):
    def __init__(self, neiview):
        super().__init__()
        self.neiview = neiview
        self._canvas_pixmap = None
        
        self._node_to_canvas_widget = {}
        self._widget_raw_coords = {}
        self._edges = None
        self._graph = None
        
        self._appearance = {
            "clusterwidget_width": None,
            "clusterwidget_height": None,
            "clusterwidget_size_scaling": None}
        
        self.horbar_last_value = 0.5
        self.verbar_last_value = 0.5
        
        self._mouse_drag_start = None
        
        self._ignore_subset_changes = False
        self._clusters = None
        
        self._show_subgroups = False
        self._mode = "mini"
        
        self._init_ui()
        
        self._font = StyledictCreator.get_font(StyledictCreator.default_styledict)
        self._font.setBold(True)
    def _init_ui(self):
        self.setupUi(self)
        
        self.horbar = self.canvas_scroller.horizontalScrollBar()
        self.verbar = self.canvas_scroller.verticalScrollBar()
        
        self.canvas.paintEvent = self._paint_canvas
        
        self.zoom_in_btn.clicked.connect(
            self.cbk_zoom_in)
        self.zoom_out_btn.clicked.connect(
            self.cbk_zoom_out)
        
        self.canvas_scroller.setWidgetResizable(True)
        
        self.horbar.valueChanged.connect(
            self.cbk_horbar_value_changed)
        self.verbar.valueChanged.connect(
            self.cbk_verbar_value_changed)
        self.horbar.rangeChanged.connect(
            self.cbk_horbar_range_changed)
        self.verbar.rangeChanged.connect(
            self.cbk_verbar_range_changed)
        
        self.groups_super_radio.toggled.connect(
            self.cbk_subgroups_radio_changed)
        self.groups_sub_radio.toggled.connect(
            self.cbk_subgroups_radio_changed)
        
        self.apply_size_changes_btn.clicked.connect(
            self.cbk_apply_appearance)
        self.rearrange_btn.clicked.connect(
            self.cbk_redraw_network)
        
        self.show_subset_standard_radio.toggled.connect(
            self.cbk_show_subset_radio_changed)
        self.show_subset_localgroup_radio.toggled.connect(
            self.cbk_show_subset_radio_changed)
        self.show_subset_other_radio.toggled.connect(
            self.cbk_show_subset_radio_changed)
        
        self.setWindowTitle(f"CluSeek {about.VERSION} - Protein Group Network")
        
        self.canvas.mousePressEvent = self.canvas_mousePressEvent
        self.canvas.mouseReleaseEvent = self.canvas_mouseReleaseEvent
        self.canvas.mouseMoveEvent = self.canvas_mouseMoveEvent
        self.canvas_scroller.wheelEvent = lambda x: x.ignore()
        
        
        self.show_subset_other_radio.setHidden(True)
        self.info_lbl.setHidden(True)
        self.info_btn.clicked.connect(self.cbk_info_btn_pressed)
        
        add_question_mark_icon(self.info_btn)
        
        # Update the UI
        self.cbk_subgroups_radio_changed(True)
        self.reset_ui()
    def get_subset_mode(self):
        if self.show_subset_standard_radio.isChecked():
            return("standard")
        elif self.show_subset_localgroup_radio.isChecked():
            return("local_group")
        elif self.show_subset_other_radio.isChecked():
            return("exact")
        else:
            assert False
    def set_subset_mode(self, mode):
        self._ignore_subset_changes = True
        if mode == "standard":
            self.show_subset_standard_radio.setChecked(True)
        elif mode == "local_group":
            self.show_subset_localgroup_radio.setChecked(True)
        elif mode == "exact":
            self.show_subset_other_radio.setChecked(True)
        else:
            raise ValueError("Invalid subset mode "
                             "(can be 'standard', 'local_group' or 'exact')")
        self._ignore_subset_changes = False
    # Data processing old
    @staticmethod
    def _create_queries(clusters, hierarchy, mode):
        # Collect headers and sequences for all lowest-level
        #   clusters. The format is [(header, sequence), (header, sequence), ...]
        #   where header is the unique cluster ID, but without the fasta >
        subclusters_ = []
        def add_cluster(c, subclusters_):
            if len(c.subclusters) > 0:
                for subc in c.subclusters.values():
                    add_cluster(subc, subclusters_)
            else:
                subclusters_.append(c)
        
        for cluster in clusters:
            add_cluster(cluster, subclusters_)
        
        if mode == "standard":
            to_show = subclusters_
        elif mode == "local_group":
            local_group_ids = {c.local_group for c in subclusters_}
            assert None not in local_group_ids
            
            to_show = set()
            for local_group_id in local_group_ids:
                to_show |= set(hierarchy.local_groups[local_group_id])
            assert all(c.type_ == "globally_aligned" for c in to_show)
        
        to_show = list(to_show)
        
        queries = []
        cluster_to_node = {}
        node_to_cluster = {}
        for i in range(len(to_show)):
            queries.append((f"{i}", to_show[i].centroid.seq))
            cluster_to_node[to_show[i]] = i
            node_to_cluster[i] = to_show[i]
        
        
        # Sort queries by length
        queries.sort(key=lambda x: len(x[1]))
        
        return(queries, cluster_to_node, node_to_cluster)
    @staticmethod
    def _create_network(node_to_cluster, alignments):
        # Create the graph in networkx
        graph = nx.Graph()
        
        for node in list(node_to_cluster):
            graph.add_node(
                node,
                cluster = node_to_cluster[node])
        
        for alignment in alignments:
            if alignment["query"] == alignment["hit"]: continue
            graph.add_edge(
                int(alignment["query"]), 
                int(alignment["hit"]),
                weight=alignment["bitscore"])
        return(graph)
    # Data processing
    @staticmethod
    def get_subgraph(graph, clusters, mode):
        to_show = set()
        
        if mode == "standard":
            for cluster in clusters:
                to_show |= cluster.get_toplevel().get_bottomlevel_subclusters()
        elif mode == "local_group":
            for cluster in clusters:
                to_show |= cluster.get_related_clusters()
        elif mode == "exact":
            to_show = set(clusters)
        
        
        subgraph = graph.subgraph([cluster.id_ for cluster in to_show])
        
        node_to_cluster = {cluster.id_: cluster for cluster in to_show}
        
        return(subgraph, node_to_cluster)
    @staticmethod
    def _create_graph_layout(graph):
        layout = nx.spring_layout(graph)
        #layout = [x for x in layout.values()]
        edges = [edge for edge in graph.edges]
        return(layout, edges)
    # Graph rendering
    def show_clusters(self, clusters, mode=None):
        if mode==None:
            self.get_subset_mode()
        else:
            self.set_subset_mode(mode)
        
        master_graph = self.neiview.neidata.cluster_hierarchy.graph
        self._graph,self._node_to_cluster = self.get_subgraph(
                                                master_graph, clusters, mode)
        
        layout,edges = self._create_graph_layout(self._graph)
        
        self._prepare_canvas(self._node_to_cluster, edges, layout)
        self._apply_graph_layout(layout, self._node_to_cluster)
        
        self._reset_canvas()
        
        self._clusters = clusters
    def _prepare_canvas(self, nodes_to_clusters, edges, layout):
        for child in self.canvas.children():
            child.setParent(None)
        self._node_to_canvas_widget = {}
        self._widget_raw_coords = {}
        self._edges = None
        for node_id in list(nodes_to_clusters):
            # Get our cluster widget
            cluster = nodes_to_clusters[node_id]
            cluster_widget = ClusterWidget(parent=self.canvas, 
                                           cluster=cluster, 
                                           neiview=self.neiview,
                                           represents="bottomlevel" if self._show_subgroups else "toplevel")
            cluster_widget.sig_floated.connect(self.cbk_clusterwidget_moved)
            self._node_to_canvas_widget[node_id] = cluster_widget
        
        # The edges are drawn and re-drawn dynamically
        self._edges = edges
    def _apply_graph_layout(self, layout, nodes_to_clusters):
        for node_id in list(nodes_to_clusters):
            cluster_widget = self._node_to_canvas_widget[node_id]
            cluster_widget.setHidden(False)
            
            # Determine position on canvas
            r_x,r_y = layout[node_id] # Raw coordinates, -1 to +1
            # We adjust these coordinates to a 0 to +1 scale so we can
            #   easily scale it to new widget sizes (which are positive)
            self._widget_raw_coords[cluster_widget] = ((r_x+1)/2,(r_y+1)/2)
    def _rescale_canvas(self, w, h):
        
        old_geometry = self.canvas.geometry()
        if w and h:
            new_width = w + self._appearance["clusterwidget_width"]
            new_height = h + self._appearance["clusterwidget_height"]
        else:
            new_width = old_geometry.width()
            new_height = old_geometry.height()
            w = new_width - self._appearance["clusterwidget_width"]
            h = new_height - self._appearance["clusterwidget_height"]
        
        if max(new_width, new_height) > 12000:
            return()
        elif min(new_width, new_height) < 100:
            return()
        
        self.canvas.setFixedSize(
            new_width, 
            new_height)
        #self.canvas_scrollable.setGeometry(
        #    #self.canvas_scrollable.geometry().width()/2 - new_width/2,
        #    #self.canvas_scrollable.geometry().height()/2 - new_height/2,
        #    0,0,
        #    new_width, 
        #    new_height)
        
        #self.canvas_scroller.adjustSize()
        
        for cluster_widget in self._node_to_canvas_widget.values():
            r_x,r_y = self._widget_raw_coords[cluster_widget]
            
            size_modifier = 1
            if self._appearance["clusterwidget_size_scaling"]:
                raw_counts = 0
                for protein in cluster_widget.cluster.proteins.values():
                    raw_counts += len(protein.fts)
                
                size_modifier = (
                    1 + raw_counts/self._appearance["clusterwidget_size_scaling"])
                
            # r_x and r_y are coordinates on a 0 to 1 scale.
            cluster_widget.setGeometry(
                r_x*w,
                r_y*h,
                self._appearance["clusterwidget_width"]*size_modifier,
                self._appearance["clusterwidget_height"]*size_modifier)
        self._redraw_canvas()
        
        self.canvas.update()
        self.canvas_scrollable.update()
        self.canvas_scroller.update()
    def _redraw_canvas(self):
        h = self.canvas.geometry().height()
        w = self.canvas.geometry().width()
        self._canvas_pixmap = qtg.QPixmap(w+1,h+1)
        self._canvas_pixmap.fill(qtc.Qt.transparent)
        
        painter = qtg.QPainter()
        painter.begin(self._canvas_pixmap)
        
        painter.setPen(qtg.QColor(128,128,128,255))
        
        lines_to_draw = []
        for nodeA,nodeB in self._edges:
            widgetA = self._node_to_canvas_widget[nodeA]
            widgetB = self._node_to_canvas_widget[nodeB]
            
            lines_to_draw.append(
                qtc.QLine(
                    widgetA.geometry().center(),
                    widgetB.geometry().center()))
        painter.drawLines(lines_to_draw)
        
        
        texts = []
        if self.exact_frequency_counts_chk.isChecked() and not self._mode=="mini":
            for cluster_widget in self._node_to_canvas_widget.values():
                
                raw_counts = 0
                for protein in cluster_widget.cluster.proteins.values():
                    raw_counts += len(protein.fts)
                
                texts.append((
                    qtc.QRect(
                        cluster_widget.geometry().bottomLeft() + qtc.QPoint(2,2),
                        qtc.QSize(cluster_widget.geometry().width(),
                                  math.ceil(3*self._appearance["clusterwidget_height"]/5))),
                    str(raw_counts)))
        self._font.setPixelSize(math.floor(3*self._appearance["clusterwidget_height"]/5))
        painter.setFont(self._font)
        painter.setPen(qtg.QColor(255,255,255,255))
        for rect,text in texts:
            painter.drawText(rect,qtc.Qt.AlignCenter,"█"*len(text))
        painter.setPen(qtg.QColor(0,0,0,255))
        for rect,text in texts:
            painter.drawText(rect,qtc.Qt.AlignCenter,text)
        
        painter.end()
    def _paint_canvas(self, event):
        event.accept()
        painter = qtg.QPainter()
        painter.begin(self.canvas)
        painter.drawPixmap(qtc.QPoint(0,0), self._canvas_pixmap)
        painter.end()
    def _reset_canvas(self):
        if self._mode == "full":
            self._rescale_canvas(1000, 1000)
        elif self._mode == "mini":
            self._rescale_canvas(
                self.geometry().width(),
                self.geometry().height())

    def show_clusters_old(self, clusters, mode="standard"):
        queries,cluster_to_node,self._node_to_cluster = \
            self._create_queries(
                clusters, 
                self.neiview.neidata.cluster_hierarchy, 
                mode=mode)
        
        self._alignments = AppRoot.cluster_manager.multiquery_blastp(
            query_seqs=queries, db_seqs=queries, chunk_size=50000, 
            evalue=self.neiview.dp["creation_settings"]["local_clustering_evalue_threshold"], 
            identity=self.neiview.dp["creation_settings"]["local_clustering_identity_threshold"], 
            bitscore=self.neiview.dp["creation_settings"]["local_clustering_bitscore_threshold"])
        
        self._graph = self._create_network(
            self._node_to_cluster, 
            self._alignments)
        
        layout,edges = self._create_graph_layout(self._graph)
        
        self._prepare_canvas(self._node_to_cluster, edges, layout)
        self._apply_graph_layout(layout, self._node_to_cluster)
        
        self._rescale_canvas(1000, 1000)
        
        self.show()
        self.activateWindow()
    def remake_network(self):
        layout,edges = self._create_graph_layout(self._graph)
        
        self._apply_graph_layout(layout, self._node_to_cluster)
        
        self._rescale_canvas(1000, 1000)
    @staticmethod
    def colorbrush_generator():
        # This generator creates a list of colors, sorts them from
        #   brightest to darkest, and then yields them on repeat.
        possible_colors = []
        for red_i in range(3,0,-1):
            for blue_i in range(3,0,-1):
                for green_i in range(3,0,-1):
                    possible_colors.append((red_i, blue_i, green_i))
        possible_colors.sort(
            key=lambda colors: max(colors),
            reverse=True)
        possible_colors = [
            qtg.QBrush(
              qtg.QColor(
                red_i*85,
                blue_i*85,
                green_i*85)) for red_i,blue_i,green_i in possible_colors]
        while True:
            for brush in possible_colors:
                yield(brush)
    # Interaction
    def wheelEvent(self, event):
        event.accept()
        if event.angleDelta().y() > 0:
            self.cbk_zoom_in()
        elif event.angleDelta().y() < 0:
            self.cbk_zoom_out()
    def canvas_mousePressEvent(self, event):
        if (event.button() == qtc.Qt.LeftButton
            and not event.modifiers()):
                event.accept()
                self._mouse_drag_start = event.pos()
    def canvas_mouseReleaseEvent(self, event):
        if (event.button() == qtc.Qt.LeftButton
            and not event.modifiers()):
                event.accept()
                self._mouse_drag_start = None
    def canvas_mouseMoveEvent(self, event):
        if self._mouse_drag_start is not None:
                dx = self._mouse_drag_start.x() - event.pos().x()
                dy = self._mouse_drag_start.y() - event.pos().y()
                
                dx = dx / self.canvas.geometry().width()
                dy = dy / self.canvas.geometry().height()
                
                hbar = self.canvas_scroller.horizontalScrollBar()
                vbar = self.canvas_scroller.verticalScrollBar()
                
                hbar.setValue(hbar.value() + dx*hbar.maximum())
                vbar.setValue(vbar.value() + dy*vbar.maximum())
    
    # UI
    def cbk_clusterwidget_moved(self, clusterwidget):
        self._redraw_canvas()
    def cbk_show_subset_radio_changed(self, checked):
        if not checked: return
        if self._ignore_subset_changes: return
        
        if self.show_subset_localgroup_radio.isChecked():
            self.show_clusters(self._clusters, mode="local_group")
        elif self.show_subset_standard_radio.isChecked():
            self.show_clusters(self._clusters, mode="standard")
        else:
            assert False, "Invalid radio button configuration."
    def set_mode_full(self):
        self._mode = "full"
        self.ui_frame.setHidden(False)
        
        self.canvas_scroller.setHorizontalScrollBarPolicy(
            qtc.Qt.ScrollBarAsNeeded)
        self.canvas_scroller.setVerticalScrollBarPolicy(
            qtc.Qt.ScrollBarAsNeeded)
        
        self.cbk_apply_appearance()
        
        
        self.exact_frequency_counts_chk.setChecked(True)
        self.cbk_apply_appearance()
    def set_mode_mini(self):
        self._mode = "mini"
        self.ui_frame.setHidden(True)
        self._appearance["clusterwidget_width"] = 5
        self._appearance["clusterwidget_height"] = 5
        
        self.canvas_scroller.setHorizontalScrollBarPolicy(
            qtc.Qt.ScrollBarAlwaysOff)
        self.canvas_scroller.setVerticalScrollBarPolicy(
            qtc.Qt.ScrollBarAlwaysOff)
        
        self._rescale_canvas(
            self.canvas_scrollable.geometry().width(), 
            self.canvas_scrollable.geometry().height())
        
    def cbk_zoom_in(self):
        self._rescale_canvas(
            round(self.canvas.geometry().width() * 1.4),
            round(self.canvas.geometry().width() * 1.4))
    def cbk_zoom_out(self):
        self._rescale_canvas(
            round(10 * self.canvas.geometry().width() / 14),
            round(10 * self.canvas.geometry().height() / 14))
    def cbk_horbar_value_changed(self, value):
        self.horbar_last_value = (
            (
                value 
                #+ math.floor(self.horbar.pageStep()/2)
            )
            / max(self.horbar.maximum(), 1))
    def cbk_verbar_value_changed(self, value):
        self.verbar_last_value = (
            (
                value 
                #+ math.floor(self.verbar.pageStep()/2)
            )
            / max(self.verbar.maximum(), 1))
    def cbk_horbar_range_changed(self, min_, max_):
        self.horbar.setValue(round(
            (self.horbar_last_value * max(self.horbar.maximum(), 1))
            #- math.floor(self.horbar.pageStep()/2)
            ))
    def cbk_verbar_range_changed(self, min_, max_):
        self.verbar.setValue(round(
            (self.verbar_last_value * max(self.verbar.maximum(), 1))
            #- math.floor(self.verbar.pageStep()/2)
            ))
    def cbk_redraw_network(self):
        self.remake_network()
    def cbk_subgroups_radio_changed(self, checked):
        if not checked: return
        if self.groups_super_radio.isChecked():
            self.set_show_subgroups(False)
        elif self.groups_sub_radio.isChecked():
            self.set_show_subgroups(True)
        else:
            assert False, "Invalid radiobutton configuration."
    def set_show_subgroups(self, value):
        if self._show_subgroups == value:
            return
        
        self._show_subgroups = value
               
        value = "bottomlevel" if value else "toplevel"
        
        for widget in self._node_to_canvas_widget.values():
            widget.set_represents(value)
        self.canvas.update()
        self.update()
    def reset_ui(self):
        self._appearance["clusterwidget_width"] = \
            self.neiview.dp["size_fixed_width"]
        self.symbol_width_spin.setValue(
            self.neiview.dp["size_fixed_width"])
        
        self._appearance["clusterwidget_height"] = \
            self.neiview.dp["size_fixed_height"]
        self.symbol_height_spin.setValue(
            self.neiview.dp["size_fixed_height"])
    def cbk_apply_appearance(self):
        self._appearance["clusterwidget_width"] = \
            self.symbol_width_spin.value()
        self._appearance["clusterwidget_height"] = \
            self.symbol_height_spin.value()
        
        if self.size_scaling_chk.isChecked():
            self._appearance["clusterwidget_size_scaling"] = \
                self.size_scaling_dspin.value()
        else:
            self._appearance["clusterwidget_size_scaling"] = \
                None
        
        self._rescale_canvas(None, None)
    def cbk_info_btn_pressed(self):
        self.info_lbl.setHidden(not self.info_lbl.isHidden())
    
    # Export
    def get_network_graph_gml(self):
        pass
        #return(self._graph.)
# Another failed context menu builder
class ContextMenuHandler():
    CONTEXT_MENU_CONSTRUCTORS = []
    #   [(priority, constructor), ...]
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
    @classmethod
    def MENU_CONSTRUCTOR(cls, constructor):
        cls.CONTEXT_MENU_CONSTRUCTORS.append(constructor)
        return(constructor)
    def _construct_menu(self):
        menu = qtw.QMenu(self)
        for constructor in sorted(self.CONTEXT_MENU_CONSTRUCTORS, 
                                  reverse=True, 
                                  key=lambda x: x[0]):
            constructor(self, menu)
        return(menu)
    def contextMenuEvent(self, event):
        event.accept()
        menu = self._construct_menu()
        menu.popup(event.globalPos())

class ExportHighlighterWidget(qtw.QScrollArea):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        
        self.neiview = None
        
        self._init_ui()
    def _init_ui(self):
        self.scrollable = qtw.QWidget()
        self.setWidget(self.scrollable)
        
        self.scrollable_lay = qtw.QHBoxLayout()
        self.scrollable.setLayout(self.scrollable_lay)
        
        self.image_label = qtw.QLabel()
        self.image_label.setScaledContents(False)
        self.scrollable_lay.addWidget(self.image_label)
        
    def _snapshot(self, neiview):
        self.neiview = neiview
        return
        pixmap = self.neiview.clusters_scrollable.grab()
        pixmap = pixmap.scaled(
            qtc.QSize(
                self.geometry().width(),
                self.geometry().width()*(
                    self.neiview.clusters_scrollable.geometry().height()
                    /self.neiview_clusters_scrollable.geometry().width())))
        self.image_label.setPixmap(pixmap)
class NeiExport(qtw.QWidget, Ui_NeiExport):
    export_values = [
        ("Gene Clusters", "cluster"), 
        ("Proteins", "protein")]
    format_values = [("GenBank file (.gb)", "gb"), 
                     ("Image (.png)", "png"),
                     ("Vector image (.svg)", "svg"), 
                     ("Excel sheet (.xlsx)", "xlsx"),
                     ("FASTA (.fsa)", "fsa"),
                     ("Network graph (.gml)", "gml")]
    formats_for_cluster = ["gb", "png", "svg", "xlsx", "fsa", "gml"]
    formats_for_protein = ["fsa", "gml"]
    xlsx_cell_contents = [
        "proteingroup_allaccessions",
        "proteingroup_size",
        "proteingroup_name",
        "proteingroup_allannotations",
        "proteingroup_annotation",
        "proteingroup_identifier",
        "protein_sequence",
        "protein_annotation",
        "protein_accession",
        "feature_length",
        "feature_position"]
    xlsx_cell_defaults = ["proteingroup_name", "protein_accession"]
    xlsx_comment_defaults = ["feature_position", "protein_annotation"]
    text_to_key = {}
    key_to_text = {}
    for text,key in (export_values+format_values):
        text_to_key[text] = key
        key_to_text[key] = text
    
    def __init__(self):
        super().__init__()
        self._init_ui()
        self.neiview = None
    def _init_ui(self):
        self.setupUi(self)
        
        for value,key in self.export_values:
            self.export_combo.addItem(value)
        
        self.cbk_export_changed()
        self.cbk_format_changed()
        
        self.export_combo.currentTextChanged.connect(self.cbk_export_changed)
        self.format_combo.currentTextChanged.connect(self.cbk_format_changed)
        self.save_btn.clicked.connect(self.cbk_save)
        self.cancel_btn.clicked.connect(self.cbk_cancel)
        
        self.xlsx_cell_contentselector = SelectorWidget(
            "Possible values", "Values to display in cells")
        self.cellcontents_lay.addWidget(self.xlsx_cell_contentselector)
        
        self.xlsx_comment_contentselector = SelectorWidget(
            "Possible values", "Values to display in comments")
        self.commentcontents_lay.addWidget(self.xlsx_comment_contentselector)
        
        for item in self.xlsx_cell_contents:
            if item not in self.xlsx_cell_defaults:
                self.xlsx_cell_contentselector.add_a_item(item)
            else:
                self.xlsx_cell_contentselector.add_b_item(item)
            if item not in self.xlsx_comment_defaults:
                self.xlsx_comment_contentselector.add_a_item(item)
            else:
                self.xlsx_comment_contentselector.add_b_item(item)
        
        self.setWindowTitle(f"CluSeek {about.VERSION} - Export")
    def open(self, neiview):
        self.neiview = neiview
        self.show()
        self.activateWindow()
        
        index = AppRoot.mainwindow.maintabs.indexOf(neiview)
        text = AppRoot.mainwindow.maintabs.tabText(index)
        
        self.exportingfrom_lbl.setText(
            f"Now exporting from <b>{text}</b>")
        return
        #pixmap = self.neiview.clusters_scrollable.grab()
        #height_ratio = (self.neiview.clusters_scrollable.geometry().width()
        #            /self.neiview.clusters_scrollable.geometry().height())
        #print("hr", height_ratio)
        #pixmap = pixmap.scaled(
        #    qtc.QSize(
        #        self.highlighter_label.geometry().width(),
        #        self.highlighter_label.geometry().width()))
        #self.highlighter_label.setPixmap(pixmap)
    @property
    def export(self):
        return(self.text_to_key[self.export_combo.currentText()])
    @property
    def format(self):
        try:
            return(self.text_to_key[self.format_combo.currentText()])
        except KeyError as e:
            if self.format_combo.currentText() == "":
                return None
            else:
                raise e
    #
    def set_available_formats(self, formats):
        self.format_combo.clear()
        for format_key in formats:
            self.format_combo.addItem(self.key_to_text[format_key])
    #
    def cbk_export_changed(self):
        if   self.export == "cluster":
            self.set_available_formats(self.formats_for_cluster)
            self.export_stack.setCurrentWidget(self.exportclusters_frame)
        elif self.export == "protein":
            self.set_available_formats(self.formats_for_protein)
            self.export_stack.setCurrentWidget(self.exportproteins_frame)
        else:
            raise ValueError(f"Invalid export combo value: {self.export}")
        #self.cbk_format_changed()
    def cbk_format_changed(self):
        if not self.format: return
        if   self.format == "gb":
            self.format_stack.setCurrentWidget(self.gbsettings_frame)
        elif self.format == "png":
            self.format_stack.setCurrentWidget(self.pngsettings_frame)
        elif self.format == "svg":
            self.format_stack.setCurrentWidget(self.svgsettings_frame)
        elif self.format == "xlsx":
            self.format_stack.setCurrentWidget(self.excelsettings_frame)
        elif self.format == "fsa":
            self.format_stack.setCurrentWidget(self.fastasettings_frame)
        elif self.format == "gml":
            if   self.export == "cluster":
                self.format_stack.setCurrentWidget(self.gmlsettings_frame)
            elif self.export == "protein":
                self.format_stack.setCurrentWidget(self.gmlprotsettings_frame)
            else:
                raise ValueError(f"Invalid export combo value: {self.export}")
    def cbk_cancel(self):
        self.setHidden(True)
    def cbk_save(self):
        self.save()
    def get_filepath(self, format):
        format_strings = {}
        for k,s in self.format_values:
            format_strings[s] = k
        
        path,suffix = qtw.QFileDialog.getSaveFileName(
                        caption="Save exported data",
                        filter=format_strings[format])
        
        return(path, suffix)
    def save(self):
        if self.export == "cluster":
            regdata_to_export = self.get_regdata()
            if not regdata_to_export:
                print("Nothing to export!")
                qtw.QMessageBox.warning(
                    None, 
                    "No data to export!", 
                    "Check your selection / export settings!")
                return
            
            if self.format == "png" or self.format == "svg":
                filepath,suffix = self.get_filepath(self.format)
                
                if not filepath: return
                
                filepath=filepath
                
                if (self.clusters_highlighted_radio.isChecked()
                  and not self.neiview.cluster_scrollable_overlay\
                      .bpselector.isHidden()):
                    highlighted = self.neiview.cluster_scrollable_overlay\
                                  .bpselector.geometry()
                else:
                    highlighted = None
                self.export_neiview_img(filepath,
                                        suffix,
                                        regdata_to_export, 
                                        highlighted=highlighted)
            elif self.format == "gb":
                #filepath = qtw.QFileDialog.getExistingDirectory(
                #    caption="Select folder to save genbank files into")
                filepath,suffix = self.get_filepath(self.format)
                
                if not filepath: return
                
                # not used
                #filepath=filepath+".gb"
                
                self.export_neiview_gb(filepath,
                                       regdata_to_export)
            elif self.format == "fsa":
                filepath,suffix = self.get_filepath(self.format)
                
                if not filepath: return
                
                filepath=filepath+".fsa"
                
                self.export_neiview_fasta(filepath, 
                                          regdata_to_export)
            elif self.format == "xlsx":
                filepath,suffix = self.get_filepath(self.format)
                
                if not filepath: return
                
                filepath=filepath+".xlsx"
                
                # args
                align = self.xlsx_align.isChecked()
                transposed = self.xlsx_transpose.isChecked()
                tags_individually = not self.xlsx_condensetags.isChecked()
                separator = self.xlsx_separator.text()
                cellcontents = self.xlsx_cell_contentselector.get_b_contents()
                commentcontents = self.xlsx_comment_contentselector.get_b_contents()
                
                self.export_neiview_xlsx(
                    filepath=filepath,
                    to_export=regdata_to_export,
                    cell_variables=cellcontents,
                    comment_variables=commentcontents,
                    align=align,
                    separator=separator,
                    transposed=transposed,
                    tags_individually=tags_individually)
            elif self.format == "gml":
                filepath,suffix = self.get_filepath(self.format)
                
                if not filepath: return
                
                filepath=filepath+".gml"
                
                self.export_neiview_gml(filepath, regdata_to_export,
                    cutoff=self.gml_cutoff_slider.value()/100,
                    metric_i=self.gml_metric_combo.currentIndex())
                
        elif self.export == "protein":
            proteins_to_export = self.get_proteins()
            if not proteins_to_export:
                print("Nothing to export!")
                qtw.QMessageBox.warning(
                    None, 
                    "No data to export!", 
                    "Check your selection / export settings!")
                
            if self.format == "fsa":
                filepath,suffix = self.get_filepath(self.format)
                
                if not filepath: return
                
                filepath = filepath+".fsa"
                
                self.export_proteins_multifasta(filepath,
                                                proteins_to_export)
            elif self.format == "gml":
                filepath,suffix = self.get_filepath(self.format)
                
                if not filepath: return
                
                filepath=filepath+".gml"
                
                self.export_proteins_network(
                    filepath,
                    proteins_to_export,
                    local_group=self.gmlprot_relatedproteins_chk.isChecked())
        
        self.filesaved_lbl.setText(f"Saved to {filepath}")
        self.setHidden(True)
    def get_regdata(self):
        to_export = []
        get_subregion = self.neiview.cluster_scrollable_overlay.bpselector\
                            .get_highlighted_subregion
        
        for regui in self.neiview.regions_all:
            # Which cluster to export
            if self.clusters_checked_radio.isChecked() and not regui.selected:
                continue
            
            # Which subregion to export
            if self.clusters_highlighted_radio.isChecked():
                reg = get_subregion(regui)
            else:
                reg = regui.reg
            
            to_export.append((reg, regui))
        if len(to_export) == 0:
            return(None)
        
        return(to_export)
    def get_proteins_old(self):
        to_export = []
        get_subregion = self.neiview.cluster_scrollable_overlay.bpselector\
                            .get_highlighted_subregion
        
        if self.proteins_selected_prgroups_radio.isChecked():
            print("pclusters is not None")
            pclusters = {}
            for pcluster in self.neiview.neidata.cluster_hierarchy.selection:
                for pt in pcluster.proteins.values():
                    if pt not in pclusters:
                        pclusters[pt] = []
                    pclusters[pt].append(pcluster)
        else:
            pclusters = None
        
        for regui in self.neiview.regions_all:
            if (self.proteins_clusters_checked_radio.isChecked()\
              and not regui.selected):
                continue
            
            if self.proteins_highlighted_radio.isChecked():
                reg = get_subregion(regui)
            else:
                reg = regui.reg
            
            for ft in reg.fts.values():
                if ft.type != "cds": 
                    continue
                
                if pclusters is not None:
                    if ft.ref not in pclusters:
                        continue
                    
                    clusters = pclusters[ft.ref]
                elif ft.ref in self.neiview.neidata.cluster_hierarchy.member_to_supercluster:
                    clusters = [self.neiview.neidata.cluster_hierarchy.member_to_supercluster[ft.ref]]
                else:
                    clusters = []
                to_export.append((ft.ref, clusters))
        
        return(to_export)
    def get_proteins(self):
        to_export = []
        get_subregion = self.neiview.cluster_scrollable_overlay.bpselector\
                            .get_highlighted_subregion
        
        if self.proteins_selected_prgroups_radio.isChecked():
            print("pclusters is not None")
            pclusters = {}
            for pcluster in self.neiview.neidata.cluster_hierarchy.selection:
                for pt in pcluster.proteins.values():
                    if pt not in pclusters:
                        pclusters[pt] = []
                    pclusters[pt].append(pcluster)
        else:
            pclusters = None
        
        for regui in self.neiview.regions_all:
            if (self.proteins_clusters_checked_radio.isChecked()\
              and not regui.selected):
                continue
            
            if self.proteins_highlighted_radio.isChecked():
                reg = get_subregion(regui)
            else:
                reg = regui.reg
            
            for ft in reg.fts.values():
                if ft.type != "cds": 
                    continue
                
                if pclusters is not None:
                    if ft.ref not in pclusters:
                        continue
                    
                    clusters = pclusters[ft.ref]
                elif ft.ref.accession in self.neiview.neidata.cluster_hierarchy.member_to_supercluster:
                    clusters = [self.neiview.neidata.cluster_hierarchy.member_to_supercluster[ft.ref.accession]]
                else:
                    clusters = []
                to_export.append((ft.ref, regui.reg, clusters))
        
        return(to_export)
        
    # Exporting data
    def export_neiview_img(self, filepath, suffix, to_export, 
                           highlighted=None, calc_resultion_only=False):
        # Assume highlighted is a QRect relative to clusters_scrollable
        maxheadwidth = max(thi[1].head.geometry().width() for thi in to_export)
        maxtailwidth = max(thi[1].tail.geometry().width() for thi in to_export)
        
        rowheight = self.neiview.dp["size_fixed_height"]+1
        tailheadgap = 10
        
        
        # thi[1] is a regui to be exported
        minoffset = min(thi[1].offset for thi in to_export)
        #maxoffset = max(thi[1].offset for thi in to_export)
        
        if highlighted:
            maxtailwidth = highlighted.width()
            minoffset = 0
        
        # Output the image dimensions instead if asked for them
        # => This is a bit hacky, the dimension function should be split off
        #    into another function (possibly).
        
        
        if calc_resultion_only:
            return(maxheadwidth+maxtailwidth-minoffset+tailheadgap,
                   rowheight * len(to_export))
        
        image = qtg.QPicture()
        painter = qtg.QPainter()
        painter.begin(image)
        
        ihead = qtg.QPicture()
        itail = qtg.QPicture()
        #itail.setBoundingRect(
        #    qtc.QRect(
        #        0, 
        #        0, 
        #        highlighted.width(), 
        #        rowheight))
        
        for i in range(len(to_export)):
            reg,regui = to_export[i]
            
            headoffset = qtc.QPoint(0, 0)
            tailoffset = qtc.QPoint(0, 0)
            if highlighted:
                tailregion = qtg.QRegion(highlighted.left(),
                                     0,
                                     highlighted.width(),
                                     regui.tail.geometry().height())
            else:
                tailregion = qtg.QRegion(minoffset, 
                                     0, 
                                     regui.tail.geometry().width()-minoffset,
                                     regui.tail.geometry().height())
            
            # qtw.QWidget.DrawChildren
            
            
            # Render head
            regui.head.render(ihead, 
                              targetOffset=headoffset,
                              renderFlags=qtw.QWidget.DrawChildren)
            
            # Render tail
            for ftui in sorted(regui.tail.ftuis, key=lambda x: x.ft.start):
                if not ftui.ft.overlaps(reg): continue
                # If the feature is not fully inside, it needs to be
                #   clipped. SVG clipping does not work in Qt
                #   so we need to render them one by one separately 
                #   separately. The redraw_image method is already used
                #   elsewhere for optimization so I make use of it here.
                
                # The geometry calculations are moderately hellish here.
                
                start = max(0, tailregion.boundingRect().left()
                                 - ftui.geometry().left())
                maxwidth = (tailregion.boundingRect().right()
                                - ftui.geometry().left() - start)
                
                feature_picture = ftui.redraw_image(
                    into_picture=True,
                    maxwidth=maxwidth,
                    start=start)
                painter.drawPicture(
                    qtc.QPoint(
                        maxheadwidth 
                          + tailheadgap
                          + ftui.geometry().left() 
                          - tailregion.boundingRect().left()
                          + 0*start,
                        rowheight*i),
                    feature_picture)
            
            painter.drawPicture(
                qtc.QPoint(0,rowheight*i), 
                ihead)
            
            painter.setClipping(True)
            
        
        painter.end()
        print(suffix)
        if "png" in suffix:
            # Determine size
            columnrows = math.floor(2**15 / rowheight)
            columns = math.ceil(len(to_export)/columnrows)
            rowwidth = maxheadwidth+maxtailwidth-minoffset+tailheadgap
            
            pixheight = len(to_export)*rowheight if columns == 1 else columnrows*rowheight
            pixwidth  = columns*rowwidth + 100*(columns-1)
            
            #
            output = qtg.QPixmap(pixwidth, pixheight)
            if self.png_transparent.isChecked():
                output.fill(qtc.Qt.transparent)
            else:
                output.fill(qtc.Qt.white)
            
            painter = qtg.QPainter()
            painter.begin(output)
            
            for coli in range(0, columns):
                partial = qtg.QPixmap(rowwidth, pixheight)
                if self.png_transparent.isChecked():
                    partial.fill(qtc.Qt.transparent)
                else:
                    partial.fill(qtc.Qt.white)
                
                painter2 = qtg.QPainter()
                painter2.begin(partial)
                
                painter2.drawPicture(qtc.QPoint(0,-pixheight*(coli)), image)
                
                painter2.end()
                painter.drawPixmap(
                    (rowwidth+100)*(coli), 0,
                    rowwidth, pixheight,
                    partial,
                    0,0,0,0
                )
                #qtc.QPoint((rowwidth+100)*(columns-1),-pixheight*(columns-1))
            painter.end()
            
            if ".png" not in filepath:
                filepath+=".png"
            x=output.save(filepath, "PNG", quality=self.png_qualslider.value())
        elif "svg" in suffix:
            pixheight = len(to_export)*rowheight
            pixwidth  = maxheadwidth+maxtailwidth-minoffset+tailheadgap
            
            if ".svg" not in filepath:
                filepath+=".svg"
            
            output = qsvg.QSvgGenerator()
            output.setSize(qtc.QSize(pixwidth, pixheight))
            output.setViewBox(qtc.QRect(0,0,pixwidth,pixheight))
            output.setFileName(filepath)
            output.setTitle("CluSeek Output")
            
            painter = qtg.QPainter()
            painter.begin(output)
            painter.drawPicture(qtc.QPoint(0,0), image)
            painter.end()
        
    def export_neiview_gb(self, filepath_main, to_export):
        stranddict = {"0": None,
                      "+": 1,
                      "-": -1}
        
        
        directory = os.path.dirname(filepath_main)
        filename_core = os.path.basename(filepath_main).rstrip(".gb")
        
        for reg,regui in to_export:
            features = []
            for ft in sorted(reg.sc.fts.values(), key=lambda x: x.start):
                # Trim features down to region size
                if ft.is_inside(reg):
                    start = ft.start - reg.start
                    stop = ft.stop - reg.start + 1
                elif ft.overlaps(reg):
                    start = max(ft.start-reg.start, 1)
                    stop = min(ft.stop-reg.start+1, reg.stop-reg.start+1)
                else:
                    continue
                
                if ft.type == "cds":
                    qualifiers = {
                        "protein_id": [f"{ft.ref.accession}.{ft.ref.version}"],
                        "translation": [ft.ref.seq if ft.ref.seq else ""],
                        "product": [ft.ref.type]}
                else:
                    qualifiers = {}
                
                sft = SeqFeature.SeqFeature(
                    location=SeqFeature.FeatureLocation(start, stop),
                    type=ft.type,
                    #location_operator=None,
                    strand=stranddict[ft.strand],
                    id="<unknown id>",
                    qualifiers=qualifiers
                )
                
                features.append(sft)
            
            
            filename = f"{filename_core}_{reg.sc.tx.sciname}_{reg.sc.accession}.{reg.sc.version}_{reg.start}-{reg.stop}.gb"
            for bad_char in "/<>:\"\\|?*":
                filename = filename.replace(bad_char, "")
            filepath = os.path.join(directory, filename)
            
            record = SeqIO.SeqRecord(
                id=f"{reg.sc.accession}.{reg.sc.version}",
                seq=Seq.Seq(reg.seq),
                name=reg.sc.accession,
                description="Sequence exported from CluSeek",
                features=features,
                annotations={
                    "molecule_type": "DNA",
                    "topology": reg.sc.annotations["topology"],
                    "taxonomy": reg.sc.annotations["taxonomy"],
                    "references": reg.sc.annotations["references"],
                    "comment": reg.sc.annotations["comment"],
                    "accessions": [reg.sc.accession, "REGION:", f"{reg.start}..{reg.stop}"]
                }
            )
            
            with open(filepath, mode="w") as outfile:
                SeqIO.write(record, outfile, "gb")
    def export_neiview_fasta(self, filepath, to_export):
        lines = []
        for reg,regui in to_export:
            try:
                [tag.id for tag in regui.tags]
            except:
                code.interact(local=locals())
            header = [reg.sc.tx.sciname, 
                      reg.sc.tx.strain,
                      f"[{','.join([str(tag.identifier) for tag in regui.tags])}]",
                      f"{reg.sc.accession}.{reg.sc.version}",
                      reg.start,
                      reg.stop]
            header = [str(item) if item else "N/A" for item in header]
            header = "; ".join(header)
            header = f">{header}"
            
            lines.append(header)
            lines.append(reg.seq)
        
        with open(filepath, mode="w") as outfile:
            outfile.write("\n".join(lines))
            outfile.flush()
    def export_neiview_xlsx(self, filepath, to_export, align=True,
                            separator=", ", cell_variables=None,
                            comment_variables=None, transposed=False,
                            tags_individually=False):
        
        # * * Setup
        
        wbk = pxl.Workbook()
        
        row = 0
        col = 0
        
        sh_seqs = wbk.active
        sh_seqs.title = "Sequences"
        sh_table = wbk.create_sheet(title="Overview Table")
        #sh_binary = wbk.create_sheet(title="Presence-absence Table")
        
        clustering_results = self.neiview.neidata.cluster_hierarchy
        
        # * Configure variables to display/defaults
        #       Mostly ballast code.
        variables_to_display = {}
        if cell_variables:
            variables_to_display["cell"] = cell_variables
        else:
            variables_to_display["cell"] = ["proteingroup_name", 
                                            "protein_accession"]
        if comment_variables:
            variables_to_display["comment"] = comment_variables
        else:
            variables_to_display["comment"] = ["feature_position", 
                                               "protein_annotation"]
        
        # * * Utilities
        
        def _row():
            return(row if not transposed else col)
        
        def _col():
            return(col if not transposed else row)
        
        
        def getter_proteingroup_allannotations(ft,cl):
            composition = cl._get_composition()
            return("\n".join([f"{key}: {composition[key]}" for key in composition]))
        
        def color(*args):
            if len(args) == 3:
                # Assume 3 integers 1-255
                r,g,b = args
                return("%02x%02x%02x" % (r,g,b))
            elif len(args) == 1 and isinstance(args[0], qtg.QColor):
                # Assume QColor
                return("%02x%02x%02x" % (
                    args[0].red(),
                    args[0].green(),
                    args[0].blue()))
            elif len(args) == 1 and args[0] is None:
                return(None)
        
        variable_getters = {
            "feature_position": 
                lambda ft,cl: f"{ft.start}:{ft.stop}",
            "feature_length": 
                lambda ft,cl: f"{ft.stop-ft.start}",
            "protein_accession": 
                lambda ft,cl: ft.ref.accession,
            "protein_annotation": 
                lambda ft,cl: ft.ref.type if ft.ref.type else "N/A",
            "protein_sequence": 
                lambda ft,cl: ft.ref.seq if ft.ref.seq else "N/A",
            "proteingroup_identifier": 
                lambda ft,cl: cl.id_,
            "proteingroup_annotation": 
                lambda ft,cl: cl.get_annotation(),
            "proteingroup_allannotations": 
                getter_proteingroup_allannotations,
            "proteingroup_name": 
                lambda ft,cl: cl.get_name(),
            "proteingroup_size": 
                lambda ft,cl: str(len(cl.proteins)),
            "proteingroup_allaccessions":
                lambda ft,cl: "+".join(cl.proteins),
        }
        
        patterns = [
            (qtc.Qt.SolidPattern, "solid"),
            (qtc.Qt.Dense1Pattern, "darkGray"),
            (qtc.Qt.Dense2Pattern, "darkGray"),
            (qtc.Qt.Dense3Pattern, "mediumGray"),
            (qtc.Qt.Dense4Pattern, "lightGray"),
            (qtc.Qt.Dense5Pattern, "gray125"),
            (qtc.Qt.Dense6Pattern, "gray0625"),
            (qtc.Qt.Dense7Pattern, "gray0625"),
            (qtc.Qt.NoBrush, "none"),
            (qtc.Qt.HorPattern, "lightHorizontal"),
            (qtc.Qt.VerPattern, "lightVertical"),
            (qtc.Qt.CrossPattern, "lightGrid"),
            (qtc.Qt.BDiagPattern, "lightDown"),
            (qtc.Qt.FDiagPattern, "lightUp"),
            (qtc.Qt.DiagCrossPattern, "lightTrellis")]
        lines=[
            (qtc.Qt.SolidLine, "thin", "medium"),
            (qtc.Qt.DashLine, "dashed", "mediumDashed"),
            (qtc.Qt.DotLine, "dotted", ""),
            (qtc.Qt.DashDotLine, "dashDot", "mediumDashDot"),
            (qtc.Qt.DashDotDotLine, "dashDotDot", "mediumDashDotDot"),
        ]
            
        def convert_pattern(pattern_to_convert):
            # Have to imitate a dict while using "is"
            #   because Qt enums can't be used as keys
            #   in a dictionary -_-
            #   (some == equal others, meaning dict will
            #   return the wrong value)
            for pattern,converted in patterns:
                if pattern is pattern_to_convert:
                    return(converted)
            return(None)
        
        def style_cluster_cell(cell, cluster):
            bg = color(cluster.styledict["background_color"])
            fg = color(cluster.styledict["foreground_color"])
            fill = cluster.styledict["foreground_pattern"]
            if (fill is None 
              or fill is qtc.Qt.NoBrush):
                fill = "solid"
                fg = bg
                bg = None
            else:
                fill = convert_pattern(
                        cluster.styledict["foreground_pattern"])
            
            # Passing a None color argument crashes the function
            #   I can't think of a more elegant way of conditionally
            #   passing the arguments here.
            kwargs = {}
            if fill:
                kwargs["fill_type"] = fill
            if fg:
                kwargs["fgColor"] = fg
            if bg:
                kwargs["bgColor"] = bg
            
            cell.fill = pxl.styles.PatternFill(**kwargs)
            
            cell.font = pxl.styles.Font(
                color=color(cluster.styledict["text_color"]),
                bold=cluster.styledict["font_bold"],
                italic=cluster.styledict["font_italic"],
                underline="single" if cluster.styledict["font_underlined"] else None,
                name=cluster.styledict["font_family"],
                size=12)
        
        # * * Write the table
        
        # * Neiview
        
        # Calculate max offset:
        if align:
            # Redo the initial steps of the loop
            max_offset = 0
            for ii in range(len(to_export)):
                reg,regui = to_export[ii]
                
                seq_fts = sorted([ft if ft.type=="cds" else None 
                                   for ft in reg.fts.values()], 
                                 key=lambda ft: ft.start)
                for i in reversed(range(0, len(seq_fts))):
                    if seq_fts[i] is None:
                        del seq_fts[i]
            
                seq_clusters = []
                for ft in seq_fts:
                    try:
                        seq_clusters.append(
                            clustering_results.member_to_subcluster[ft.ref.accession])
                    except KeyError:
                        print("Could not find protein group for feature while exporting .xlsx file")
                        seq_clusters.append(None)
                if regui.last_align:
                    found_cluster = False
                    for i in range(len(seq_clusters)):
                        if seq_clusters[i] and regui.last_align in seq_clusters[i].get_lineage():
                            found_cluster = True
                            break
                    # It is possible to not find the aligned cluster, as it could
                    #   be outside of this region.
                    offset = i if found_cluster else 0
                else:
                    offset = 0
                
                if offset > max_offset:
                    max_offset = offset
        
        # Calculate maxtags
        max_tags = 0
        if tags_individually:
            for reg,regui in to_export:
                if len(regui.tags) > max_tags:
                    max_tags = len(regui.tags)
        
        # Fill in the values for each region
        for ii in range(len(to_export)):
            col = ii+1
            reg,regui = to_export[ii]
            
            seq_fts = sorted(
                [ft if ft.type=="cds" else None for ft in reg.fts.values()], 
                key=lambda ft: ft.start,
                reverse=regui.reversed)
            for i in reversed(range(0, len(seq_fts))):
                if seq_fts[i] is None:
                    del seq_fts[i]
            
            seq_clusters = []
            for ft in seq_fts:
                try:
                    seq_clusters.append(
                        clustering_results.member_to_supercluster[ft.ref.accession])
                except KeyError:
                    print("Could not find protein group for feature while exporting .xlsx file")
                    seq_clusters.append(None)
            
            
            # Align to last aligned if applicable
            if align and regui.last_align:
                found_cluster = False
                for i in range(len(seq_clusters)):
                    if seq_clusters[i] and regui.last_align in seq_clusters[i].get_lineage():
                        found_cluster = True
                        break
                # It is possible to not find the aligned cluster, as it could
                #   be outside of this region.
                offset = max_offset-i if found_cluster else 0
            else:
                offset = 0
            
            
            # Write the header
            row = 1
            sh_seqs.cell(
                column = _col(), row=_row(),
                value  = reg.sc.tx.sciname)
            
            row += 1
            sh_seqs.cell(
                column = _col(), row=_row(),
                value  = reg.sc.accession)
            
            if tags_individually:
                for tag in regui.tags:
                    row += 1
                    cell = sh_seqs.cell(
                        column=_col(), row=_row(),
                        value=tag.data["identifier"])
                    cell.comment = pxl.comments.Comment(tag.data["name"], "")
                    cell.fill = pxl.styles.PatternFill(
                        patternType="solid",
                        fgColor=color(tag.data["color_background"]))
                    cell.font = pxl.styles.Font(
                        color=color(tag.data["color_text"]))
                
                # Add empty space so sequences start at same point
                row += max_tags - len(regui.tags)
            else:
                row += 1
                cell = sh_seqs.cell(
                    column=_col(), row=_row(),
                    value=separator.join(
                        [tag.data["identifier"] for tag in regui.tags]))
            
            header_end = row
            
            # Apply sequence offset
            row += offset
            
            # Write sequence itself (=tail)
            for i in range(len(seq_fts)):
                row += 1
                
                feature = seq_fts[i]
                cluster = seq_clusters[i]
                
                cell = sh_seqs.cell(
                    column=_col(), row=_row())
                
                if cluster:
                    value = separator.join(
                        [variable_getters[varname](feature,cluster) \
                         for varname in variables_to_display["cell"]])
                    cell.value = value
                    
                    comment = "\n".join(
                        [variable_getters[varname](feature,cluster) \
                        for varname in variables_to_display["comment"]])
                    cell.comment = pxl.comments.Comment(comment, "")
                    
                    style_cluster_cell(cell, cluster)
                else:
                    cell.value = "N/A"
                    cell.comment = pxl.comments.Comment(
                        feature.ref.accession, "")
                
        
        # Freeze the header pane
        row = header_end+1
        col = 1
        topleft = sh_seqs.cell(column=_col(), row=_row())
        sh_seqs.freeze_panes = topleft
        
        
        # * Protein Groups
        row = 0
        col = 0
        # Get overview table
        otable = self.neiview.proteingroupstable.cluster_table
        # Copy headers
        for col in range(0, otable.columnCount()):
            item = otable.horizontalHeaderItem(col)
            sh_table.cell(column=col+1, row=1, value=item.text())
        # Copy cells
        for row in range(0, otable.rowCount()):
            for col in range(0, otable.columnCount()):
                item = otable.item(row, col)
                if item is None:
                    # If it's not an item, we can assume it's
                    #   a ColourableButton until more widgets are inserted
                    item = otable.cellWidget(row, col)
                    
                    cluster = item.cluster
                    value = cluster.get_name()
                    
                    cell = sh_table.cell(column=col+1, row=row+2, value=value)
                    
                    style_cluster_cell(cell, cluster)
                else:
                    value = item.text()
                    cell = sh_table.cell(column=col+1, row=row+2, value=value)
        
        wbk.save(filename=filepath)
    def export_neiview_gml(self, filepath, to_export, cutoff, metric_i):
        metrics = [
            self.neiview.network_regions_by_jaccard,
            #self.neiview.network_regions_by_cluster_order
        ]
        metric = metrics[metric_i]
        
        graph = metric(to_export, cutoff)
        nx.write_gml(graph, filepath)
    def export_proteins_network(self, filepath, to_export, local_group):
        master_graph = self.neiview.neidata.cluster_hierarchy.graph
        
        clusters = []
        proteins = set()
        for protein,region,cluster_bundle in to_export:
            proteins.add(protein)
            clusters.extend(cluster_bundle)
        clusters = set(clusters)
        print(" * * * ")
        print(clusters)
        
        if local_group:
            mode = "local_group"
        else:
            mode = "standard"
        
        subgraph,node_to_cluster = ClusterNetworkViewer2.get_subgraph(
                                        master_graph,
                                        clusters,
                                        mode)
        subgraph = nx.Graph(subgraph)
        
        for id_ in node_to_cluster:
            cluster = node_to_cluster[id_]
            if not local_group:
                relevant_proteins = set(cluster.proteins.values())&proteins
            else:
                relevant_proteins = set(cluster.proteins.values())
            
            if len(relevant_proteins) == 0:
                subgraph.remove_node(id_)
                continue
            
            subgroup_data = {
                "GroupID": int(cluster.get_toplevel().id_),
                "GroupName": cluster.get_toplevel().get_name(),
                "SubgroupID": int(cluster.id_),
                "SubgroupName": cluster.get_name(),
                "SubgroupSize": len(relevant_proteins),
                "SubgroupProteins": ", ".join([protein.accession for protein in relevant_proteins]),
                "SubgroupTopAnnotation": cluster._get_composition().most_common()[0][0],
                "SubgroupAnnotations": ", ".join([protein.type for protein in relevant_proteins]),
            }
            subgraph.nodes[id_].update(subgroup_data)
            subgraph.nodes[id_]["SubgroupCentroidSequence"] = subgraph.nodes[id_]["sequence"]
            del subgraph.nodes[id_]["sequence"]
        
        for edge_id in list(subgraph.edges):
            if edge_id[0] == edge_id[1]:
                subgraph.remove_edge(*edge_id)
                continue
            evalue = subgraph.edges[edge_id]["evalue"]
            subgraph.edges[edge_id]["evalue"] = round(math.log(evalue, 10) if evalue != 0 else 0)
        nx.write_gml(subgraph, filepath)
        
    def export_proteins_multifasta(self, filepath, to_export):
        output = []
        noseq = []
        for protein,region,clusters in to_export:
            if not protein.seq:
                noseq.append(protein)
                continue
            sequence=protein.seq
            subcluster = self.neiview.neidata.cluster_hierarchy.member_to_subcluster[protein.accession]
            header=(f">{protein.accession}\t{region.sc.tx.sciname}\t{region.sc.accession}\t"
                    f"{'|'.join([cluster.name if cluster.name else cluster.id_ for cluster in clusters])}\t"
                    f"{protein.type}")
            #header=(f">{'|'.join([cluster.name if cluster.name else cluster.id_ for cluster in clusters])};"
            #       f"{subcluster.id_};{subcluster.centroid.accession};{taxon.sciname};{protein.accession};{protein.type}")
            output.append(header)
            output.append(sequence)
        
        with open(filepath, "w") as file:
            file.write("\n".join(output))
        if noseq:
            qtw.QMessageBox.warning(
                None, 
                "Missing sequences", 
                f"The aminoacid sequences for {len(noseq)}/{len(noseq)+len(output)} "
                 "are not available locally and were excluded from the multifasta file. "
                 "Consider using the CSV output instead if you wish to see them.")
    def export_proteins_csv(self):
        to_export = []
        for cluster in self.hierarchy.clusters_all.values():
            if cluster.get_selected():
                to_export.append(cluster)
        
        filepath,suffix = qtw.QFileDialog.getSaveFileName(
                        caption="Save output",
                        filter="Comma separated values(*.csv)")
        if not filepath:
            return
        
        output = []
        output.append(["Protein group ID", "Protein group name", "Organism", "Protein accession", "Protein annotation", "Sequence"])
        for cluster in to_export:
            for protein in cluster.proteins.values():
                if not protein.seq:
                    noseq.append(protein)
                    continue
                taxa = [ft.sc.tx for ft in protein.fts.values()]
                for taxon in taxa:
                    row[cluster.id_,
                        cluster.name,
                        taxon,
                        protein.accession]
                row=[protein.accession, 
                     cluster.id_, 
                     cluster.name if cluster.name else "N/A",
                     protein.type, 
                     protein.seq if protein.seq else "N/A"]
                output.append(row)
        
        with open(filepath, "w", newline="") as file:
            writer = csv.writer(file)
            writer.writerows(output)
class NeiConfig(qtw.QWidget, Ui_NeiConfig):
    sig_closed = qtc.Signal()
    def __init__(self, neiview):
        super().__init__()
        
        self.neiview = neiview
        
        self._init_ui()
        
        self.load_values()
    def _init_ui(self):
        self.setupUi(self)
        self.io_apply_btn.clicked.connect(self.cbk_apply)
        self.io_cancel_btn.clicked.connect(self.cbk_cancel)
    def load_values(self):
        if self.neiview.dp["size_mode"] == "proportional":
            self.cds_repr_radio.setChecked(True)
        elif self.neiview.dp["size_mode"] == "fixed":
            self.cds_fixed_radio.setChecked(True)
        
        self.cds_width_spin.setValue(self.neiview.dp["size_fixed_width"])
        self.cds_height_spin.setValue(self.neiview.dp["size_fixed_height"])
        self.cds_bpscaling_spin.setValue(self.neiview.dp["bp_per_pixel"])
        
        labels = self.neiview.dp["show_labels"]
        if labels == "all":
            self.pglabels_all_radio.setChecked(True)
        elif labels == "custom":
            self.pglabels_custom_radio.setChecked(True)
        elif labels == "none":
            self.pglabels_none_radio.setChecked(True)
        
        self.headers_taxon_chk.setChecked(
            "taxon" in self.neiview.dp["header_elements_shown"])
        self.headers_region_chk.setChecked(
            "region" in self.neiview.dp["header_elements_shown"])
        self.headers_tags_chk.setChecked(
            "tags" in self.neiview.dp["header_elements_shown"])
        
        if self.neiview.dp["arrow_shape"] == "notched":
            self.arrowshape_notched_radio.setChecked(True)
        elif self.neiview.dp["arrow_shape"] == "standard":
            self.arrowshape_standard_radio.setChecked(True)
        elif self.neiview.dp["arrow_shape"] == "compact":
            self.arrowshape_compact_radio.setChecked(True)
    def apply_values(self):
        if self.cds_repr_radio.isChecked():
            self.neiview.dp["size_mode"] = "proportional"
        elif self.cds_fixed_radio.isChecked():
            self.neiview.dp["size_mode"] = "fixed"
        
        cds_width = self.cds_width_spin.value()
        cds_height = self.cds_height_spin.value()
        cds_bpscaling = self.cds_bpscaling_spin.value()
        self.neiview.dp["size_fixed_width"] = cds_width
        self.neiview.dp["size_fixed_height"] = cds_height
        self.neiview.dp["bp_per_pixel"] = cds_bpscaling
        
        if self.pglabels_all_radio.isChecked():
            labels = "all"
        elif self.pglabels_custom_radio.isChecked():
            labels = "custom"
        elif self.pglabels_none_radio.isChecked():
            labels = "none"
        self.neiview.dp["show_labels"] = labels
        
        header_elements_checked = [
            (self.headers_taxon_chk, "taxon"),
            (self.headers_region_chk, "region"),
            (self.headers_tags_chk, "tags")
        ]
        toshow,tohide = [],[]
        for chk,key in header_elements_checked:
            if chk.isChecked():
                toshow.append(key)
            else:
                tohide.append(key)
        
        self.neiview.change_header_elements(toshow, tohide)
        
        if self.arrowshape_notched_radio.isChecked():
            self.neiview.dp["arrow_shape"] = "notched"
        elif self.arrowshape_standard_radio.isChecked():
            self.neiview.dp["arrow_shape"] = "standard"
        elif self.arrowshape_compact_radio.isChecked():
            self.neiview.dp["arrow_shape"] = "compact"
        
        # Finally
        self.neiview.update_geometry()
        self.neiview.force_repaint()
    def cbk_cancel(self):
        self.setHidden(True)
        self.sig_closed.emit()
    def cbk_apply(self):
        self.apply_values()
        self.setHidden(True)
        self.sig_closed.emit()
    def cbk_settings_changed(self):
        self.load_values()

class AboutNeiview(qtw.QWidget, Ui_AboutNeiview):
    def __init__(self, neiview):
        super().__init__()
        self.neiview = neiview
        
        self._init_ui()
        
        self.neiview.sig_neiview_loaded.connect(self.update_about)
    def _init_ui(self):
        self.setupUi(self)
    def update_about(self):
        str_general = [
            f"Display name: {self.neiview.displayname}",
            f"Internal identifier: {self.neiview.id_}",
            f"Genetic regions (clusters) total: {len(self.neiview.regions_all)}",]
        str_general = "\n".join(str_general)
        
        if (self.neiview.dp.get("creation_settings") is not None
          and "creation_data_missing" not in self.neiview.dp["creation_settings"]):
            str_creation_params = []
            for key in sorted(list(self.neiview.dp["creation_settings"])):
                if key in ["marker_accsets", "about_rules"]: continue
                
                value = self.neiview.dp["creation_settings"][key]
                str_creation_params.append(f"{key.replace('_', ' ')}: {value}")
            str_creation_params.append("")
            str_creation_params.append("Rules:")
            for aboutstring in self.neiview.dp["creation_settings"]["about_rules"]:
                str_creation_params.append("    "+aboutstring)
            
            str_creation_params = "\n".join(str_creation_params)
        else:
            str_creation_params = "<Creation data missing>"
        
        self.general_info_lbl.setText(str_general)
        self.creation_info_lbl.setText(str_creation_params)

class NeiView(qtw.QWidget, Ui_NeiView):
    @staticmethod
    def get_default_display_policy():
        default_display_policy = {
            "size_mode": "proportional", # or fixed
            "size_fixed_width": 50,
            "size_fixed_height": 21,
            "bp_per_pixel": 33,
            "all_tags": {},
            #taxon,region,tags
            "header_elements_shown": ["taxon","region","tags"],
            "restriction_rules": [],
            "displayname": "",
            "highlight_rare": False,
            "show_labels": "all", #str: all, custom, none
            "neiredraw_id": 0,
            "creation_settings": None,
            "cluster_graph": None,
            "arrow_shape": "standard", # notched, standard, compact
        }
        return(default_display_policy)
    is_initialized = False

    sig_neiview_loaded = qtc.Signal()
    sig_selection_manager_created = qtc.Signal()
    sig_clusters_scrollable_mousepress = qtc.Signal(object)
    def __init__(self, neidata, id_, displayname=None):
        super().__init__()
        
        self.display_policy = self.get_default_display_policy()
        self.neidata = neidata
        self.regions_all = []
        self.regions_displayed = []
        self.regions_visible = set()
        self.regions_selection_anchor = None
        
        # Selection data for reguis is saved in the regui itself.
        #   Reguis are selected via the normal method while being loaded,
        #   therefore regions_selected is not saved.
        self.regions_selected = []
        
        self.id_ = id_
        
        self.taginfowidget = RegionUI.TagInfoWidget(self)
        self.neiconfig = NeiConfig(self)
        self.selection_manager = None # See self._init_selection_manager
        
        self.not_shown = True
        self._await_ready_iterations_left = -1
        
        self._handle_moving = False
        self._handle_pos_last = None
        self._handle_move_time = None
        
        self.dontdraw_last_update = 0
        
        self.__init_ui()
        
        # This displayname needs to be defined only after UI is initialized
        self.displayname = displayname
    def __init_ui(self):
        self.setupUi(self)
        
        self.clusters_scrollable_lay.setAlignment(qtc.Qt.AlignTop)
        self.headers_scrollable_lay.setAlignment(qtc.Qt.AlignTop)
        
        self.displayname_ledit.editingFinished.connect(
            self.cbk_displayname_ledit_changed)
        self.displayname_ledit.setText(
            self.displayname)
        
        self.delete_btn.clicked.connect(
            self.cbk_delete)
        
        self.destroyed.connect(self.cbk_deleted)
        
        # * * Misc widgets
        # * Drag and drop indicator
        self.drag_drop_indicator = self.DragDropIndicator(self.headers_scrollable)
        self.drag_drop_indicator.setHidden(True)
        
        self._init_hidebuttons()
        
        self._init_slider()

        # * Cluster overlay
        self.cluster_overlay = self.ClusterOverlay(self.clusters_slide, 
                                                   self.clusters_scroller, 
                                                   self)
        
        # * Cluster scoller overlay
        self.cluster_scrollable_overlay = self.ClusterScrollableOverlay(
                                            self.clusters_scrollable_container,
                                            self.clusters_scrollable,
                                            self)
        
        # * * Scrollbar chaos
        self.horizontal_scrollbar.setHidden(True)
        self.vertical_scrollbar.setHidden(True)
        
        self.clusters_scroller.setHorizontalScrollBar(self.horizontal_scrollbar)
        self.clusters_scroller.setVerticalScrollBar(self.vertical_scrollbar)
        self.headers_scroller.setVerticalScrollBar(self.vertical_scrollbar)
        
        self.headers_scroller.setHorizontalScrollBar(self.headers_scrollbar)
        self.headers_slide_lay.addWidget(self.headers_scrollbar, 1, 0, 1, 2)
        
        #
        self.headers_scroller.setSizeAdjustPolicy(qtw.QAbstractScrollArea.AdjustToContents)
        
        #self.horizontal_scrollbar.parentWidget().layout().removeWidget(
        #    self.horizontal_scrollbar)
        #self.vertical_scrollbar.parentWidget().layout().removeWidget(
        #    self.vertical_scrollbar)
        
        #self.neiview_lay.addWidget(self.horizontal_scrollbar, 3, 4, 1, 1)
        #self.neiview_lay.addWidget(self.vertical_scrollbar, 2, 5, 1, 1)
        self.clusters_slide_lay.addWidget(self.horizontal_scrollbar, 2, 0, 1, 1)
        self.clusters_slide_lay.addWidget(self.vertical_scrollbar, 0, 1, 1, 1)
        #self.horizontal_scrollbar.setHidden(False)
        #self.vertical_scrollbar.setHidden(False)
        
        # Update clusters_scroller scrollbar when it is resized
        #self.clusters_scroller.setSizeAdjustPolicy(qtw.QAbstractScrollArea.AdjustToContents)

        # * * Callbacks
        #self.show_lefttabs_btn.clicked.connect(self.cbk_toggle_lefttabs)
        #self.show_righttabs_btn.clicked.connect(self.cbk_toggle_righttabs)
        
        # Depricated save button
        #self.tool_save.clicked.connect(self.cbk_save_neiview)
        
        self.tool_export.clicked.connect(self.cbk_export)
        self.config_btn.clicked.connect(self.cbk_config)
        
        self.tool_toggle_sizemode.setCheckable(True)
        self.tool_toggle_sizemode.toggled.connect(self.cbk_sizemode_toggled)
        
        self.searchbar_ledit.textChanged.connect(self.cbk_search_string_changed)
        
        self.vertical_scrollbar.rangeChanged.connect(
            self.on_vert_scrollbar_range_changed)
        self.vertical_scrollbar.valueChanged.connect(
            self.on_vert_scrollbar_moved)
        
        self.highlightrare_btn.clicked.connect(
            self.cbk_highlight_rare)
        
        def paintEvent_white_background(event):
            painter = qtg.QPainter()
            painter.begin(self.clusters_scrollable)
            
            h = self.clusters_scrollable.geometry().height()-1
            w = self.clusters_scrollable.geometry().width()-1
            
            painter.setBrush(qtg.QBrush(qtg.QColor(255,255,255,255)))
            
            painter.drawRect(qtc.QRectF(0,0,w,h))
            
            painter.end()
        self.clusters_scrollable.paintEvent = paintEvent_white_background
        
        # Hooking up press events to cluster scrollable
        self.clusters_scrollable.mousePressEvent = \
            self.cluster_scrollable_mouse_press
        
        # Selection manager
        self.sig_neiview_loaded.connect(self._init_selection_manager)
        
        # Info window 
        self.frame_info = InfoFrame(self)
        self.lefttabs_tabs.addTab(self.frame_info, "Info")
        
        # StyleDictCreator is created only after data is loaded
        self.sig_neiview_loaded.connect(
            self._init_styledictcreator)
        
        # About neiview
        self.about_neiview = AboutNeiview(self)
        self.lefttabs_tabs.addTab(self.about_neiview, "General Info")
        
        # Sort clusters menubutton
        self._init_sortingmenu()
        
        # Cluster legend
        self.legend = NeiView.Legend()
        self.legend.setHidden(True)
        self.legend_btn.clicked.connect(
            self.cbk_show_cluster_legend)
        add_question_mark_icon(self.legend_btn)
        
        # Protein groups table
        self.proteingroupstable = ProteinGroupTable(self)
        #self.proteingroupstable.toolbar_row.addWidget(
        #    self.show_righttabs_btn)
        self.right_slide.layout().addWidget(
            self.proteingroupstable,
            0,0)
        
        self._init_header_background_context_menu()
        
        # On load
        self.sig_neiview_loaded.connect(self.reset_view)
        self.sig_neiview_loaded.connect(self._on_load)
        self.tool_hide3.click()
        self.tool_hide4.click()
    def _init_sortingmenu(self):
        sorting_menu = qtw.QMenu(self)
        
        #ACTION: Sort by jaccard
        action = sorting_menu.addAction("Sort by shared protein groups (Jaccard index)")
        action.triggered.connect(self.sort_by_jaccard)
        
        #ACTION: Sort by alphabet
        action = sorting_menu.addAction("Sort alphabetically by taxon name")
        action.triggered.connect(self.sort_by_alphabet)
        
        self.sort_btn.setMenu(sorting_menu)
    def _init_styledictcreator(self):
        self.styledictcreator = StyledictCreator(self)
    def _init_header_background_context_menu(self):
        def show_all_headers():
            for regui in self.regions_all:
                regui.set_displayed(True)
        
        def contextMenuEvent(event):
            event.accept()
            
            menu = qtw.QMenu(self)
            
            # ACTION: Show all
            action = menu.addAction("Show all hidden")
            action.triggered.connect(show_all_headers)
            
            
            # Generate menu at click location
            menu.popup(event.globalPos())
        self.headers_scrollable.contextMenuEvent = contextMenuEvent
    def _on_load(self):
        self.is_initialized = True
    def _init_slider(self):
        self.handle_slide.setCursor(qtc.Qt.SizeHorCursor)
        
        def mousePressEvent(event):
            if not event.button() == qtc.Qt.LeftButton:
                event.ignore()
                return
            event.accept()
            self._handle_moving = True
            self._handle_pos_last = event.globalPos().x()
            self._handle_move_time = time.time()
        def mouseReleaseEvent(event):
            if not self._handle_moving:
                event.ignore()
                return
            event.accept()
            self._handle_moving = False
            self._handle_pos_last = None
        def mouseMoveEvent(event):
            if not self._handle_moving or (time.time() - self._handle_move_time) < 0.1:
                event.ignore()
                return
            event.accept()
            self._handle_move_time = time.time()
            handle_pos=event.globalPos().x()
            diff=handle_pos - self._handle_pos_last
            self.headers_slide.setFixedWidth(
                max(self.headers_slide.geometry().width() + diff,0))
            
            self._handle_pos_last=handle_pos
        
        self.handle_slide.mousePressEvent = mousePressEvent
        self.handle_slide.mouseReleaseEvent = mouseReleaseEvent
        self.handle_slide.mouseMoveEvent = mouseMoveEvent
    def _init_selection_manager(self):
        # Selection manager
        self.selection_manager = SelectionManager(
            neiview=self)
        #self.neidata.cluster_hierarchy.sig_cluster_selected.connect(
        #    self.selection_manager.cbk_selection_changed)
        #self.neidata.cluster_hierarchy.sig_styledicts_changed.connect(
        #    self.cbk_neiview_repaint_needed)
        self.righttabs_tabs.addTab(self.selection_manager, "Selection")
        self.sig_selection_manager_created.emit()
    @property 
    def dp(self):
        return(self.display_policy)
    @property
    def restriction_rules(self):
        return(self.display_policy["restriction_rules"])
    @property
    def displayname(self):
        return(self.dp["displayname"])
    @displayname.setter
    def displayname(self, value):
        self.dp["displayname"] = value
        AppRoot.mainwindow.maintabs.setTabText(
                AppRoot.mainwindow.maintabs.indexOf(self), 
                value)
        self.displayname_ledit.setText(value)
    def cbk_displayname_ledit_changed(self):
        if self.displayname_ledit.text() != self.displayname\
          and self.displayname_ledit.text() != "":
            self.displayname = self.displayname_ledit.text()
    def showEvent(self, event):
        event.ignore()
        if self.not_shown:
            self.not_shown = False
            def hack():
                regui = self.regions_all[0]
                regui.set_offset(regui.offset+1)
                def hack2():
                    regui.set_offset(regui.offset-1)
                AppRoot.timer.after_delay_do(hack2)
                self.reset_view()
            AppRoot.timer.after_delay_do(hack)
    # Core UI callbacks
    def cbk_toggle_lefttabs(self):
        self.lefttabs_tabs.setHidden(not self.lefttabs_tabs.isHidden())
    def cbk_toggle_righttabs(self):
        self.righttabs_tabs.setHidden(not self.lefttabs_tabs.isHidden())
    def cbk_search_string_changed(self):
        self.restrict_to_search(self.searchbar_ledit.text())
    def cbk_sizemode_toggled(self, bool_):
        if bool_:
            self.use_fixed_widths()
        else:
            self.use_proportional_widths()
    def _init_hidebuttons(self):
        for i in range(1,5): # 1 to 4
            btn = getattr(self, f"tool_hide{i}")
            btn.setCheckable(True)
            btn.toggle()
            btn.toggled.connect(getattr(self, f"cbk_hide{i}"))
    def cbk_hide1(self, checked):
        self.lefttabs_tabs.setHidden(not checked)
    def cbk_hide2(self, checked):
        self.clusterview_frame.setHidden(not checked)
    def cbk_hide3(self, checked):
        self.right_slide.setHidden(not checked)
    def cbk_hide4(self, checked):
        self.righttabs_tabs.setHidden(not checked)
    def cbk_export(self):
        AppRoot.neiexport.open(self)
        AppRoot.neiexport.activateWindow()
    def cbk_delete(self):
        # Deletes the neiview from the UI.
        #   Does not reliably purge everything, but it does its best.
        #   Still leaks memory.
        answer = qtw.QMessageBox.question(None,
            "Are you sure?",
            "This will permanently remove this cluster view and any changes"
            " you have made to it.",
            buttons=qtw.QMessageBox.Yes | qtw.QMessageBox.No)
        if answer == qtw.QMessageBox.Yes:
            
            to_purge = [self]
            for obj in to_purge:
                if isinstance(obj, qtc.QObject) and len(obj.__dict__) > 0:
                    to_purge.extend(obj.children())
            
            to_purge_set = weakref.WeakSet(to_purge)
            
            #
            
            AppRoot.dbman.purge_neidata(
                self.id_)
            AppRoot.mainwindow.maintabs.removeTab(
                AppRoot.mainwindow.maintabs.indexOf(
                    self))
            self.setParent(None)
            self.neidata.neiview = None
            self.deleteLater()
            
            #
            
            while to_purge:
                obj = to_purge.pop(-1)
                
                keys_to_purge = []
                
                for key in obj.__dict__.keys():
                    value = obj.__dict__[key]
                    if not isinstance(value, qtc.QObject):
                        continue
                    
                    if key=="neiview" or value in to_purge_set:
                        keys_to_purge.append(key)
                
                [delattr(obj, key) for key in keys_to_purge]
                
                obj.deleteLater()
            del keys_to_purge
            
            gc.collect()
            
            def really_after():
                gc.collect()
            AppRoot.timer.after_delay_do(really_after, 2000)
    def cbk_show_cluster_legend(self):
        self.legend.show()
        self.legend.activateWindow()
    @staticmethod
    def cbk_deleted():
        pass
    def cluster_scrollable_mouse_press(self, event):
        if event.buttons() == qtc.Qt.LeftButton:
            event.accept()
            self.neidata.cluster_hierarchy.clear_selection()
        else:
            event.ignore()
        self.sig_clusters_scrollable_mousepress.emit(event)
    def hideEvent(self, event):
        AppRoot.neiexport.hide()
        self.taginfowidget.hide()
        self.neiconfig.hide()
    
    def cbk_neiview_repaint_needed(self):
        for regui in self.regions_visible:
            # Using a region method isn't as great performance-wise, 
            #   since it adds one more method invocation
            #   for a large scale time-sensitive operation, but code is
            #   more maintaible so whatever.
            regui.redraw_clusterwidgets()
            
            #for ftui in regui.tail.ftuis:
            #    ftui.update()
            #    ftui.update_tooltip()
    def force_repaint(self):
        self.dp["neiredraw_id"] += 1
        if self.dp["neiredraw_id"] > 50000:
            self.dp["neiredraw_id"] = 0
        self.cbk_neiview_repaint_needed()
    
    def cbk_clunames_changed(self):
        self.cbk_neiview_repaint_needed()
    def cbk_config(self):
        self.neiconfig.show()
        self.neiconfig.activateWindow()
        self.neiconfig.load_values()
    
    def get_new_tag(self):
        # Assign new ID
        id_=0
        for tag_id in self.display_policy["all_tags"]:
            if tag_id >= id_:
                id_ = tag_id + 1
        
        # Create it and register it
        tag = RegionUI.Tag(id_)
        self.display_policy["all_tags"][id_] = tag
        
        return(tag)
    def remove_tag(self, tag):
        for regui in list(tag.reguis):
            regui.remove_tag(tag)
        
        del self.display_policy["all_tags"][tag.data["id"]]
    
    def open_taginfowidget(self, tag, globalX=None, globalY=None):
        self.taginfowidget.load_tag(tag)
        self.taginfowidget.show()
        self.taginfowidget.activateWindow()
        if globalX and globalY:
            self.taginfowidget.setGeometry(globalX, globalY, 250, 300)
    #
    def _await_ready(self):
        if self.horizontal_scrollbar.maximum() == 0 and self._await_ready_iterations_left > 0:
            self._await_ready_iterations_left -= 1
            AppRoot.timer.after_delay_do(self._await_ready, 400)
        
        AppRoot.timer.after_delay_do(self.sig_neiview_loaded.emit, 400)
    def init_displayed_data(self, jaccard=True):
        # Update settings
        self.dp["creation_settings"] = self.neidata.settings
        
        
        # TODO: Set width of clusters_scrollable to 3x longest region's length.
        
        # Testing only
        for region in sorted(self.neidata.regions_to_display, 
                             key=lambda x: x.sc.tx.sciname):
            regionUI = RegionUI(self, region)
            self.regions_all.append(regionUI)
            self.regions_displayed.append(regionUI)
            
            self.clusters_scrollable_lay.addWidget(regionUI.tail)
            self.headers_scrollable_lay.addWidget(regionUI.head)
            
            # We want regions that aren't displayed to stay hidden
            #   and for now, we don't know if these will be displayed
            #   and in what order.
            #region.set_hidden(True)
        
        centre = self.get_centre()
        for regui in self.regions_all:
            regui.set_offset(centre - (regui.reg.length()/self.display_policy["bp_per_pixel"])/2)
            regionUI.update_tail_geometry()
        
        self.drag_drop_indicator.setParent(None)
        self.drag_drop_indicator.setParent(self.headers_scrollable)
        
        # Post-creation tasks
        
        if jaccard:
            self.sort_by_jaccard()
        
        if self.dp["creation_settings"]["loading_existing_session"] is False:
            self.highlight_markers()
        
        
        # This is here solely that the current version of the
        #   graph gets saved, because we don't save NeiData or
        #   ClusterHierarchy objects, and doing so
        #   would break save file compatbility again.
        if self.neidata.cluster_hierarchy.graph:
            self.dp["cluster_graph"] = self.neidata.cluster_hierarchy.graph
        
        AppRoot.mainwindow.maintabs.setCurrentWidget(self)
        AppRoot.timer.after_delay_do(self._await_ready)
    
    def temp_show_all(self):
        # Does this get used?
        for region in self.regions_all:
            region.set_displayed(True)
    # Content visibility limits
    def on_vert_scrollbar_range_changed(self, min_, max_):
        pass
    def on_vert_scrollbar_moved(self, position):
        if time.time() - self.dontdraw_last_update < 0.2: return
        self.dontdraw_last_update = time.time()
        # .maximum() can be 0!
        vbmax = self.vertical_scrollbar.maximum()
        scroll_percentage = (position/vbmax) if vbmax != 0 else 0
        viewed_area_start = ((self.clusters_scrollable.geometry().height()
                              -self.clusters_scroller.geometry().height())
                             *scroll_percentage)
        viewed_area_stop  = (viewed_area_start 
                             + self.clusters_scroller.geometry().height())
        
        viewrect = qtc.QRect(0, 
                             viewed_area_start, 
                             self.geometry().width(),
                             viewed_area_stop-viewed_area_start)
        
        for regui in self.regions_all:
            if regui.hidden: continue
            regui.set_dontdraw(
                not (viewed_area_start 
                     < regui.tail.geometry().top() 
                     < viewed_area_stop))
    
    # = Display rules
    def change_header_elements(self, toshow, tohide):
        valid={"taxon","region","tags"}
        toshow=set(toshow)
        tohide=set(tohide)
        for item in toshow:
            assert item in valid
            if item not in self.dp["header_elements_shown"]:
                self.dp["header_elements_shown"].append(item)
        for item in tohide:
            assert item in valid
            if item in self.dp["header_elements_shown"]:
                self.dp["header_elements_shown"].remove(item)
        self.update_headers_content()
    def update_headers_content(self):
        displayed = self.dp["header_elements_shown"]
        taxon = "taxon" not in displayed
        region = "region" not in displayed
        tags = "tags" not in displayed
        
        for regui in self.regions_all:
            regui.head.set_taxon_hidden(taxon)
            regui.head.set_region_hidden(region)
            regui.head.set_tags_hidden(tags)
    
    def use_proportional_widths(self): #TODO:
        self.display_policy["size_mode"] = "proportional"
        self.update_geometry()
    def use_fixed_widths(self): #TODO:
        self.display_policy["size_mode"] = "fixed"
        self.update_geometry()
    def update_geometry(self):
        for region in self.regions_all:
            region.update_tail_geometry()
            region.update_head_geometry()
    
    # = Subset changes
    def restrict_to_search(self, search_string):
        for regionui in self.regions_displayed:
            if search_string.lower() in regionui.reg.sc.tx.sciname.lower() \
              or search_string.lower() in regionui.reg.sc.accession.lower() \
              or search_string.lower() in regionui.reg.sc.fts \
              or any(search_string.lower() in tag.data["name"].lower() for tag in regionui.tags)\
              or any(search_string.lower() in tag.data["identifier"].lower() for tag in regionui.tags):
                regionui.set_hidden(False)
            else:
                regionui.set_hidden(True)
    
    # = Utility methods 1
    def get_centre(self):
        bp_per_pixel = self.display_policy["bp_per_pixel"]
        if self.display_policy["size_mode"] == "proportional":
            centre = max(regionui.reg.length() for regionui in self.regions_all)\
                     /bp_per_pixel
        elif self.display_policy["size_mode"] == "fixed":
            centre = max(len(regionui.tail.ftuis) for regionui in self.regions_all)\
                     *(self.display_policy["size_fixed_width"]+1)
        return(centre)
    def align_regions_to_cluster(self, cluster):
        centre = self.get_centre()
        for regionui in self.regions_all:
            regionui.align_to(cluster, centre)
        self.scroll_to(x=centre)
    def scroll_to(self, x=None, y=None):
        if x is None:
            if self.horizontal_scrollbar.maximum() == self.horizontal_scrollbar.maximum():
                x = 0
            else:
                x=(self.clusters_scrollable.geometry().width()
                    *(self.horizontal_scrollbar.value() - self.horizontal_scrollbar.minimum())
                    /(self.horizontal_scrollbar.maximum() - self.horizontal_scrollbar.minimum()))
        if y is None:
            if self.vertical_scrollbar.maximum() == self.vertical_scrollbar.maximum():
                y = 0
            else:
                y=(self.clusters_scrollable.geometry().height()
                    *(self.vertical_scrollbar.value() - self.vertical_scrollbar.minimum())
                    /(self.vertical_scrollbar.maximum() - self.vertical_scrollbar.minimum()))
            
        self.clusters_scroller.ensureVisible(x,y)
    def move_region_to(self, moved_regui, target, move_below=False):
            if move_below is None:
                move_below = moved_regui.head.geometry().top() < target.head.geometry().top()
            # Move head
            self.headers_scrollable_lay.removeWidget(
                moved_regui.head)
            self.headers_scrollable_lay.insertWidget(
                target.head.parentWidget().layout().indexOf(target.head)+move_below,
                moved_regui.head)
            
            # Move tail
            self.clusters_scrollable_lay.removeWidget(
                moved_regui.tail)
            self.clusters_scrollable_lay.insertWidget(
                self.clusters_scrollable_lay.indexOf(target.tail)
                    +move_below,
                moved_regui.tail)
            
            # Move abstract background data
            self.regions_displayed.remove(moved_regui)
            self.regions_displayed.insert(
                self.regions_displayed.index(target)
                    +move_below,
                moved_regui)
            
            self.regions_all.remove(moved_regui)
            self.regions_all.insert(
                self.regions_all.index(target)
                    +move_below,
                moved_regui)
    def move_regions_to(self, reguis_to_move, target, move_below=False):
        for regui in reguis_to_move:
            if regui is target: continue
            self.move_region_to(regui, target, move_below)
    # Depricated, remove later and replace uses
    def get_selected_reguis(self):
        return(list(self.regions_selected))
    def reset_view(self):
        # Centre view
        c = self.get_centre()/self.clusters_scrollable.geometry().width()
        self.horizontal_scrollbar.setValue(
            c * self.horizontal_scrollbar.maximum())
        
        # Set header tab width
        allheaders = sorted([region.head.geometry().width() 
                            for region in self.regions_visible])
        try:
            self.headers_slide.setFixedWidth(
                allheaders[round(len(allheaders)*0.9)-1])
        except:
            print("Error resetting view:", 
                  len(allheaders)*0.9-1, 
                  len(allheaders))
    
    # No longer in use, delete at will
    def display_cluster_network(self):
        self.cluster_network.show_clusters(list(self.neidata.cluster_hierarchy.selection))
    # = Whole dataset categorization
    def highlight_markers(self):
        hierarchy = self.neidata.cluster_hierarchy
        accset_names = list(self.dp["creation_settings"]["marker_accsets"])
        accsets = list(self.dp["creation_settings"]["marker_accsets"][name] for name in accset_names)

        # Iterate over all accets to see which clusters to belong to which ones
        #   (with one cluster being able to be part of many)
        cluster_to_accsets = {}
        for i in range(len(accsets)):
            accset = accsets[i]
            for cluster in {hierarchy.member_to_supercluster.get(acc) for acc in accset}:
                if cluster is None: continue
                if cluster not in cluster_to_accsets:
                    # The signature we use is a bit clunkier but should make it
                    #   impossible for an error to occur due to identically named
                    #   markers. I don't think that scenario is particularly likely,
                    #   admittedly.
                    cluster_to_accsets[cluster] = [None] * len(accsets)
                cluster_to_accsets[cluster][i] = True
        
        # Group clusters based on the specific combination of accsets
        marker_groups = {}
        for cluster in cluster_to_accsets:
            signature = tuple(cluster_to_accsets[cluster])
            
            if signature not in marker_groups:
                marker_groups[signature] = []
            
            marker_groups[signature].append(cluster)
        
        # Create the actual groups and name/color them appropriately
        for signature in marker_groups:
            # Figure out what accset_names we have on our hands
            translated_signature = []
            for i in range(len(signature)):
                if signature[i]:
                    translated_signature.append(accset_names[i])
            
            clusters = marker_groups[signature]

            marker_cluster = hierarchy.group(clusters)
            marker_cluster.set_type("user_defined")
            marker_cluster.set_styledict_config({"background_color": qtg.QColor(255,0,0)})
            marker_cluster.set_name(" + ".join(self.neidata.dataset.aliases[accset_name] for accset_name in translated_signature))
            marker_cluster.update_styledict()
        
        return
    def cbk_highlight_rare(self):
        if self.dp["highlight_rare"]:
            for cluster in self.neidata.cluster_hierarchy.clusters_all.values():
                #cluster.unset_autostyle("rare", delay_update=True)
                cluster.unset_autostyle("rare")
            #self.neidata.cluster_hierarchy.sig_styledicts_changed.emit()
        else:
            self.gray_out_rares(0.20)
            self.black_out_singletons()
        self.dp["highlight_rare"] = not self.dp["highlight_rare"]
    def black_out_singletons(self, singleton_style_config=None):
        if singleton_style_config is None:
            singleton_style_config = StyledictCreator.styledict_singletoncluster
        self.gray_out_rares(
            threshold=1, 
            rare_style_config=singleton_style_config,
            percent_threshold=False)
    def gray_out_rares(self, threshold, rare_style_config=None, 
      percent_threshold=True):
        # Init
        regions = [regui.reg for regui in self.regions_all]
        
        if rare_style_config is None:
            rare_style_config = StyledictCreator.styledict_rarecluster
        
        if percent_threshold:
            threshold = len(regions)*threshold
        
        # Get unique counts per region
        unique_cluster_counts = Counter()
        for region in regions:
            represented_clusters = set()
            for feature in region.fts.values():
                if feature.type.upper() != "CDS": continue
                represented_clusters.add(
                    self.neidata.cluster_hierarchy.member_to_supercluster.get(
                        feature.ref.accession))
            for represented_cluster in represented_clusters:
                if represented_cluster is None: continue
                unique_cluster_counts[represented_cluster] += 1
        
        # Evaluate the final count and assign highlights
        i=0
        toplevel_clusters = set(self.neidata.cluster_hierarchy.clusters.values())
        for cluster in unique_cluster_counts:
            if cluster not in toplevel_clusters: continue
            if unique_cluster_counts[cluster] <= threshold:
                i+=1
                #cluster.set_autostyle("rare", rare_style_config, delay_update=True)
                cluster.set_autostyle("rare", rare_style_config)
        #self.neidata.cluster_hierarchy.sig_styledicts_changed.emit()
        
    # = Sorting functions
    def sort_by_alphabet(self):
        self.regions_all.sort(key=lambda x: x.reg.sc.tx.sciname)
        self.update_regui_order()
    def sort_by_taxonomy(self): #TODO:
        # :)
        pass
    def sort_by_jaccard(self): #TODO:
        #threshold = 0.85
        unchecked = sorted(list(self.regions_all), key=lambda x: x.reg.length(), reverse=False)
        checked = []
        current = unchecked[0]
        unchecked.remove(current)
        checked.append(current)
        current_clusters = {
            self.neidata.cluster_hierarchy.member_to_supercluster.get(ft.ref.accession)\
              for ft in current.reg.fts.values()}
        if None in current_clusters:
            current_clusters.remove(None)
        while unchecked:
            best = None
            best_score = -1
            best_clusters = None
            for next in unchecked:
                next_clusters = {
                    self.neidata.cluster_hierarchy.member_to_supercluster.get(ft.ref.accession)\
                      for ft in next.reg.fts.values()}
                if None in next_clusters:
                    next_clusters.remove(None)
                try:
                    score = (len(current_clusters & next_clusters) 
                             / len(current_clusters | next_clusters))
                except ZeroDivisionError:
                    score = 0
                    print(f"Region {next.reg.sc.accession} has no protein groups!")
                if score > best_score:
                    best = next
                    best_score = score
                    best_clusters = next_clusters
                    #if score > threshold:
                    #    break
            if best is None:
                best = unchecked[0]
            
            checked.append(best)
            unchecked.remove(best)
            
            current = best
            current_clusters = best_clusters
        
        #
        self.regions_all = checked
        self.update_regui_order()
    def sort_by_cluster_order(self):
        #graph, regiddict = self.network_regions_by_cluster_order(
        #    [regui,regui.reg for regui in self.regions_all],
        #    0)
        pass
    
    #nonfunctional
    def network_regions_by_cluster_order(self, tosort, cutoff):
        def similarity_index(clustersets, regA, regB):
            def nearest_range(last_match, target_set):
                # An expanding range in both directions.
                #   so   3, 2, 4, 1, 5, 0, 6
                #   until it hits edge of array
                i=1
                max_val = len(target_set)-1
                min_val = 0
                while True:
                    if last_match+i <= max_val:
                        yield(last_match+i, i)
                    if last_match-i >= min_val:
                        yield(last_match-i, i)
                    if last_match+i > max_val and last_match-i < min_val:
                        break
                    
                    i += 1
            gaps = 0
            listA = clustersets[regA]
            listB = clustersets[regB]
            
            last_match = 0
            matched = set()
            n_gaps = 0
            for iA in range(0, len(listA)):
                match = False
                for iB,distance in nearest_range(last_match, listB):
                    #if iB in already_matched: continue
                    if listA[iA] is listB[iB]:
                        matched.add(iB)
                        n_gaps += bool(distance)
                        match = True
                        break
                if not match:
                    n_gaps += 1
            
            return(1 - (n_gaps / len(set(listA) | set(listB))))
                
        graph = nx.Graph()
        
        clustersets = {}
        regids = {}
        i=0
        for reg,regui in tosort:
            tags = sorted([tag.data["identifier"] for tag in regui.tags])
            print(reg.start, reg.stop)
            graph.add_node(
                str(i),
                taxid=reg.sc.tx.taxid, 
                sciname=reg.sc.tx.sciname, 
                contig_accession=reg.sc.accession, 
                region_start=reg.start, 
                region_stop=reg.stop,
                tags=", ".join(tags))
            regids[reg] = str(i)
            clustersets[reg] = [self.neidata.cluster_hierarchy\
                                    .member_to_supercluster.get(ft.ref.accession)
                                    for ft in reg.fts.values()]
            i+=1
        del i
        
        for iA in range(len(tosort)):
            regA,reguiA = tosort[iA]
            for iB in range(iA+1, len(tosort)):
                regB,reguiB = tosort[iB]
                if regA is regB:
                    continue
                similarity=similarity_index(clustersets, regA, regB)
                if similarity < cutoff:
                    continue
                graph.add_edge(regids[regA], regids[regB], 
                               similarity=similarity)
        
        return(graph) #regids)
        
    def network_regions_by_jaccard(self, tosort, cutoff):
        def basic_jaccard_index(clustersets, _regA, _regB):
            return( len(clustersets[_regA] & clustersets[_regB])
                    / len(clustersets[_regA] | clustersets[_regB]) )
        graph = nx.Graph()
        
        clustersets = {}
        regids = {}
        i=0
        for reg,regui in tosort:
            tags = sorted([tag.data["identifier"] for tag in regui.tags])
            print(reg.start, reg.stop)
            graph.add_node(
                str(i),
                taxid=reg.sc.tx.taxid, 
                sciname=reg.sc.tx.sciname, 
                contig_accession=reg.sc.accession, 
                region_start=reg.start, 
                region_stop=reg.stop,
                tags=", ".join(tags))
            regids[reg] = str(i)
            clustersets[reg] = set([self.neidata.cluster_hierarchy\
                                    .member_to_supercluster.get(ft.ref.accession)
                                    for ft in reg.fts.values()])
            i+=1
        del i
        
        # Compare each element with each other element. Jaccard index is
        #   commutative so we only do comparisons we haven't done yet.
        for iA in range(len(tosort)):
            regA,reguiA = tosort[iA]
            for iB in range(iA+1, len(tosort)):
                regB,reguiB = tosort[iB]
                if regA is regB:
                    continue
                similarity=basic_jaccard_index(clustersets, regA, regB)
                if similarity < cutoff:
                    continue
                graph.add_edge(regids[regA], regids[regB], 
                               similarity=similarity)
        
        return(graph)
    def sort_by_list(self, ordered):
        
        # Order what we can by the list, copy the rest.
        new_order = []
        for regui in ordered:
            self.regions_all.remove(regui)
            new_order.append(regui)
        for regui in self.regions_all:
            new_order.append(regui)
        self.regions_all = new_order
        
        self.update_regui_order()
    def update_regui_order(self):
        for regui in self.regions_all:
            self.clusters_scrollable_lay.removeWidget(regui.tail)
            self.headers_scrollable_lay.removeWidget(regui.head)
        for regui in self.regions_all:
            self.clusters_scrollable_lay.addWidget(regui.tail)
            self.headers_scrollable_lay.addWidget(regui.head)
    
    # = Classes
    class DragDropIndicator(qtw.QWidget):
        def __init__(self, header_scrollable):
            super().__init__(header_scrollable)
            self.header_scrollable = header_scrollable
        def paintEvent(self, event):
            painter = qtg.QPainter()
            painter.begin(self)
            #painter.setCompositionMode(qtg.QPainter.CompositionMode_Difference)
            
            painter.setBrush(qtg.QBrush(qtg.QColor(128,128,128,255)))
            
            painter.drawRect(0, 0, self.geometry().width(), 3)
            
            painter.end()
        def jump_to(self, target_header, draw_on_top_edge):
            if draw_on_top_edge:
                self.setGeometry(0, 
                                 max(1, target_header.geometry().top() - 2),
                                 self.header_scrollable.geometry().width(), 
                                 3)
            else:
                self.setGeometry(0, 
                                 max(1, 
                                     target_header.geometry().top()
                                       + target_header.geometry().height()
                                       - 1),
                                 self.header_scrollable.geometry().width(), 
                                 3)
    
    # Overlays
    class ClusterScrollableOverlay(qtw.QWidget):
        def __init__(self, container, target, neiview):
            super().__init__(container)
            self.container = container
            self.target = target
            self.neiview = neiview
            self.setAttribute(qtc.Qt.WA_TransparentForMouseEvents, True)
            self._init_ui()
        def _init_ui(self):
            self.bpselector = self.BPSelector(
                self,
                self.neiview.clusters_scrollable,
                self.neiview.clusters_scroller,
                self.neiview)
            self.target.resizeEvent = self.on_target_resize
            self.on_target_resize()
        def on_target_resize(self, event=None):
            self.setGeometry(self.target.geometry())
            
            self.bpselector.hide_move()
            
            w = self.geometry().width()
            h = self.geometry().height()
        class BPSelector(qtw.QWidget):
            update_delay = 0.025
            initial_move_delay = 0.2
            scroll_speed = 800*update_delay
            
            sig_bpselection_changed = qtc.Signal(bool)
            def __init__(self, parent, target, viewport, neiview):
                super().__init__(parent)
                
                self.mode = "replace"
                # Mode 0: Replace
                # Mode 1: Add
                
                self.target = target
                self.neiview = neiview
                self.viewport = viewport
                
                self.move_triggered = False
                self.moving = False
                self.move_startpos = None
                self.time_move_start = None
                self.time_last_moved = None
                
                neiview.sig_clusters_scrollable_mousepress.connect(
                    self.on_press)
                self.target.mouseMoveEvent = self.on_move
                self.target.mouseReleaseEvent = self.on_release
                
                self.bplabel = qtw.QLabel("")
                self.bpwidget = qtw.QWidget()
                self.bpwidget.setAutoFillBackground(True)
                lay = qtw.QVBoxLayout()
                lay.setContentsMargins(1,1,1,1)
                lay.addWidget(self.bplabel)
                self.bpwidget.setLayout(lay)
                self.bpwidget.setParent(parent)
                self.bpwidget.setHidden(True)
                
                self.setAttribute(qtc.Qt.WA_TransparentForMouseEvents, False)
                self.setHidden(True)
            def on_press(self, event):
                if (event.button() != qtc.Qt.LeftButton
                  or event.modifiers() != NEI_REGION_SELECT_MODIFIER):
                    return(None)
                self.hide_move()
                
                #if self.mode == "replace":
                #    self.neiview.neidata.cluster_hierarchy.clear_selection()
                
                self.move_triggered = True
                self.move_startpos = event.pos()
                self.time_move_start = time.time()
                self.time_last_moved = time.time()
                
            def on_move(self, event):
                if not self.move_triggered or (time.time() - self.time_last_moved) < self.update_delay:
                    return
                self.time_last_moved = time.time()
                if not self.move_visible:
                    if (time.time() - self.time_move_start) < self.initial_move_delay:
                        return
                    else:
                        pass
                        self.show_move()
                
                self.moving = True
                
                startX = min(self.move_startpos.x(), event.x())
                stopX = max(self.move_startpos.x(), event.x())
                startY = min(self.move_startpos.y(), event.y())
                stopY = max(self.move_startpos.y(), event.y())
                
                #self.setGeometry(
                #    startX,
                #    startY,
                #    stopX-startX,
                #    stopY-startY)
                
                self.setGeometry(
                    startX,
                    0,
                    stopX-startX,
                    self.target.geometry().height())
                
                # Width indicator
                bps = (stopX - startX)*self.neiview.dp["bp_per_pixel"]
                self.bpwidget.setGeometry(
                    event.x()-self.bpwidget.sizeHint().width()-2, 
                    event.y()-self.bpwidget.sizeHint().height()-2, 
                    self.bpwidget.sizeHint().width(),
                    self.bpwidget.sizeHint().height())
                self.bplabel.setText(f"<b>{bps} bp</b>")
                
                # Autoscroll
                w = self.viewport.geometry().width()
                h = self.viewport.geometry().height()
                
                pmin = self.viewport.mapToGlobal(qtc.QPoint(w*0.1, h*0.1))
                pmax = self.viewport.mapToGlobal(qtc.QPoint(w*0.9, h*0.9))
                
                if event.globalX() > pmax.x():
                    self.scroll_viewport(1,0)
                if event.globalX() < pmin.x():
                    self.scroll_viewport(-1,0)
                if event.globalY() > pmax.y():
                    self.scroll_viewport(0,1)
                if event.globalY() < pmin.y():
                    self.scroll_viewport(0,-1)
                
                # Let everybody know we moved the bpselection
                self.sig_bpselection_changed.emit(True)
            def scroll_viewport(self, dx, dy):
                if dx != 0:
                    hor = self.viewport.horizontalScrollBar()
                    hor.setValue(hor.value() + dx*hor.singleStep())
                if dy != 0:
                    ver = self.viewport.verticalScrollBar()
                    ver.setValue(ver.value() + dy*ver.singleStep())
            def select_touched_clusters(self):
                clusters = set()
                for tail in self.neiview.clusters_scrollable.children():
                    # Tails have comparable coordinates to BPOverlay due to their
                    #   parents overlapping and having the same size
                    if not tail.geometry().intersects(self.geometry()):
                        continue
                    
                    # The children we have to remap to parent
                    tail_offset = tail.geometry().topLeft()
                    
                    for ftwidget in tail.children():
                        if not isinstance(ftwidget, TranscriptUI):
                            continue
                        if not ftwidget.geometry().translated(tail_offset).intersects(
                                    self.geometry()):
                            continue
                        
                        clusters.add(ftwidget.get_toplevel())
                
                for cluster in clusters:
                    cluster.set_selected(True)
                
            def on_release(self, event):
                event.ignore()
                if self.move_triggered:
                    self.move_triggered = False
                if self.moving:
                    self.moving = False
                    #self.select_touched_clusters()
                #self.hide_move()
            def show_move(self):
                self.move_visible = True
                self.setHidden(False)
                self.bpwidget.setHidden(False)
            def hide_move(self):
                self.move_visible = False
                self.setHidden(True)
                self.bpwidget.setHidden(True)
                
                # Let everybody know bpselection is no longer visible
                self.sig_bpselection_changed.emit(False)
            def paintEvent(self, event):
                painter = qtg.QPainter()
                painter.begin(self)
                
                h = self.geometry().height()-1
                w = self.geometry().width()-1
                
                painter.setPen(
                    qtg.QColor(0,0,0,128))
                
                painter.setBrush(
                    qtg.QBrush(
                        qtg.QColor(128,128,128,32)))
                
                painter.drawRect(qtc.QRectF(0,0,w,h))
                
                painter.end()
            def get_highlighted_subregion(self, regui):
                if self.isHidden():
                    return(None)
                
                pixstart = round(self.geometry().left() - regui.offset)
                pixstop = round(self.geometry().right()+1 - regui.offset)
                
                start = pixstart * self.neiview.dp["bp_per_pixel"]
                stop = pixstop * self.neiview.dp["bp_per_pixel"]
                
                if regui.reversed:
                    start,stop = (
                        regui.reg.length() - stop,
                        regui.reg.length() - start)
                
                newreg = dframe.GeneticRegion(
                    dataset=regui.reg.dataset,
                    scaffold=regui.reg.sc,
                    start=start+regui.reg.start,
                    # I sincerely don't know if this should be a -1 but
                    #   this way the number of bps on the selection corresponds
                    #   to the length of the seq.
                    # I assume it's not an issue because it is a drag select,
                    #   but if this function ever gets used for precision selection,
                    #   this may become na issue.
                    stop=stop+regui.reg.start - 1,
                    seq=regui.reg.seq[start:stop])
                
                return(newreg)
            def copy_as_image(self):
                pixmap = qtg.QPixmap()
                self.neiview.clusters_scrollable.render(
                    pixmap,
                    qtc.QPoint(0,0),
                    qtg.QRegion(self.geometry()))
                
                clipboard = AppRoot.qtapp.clipboard()
                clipboard.clear()
                clipboard.setPixmap(pixmap)
            def contextMenuEvent(self, event):
                return()
                event.accept()
                menu = qtw.QMenu(self)
                menu._open_at = event.globalPos()
                
                #ACTION: Copy as image
                action = menu.addAction("Copy as image")
                action.triggered.connect(self.cbk_customize)
                
                menu.popup(event.globalPos()) 
    class ClusterOverlay(qtw.QWidget):
        scale_rect = (0.87, 0.95)
        def __init__(self, container, target, neiview):
            super().__init__(container)
            self.container = container
            self.target = target
            self.neiview = neiview
            self.setAttribute(qtc.Qt.WA_TransparentForMouseEvents, True)
            self._init_ui()
            
        def _init_ui(self):
            self.pretend_this_is_scale = NeiView.ScaleIndicator(self.neiview, self)
            #self.pretend_this_is_scale = qtw.QPushButton("bla", self)
            
            if self.target.resizeEvent:
                self.target_legacy_resize = self.target.resizeEvent
                print("yep")
            else:
                self.target_legacy_resize = None
            self.target.resizeEvent = self.on_target_resize
            self.on_target_resize()
            
        def on_target_resize(self, event=None):
            self.target_legacy_resize(event)
            
            self.setGeometry(self.target.geometry())
            
            w = self.geometry().width()
            h = self.geometry().height()
            
            scale_width = max(70, (1-self.scale_rect[0]-0.01)*w)
            scale_height = max(35, (1-self.scale_rect[1]-0.01)*h)
            
            self.pretend_this_is_scale.setGeometry(
                w - (scale_width+(0.01*w)),
                h - (scale_height+(0.01*h)),
                scale_width,
                scale_height)
        #def paintEvent(self, event):
        #    painter = qtg.QPainter()
        #    painter.begin(self)
        #    
        #    h = self.geometry().height()-1
        #    w = self.geometry().width()-1
        #    
        #    # Draw background
        #    painter.setBrush(qtg.QBrush(qtg.QColor(255,128,64,48)))
        #    
        #    painter.drawRect(qtc.QRectF(0,0,w,h))
        #    
        #    painter.end()
    class ScaleIndicator(qtw.QWidget):
        #scale_breakpoints = (10, 100, 200, 500, 1000, 2000, 5000, 10000, 20000, 50000, 100000)
        scale_breakpoints = ([i*10 for i in range(1,10)]
                             +[i*100 for i in range(1,10)]
                             +[i*1000 for i in range(1,10)]
                             +[i*10000 for i in range(1,10)]
                             +[i*100000 for i in range(1,10)])
        def __init__(self, neiview, parent=None):
            super().__init__(parent)
            self.neiview = neiview
        def paintEvent(self, event):
            painter = qtg.QPainter()
            painter.begin(self)
            
            h = self.geometry().height()-1
            w = self.geometry().width()-1
            
            
            # Find how big the bar should be
            for breakpoint in self.scale_breakpoints[::-1]:
                if (breakpoint/self.neiview.display_policy["bp_per_pixel"]) <= w*0.96:
                    break
            
            breakpoint_pixels = breakpoint/self.neiview.display_policy["bp_per_pixel"]
            
            # Draw background
            painter.setBrush(qtg.QBrush(qtg.QColor(255,255,255,128)))
            
            painter.drawRect(qtc.QRectF(0,0,w,h))
            
            # Draw bar and text
            painter.setBrush(qtg.QBrush(qtg.QColor(0,0,0,255)))
            
            painter.drawRect(qtc.QRectF(0.5*(w-breakpoint_pixels),
                                        0.6*h,
                                        breakpoint_pixels,
                                        0.2*h))
            painter.drawText(qtc.QRect(0.5*(w-breakpoint_pixels),
                                       0.1*h,
                                       breakpoint_pixels,
                                       0.4*h), 
                             qtc.Qt.AlignCenter,
                             str(breakpoint))
            
            painter.end()
    class Legend(qtw.QWidget, Ui_NeiviewLegend):
        def __init__(self):
            super().__init__()
            self._init_ui()
        def _init_ui(self):
            self.setupUi(self)
            self.setWindowTitle(f"CluSeek {about.VERSION} - Cluster View controls")
            
class HeaderWidget(qtw.QWidget):
    clicked = qtc.Signal()
    def __init__(self, text):
        super().__init__()
        self._init_ui()
        self.label.setText(text)
    def _init_ui(self):
        #self.setFrameShape(qtw.QFrame.Panel)
        #self.setFrameShadow(qtw.QFrame.Raised)
        
        self.lay = qtw.QHBoxLayout()
        self.lay.setContentsMargins(2,0,2,0)
        self.lay.setSpacing(2)
        self.setLayout(self.lay)
        
        self.label = qtw.QLabel()
        self.lay.addWidget(self.label)
class DummyEvent():
    def __init__(self, buttons, modifiers):
        self._buttons = buttons
        self._modifiers = modifiers
    def buttons(self):
        return(self._buttons)
    def modifiers(self):
        return(self._modifiers)
    def accept(self):
        pass
    def ignore(self):
        pass
class RegionUI(qtc.QObject):
    sig_selection_changed = qtc.Signal(object, bool)
    # TODO: no function for setting an offset, calculating best offset 
    def __init__(self, neiview, region):
        super().__init__()
        
        self.neiview = neiview
        self.reg = region
            
        self.reversed = False
        self.selected = False
        self.offset   = 0
        
        self.tags = []
        self.displayed = True
        self.hidden = False
        self._dontdraw = True
        
        self.last_align = None
        
        self._init_ui()
    def _init_ui(self):
        self.head = self.RegionUI_Head(self)
        self.tail = self.RegionUI_Tail(self)
        
    # = Geometry management
    def update_tail_geometry(self):
        self.tail.update_tail_geometry()
    def update_head_geometry(self):
        self.head.update_head_geometry()
    # TODO: set_reversed and reverse have duplicate code that should be merged
    #   reverse should be a contxetual wrapper around set_reversed
    #   For now it works. The align_to function calls set_reversed, so it
    #   not calling align is actually desirable.
    def reverse(self):
        self.reversed = not self.reversed
        self.update_tail_geometry()
        if self.last_align:
            def temp():
                self.align_to(self.last_align, no_flip=True)
            AppRoot.timer.after_delay_do(temp)
    def set_reversed(self, value):
        if self.reversed == value: return
        self.reversed = value
        self.update_tail_geometry()
        self.redraw_clusterwidgets()
    def update_scale(self):
        pass
    def toggle_select(self):
        self.set_selected(not self.selected)
    def set_selected(self, is_selected):
        if self.selected == is_selected: return
        #Cannot selected regions that are not displayed.
        if is_selected and not self.displayed: return
        self.selected = is_selected
        if is_selected:
            assert self not in self.neiview.regions_selected
            self.neiview.regions_selected.append(self)
        else:
            self.neiview.regions_selected.remove(self)
        #self.head.cbk_regui_selection_changed()
        self.update_visibility()
        self.sig_selection_changed.emit(self, self.selected)
    
    # = Visibility
    def set_hidden(self, hide):
        self.hidden = hide
        if hide and self.selected:
            self.set_selected(False)
        self.update_visibility()
    def set_displayed(self, display):
        if display == self.displayed: return
        if not display and self.selected:
            # Regions that are not displayed cannot be selected
            self.set_selected(False)
        
        self.displayed = display
        if display:
            assert self not in self.neiview.regions_displayed
            self.neiview.regions_displayed.append(self)
        else:
            self.neiview.regions_displayed.remove(self)
        
        self.update_visibility()
    def update_visibility(self):
        # If you're changing this logic, check if the header paintevent
        #   logic needs changing as well.
        if self.selected:
            # Show if selected
            self.tail.setHidden(False)
            self.head.setHidden(False)
            self.neiview.regions_visible.add(self)
        elif not self.displayed or self.hidden:
            # Hide if not displayed AND/OR explicitly hidden
            self.tail.setHidden(True)
            self.head.setHidden(True)
            if self in self.neiview.regions_visible:
                self.neiview.regions_visible.remove(self)
        else:
            # Show if displayed and not hidden
            self.tail.setHidden(False)
            self.head.setHidden(False)
            self.neiview.regions_visible.add(self)
    @property
    def dontdraw(self):
        return(self._dontdraw)
    def set_dontdraw(self, dontdraw):
        if dontdraw == self.dontdraw: return
        self._dontdraw = dontdraw
        self.tail._set_dontdraw(dontdraw)
    def set_offset(self, offset):
        self.offset = offset
        self.update_tail_geometry()
    def align_to(self, cluster, centre=None, no_flip=False):
        if not centre:
            centre = self.neiview.get_centre()
        for ftui in self.tail.ftuis:
            # TODO: May need to be updated after refactoring
            if isinstance(ftui, TranscriptUI):
                if cluster in ftui.cluster.get_lineage():
                    if not no_flip:
                        self.set_reversed(ftui.ft.strand == "-")
                    if self.neiview.display_policy["size_mode"] == "proportional":
                        if self.reversed:
                            start = int((ftui.tail.regui.reg.stop-ftui.ft.stop)
                                         / self.neiview.display_policy["bp_per_pixel"])
                        else:
                            start = int((ftui.ft.start-ftui.tail.regui.reg.start)
                                         / self.neiview.display_policy["bp_per_pixel"])
                    elif self.neiview.display_policy["size_mode"] == "fixed":
                        index = sorted(self.tail.ftuis, 
                                       key=lambda x: x.ft.start).index(ftui)
                        if self.reversed:
                            index = len(self.reg.fts)-index-1
                        start = (self.neiview.display_policy["size_fixed_width"]+1)*index
                    self.set_offset(centre-start)
                    self.last_align = cluster
                    return
        self.last_align = None
   #
    def add_tag(self, tag):
        self.tags.append(tag)
        tag.reguis.append(self)
        self.head.add_tagwidget(tag)
    def remove_tag(self, tag):
        if tag not in self.tags:
            raise ValueError("Tag not found in Region UI")
        self.tags.remove(tag)
        tag.reguis.remove(self)
        self.head.remove_tagwidget(tag)
    def redraw_clusterwidgets(self):
        for ftui in self.tail.ftuis:
            ftui.update()
            ftui.update_tooltip()
   #def contextMenuEvent(self, event):
    #    pass
        #TODO: #STARTHERE:
    # = Subordinate classes

    class TagInfoWidget(qtw.QWidget, Ui_TagInfoWidget):
        def __init__(self, neiview):
            self.tag = None
            self.data = None
            self.neiview = neiview
            self.init_ui()
        def init_ui(self):
            super().__init__()
            self.setupUi(self)
            
            def paintevent_text(event):
                painter = qtg.QPainter()
                painter.begin(self.colorpicker_text)
                
                h = self.colorpicker_text.geometry().height()
                w = self.colorpicker_text.geometry().width()-1
                
                painter.setBrush(
                    qtg.QBrush(
                        self.data["color_background"], 
                        bs=qtc.Qt.SolidPattern))
                
                painter.drawRect(0,0,w,h)
                
                # Draw text
                font = qtg.QFont()
                font.setBold(True)
                font.setPixelSize(math.ceil(h*0.5))
                painter.setFont(font)
                
                painter.setPen(self.data["color_text"])
                painter.drawText(qtc.QRectF(0, 0, w, h),
                                 qtc.Qt.AlignCenter,
                                 "Set Text Color")
                
                painter.end()
            def paintevent_background(event):
                painter = qtg.QPainter()
                painter.begin(self.colorpicker_background)
                
                h = self.colorpicker_background.geometry().height()
                w = self.colorpicker_background.geometry().width()-1
                
                painter.setBrush(
                    qtg.QBrush(
                        self.data["color_background"], 
                        bs=qtc.Qt.SolidPattern))
                
                painter.drawRect(0,0,w,h)
                
                # Draw text
                font = qtg.QFont()
                font.setBold(True)
                font.setPixelSize(math.ceil(h*0.5))
                painter.setFont(font)
                
                painter.setPen(self.data["color_text"])
                painter.drawText(qtc.QRectF(0, 0, w, h),
                                 qtc.Qt.AlignCenter,
                                 "Set Background Color")
                
                painter.end()
            
            self.colorpicker_text.paintEvent = paintevent_text
            self.colorpicker_background.paintEvent = paintevent_background
            
            # Btns
            self.apply_btn.clicked.connect(self.cbk_apply)
            self.cancel_btn.clicked.connect(self.cbk_cancel)
            self.delete_btn.clicked.connect(self.cbk_delete)
            
            self.colorpicker_text.mousePressEvent = self.cbk_set_color_text_release
            self.colorpicker_background.mousePressEvent = self.cbk_set_color_background_release
        def refresh_colorwidgets(self):
            self.colorpicker_text.update()
            self.colorpicker_background.update()
        def cbk_set_color_text_release(self, event):
            if event.buttons() != qtc.Qt.LeftButton:
                return
            
            old_color = self.data["color_text"]
            color = qtw.QColorDialog.getColor(initial=old_color)
            if old_color == color or not color.isValid():
                return
            
            self.data["color_text"] = color
            self.refresh_colorwidgets()
        def cbk_set_color_background_release(self, event):
            if event.buttons() != qtc.Qt.LeftButton:
                return
            
            old_color = self.data["color_background"]
            color = qtw.QColorDialog.getColor(initial=old_color)
            if old_color == color or not color.isValid():
                return
            
            self.data["color_background"] = color
            self.refresh_colorwidgets()
        def cbk_apply(self):
            self.update_data()
            valid = self.data_valid()
            if valid is True:
                self.apply_changes()
                self.close()
            elif valid is False:
                self.neiview.remove_tag(self.tag)
                self.close()
            elif valid is None:
                return
        def cbk_delete(self):
            answer = qtw.QMessageBox.question(
                None,
                "Delete tag?",
                "Are you sure you want to delete this tag? "
                "This will remove ALL instances of the tag. "
                "If you wish to remove just one instance, "
                "right click the undesired tag and select 'Untag' instead.",
                qtw.QMessageBox.Yes | qtw.QMessageBox.No,
                qtw.QMessageBox.No)
            if answer is not qtw.QMessageBox.Yes:
                return
            
            self.neiview.remove_tag(self.tag)
            self.close()
        def cbk_cancel(self):
            self.close()
        def load_tag(self, tag):
            self.tag = tag
            self.data = dict(self.tag.data)
            
            self.refresh_colorwidgets()
            self.name_ledit.setText(self.data["name"])
            self.hidden_chk.setCheckState(
                qtc.Qt.Checked if self.data["hidden"] else qtc.Qt.Unchecked)
            self.ident_ledit.setText(self.data["identifier"])
            self.descr_tedit.setText(self.data["description"])
        def closeEvent(self, event):
            if self.tag.not_initialized():
                self.neiview.remove_tag(self.tag)
        def close(self):
            if self.tag.not_initialized():
                self.neiview.remove_tag(self.tag)
            self.setHidden(True)
        def update_data(self):
            self.data["name"] = self.name_ledit.text()
            self.data["identifier"] = self.ident_ledit.text()
            self.data["hidden"] = self.hidden_chk.isChecked()
            self.data["description"] = self.descr_tedit.toHtml()
            # Colors update automatically
        def apply_changes(self):
            self.update_data()
            self.tag.data = self.data
        def data_valid(self):
            # Identifier must be 1) unique, and 2) not be whitespace/empty string
            duplicate_identifier=False
            for other_tag in self.neiview.display_policy["all_tags"].values():
                # If identifier strings match but tag ID's dont
                if other_tag.data["identifier"] == self.tag.data["identifier"]\
                  and other_tag.data["id"] != self.tag.data["id"]:
                    duplicate_identifier=True
                    break
            if not self.data["identifier"].replace(" ","").replace("\t","")\
              or duplicate_identifier:
                btn_return = qtw.QMessageBox.Cancel
                btn_discard = qtw.QMessageBox.Discard
                answer = qtw.QMessageBox.question(
                    None,
                    "Discard tag?",
                    "A valid tag needs to have unique non-empty identifier.",
                    btn_discard | btn_return,
                    btn_discard)
                if answer is btn_return:
                    return None
                elif answer is btn_discard:
                    return(False)
                else:
                    assert False, f"Invalid btn answer, was {answer}, type {type(answer)}"
            else:
                return(True)
    class TagWidget(qtw.QWidget):
        def __init__(self, regui, tag):
            super().__init__()
            
            self.tag = tag
            self.regui = regui
            self.neiview = regui.neiview
            
            self.init_ui()
        def init_ui(self):
            self.lay = qtw.QHBoxLayout()
            self.label = qtw.QLabel("")
            self.label.setAlignment(qtc.Qt.AlignCenter)
            
            self.lay.addWidget(self.label)
            self.setLayout(self.lay)
        def paintEvent(self, event):
            # Begin
            painter = qtg.QPainter()
            painter.begin(self)
            
            h = self.geometry().height()
            w = self.geometry().width()-1
            
            # Prepare text, resize
            
            font = qtg.QFont()
            font.setBold(True)
            font.setPixelSize(math.ceil(h*0.5))
            painter.setFont(font)
            painter.setPen(self.tag.data["color_text"])
            
            brect = painter.boundingRect(qtc.QRectF(2,2, w-2, h-2),
                                 qtc.Qt.AlignCenter,
                                 self.tag.data["identifier"])
            
            self.setFixedWidth(brect.width()+5)
            
            w = self.geometry().width()-1
            
            # Paint background
            painter.setBrush(
                qtg.QBrush(
                    self.tag.data["color_background"], 
                    bs=qtc.Qt.SolidPattern))
            
            painter.drawRect(2,2,w-2,h-2)
            
            # Paint text
            painter.drawText(qtc.QRectF(2, 2, w-2, h-2),
                             qtc.Qt.AlignCenter,
                             self.tag.data["identifier"])
            painter.setPen(self.tag.data["color_text"])
            painter.end()
        def cbk_untag(self):
            self.regui.remove_tag(self.tag)
        def cbk_delete_tag(self):
            self.regui.neiview.taginfowidget.load_tag(self.tag)
            self.regui.neiview.taginfowidget.cbk_delete()
        def mousePressEvent(self, event):
            if event.buttons() == qtc.Qt.LeftButton:
                event.accept()
                self.neiview.open_taginfowidget(self.tag, event.globalX(), event.globalY())
        def contextMenuEvent(self, event):
            # Send this event up the chain
            event.accept()
            menu = qtw.QMenu(self)
            
            #ACTION: Remove tag
            action = menu.addAction("Untag")
            action.triggered.connect(self.cbk_untag)
            
            #ACTION: Delete tag
            action = menu.addAction("Delete Tag")
            action.triggered.connect(self.cbk_delete_tag)
            
            #ACTION: 
            
            # Generate menu at click location
            menu.popup(event.globalPos())
    class Tag():
        default_data = {
            "id": None,
            "name": "",
            "identifier": "",
            "hidden": False,
            "description": "",
            "color_text": qtg.QColor(0,0,0,255),
            "color_background": qtg.QColor(255,255,255,255),
        }
        def __init__(self, id):
            self.data = dict(self.default_data)
            self.data["id"] = id
            self.reguis = []
        def not_initialized(self):
            if self.data["identifier"] == "":
                return(True)
        def update_data(self, data):
            self.data = data
            for regui in self.reguis:
                regui.head.update_tagwidgets()
        @property
        def id(self):
            return(self.data["id"])
        @id.setter
        def id(self, val):
            self.data["id"] = val
        @property
        def name(self):
            return(self.data["name"])
        @name.setter
        def name(self, val):
            self.data["name"] = val
        @property
        def identifier(self):
            return(self.data["identifier"])
        @identifier.setter
        def identifier(self, val):
            self.data["identifier"] = val
        @property
        def hidden(self):
            return(self.data["hidden"])
        @hidden.setter
        def hidden(self, val):
            self.data["hidden"] = val
        @property
        def description(self):
            return(self.data["description"])
        @description.setter
        def description(self, val):
            self.data["description"] = val

    class RegionUI_Head(qtw.QWidget):
        def __init__(self, regionui):
            super().__init__()
            
            self.regui = regionui
            self.tagwidgets = {}
            
            self._init_ui()
        def _init_ui(self):
            
            
            #
            self.wrapper_layout = qtw.QHBoxLayout()
            self.wrapper_layout.setContentsMargins(1,1,1,1)
            self.setLayout(self.wrapper_layout)
            
            # Frame
            self.head_frame = qtw.QFrame()
            self.head_frame.setFrameShape(qtw.QFrame.Panel)
            self.head_frame.setFrameShadow(qtw.QFrame.Raised)
            self.wrapper_layout.addWidget(self.head_frame)
            
            # Layout:
            self.head_layout = qtw.QHBoxLayout()
            self.head_layout.setAlignment(qtc.Qt.AlignLeft)
            self.head_layout.setContentsMargins(0,1,0,1)
            self.head_layout.setSpacing(4)
            
            self.head_frame.setLayout(self.head_layout)
            
            # The notorious buttons themselves:
            #self.select_chk = qtw.QCheckBox()
            self.taxon_btn = HeaderWidget(self.regui.reg.sc.tx.sciname)
            self.taxon_btn.setContextMenuPolicy(qtc.Qt.NoContextMenu)
            
            self.region_btn = HeaderWidget(self.regui.reg.sc.accession)
            self.region_btn.setContextMenuPolicy(qtc.Qt.NoContextMenu)
            
            #self.select_chk.stateChanged.connect(self.cbk_select_chk_toggled)
            
            #self.head_layout.addWidget(self.select_chk)
            self.head_layout.addWidget(self.taxon_btn)
            self.head_layout.addWidget(self.region_btn)
            
            self.setAcceptDrops(True)
            
            # Adjust displayed items based on displaypolicy
            self.set_taxon_hidden(
                "taxon" not in self.regui.neiview.dp["header_elements_shown"])
            self.set_region_hidden(
                "region" not in self.regui.neiview.dp["header_elements_shown"])
            self.set_tags_hidden(
                "tags" not in self.regui.neiview.dp["header_elements_shown"])
            
            self.regui.sig_selection_changed.connect(
                self.cbk_region_selection_changed)
            
            self.update_head_geometry()
        def cbk_info(self):
            self.regui.neiview.frame_info.display_info(
                region=self.regui.reg, 
                sequence=self.regui.reg.sc, 
                taxon=self.regui.reg.sc.tx)
        #def cbk_select_chk_toggled(self):
        #    self.regui.set_selected(self.select_chk.isChecked())
        #def cbk_regui_selection_changed(self):
        #f    self.select_chk.setChecked(self.regui.selected)
        # = Switching things around
        # When this widget is moved
        def mousePressEvent(self, event):
            if event.buttons() == qtc.Qt.LeftButton:
                self.cbk_info()
                if event.modifiers() == NEI_MOVEMENT_MODIFIER:
                    event.accept()
                    
                    drag = qtg.QDrag(self)
                    
                    mime_data = qtc.QMimeData()
                    mime_data.dragged_header = self
                    
                    drag.setMimeData(mime_data)
                    drag.setHotSpot(event.pos() - self.rect().topLeft())
                    
                    drop_action = drag.start(qtc.Qt.MoveAction)
                    
                    #self.regui.tail.setHidden(True)
                    #self.setHidden(True)
                    return
                elif event.modifiers() == NEI_REPLACE_SELECT_AND_DRAG_MODIFIER:
                    event.accept()
                    
                    self.regui.neiview.regions_selection_anchor = self
                    
                    previously_selected = (len(self.regui.neiview.regions_selected)==1 and self.regui in self.regui.neiview.regions_selected)
                    for regui in list(self.regui.neiview.regions_selected):
                        regui.set_selected(False)
                    
                    self.regui.set_selected(not previously_selected)
                    return
                elif event.modifiers() == NEI_REGION_SELECT_MODIFIER:
                    event.accept()
                    
                    for regui in list(self.regui.neiview.regions_selected):
                        regui.set_selected(False)
                    
                    if not self.regui.neiview.regions_selection_anchor:
                        self.regui.neiview.regions_selection_anchor = self
                        self.regui.set_selected(True)
                        return
                    
                    index1 = self.regui.neiview.headers_scrollable_lay.indexOf(
                        self.regui.neiview.regions_selection_anchor)
                    index2 = self.regui.neiview.headers_scrollable_lay.indexOf(
                        self)
                    
                    start = min(index1, index2)
                    stop = max(index1, index2)
                    
                    to_select = []
                    for i in range(start, stop+1, 1):
                        to_select.append(
                            self.regui.neiview.headers_scrollable_lay.itemAt(i).widget())
                    for widget in to_select:
                        if not widget.regui.displayed: continue
                        widget.regui.set_selected(True)
                    return
                elif event.modifiers() == NEI_ADD_SELECT_AND_DRAG_MODIFIER:
                    event.accept()
                    
                    if self.regui.selected:
                        self.regui.set_selected(False)
                    else:
                        self.regui.set_selected(True)
                        self.regui.neiview.regions_selection_anchor = self
                    #self.regui.toggle_select()
                    return
                else:
                    event.ignore()
            else:
                event.ignore()
        def mouseMoveEvent(self, event):
            pass
        #
        def set_taxon_hidden(self, value):
            self.taxon_btn.setHidden(value)
        def set_region_hidden(self, value):
            self.region_btn.setHidden(value)
        def set_tags_hidden(self, value):
            for tagwidget in self.tagwidgets.values():
                tagwidget.setHidden(value)
        def update_head_geometry(self):
            self.setFixedHeight(
                self.regui.neiview.display_policy["size_fixed_height"]+3)
        # Reactions to drag events of other widgets
        def dragEnterEvent(self, event):
            if hasattr(event.mimeData(), "dragged_header"):
                event.accept()
                
                if event.mimeData().dragged_header is self:
                    self.regui.neiview.drag_drop_indicator.setHidden(True)
                    return
                
                moving_down = event.mimeData().dragged_header.geometry().top() < self.geometry().top()
                
                self.regui.neiview.drag_drop_indicator.setHidden(False)
                self.regui.neiview.drag_drop_indicator.jump_to(self, not moving_down)
            else:
                event.ignore()
        def dragLeaveEvent(self, event):
            pass
        def dropEvent(self, event):
            if not hasattr(event.mimeData(), "dragged_header"): return
            
            moved_regui = event.mimeData().dragged_header.regui
            
            if moved_regui is not self.regui:
                self.regui.neiview.move_region_to(moved_regui,self.regui,None)
            else:
                print("Move cancelled: Can't move to self")
            
            # ???
            self.regui.neiview.drag_drop_indicator.setHidden(True)
            event.setDropAction(qtc.Qt.MoveAction)
            event.accept()
        def add_tagwidget(self, tag):
            tagwidget = RegionUI.TagWidget(self.regui, tag)
            self.tagwidgets[tag] = tagwidget
            self.head_layout.addWidget(tagwidget)
        def remove_tagwidget(self, tag):
            self.head_layout.removeWidget(self.tagwidgets[tag])
            self.tagwidgets[tag].setParent(None)
            self.tagwidgets[tag].setHidden(True)
            del self.tagwidgets[tag]
        def update_tagwidgets(self):
            for tagwidget in self.tagwidgets.values():
                tagwidget.update()
        def cbk_region_selection_changed(self, region, selected):
            if selected is True:
                self.head_frame.setFrameShadow(qtw.QFrame.Sunken)
                self.head_frame.setAutoFillBackground(True)
                self.head_frame.setBackgroundRole(qtg.QPalette.Highlight)
                
                self.regui.tail.setAutoFillBackground(False)
                self.regui.tail.setBackgroundRole(qtg.QPalette.Highlight)
            elif selected is False:
                self.head_frame.setFrameShadow(qtw.QFrame.Raised)
                self.head_frame.setBackgroundRole(qtg.QPalette.Window)
                
                self.regui.tail.setAutoFillBackground(False)
                self.regui.tail.setBackgroundRole(qtg.QPalette.Window)
        def cbk_add_tag_new(self):
            tag = self.regui.neiview.get_new_tag()
            self.regui.neiview.open_taginfowidget(tag)
            for regui in self.regui.neiview.regions_selected:
                if tag not in regui.tags:
                    regui.add_tag(tag)
        def cbk_select_all(self):
            for regui in self.regui.neiview.regions_all:
                regui.set_selected(True)
        def cbk_select_visible(self):
            for regui in self.regui.neiview.regions_visible:
                regui.set_selected(True) 
        def cbk_unselect_all(self):
            for regui in self.regui.neiview.get_selected_reguis():
                regui.set_selected(False)
        
        # For following methods, make sure to iterate over a copy,
        #   as the original will be changed during iteration.
        def cbk_select_displayed(self):
            for regui in list(self.regui.neiview.regions_displayed):
                regui.set_selected(True)
        def cbk_unselect_displayed(self):
            for regui in list(self.regui.neiview.regions_displayed):
                regui.set_selected(False)
        def cbk_select_undisplayed(self):
            for regui in set(self.regui.neiview.regions_all).difference(
              self.regui.neiview.regions_displayed):
                regui.set_selected(True)
        def cbk_unselect_undisplayed(self):
            for regui in set(self.regui.neiview.regions_all).difference(
              self.regui.neiview.regions_displayed):
                regui.set_selected(False)
        def cbk_invert_selection(self):
            for regui in list(self.regui.neiview.regions_all):
                regui.toggle_select()
        #
        def cbk_unhide_all(self):
            for regui in self.regui.neiview.regions_all:
                regui.set_displayed(True)
        def cbk_undisplay_selected(self):
            for regui in self.regui.neiview.get_selected_reguis():
                regui.set_displayed(False)
                regui.set_selected(False)
        def cbk_display_selected(self):
            for regui in (set(self.regui.neiview.regions_all)
                          - set(self.regui.neiview.regions_selected)):
                regui.set_displayed(False)
                regui.set_selected(False)
        #
        def get_callback_for_adding_tag(self, tag):
            def cbk():
                for regui in self.regui.neiview.regions_selected:
                    if tag not in regui.tags:
                        regui.add_tag(tag)
            return(cbk)
        def get_callback_for_selecting_by_tag(self, tag):
            def cbk():
                for regui in self.regui.neiview.regions_all:
                    if tag not in regui.tags: continue
                    regui.set_selected(True)
            return(cbk)
        def get_callback_for_unselecting_by_tag(self, tag):
            def cbk():
                for regui in self.regui.neiview.regions_all:
                    if tag not in regui.tags: continue
                    regui.set_selected(False)
            return(cbk)
        def cbk_insert_selected_above(self):
            self.regui.neiview.move_regions_to(
                self.regui.neiview.get_selected_reguis(),
                self.regui,
                False)
        def cbk_insert_selected_below(self):
            self.regui.neiview.move_regions_to(
                self.regui.neiview.get_selected_reguis(),
                self.regui,
                True)
        def cbk_remove_all_tags(self):
            for regui in self.regui.neiview.regions_selected:
                for tag in list(regui.tags):
                    regui.remove_tag(tag)
        def cbk_flip_selected(self):
            for regui in self.regui.neiview.regions_selected:
                print(regui.reg.sc.accession)
                regui.reverse()
        def get_callback_for_removing_tag(self, tag):
            def cbk():
                for regui in self.regui.neiview.regions_selected:
                    if tag in regui.tags:
                        regui.remove_tag(tag)
            return(cbk)
        def contextMenuEvent(self, event):
            # Select cluster if it is right clicked
            #if not self.regui.selected:
            #    fake = DummyEvent(
            #        qtc.Qt.LeftButton,
            #        NEI_REPLACE_SELECT_AND_DRAG_MODIFIER)
            #    self.mousePressEvent(fake)
            event.accept()
            
            menu = qtw.QMenu(self)
            
            # SUBMENU: Tag selected
            submenu = menu.addMenu("Tag selected ...")
            
            action = submenu.addAction("New tag")
            action.triggered.connect(self.cbk_add_tag_new)
            
            submenu.addSeparator()
            
            for tag in self.regui.neiview.display_policy["all_tags"].values():
                action = submenu.addAction(tag.data["identifier"])
                action.setToolTip(tag.data["name"])
                action.triggered.connect(
                    self.get_callback_for_adding_tag(tag))
            
            no_selection = len(self.regui.neiview.get_selected_reguis())==0
            
            # SUBMENU: Untag selected
            submenu = menu.addMenu("Untag selected ...")
            
            action = submenu.addAction("All tags")
            action.triggered.connect(self.cbk_remove_all_tags)
            
            submenu.addSeparator()
            
            for tag in self.regui.tags:
                action = submenu.addAction(tag.data["identifier"])
                action.setToolTip(tag.data["name"])
                action.triggered.connect(
                    self.get_callback_for_removing_tag(tag))
            
            # SEPARATOR
            menu.addSeparator()
            
            # SUBMENU: Select
            submenu = menu.addMenu("Select ...")
            
            # ACTION: Select all
            #ACTION: Select displayed
            action = submenu.addAction("All")
            action.triggered.connect(self.cbk_select_displayed)
            
            action = submenu.addAction("Visible")
            action.triggered.connect(self.cbk_select_visible)
            
            #ACTION: Select non-displayed
            #action = submenu.addAction("Hidden")
            #action.triggered.connect(self.cbk_select_undisplayed)
            
            # SUBSUBMENU: Select by tag
            subsubmenu = submenu.addMenu("By tag")
            
            for tag in self.regui.neiview.display_policy["all_tags"].values():
                action = subsubmenu.addAction(tag.data["identifier"])
                action.setToolTip(tag.data["name"])
                action.triggered.connect(
                    self.get_callback_for_selecting_by_tag(tag))
            
            # SUBMENU: Unselect
            submenu = menu.addMenu("Unselect ...")
            
            # ACTION: Unselect all
            action = submenu.addAction("All")
            action.triggered.connect(self.cbk_unselect_all)
            if no_selection: action.setDisabled(True)
            
            # ACTION: Unselect displayed
            action = submenu.addAction("Displayed")
            action.triggered.connect(self.cbk_unselect_displayed)
            
            # ACTION: Unselect non-displayed
            action = submenu.addAction("Hidden")
            action.triggered.connect(self.cbk_unselect_undisplayed)
            
            # SUBSUBMENU: Unselect by tag
            subsubmenu = submenu.addMenu("By tag")
            
            for tag in self.regui.neiview.display_policy["all_tags"].values():
                action = subsubmenu.addAction(tag.data["identifier"])
                action.setToolTip(tag.data["name"])
                action.triggered.connect(
                    self.get_callback_for_unselecting_by_tag(tag))
            
            # ACTION: Invert selection
            action = menu.addAction("Invert selection")
            action.triggered.connect(self.cbk_invert_selection)
            
            # ACTION: Move selected above
            action = menu.addAction("Insert selected above")
            action.triggered.connect(self.cbk_insert_selected_above)
            if no_selection: action.setDisabled(True)
            
            # ACTION: Move selected below
            action = menu.addAction("Insert selected below")
            action.triggered.connect(self.cbk_insert_selected_below)
            if no_selection: action.setDisabled(True)
            
            # SEPARATOR
            menu.addSeparator()
            
            # ACTION: Flip selected
            action = menu.addAction("Flip selected")
            action.triggered.connect(self.cbk_flip_selected)
            if no_selection: action.setDisabled(True)
            
            # SEPARATOR
            menu.addSeparator()
            
            # ACTION: Hide selected
            action = menu.addAction("Hide selected")
            action.triggered.connect(self.cbk_undisplay_selected)
            if no_selection: action.setDisabled(True)
            
            # ACTION: Show only selected
            action = menu.addAction("Show only selected")
            action.triggered.connect(self.cbk_display_selected)
            if no_selection: action.setDisabled(True)
            
            # ACTION: Show all hidden
            action = menu.addAction("Show all hidden")
            action.triggered.connect(self.cbk_unhide_all)
            
            # Generate menu at click location
            menu.popup(event.globalPos())
            
    class RegionUI_Tail(qtw.QWidget):
        #TODO: Manual offset adjustment
        def __init__(self, regionui):
            super().__init__()
            
            self.regui = regionui
            
            self.ftuis = []
            
            self._dragging = False
            
            self.init_ui()
        def init_ui(self):
            for ft in self.regui.reg.fts.values():
                try:
                    cluster = self.regui.neiview.neidata.\
                                cluster_hierarchy.member_to_subcluster[ft.ref.accession]
                    self.ftuis.append(
                        TranscriptUI(
                            self, 
                            ft, 
                            cluster, 
                            tail=self, 
                            represents="toplevel"))
                except:
                    self.ftuis.append(FeatureUI(self, ft, text="N/A"))
            self.update_tail_geometry()
        def _set_dontdraw(self, dontdraw):
            pass
        def update_tail_geometry(self):
            if self.regui.neiview.display_policy["size_mode"] == "proportional":
                self._update_tail_geometry_proportional()
            elif self.regui.neiview.display_policy["size_mode"] == "fixed":
                self._update_tail_geometry_fixed()
            else:
                assert False, ("Invalid display_policy[size_mode]: "
                               f"{self.regui.neiview.display_policy['size_mode']}")
        def _update_tail_geometry_fixed(self):
            ftuis = sorted(self.ftuis, 
                           key=lambda x: x.ft.start, 
                           reverse=self.regui.reversed)
            for i in range(len(ftuis)):
                ftuis[i].setGeometry((self.regui.neiview.display_policy["size_fixed_width"]+1)*i\
                                        +self.regui.offset,
                                     1,
                                     self.regui.neiview.display_policy["size_fixed_width"],
                                     self.regui.neiview.display_policy["size_fixed_height"])
                ftuis[i].set_reversed(self.regui.reversed)
            
            self.setFixedSize(
                len(ftuis)\
                  *(self.regui.neiview.display_policy["size_fixed_width"]+1)\
                  + self.regui.offset,
                self.regui.neiview.display_policy["size_fixed_height"]+2)
            # TODO: set min max sizes
            #       note: not sure what this means
        def _update_tail_geometry_proportional(self):
            bp_per_pixel = self.regui.neiview.display_policy["bp_per_pixel"]
            region_length = self.regui.reg.length() / bp_per_pixel
            for ftui in self.ftuis:
                # TODO: Implement vertical offset.
                start = int((ftui.ft.start-self.regui.reg.start) / bp_per_pixel)
                stop = int((ftui.ft.stop-self.regui.reg.start) / bp_per_pixel)
                length = stop-start
                if self.regui.reversed:
                    start = region_length - stop
                    ftui.set_reversed(True)
                else:
                    ftui.set_reversed(False)
                ftui.setGeometry(start+self.regui.offset, 1, length, 
                                 self.regui.neiview.display_policy["size_fixed_height"])
            self.setFixedSize(
                region_length+self.regui.offset,
                self.regui.neiview.display_policy["size_fixed_height"]+2)
            
        
        # Dragging functionality
        def mousePressEvent(self, event):
            if event.button() == qtc.Qt.LeftButton and event.modifiers() == NEI_MOVEMENT_MODIFIER:
                self._move_origin = event.globalPos().x()
                self._move_origin_time = time.time()
                self._move_origin_first_click_time = time.time()
                self._dragging = True
                event.accept()
            else:
                event.ignore()
        def mouseDoubleClickEvent(self, event):
            if event.modifiers() == NEI_MOVEMENT_MODIFIER:
                self.regui.reverse()
                event.accept()
            else:
                event.ignore()
        def mouseReleaseEvent(self, event):
            if self._dragging:
                self._dragging = False
            event.ignore()
        def mouseMoveEvent(self, event):
            if self._dragging\
              and event.buttons() == qtc.Qt.LeftButton\
              and event.modifiers() == NEI_MOVEMENT_MODIFIER:
                #if ((time.time() - self._move_origin_time) < 0.1) or ((time.time() - self._move_origin_first_click_time) < 0.6):
                if (time.time() - self._move_origin_time) < 0.1:
                    return
                curr_pos = event.globalPos().x()
                difference = curr_pos - self._move_origin
                self.regui.set_offset(max(self.regui.offset+difference, 0))
                self._move_origin = curr_pos
                event.accept()
            elif self._dragging:
                self._dragging = False
                event.ignore()
            event.ignore()
        # Events
        def paintEvent(self, event):
            if self.regui.neiview.display_policy["size_mode"] == "fixed": return
            
            # Region length is longer than the actual length of the sequence
            #   we want to show only the length of the available sequence
            bp_per_pixel = self.regui.neiview.display_policy["bp_per_pixel"]
            region_length = self.regui.reg.length() / bp_per_pixel
            seq_length = len(self.regui.reg.seq) / bp_per_pixel
            
            if not self.regui.reversed:
                sequence = (self.regui.offset, self.regui.offset+seq_length)
            else:
                sequence = (self.regui.offset+region_length-seq_length, 
                                 self.regui.offset+region_length)
                
            painter = qtg.QPainter()
            painter.begin(self)
            painter.setBrush(qtg.QBrush(qtg.QColor(148,148,148,255)))
            painter.setPen(qtg.QColor(148,148,148,255))
            
            
            h = self.geometry().height()
            w = self.geometry().width()
            
            painter.drawLine(
                sequence[0], 
                h/2, 
                sequence[1],
                h/2)
            
            painter.end()
            
        #def paintEvent(self, event):
        #    painter = qtg.QPainter()
        #    painter.begin(self)
        #    
        #    h = self.geometry().height()
        #    w = self.geometry().width()
        #    
        #    painter.setBrush(qtg.QBrush(qtg.QColor(255,225,225,0)))
        #    painter.setPen(qtg.QColor(0,0,0,0))
        #    
        #    painter.drawRect(qtc.QRectF(0,0,w,h))
        #    
        #    painter.end()
class FeatureUI(qtw.QWidget):
    # This is primarily a graphical object.
    #   It draws an arrow within its limits using the brushes provided.
    #   This arrow may be facing left or right (reversed).
    
    # TODO: Implement an infowindow cascade to show information about
    #       the feature, as well as any possible subclasses.
    # TODO: Consider posibility of a FeatureUI without a regui
    sig_removeme = qtc.Signal(object)
    
    max_head_size = 14
    def __init__(self, parent, feature, neiview=None, tail=None, text="", styledict=None):
        super().__init__(parent)
        
        self.ft = feature
        self.styledict = styledict if styledict else StyledictCreator.default_styledict
        self.text = text
        self.tail = tail
        self._removable = False
        self.neiview = neiview
        
        self._image = None
        # This is not the same as strand
        self.reversed = False
        
        self.last_redraw_id = -1
        self.last_neiredraw_id = -1
        self.last_redraw_reversed = self.reversed
        self.resized = False
    
    @property
    def has_custom_text(self):
        return(True)
    
    def set_reversed(self, value):
        if self.ft.strand == "+":
            self.reversed = value
        elif self.ft.strand == "-":
            self.reversed = not value
        #TODO: Redraw, if it doesn't automatically.
    def set_styledict(self, styledict):
        self.styledict = styledict
    def reverse(self):
        # Convenience method
        self.set_reversed(not self.reversed)
    
    def update_tooltip(self):
        self.setToolTip(self.text)
    # = = = Rendering infrastructure
    # make_shape_arrow
    @staticmethod
    def make_shape_arrow_notched(w, h, reversed):
        # Refpoints
        head = min(w, math.floor(h/2))+1
        tail = max(w-head+1, 1)
        butt = min(w/4, min(tail-1, math.floor(h/4)))
        
        h0   = 0
        h25  = math.floor(h*0.25)
        h49  = math.floor(h/2)
        h51  = math.ceil(h/2)
        h75  = h  - h25
        h100 = h
        
        if w < 5:
            h49 = h25
            h51 = h75
        
        # Create Polygon
        polygon = qtg.QPolygonF()
        if reversed:
            if tail>0:
                points = [(w,      h25),
                          (w-butt, h49),
                          (w-butt, h51),
                          (w,      h75),
                          (head,   h75),
                          (head+1,   h100+1),
                          (1,      h51),
                          (1,      h49),
                          (head,   h0),
                          (head,   h25)]
            else:
                points = [(head,   h75),
                          (head+1,   h100+1),
                          (1,      h51),
                          (1,      h49),
                          (head,   h0),
                          (head,   h25)]
                
        else:
            if tail>0:
                points = [(1,    h25),
                          (tail, h25),
                          (tail, h0),
                          #(w,    h/2),
                          (w,    h49),
                          (w,    h51),
                          (tail, h100),
                          (tail, h75),
                          (1,    h75),
                          (butt, h51),
                          (butt, h49)]
            else:
                points = [(tail, h25),
                          (tail, h0),
                          #(w,    h/2),
                          (w,    h49),
                          (w,    h51),
                          (tail, h100),
                          (tail, h75)]
        
        points = [(int(x[0]), int(x[1])) for x in points]
        
        [polygon.append(qtc.QPointF(*point)) for point in points]
        
        return(polygon)
    @staticmethod
    def make_shape_arrow_standard(w, h, reversed):
        # Refpoints
        head = min(w, math.floor(h/2))+1
        tail = max(w-head+1, 1)
        butt = min(w/4, min(tail-1, math.floor(h/4)))
        
        h0   = 0
        h25  = math.floor(h*0.25)
        h49  = math.floor(h/2)
        h51  = math.ceil(h/2)
        h75  = h  - h25
        h100 = h
        
        if w < 5:
            h49 = h25
            h51 = h75
        
        # Create Polygon
        polygon = qtg.QPolygonF()
        if reversed:
            if tail>0:
                points = [(w,      h25),
                          #(w-butt, h49), # notch
                          #(w-butt, h51), # notch
                          (w,      h75),
                          (head,   h75),
                          (head+1,   h100+1),
                          #(0,      h/2),
                          (1,      h51),
                          (1,      h49),
                          (head,   h0),
                          (head,   h25)]
            else:
                points = [
                          #(head,   h75),
                          (head+1,   h100+1),
                          #(0,      h/2),
                          (1,      h51),
                          (1,      h49),
                          (head,   h0),
                          #(head,   h25)
                         ]
                
        else:
            if tail>0:
                points = [(1,    h25),
                          (tail, h25),
                          (tail, h0),
                          #(w,    h/2),
                          (w,    h49),
                          (w,    h51),
                          (tail, h100),
                          (tail, h75),
                          (1,    h75),
                          #(butt, h51),
                          #(butt, h49)
                         ]
            else:
                points = [
                          #(tail, h25),
                          (tail, h0),
                          #(w,    h/2),
                          (w,    h49),
                          (w,    h51),
                          (tail, h100),
                          #(tail, h75)
                         ]
        
        points = [(int(x[0]), int(x[1])) for x in points]
        
        [polygon.append(qtc.QPointF(*point)) for point in points]
        
        return(polygon)
    @staticmethod
    def make_shape_arrow_compact(w, h, reversed):
        # Refpoints
        head = min(w, math.floor(h/2))+1
        tail = max(w-head+1, 1)
        butt = min(w/4, min(tail-1, math.floor(h/4)))
        
        h0   = 0
        h25  = math.floor(h*0.25)
        h49  = math.floor(h/2)
        h51  = math.ceil(h/2)
        h75  = h  - h25
        h100 = h
        
        if w < 5:
            h49 = h25
            h51 = h75
        
        # Create Polygon
        polygon = qtg.QPolygonF()
        if reversed:
            if tail>0:
                points = [(w,      h0), #(top right)
                          (head,   h0),
                          (1,      h49),
                          (1,      h51),
                          (head,   h100),
                          (w,      h100)]
            else:
                points = [(head,   h0),
                          (1,      h49),
                          (1,      h51),
                          (head,   h100)]
                
        else:
            if tail>0:
                points = [(1,    h0), #(top left)
                          (tail, h0),
                          (w,    h49),
                          (w,    h51),
                          (tail, h100),
                          (1,    h100)]
            else:
                points = [(tail, h0),
                          (w,    h49),
                          (w,    h51),
                          (tail, h100)]
        
        points = [(int(x[0]), int(x[1])) for x in points]
        
        [polygon.append(qtc.QPointF(*point)) for point in points]
        
        return(polygon)
    @staticmethod
    def make_shape_box(w, h, reversed):
        polygon = qtg.QPolygonF()
        
        points = [(1,1),
                  (w,1),
                  (w,h-1),
                  (1,h-1)]
        
        points = [(int(x[0]), int(x[1])) for x in points]
        
        [polygon.append(qtc.QPointF(*point)) for point in points]
        
        return(polygon)
    
    make_shape = make_shape_arrow_standard
    
    def resizeEvent(self, event):
        self.resized = True
    # Default shape:
    def paintEvent(self, event):
        # I apologise for this conditional monster.
        if (self.resized
          or self.last_redraw_id != self.styledict["redraw_id"]
          or ((self.last_neiredraw_id != self.tail.regui.neiview.dp["neiredraw_id"])
               if self.tail is not None else False)
          or self.last_redraw_reversed != self.reversed
          or not self._image):
            self._image = self.redraw_image()
            self.last_redraw_id = self.styledict["redraw_id"]
            if self.tail:
                self.last_neiredraw_id = self.tail.regui.neiview.dp["neiredraw_id"]
            self.last_redraw_reversed = self.reversed
            if self.resized: self.resized = False
        painter = qtg.QPainter()
        painter.begin(self)
        painter.drawPixmap(qtc.QPoint(0,0), self._image)
        painter.end()
    
    def redraw_image(self, into_picture=False, maxwidth=float("inf"), start=0):
        h = self.geometry().height()-1
        w = self.geometry().width()-1
        
        # We need this for rendering features into svg. Pain.
        if not into_picture:
            image = qtg.QPixmap(w+1,h+1)
            image.fill(qtc.Qt.transparent)
        else:
            image = qtg.QPicture()
        
        painter = qtg.QPainter()
        painter.begin(image)
        
        polygon = self.make_shape(w,h,self.reversed)
        
        limiter = qtc.QRect(start, 0, min(maxwidth, w), h)
        polygon = polygon.intersected(qtg.QPolygon(limiter))
        
        # Set line
        painter.setPen(StyledictCreator.get_pen(self.styledict))
        
        # Draw background
        #painter.setPen(self.styledict["foreground_pen"])
        painter.setBrush(StyledictCreator.get_background_brush(self.styledict))
        painter.drawPolygon(polygon)
        
        # Draw pattern
        foreground_brush = StyledictCreator.get_foreground_brush(self.styledict)
        if foreground_brush:
            painter.setBrush(foreground_brush)
            painter.drawPolygon(polygon)
        
        if (self.tail is None
          or self.tail.regui.neiview.dp["show_labels"] == "all"
          or (self.tail.regui.neiview.dp["show_labels"] == "custom" 
              and self.has_custom_text)):
            # Draw text
            font = StyledictCreator.get_font(self.styledict)
            if self.get_neiview().dp["arrow_shape"] == "compact":
                font.setPixelSize(math.ceil(h-2))
            else:
                font.setPixelSize(math.ceil(h/2))
            painter.setFont(font)
            
            #painter.setFont(self.styledict["font"])
            #painter.setPen(qtc.Qt.white)
            #painter.drawText(qtc.QRectF(1, 0.25*h+1, w+1, 0.75*h+1), 
            #                 qtc.Qt.AlignCenter,
            #                 self.text)
            painter.setPen(self.styledict["text_color"])
            painter.drawText(qtc.QRectF(start, 0, min(maxwidth, w), h),
                             qtc.Qt.AlignCenter,
                             self.text)
        
        
        # Conclude
        painter.end()
        return(image)
    
    def set_removable(self, removable):
        self._removable = removable
    # = = = UI Interactions (mouse events mainly)
    def cbk_infoclick(self, invert_cluster=False):
        # This needs to be done after a delay because events initiated by this
        #   function can result in the destruction of the FeatureUI they were
        #   initiated from before said function concludes.
        # Therefore we have the timer do our dirty work instead.
        def infoclick_after_delay():
            if self.tail:
                self.tail.regui.neiview.frame_info.display_quick(self, invert_cluster)
            elif self.neiview:
                self.neiview.frame_info.display_quick(self, invert_cluster)
        AppRoot.timer.after_delay_do(infoclick_after_delay, 1)
    
    def get_neiview(self):
        if self.tail:
            return(self.tail.regui.neiview)
        elif self.neiview:
            return(self.neiview)
    
    def _removeme(self):
        self.sig_removeme.emit(self)
class TranscriptUI(FeatureUI):
    cluster_clicked = qtc.Signal(object)
    sig_reparented = qtc.Signal(object)
    sig_floated = qtc.Signal(object)
    def __init__(self, parent, feature, cluster, tail, represents=None, neiview=None):
        super().__init__(parent, feature, tail=tail, neiview=neiview)
        
        self.cluster = cluster
        self.cluster.add_clusterui(self)
        self._last_press = 0
        self._to_unselect = None
        
        self._text_override = None
        self._release_timer_id=None
        
        if represents is None:
            raise Exception
        
        # represents may be "toplevel" or "bottomlevel"
        self.represents = represents
        
        self.__init_ui()
    def __init_ui(self):
        self.update_tooltip()
        self.setAcceptDrops(True)
    @property
    def styledict(self):
        return(self.cluster.styledict)
    @property
    def text(self):
        if self._text_override:
            return(self._text_override)
        else:
            return(self.get_toplevel().get_name())
    def set_text_override(self, text):
        self._text_override = text
    
    def set_represents(self, value):
        if value != self.represents:
            self.represents = value
            self._image = self.redraw_image()
            self.update()
    
    @property
    def make_shape(self):
        shapes = {"notched": self.make_shape_arrow_notched,
                  "standard": self.make_shape_arrow_standard,
                  "compact": self.make_shape_arrow_compact}
        return(shapes[self.get_neiview().dp["arrow_shape"]])
    
    @property
    def has_custom_text(self):
        if self.get_toplevel().name:
            #print("transcript: has custom text")
            return(True)
        else:
            #print("transcript: has no custom text")
            return(False)
    #TODO: We shouldn't need these useless setters.
    #      Figure out and fix the inheritance here, maybe via a virtual superclass.
    @styledict.setter
    def styledict(self, value):
        #print("Attempted to set styledict for TranscriptUI, ignored.")
        return
    @text.setter
    def text(self, value):
        return
    def update_tooltip(self):
        self.setToolTip(self.text)
    #
    def mouseReleaseEvent(self, event):
        if self._release_timer_id:
            AppRoot.timer.kill_timer(self._release_timer_id)
            self._release_timer_id = None
        if event.button() == qtc.Qt.LeftButton:
            if event.modifiers() == NEI_REGION_SELECT_MODIFIER:
                event.ignore()
                return(None)
            if (event.modifiers() == NEI_REPLACE_SELECT_AND_DRAG_MODIFIER
              and (time.time() - self._last_press) < NEI_HOLD_ACTIVATE_DELAY):
                event.accept()
                
                #print("Release Replace")
                #for cluster in self._to_unselect:
                #    cluster.set_selected(False)
                #self._to_unselect = None
                return(None)
                
            elif (event.modifiers() == NEI_ADD_SELECT_AND_DRAG_MODIFIER
              and (time.time() - self._last_press) < NEI_HOLD_ACTIVATE_DELAY):
                event.accept()
                
                if self._to_unselect:
                    for cluster in self._to_unselect:
                        cluster.set_selected(False)
                    self._to_unselect = None
                return(None)
        event.ignore()
        return(None)
    
    def mousePressEvent(self, event):
        self.mouse_press(event, doubleclick=False)
    def mouse_press(self, event, doubleclick=False):
        if doubleclick:
            print("Double click!")
        pos_ = event.pos()
        self._last_press = time.time()
        
        mycluster = self.get_toplevel() if not doubleclick else self.get_not_represents()
        
        if event.modifiers() == NEI_MOVEMENT_MODIFIER:
            event.ignore()
            return
        if event.button() == qtc.Qt.LeftButton:
            if not event.modifiers():
                self.cbk_infoclick()
                self.cluster_clicked.emit(mycluster)
            pass
            
            if event.modifiers() == NEI_REPLACE_SELECT_AND_DRAG_MODIFIER:
                event.accept()
                
                #self._to_unselect = set(mycluster().hierarchy.selection)
                
                #def select_and_drag():
                mycluster.hierarchy.clear_selection()
                mycluster.set_selected(True)
                
                on_fail = "unselect"
                self._release_timer_id = AppRoot.timer.after_delay_do(
                    lambda: self._begin_drag_if_not_released(pos_, on_fail), 
                    NEI_HOLD_ACTIVATE_DELAY)
            elif event.modifiers() == NEI_ADD_SELECT_AND_DRAG_MODIFIER:
                event.accept()
                
                if mycluster.get_selected():
                    self._to_unselect = {mycluster}
                    on_fail = "unselect"
                else:
                    self._to_unselect = {}
                    mycluster.set_selected(True)
                    on_fail = None
                
                self._release_timer_id = AppRoot.timer.after_delay_do(
                    lambda: self._begin_drag_if_not_released(pos_, on_fail), 
                    NEI_HOLD_ACTIVATE_DELAY)
            elif event.modifiers() == NEI_MOVEMENT_MODIFIER:
                event.ignore()
            elif event.modifiers() == NEI_REGION_SELECT_MODIFIER:
                event.ignore()
            else:
                event.ignore()
        # Asks parent to remove it.
        #elif (self._removable 
        #  and event.button() == qtc.Qt.RightButton
        #  and event.modifiers() == NEI_REPLACE_SELECT_AND_DRAG_MODIFIER):
        #    event.accept()
        #    
        #    AppRoot.timer.after_delay_do(self._removeme)
        else:
            event.ignore()
        # Old code:
        # Need to ignore the event so we can catch drag events
        #   on parent widgets
        #self._timer_id = AppRoot.timer.after_delay_do(
        #    self.cbk_set_color, 250)
        pass
    def _begin_drag_if_not_released(self, pos, on_fail):
        self._begin_drag(pos, on_fail)
        self._release_timer_id = None
    def dropEvent(self, event):
        if hasattr(event.mimeData(), "dragged_clusterwidget"):
            if event.mimeData().on_fail == "unselect":
                event.mimeData().dragged_clusterwidget.setSelected(False)
    def _begin_drag(self, position, on_fail):
        drag = qtg.QDrag(self.get_neiview())
        
        mime_data = qtc.QMimeData()
        mime_data.dragged_clusterwidget = self
        mime_data.on_fail = on_fail
        
        drag.setMimeData(mime_data)
        drag.setHotSpot(position - self.rect().topLeft())
        
        drop_action = drag.start(qtc.Qt.MoveAction)
    # Debug:
    def cbk_customize(self):
        self.get_neiview().styledictcreator.load(self.get_toplevel())
        self.get_neiview().styledictcreator.show()
        self.get_neiview().styledictcreator.activateWindow()
    def cbk_toggle_select(self):
        self.get_toplevel().toggle_selected()
    def cbk_align_to(self):
        # TODO: Not kosher, needs rework to explicitly reference neiview
        self.tail.regui.neiview.align_regions_to_cluster(self.get_toplevel())
    def cbk_set_color(self):
        self.get_neiview().styledictcreator.quick_set_color(self.get_toplevel())
    def cbk_apply_last(self):
        self.get_neiview().styledictcreator.apply_last_style(self.get_toplevel())
    
    def cbk_group_selected(self):
        self.get_neiview().neidata.cluster_hierarchy.group_selected_clusters()
    def cbk_ungroup_selected(self):
        self.get_neiview().neidata.cluster_hierarchy.ungroup_selected_clusters()
    def cbk_select_subgroup(self):
        self.get_not_represents().hierarchy.clear_selection()
        self.get_not_represents().set_selected(True)
    def cbk_clear_customizations(self):
        answer = qtw.QMessageBox.question(None,
            "Are you sure?",
            "This will clear customizations from ALL selected groups"
            " and their subgroups.",
            buttons=qtw.QMessageBox.Yes | qtw.QMessageBox.No)
        if answer == qtw.QMessageBox.Yes:
            for selected_cluster in self.cluster.hierarchy.selection:
                selected_cluster.clear_all_styledicts()
    def cbk_removeme(self):
        AppRoot.timer.after_delay_do(self._removeme)
    def contextMenuEvent(self, event):
        if not (set(self.get_toplevel().get_lineage()) 
                & set(self.get_toplevel().hierarchy.selection)):
            self.get_toplevel().hierarchy.clear_selection()
            self.get_toplevel().set_selected(True)
            self.cbk_infoclick()
        #if event.modifiers() == :
        #    print(dir(event))
        #    event.ignore()
        #    return
        # Generate menu at click location
        event.accept()
        menu = qtw.QMenu(self)
        menu._open_at = event.globalPos()
        
        #ACTION: Select subcluster
        action = menu.addAction("Select only subgroup")
        action.triggered.connect(self.cbk_select_subgroup)
        
        #SEPARATOR
        action = menu.addSeparator()
        
        #ACTION: Align to
        if self.tail:
            action = menu.addAction("Align to")
            action.triggered.connect(self.cbk_align_to)
        
        #ACTION: Flip
        #if self.tail:
        #    action = menu.addAction("Flip sequence")
        #    action.triggered.connect(self.tail.regui.reverse)
        
        #SEPARATOR
        action = menu.addSeparator()
        
        #ACTION: Quick color
        action = menu.addAction("Set color")
        action.triggered.connect(self.cbk_set_color)
        
        #ACTION: Apply last
        action = menu.addAction("Apply last used customization")
        action.triggered.connect(self.cbk_apply_last)
        
        #ACTION: Set Selected
        #if self.get_toplevel().get_selected():
        #    action = menu.addAction("Unselect protein group")
        #else:
        #    action = menu.addAction("Select protein group")
        #action.triggered.connect(self.cbk_toggle_select)
        
        #SEPARATOR
        action = menu.addSeparator()
        
        action = menu.addAction("Group selected")
        action.triggered.connect(self.cbk_group_selected)
        
        action = menu.addAction("Ungroup selected")
        action.triggered.connect(self.cbk_ungroup_selected)
        
        
        #SEPARATOR
        action = menu.addSeparator()
        
        #ACTION: Customize
        action = menu.addAction("Customize")
        action.triggered.connect(self.cbk_customize)
        
        #ACTION: Clear customizations
        action = menu.addAction("Clear customizations")
        action.triggered.connect(self.cbk_clear_customizations)
        
        #ACTION: Info
        #action = menu.addAction("Information")
        #action.triggered.connect(self.cbk_infoclick)
        
        if self._removable:
            # SEP
            action = menu.addSeparator()
            
            # Remove from clustercontainerwidget
            action = menu.addAction("Remove selected from container")
            action.triggered.connect(self.cbk_removeme)
        
        # FINISH
        menu.popup(event.globalPos()) 
    
    # Change to get_represents()
    def get_toplevel(self):
        if self.represents == "toplevel":
            return(self.cluster.get_toplevel())
        elif self.represents == "bottomlevel":
            return(self.cluster)
    def get_not_represents(self):
        if self.represents == "toplevel":
            return(self.cluster)
        elif self.represents == "bottomlevel":
            return(self.cluster.get_toplevel())
class ClusterWidget(TranscriptUI):
        make_shape = staticmethod(FeatureUI.make_shape_box)
        def __init__(self, parent, cluster, neiview=None, represents="bottomlevel"):
            super().__init__(parent=parent, feature=None, cluster=cluster, tail=None,
                             represents=represents, neiview=neiview)
        @property
        def has_custom_text(self):
            return(True)

class ProteinGroupTable(qtw.QWidget, Ui_ProteinGroupTable):
    # TODO: Maybe do everything with QTableView.setRowHidden(row, hide)
    def __init__(self, neiview):
        super().__init__()
        
        self.hierarchies = neiview.neidata.cluster_hierarchy
        self.neiview = neiview
        self.neiview.sig_neiview_loaded.connect(
            self.cbk_on_neiview_loaded)
        
        self._cluster_refresh_needed = False
        self._cluster_refresh_toadd = set()
        self._cluster_refresh_toremove = set()
        
        self._column_widths_set = False
        
        self.init_ui()
    def init_ui(self):
        self.setupUi(self)
        print("ProteinGroupTable UI set up")
        
        # Callbacks
        self.displayed_groups_combo.currentIndexChanged.connect(
            self.cbk_on_displayed_groups_changed)
        self.cluster_table.itemClicked.connect(self.cbk_item_changed)
        self.search_category_combo.currentIndexChanged.connect(
            self.cbk_searchbar_changed)
        self.search_bar.textChanged.connect(self.cbk_searchbar_changed)
        
        self.copy_action = qtw.QAction("Copy selected")
        self.copy_action.triggered.connect(
            self.copy_highlighted_into_clipboard)
        self.copy_action.setShortcut(qtg.QKeySequence.Copy)
        self.copy_action.setShortcutContext(qtc.Qt.WidgetWithChildrenShortcut)
        self.addAction(self.copy_action)
        self.info_lbl.setHidden(True)
        self.info_btn.clicked.connect(
                self.cbk_info_btn_pressed)
        
        add_question_mark_icon(self.info_btn)
    def cbk_on_neiview_loaded(self):
        self.hierarchies = self.neiview.neidata.cluster_hierarchy
        
        self.hierarchies.sig_cluster_removed.connect(
            self._cbk_on_cluster_removed)
        self.hierarchies.sig_cluster_added.connect(
            self._cbk_on_cluster_added)
        self.hierarchies.sig_clusters_grouped.connect(
            self._cbk_on_clusters_grouped_or_ungrouped)
        self.hierarchies.sig_clusters_ungrouped.connect(
            self._cbk_on_clusters_grouped_or_ungrouped)
        
        self._init_table()
        
        self.cbk_on_displayed_groups_changed()
    def cbk_on_displayed_groups_changed(self):
        self.clear_table()
        self.add_clusters(
            self._select_subset_from(
                self.hierarchies.clusters_all.values(),
                self._get_subset_type()))
    def cbk_searchbar_changed(self):
        self.set_visible_clusters(self.search_clusters())
    def search_clusters(self):
        query = self.search_bar.text()
        field = self.search_category_combo.currentText()
        hits = []
        
        getters = {
            "Group identifiers": lambda cl: cl.id_+cl.get_name(),
            "All annotations": lambda cl: cl._metadata["annotation"]+" ".join(cl._get_composition()),
            "Top annotation": lambda cl: cl.get_annotation(),
        }
        getters["All fields"] = lambda cl: getters["Group identifiers"](cl)+getters["All annotations"](cl)
        
        get_fields = getters[field]
        
        for row_i in range(self.cluster_table.rowCount()):
            cluster = self.cluster_table.cellWidget(row_i, 1).cluster
            if query in get_fields(cluster):
                hits.append(cluster)
        
        return(hits)
    def set_visible_clusters(self, clusters):
        for row_i in range(self.cluster_table.rowCount()):
            self.cluster_table.setRowHidden(
                row_i, 
                self.cluster_table.cellWidget(row_i, 1).cluster not in clusters)
    def _get_subset_type(self):
        # toplevel, user_created, locally_aligned, globally_aligned, selected
        subset_type = ""
        text = self.displayed_groups_combo.currentText()
        if text == "Groups":
            subset_type="toplevel"
        elif text=="User-created":
            subset_type="user_created"
        #elif text=="Local alignment-clustered":
        #    subset_type="locally_aligned"
        elif text=="Subgroups":
            subset_type="globally_aligned"
        #elif text=="Selected":
        #    subset_type="selected"
        else:
            assert False, "Invalid value for displayed_groups_combo."
        return(subset_type)
    def _select_subset_from(self, clusters, subset_type):
        if subset_type == "toplevel":
            subset = {cluster.get_toplevel() for cluster in clusters}
        elif subset_type in ("user_created", "locally_aligned", "globally_aligned"):
            subset = {cluster if cluster.type_ == subset_type \
                        else "dummy" for cluster in clusters}
            # Using the string "dummy" to catch possible errors with
            #   None clusters getting passed through arguments.
            subset.remove("dummy")
        elif subset_type == "selected":
            pass
        
        subset = list(subset)
        for i in range(len(subset)-1, -1, -1):
            if subset[i].removed:
                del subset[i]
        
        return(subset)
    def _generate_rows(self, clusters, subset_of_scaffolds):
        for cluster in clusters:
            #
            counts = Counter()
            for protein in cluster.proteins.values():
                for ft in protein.fts.values():
                    counts[ft.sc.accession] += 1
            subset_accessions = {region.sc.accession for region \
                                 in subset_of_scaffolds}
            for key in list(counts):
                if key not in subset_accessions:
                    del counts[key]
            
            counts_unique = len(counts)
            counts_total = sum(counts.values())
            duplications = Counter()
            for value in counts.values():
                duplications[value] += 1
            
            counts_duplications = []
            try:
                for i in range(1, max(duplications.keys())+1):
                    if i in duplications:
                        counts_duplications.append(f"{duplications[i]}x{i}")
                    #else:
                    #    counts_duplications.append(f"0x{i}")
            except ValueError as e:
                print("ERROR: Protein group has no instances?!")
                counts_duplications.append(f"error")
            
            counts_duplications = ", ".join(str(x) for x in counts_duplications)
            
            #
            widgets = (self.SelectionCheckbox(self, cluster), # Checkmark
                       ClusterWidget(None, cluster, neiview=self.neiview)) # ClusterWidget
            
            values = (
                cluster.get_annotation(), # Annotation
                counts_unique, # Unique Counts
                counts_total, # Total Counts
                counts_duplications, # Duplications
                )
            text_fields = []
            for value in values:
                text_field = qtw.QTableWidgetItem()
                text_field.setData(qtc.Qt.DisplayRole, value)
                text_field.setFlags(qtc.Qt.ItemIsSelectable | qtc.Qt.ItemIsEnabled)
                text_fields.append(text_field)
            text_fields = tuple(text_fields)
            yield(widgets+text_fields)
    def _init_table(self):
        # Disable sorting
        self.cluster_table.setSortingEnabled(False)
        
        # Init table
        self.cluster_table.setColumnCount(6)
        self.cluster_table.setRowCount(0)
        self.cluster_table.setHorizontalHeaderLabels(
            ["Selected", "ID", "Annotation", "Unique\nCounts", 
            "Total\nCounts", "Duplication\nCounts"])
        
        # Set the width of 0th column to 20
        # (Doesn't work right now)
        sh = self.cluster_table.horizontalHeaderItem(0).sizeHint()
        sh.setWidth(20)
        self.cluster_table.horizontalHeaderItem(0).setSizeHint(sh)
        
        # Re-enable sorting
        self.cluster_table.setSortingEnabled(True)
    def clear_table(self):
        # Disable sorting
        self.cluster_table.setSortingEnabled(False)
        self.cluster_table.clear()
        self._init_table()
        
        # Re-enable sorting
        self.cluster_table.setSortingEnabled(True)
    def add_clusters(self, clusters):
        # Disable sorting
        self.cluster_table.setSortingEnabled(False)
        
        # Set starting row
        row_i = self.cluster_table.rowCount() - 1
        #TODO: Replace self.neiview.neidata.regions_to_display with
        #      an argument that changes depending on visible clusters.
        # Note: As of rewrite, the above may not apply.   
        
        # Add new rows to accomodate addition
        self.cluster_table.setRowCount(
            self.cluster_table.rowCount() + len(clusters) - 1)
        
        for row in self._generate_rows(
          clusters, self.neiview.neidata.regions_to_display):
            for col_i in range(len(row)):
                item = row[col_i]
                if isinstance(item, qtw.QTableWidgetItem):
                    self.cluster_table.setItem(row_i, col_i, item)
                else:
                    self.cluster_table.setIndexWidget(
                        self.cluster_table.model().index(row_i,col_i), item)
            row_i += 1
        
        # Set column widths if they haven't been yet.
        if not self._column_widths_set:
            self.cluster_table.horizontalHeader().resizeSections(
                qtw.QHeaderView.ResizeToContents)
            self.cluster_table.horizontalHeader().resizeSection(
                1,
                self.neiview.dp["size_fixed_width"])
            self.cluster_table.horizontalHeader().resizeSection(
                2,
                int(round(self.cluster_table.horizontalHeader().sectionSize(2)/2)))
        
        # Re-enable sorting
        self.cluster_table.setSortingEnabled(True)
    def remove_clusters(self, clusters):
        # This function removes any clusters passed as arguments, but
        #   does not raise an error if a specific cluster is not
        #   found.
        
        # Disable sorting
        self.cluster_table.setSortingEnabled(False)
        
        clusters = set(clusters)
        for row_i in range(self.cluster_table.rowCount()-1, -1, -1):
            if self.cluster_table.cellWidget(row_i, 1).cluster in clusters:
                self.cluster_table.removeRow(row_i)
            
            
        # Re-enable sorting
        # This implicitly sorts the table without having to return
        #   control to main loop, meaning we can go straight to
        #   adding new clusters without any pauses.
        self.cluster_table.setSortingEnabled(True)
    def cbk_item_changed(self, item):
        if isinstance(item, self.SelectionCheckbox):
            item.cbk_checkstate_changed()
    def _cbk_on_cluster_removed(self, cluster):
        self._cluster_refresh_needed = True
        self._cluster_refresh_toremove.add(cluster)
        AppRoot.timer.after_delay_do(self._cluster_refresh, 100)
    def _cbk_on_cluster_added(self, cluster):
        self._cluster_refresh_needed = True
        self._cluster_refresh_toadd.add(cluster)
        AppRoot.timer.after_delay_do(self._cluster_refresh, 100)
    def _cbk_on_clusters_grouped_or_ungrouped(self, clusters, parents):
        self._cluster_refresh_needed = True
        
        self._cluster_refresh_toremove.update(clusters)
        self._cluster_refresh_toremove.update(parents)
        
        self._cluster_refresh_toadd.update(clusters)
        self._cluster_refresh_toadd.update(parents)
        
        AppRoot.timer.after_delay_do(self._cluster_refresh, 100)
    def _cluster_refresh(self):
        if self._cluster_refresh_needed:
            if self._cluster_refresh_toremove:
                # 
                self.remove_clusters(self._cluster_refresh_toremove)
            
            # Filter out clusters that don't belong
            to_add = self._select_subset_from(
                clusters=self._cluster_refresh_toadd,
                subset_type=self._get_subset_type())
            if to_add:
                self.add_clusters(to_add)
            
            self._cluster_refresh_needed = False
            self._cluster_refresh_toremove.clear()
            self._cluster_refresh_toadd.clear()
    def copy_highlighted_into_clipboard(self):
        output = list()
        ranges_ = self.cluster_table.selectedRanges()
        if not ranges_: return
        for range_ in ranges_:
            for row in range(range_.topRow(), range_.bottomRow()+1):
                output.append(list())
                for col in range(range_.leftColumn(), range_.rightColumn()+1):
                    item = self.cluster_table.item(row, col)
                    if item is None:
                        # If it's not an item, we can assume it's
                        #   a ColourableButton until more widgets are inserted
                        item = self.cluster_table.cellWidget(row, col)
                        
                        cluster = item.cluster
                        value = cluster.get_name()
                        
                        output[-1].append(value)
                    else:
                        value = item.text()
                        
                        output[-1].append(value)
        del col
        del row
        clipboard = AppRoot.qtapp.clipboard()
        clipboard.clear()
        clipboard.setText("\n".join(["\t".join(row) for row in output]))
    def cbk_info_btn_pressed(self):
        self.info_lbl.setHidden(not self.info_lbl.isHidden())
    # TODO: Subclass ClusterWidget and qtw.QTableWidgetItem for a pureblood
    #       list item widget that does all the things we want it to.
    class SelectionCheckbox(qtw.QTableWidgetItem):
        def __init__(self, topclusters, cluster, *args, **kwargs):
            super().__init__(*args, **kwargs)
            self.setFlags(qtc.Qt.ItemIsSelectable | qtc.Qt.ItemIsUserCheckable
                          | qtc.Qt.ItemIsEnabled)
            self.topclusters = topclusters
            self.cluster = cluster
            
            self.setCheckState(qtc.Qt.Unchecked)
            def cbk_cluster_state_changed(state):
                try: 
                    self.checkState()
                except RuntimeError:
                    cluster.sig_cluster_selected.disconnect(
                        cbk_cluster_state_changed)
                    return
                if state == bool(self.checkState()): return
                if state:
                    self.check()
                else:
                    self.uncheck()
                self.cbk_update_checkmark()
            self.cluster.sig_cluster_selected.connect(
                cbk_cluster_state_changed)
            
            # This is an absolutely disgusting hack to make
            #   utilizing whitespace to sort TableWidgetItems
            #   already containing something else
            self.setData(qtc.Qt.DisplayRole, "")
            self.cbk_update_checkmark()
        # Due to an oversight I cannot understand atm, this callback is defined
        #   in __init__, because if it s a method, the connection to
        #   the signal fails.
        #def cbk_cluster_state_changed(self, state):
        #    if state == self.checkState(): return
        #    if state:
        #        self.check()
        #    else:
        #        self.uncheck()
        def cbk_checkstate_changed(self):
            state = self.checkState()
            if state is qtc.Qt.Checked:
                self.cluster.set_selected(bool(state))
            elif state is qtc.Qt.Unchecked:
                self.cluster.set_selected(bool(state))
            self.cbk_update_checkmark()
        def cbk_update_checkmark(self):
            state = self.checkState()
            if state is qtc.Qt.Checked:
                self.setText("Yes")
            elif state is qtc.Qt.Unchecked:
                self.setText("No")
        def check(self):
            self.setCheckState(qtc.Qt.Checked)
            # See init
            #self.setData(qtc.Qt.DisplayRole, ":)")
        def uncheck(self):
            self.setCheckState(qtc.Qt.Unchecked)
            # See init
            # self.setData(qtc.Qt.DisplayRole, ":(")
#
class SelectorWidget(qtw.QWidget, Ui_SelectorWidget):
    def __init__(self, name_a, name_b, items_a=[], items_b=[], *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.init_ui(name_a, name_b, items_a, items_b)
    def init_ui(self, name_a, name_b, items_a, items_b):
        self.setupUi(self)
        
        self.ui_listA_label.setText(name_a)
        self.ui_listB_label.setText(name_b)
        
        self.ui_btn_AtoB.clicked.connect(self.cbk_atob)
        self.ui_btn_BtoA.clicked.connect(self.cbk_btoa)
        self.ui_btn_Bup.clicked.connect(self.cbk_bup)
        self.ui_btn_Bdown.clicked.connect(self.cbk_bdown)
        
        for item in items_a:
            self.add_a_item(item)
        for item in items_b:
            self.add_b_item(item)
    def cbk_atob(self):
        # An overly elaborate one liner.
        # currentItem() returns an object,
        #   but takeItem needs row
        self.ui_listB.addItem( # Move item
            self.ui_listA.takeItem( # Get item from row and remove from A
                self.ui_listA.row( # Get row of current item
                    self.ui_listA.currentItem() # Get current item
                    )))
    def cbk_btoa(self):
        # An overly elaborate one liner.
        # currentItem() returns an object,
        #   but takeItem needs row
        self.ui_listA.addItem( # Move item
            self.ui_listB.takeItem( # Get item from row and remove from B
                self.ui_listB.row( # Get row of current item
                    self.ui_listB.currentItem() # Get current item
                    )))
        self.ui_listA.sortItems()
    def cbk_bup(self):
        # Take item from list, put it one higher.
        row = self.ui_listB.row(self.ui_listB.currentItem())
        if row==0: return #Don't move if it's at the top
        item = self.ui_listB.takeItem(row)
        self.ui_listB.insertItem(row-1, item)
        self.ui_listB.setCurrentRow(row-1)
    def cbk_bdown(self):
        # Take item from list, put it one lower.
        
        #Get Row
        row = self.ui_listB.row(self.ui_listB.currentItem())
        
        #Don't move down if there's nothing lower
        if not self.ui_listB.item(row+1): 
            return
        item = self.ui_listB.takeItem(row)
        self.ui_listB.insertItem(row+1, item)
        self.ui_listB.setCurrentRow(row+1)
    def get_a_contents(self):
        contents = []
        for i in range(self.ui_listA.count()):
            contents.append(self.ui_listA.item(i).text())
        return(contents)
    def get_b_contents(self):
        contents = []
        for i in range(self.ui_listB.count()):
            contents.append(self.ui_listB.item(i).text())
        return(contents)
    
    #def get_a_contents(self):
    
    #def get_b_contents(self):
    
    def add_a_item(self, label):
        self.ui_listA.addItem(label)
    def add_b_item(self, label):
        self.ui_listB.addItem(label)
#
class ClusterContainerWidget(qtw.QWidget):
    sig_cluster_added = qtc.Signal(object)
    sig_cluster_removed = qtc.Signal(object)
    def __init__(self, neiview, stretchable=True, contents_accept_drops=True,
                       row_size=5):
        super().__init__()
        self.row_size = row_size
        self.init_ui(stretchable)
        self.neiview = neiview
        
        #Remove later
        #hierarchy.sig_styledicts_changed.connect(self.cbk_styledicts_changed)
        
        self.contents_accept_drops = contents_accept_drops
    @property
    def hierarchy(self):
        return(self.neiview.neidata.cluster_hierarchy)
    def init_ui(self, stretchable):
        self.setAcceptDrops(True)
        
        self.lay = qtw.QGridLayout()
        self.lay.setContentsMargins(1,1,1,1)
        self.lay.setSpacing(1)
        
        if stretchable:
            self.meta_lay = qtw.QVBoxLayout()
            self.meta_lay.setContentsMargins(0,0,0,0)
            self.meta_lay.setSpacing(0)
            self.setLayout(self.meta_lay)
            
            self.meta_lay.addLayout(self.lay)
        
            self.meta_lay.addStretch()
        else:
            self.setLayout(self.lay)
        
        for i in range(5):
            self.lay.setColumnMinimumWidth(i,
                NeiView.get_default_display_policy()["size_fixed_width"])
        
        self.setMinimumHeight(
            NeiView.get_default_display_policy()["size_fixed_height"])
    def add_cluster(self, cluster):
        if cluster not in self.hierarchy.clusters_all.values():
            self.cbk_on_invalid_cluster(cluster)
            return
        if cluster in self.get_clusters():
            return
        cwidget = ClusterWidget(self, cluster, neiview=self.neiview)
        
        cwidget.setAcceptDrops(self.contents_accept_drops)
        cwidget.set_removable(not self.contents_accept_drops)
        cwidget.sig_removeme.connect(self.cbk_widget_removeme)
        
        cwidget.sig_reparented.connect(
            self.cbk_on_clusterwidget_reparented)
        cwidget.setMaximumSize(
            NeiView.get_default_display_policy()["size_fixed_width"],
            NeiView.get_default_display_policy()["size_fixed_height"])
        cwidget.setMinimumSize(
            NeiView.get_default_display_policy()["size_fixed_width"],
            NeiView.get_default_display_policy()["size_fixed_height"])
        
        index = self.lay.count()
        row = (index-index%self.row_size)/self.row_size
        col = index%self.row_size
        self.lay.addWidget(cwidget,
                           row,
                           col)
        self._update_size()
        self.sig_cluster_added.emit(cluster)
    def remove_cluster(self, cluster):
        to_remove = []
        for index in range(self.lay.count()):
            widget = self.lay.itemAt(index).widget()
            if widget.cluster is cluster:
                to_remove.append(widget)
        for widget in to_remove:
            self.lay.removeWidget(widget)
            widget.setParent(None)
        self.resort_widgets()
        self.sig_cluster_removed.emit(cluster)
    def resort_widgets(self):
        widgets = []
        for index in range(self.lay.count()):
            widget = self.lay.itemAt(index).widget()
            widgets.append(widget)
        for index in range(len(widgets)):
            self.lay.addWidget(widgets[index], 
                               (index-index%self.row_size)/self.row_size,
                               index%self.row_size)
        self._update_size()
    def _update_size(self):
        index = self.lay.count()
        row = max((index-index%self.row_size)/self.row_size, 1)
        col = index%self.row_size
        self.setMinimumHeight((3+NeiView.get_default_display_policy()["size_fixed_height"])*row)
    def get_clusters(self):
        clusters = set()
        for index in range(self.lay.count()):
            widget = self.lay.itemAt(index).widget()
            clusters.add(widget.cluster)
        return(clusters)
    def clear(self):
        for cluster in list(self.get_clusters()):
            self.remove_cluster(cluster)
    def cbk_on_invalid_cluster(self, cluster):
        qtw.QMessageBox.warning(None,
            f"Invalid protein group {cluster.id_}",
            "Protein group may belong to a different neighborhood view than this container.")
    def cbk_widget_removeme(self, widget):
        for cluster in (set(self.get_clusters()) 
                            & self.neiview.neidata.cluster_hierarchy.selection):
            self.remove_cluster(cluster)
    
    def cbk_on_clusterwidget_reparented(self, clusterwidget):
        self.remove_cluster(clusterwidget.get_toplevel())
        #self.resort_widgets()
        #clusterwidget.sig_reparented.disconnect(
        #    self.cbk_on_clusterwidget_reparented)
    #
    def cbk_styledicts_changed(self):
        for index in range(self.lay.count()):
            widget = self.lay.itemAt(index).widget()
            widget.update()
    def paintEvent(self, event):
        painter = qtg.QPainter()
        painter.begin(self)
        
        h = self.geometry().height()
        w = self.geometry().width()
        
        painter.setBrush(qtg.QBrush(qtg.QColor(255,225,225,255)))
        painter.setPen(qtg.QColor(0,0,0,0))
        
        painter.drawRect(qtc.QRectF(0,0,w,h))
        
        painter.end()
    def dragEnterEvent(self, event):
        if hasattr(event.mimeData(), "dragged_clusterwidget"):
            event.accept()
        else:
            event.ignore()
    def dropEvent(self, event):
        if hasattr(event.mimeData(), "dragged_clusterwidget"):
            event.accept()
            dragged_clusterwidget = event.mimeData().dragged_clusterwidget
            
            if dragged_clusterwidget.get_toplevel().get_selected():
                to_add = set(self.hierarchy.selection)
            else:
                to_add = {dragged_clusterwidget.get_toplevel()}
            
            own_clusters = set(self.get_clusters())
            for cluster in to_add - own_clusters:
                self.add_cluster(cluster)
#
class SelectionManager(qtw.QWidget, Ui_SelectionManager):
    def __init__(self, neiview):
        super().__init__()
        
        self.neiview = neiview
        self.hierarchy = neiview.neidata.cluster_hierarchy
        self.doomed_widget = None
        
        self.init_ui()
    def init_ui(self):
        self.setupUi(self)
        
        for name in self.options:
            self.rule_combo.addItem(name)
        
        self.add_rule_btn.clicked.connect(self.cbk_add_rule)
        
        self.select_btn.clicked.connect(self.cbk_select)
        self.show_btn.clicked.connect(self.cbk_show)
        
        self.setAcceptDrops(True)
    def load_initial_rules(self):
        self.add_rulewidget(self.Rule_All(self.neiview))
        self.add_rulewidget(self.Rule_None(self.neiview))
        self.add_rulewidget(self.Rule_AtLeast(self.neiview))
        self.add_rulewidget(self.Rule_AtMost(self.neiview))
    def dragEnterEvent(self, event):
        if hasattr(event.mimeData(), "dragged_clusterwidget")\
          and event.mimeData().dragged_clusterwidget._removable:
            event.accept()
        else:
            event.ignore()
    def dropEvent(self, event):
        if hasattr(event.mimeData(), "dragged_clusterwidget")\
          and event.mimeData().dragged_clusterwidget._removable:
            event.accept()
            dragged_clusterwidget = event.mimeData().dragged_clusterwidget
            dragged_clusterwidget.setParent(None)
            dragged_clusterwidget.sig_reparented.emit(dragged_clusterwidget)
            dragged_clusterwidget.setHidden(True)
            self.doomed_widget = dragged_clusterwidget
    def get_rule_results(self):
        results = []
        evaluators = self.get_evaluators()
        while None in evaluators:
            evaluators.remove(None)
        for regionui in self.neiview.regions_all:
            if all(evaluator(regionui.reg) for evaluator in evaluators):
                results.append(regionui)
        return(results)
    def cbk_select(self):
        for regui in self.neiview.regions_all:
            regui.set_selected(False)
        for regui in self.get_rule_results():
            regui.set_selected(True)
    def cbk_show(self):
        results = set(self.get_rule_results())
        for regui in self.neiview.regions_all:
            if regui in results:
                regui.set_displayed(True)
            else:
                regui.set_displayed(False)
    def get_evaluators(self):
        evaluators = []
        for rule in self.neiview.display_policy["restriction_rules"]:
            evaluators.append(rule.get_evaluator())
        return(evaluators)
    def cbk_cluster_added_to_selection(self, cluster):
        if not cluster.get_selected():
            cluster.set_selected(True)
    def cbk_cluster_removed_from_selection(self, cluster):
        if cluster.get_selected():
            cluster.set_selected(False)
    def cbk_add_rule(self):
        # Create rulewidget
        rule_name = self.rule_combo.currentText()
        rule_type = self.options[rule_name]
        rulewidget = rule_type(self.neiview)
        self.add_rulewidget(rulewidget)
    def add_rulewidget(self, rulewidget):
        self.rules_lay.insertWidget(self.rules_lay.count()-1, rulewidget)
    class RuleWidget(qtw.QWidget):
        message = "N/A"
        displaytype = None
        sig_rule_removed = qtc.Signal(object)
        def __init__(self, neiview):
            super().__init__()
            self.neiview = neiview
            self.hierarchy = neiview.neidata.cluster_hierarchy
            
            self.data = {}
            
            assert self not in self.neiview.display_policy["restriction_rules"]
            self.neiview.display_policy["restriction_rules"].append(self)
            
            self.destroyed.connect(self.cbk_destroyed)
        def is_complete(self):
            pass
        def get_evaluator(self, region):
            pass
        def cbk_remove(self):
            self.sig_rule_removed.emit(self)
            self.setParent(None)
            self.cbk_destroyed()
        def cbk_destroyed(self):
            self.neiview.display_policy["restriction_rules"].remove(self)
            self.destroyed.disconnect(self.cbk_destroyed)
            assert self not in self.neiview.display_policy["restriction_rules"]
        def _make_remove_button(self):
            remove_btn = qtw.QPushButton("x")
            remove_btn.clicked.connect(self.cbk_remove)
            remove_btn.setFixedSize(20,20)
            return(remove_btn)
        def _clone_clusterwidget(self, transcriptui):
            cloned_widget = ClusterWidget(self, transcriptui.cluster)
            cloned_widget.setFixedSize(
                NeiView.get_default_display_policy()["size_fixed_width"],
                NeiView.get_default_display_policy()["size_fixed_height"])
            cloned_widget.setAcceptDrops(False)
            cloned_widget.set_removable(True)
            return(cloned_widget)
        # For saving/loading
        #   These methods assume that a given rulewidget subclass has exactly
        #   one ClusterContainerWidget called cluster_container in which it
        #   keeps all its information.
        #   Other kinds of setups will require reimplementing the methods.
        def _to_dict(self):
            # Debug try-except clause
            clusters = [str(cluster.id_) for cluster in \
                self.cluster_container.get_clusters()]
            data_ = dict(self.data)
            data_.update({"cluster_ids": clusters})
            return(data_)
        @classmethod
        def _from_dict(cls, data_, neiview):
            new = cls(neiview)
            cluster_ids = list(data_["cluster_ids"])
            del data_["cluster_ids"]
            
            new.data.update(data_)
            for id_ in cluster_ids:
                new.cluster_container.add_cluster(
                    neiview.neidata.cluster_hierarchy.clusters_all[id_])
            
            if hasattr(new, "spinbox_value"):
                new.spinbox_value = new.spinbox_value
            return(new)
    class Rule_All(RuleWidget):
        message = "Must contain ALL: "
        def __init__(self, neiview):
            super().__init__(neiview)
            self.init_ui()
        def init_ui(self):
            self.setAcceptDrops(True)
            
            self.lay = qtw.QGridLayout()
            self.lay.setContentsMargins(0,0,0,0)
            self.setLayout(self.lay)
            
            self.lay.addWidget(self._make_remove_button(), 0, 0)
            
            self.lay.addWidget(qtw.QLabel(self.message), 0, 1)
            
            self.cluster_container = ClusterContainerWidget(
                self.neiview,
                stretchable=False,
                contents_accept_drops=False,
                row_size=4)
            self.lay.addWidget(self.cluster_container, 1, 0, 1, -1)
        def dragEnterEvent(self, event):
            self.cluster_container.dragEnterEvent(event)
        def dropEvent(self, event):
            self.cluster_container.dropEvent(event)
        def _get_clusters(self):
            return(self.cluster_container.get_clusters())
        def get_evaluator(self):
            clusters = self._get_clusters()
            if len(clusters) < 1:
                return(None)
            def evaluator(region):
                proteins = {ft.ref if ft.type == "cds" else None for ft in region.fts.values()}
                passed = True
                for cluster in clusters:
                    if not proteins.intersection(cluster.proteins.values()):
                        passed = False
                        break
                return(passed)
            return(evaluator)
    class Rule_None(Rule_All):
        message = "Must NOT contain "
        if_true_return = True
        def __init__(self, neiview):
            super().__init__(neiview)
        def get_evaluator(self):
            clusters = self._get_clusters()
            if len(clusters) < 1:
                return(None)
            def evaluator(region):
                proteins = {ft.ref if ft.type == "cds" else None for ft in region.fts.values()}
                passed = True
                for cluster in clusters:
                    if proteins.intersection(cluster.proteins.values()):
                        passed = False
                        break
                return(passed)
            return(evaluator)
    class Rule_AtLeast(RuleWidget):
        message = "Must contain at least "
        def __init__(self, neiview):
            super().__init__(neiview)
            self.init_ui()
        def init_ui(self):
            self.setAcceptDrops(True)
            
            self.lay = qtw.QGridLayout()
            self.lay.setContentsMargins(0,0,0,0)
            self.setLayout(self.lay)
            
            self.lay.addWidget(self._make_remove_button(), 0, 0)
            
            self.lay.addWidget(qtw.QLabel(self.message), 0, 1)
            self.config_spinbox = qtw.QSpinBox()
            self.config_spinbox.valueChanged.connect(
                self.cbk_config_spinbox_changed)
            self.config_spinbox.setMinimum(1)
            self.config_spinbox.setValue(1)
            self.lay.addWidget(self.config_spinbox, 0, 2)
            
            self.lay.addWidget(qtw.QLabel(" of "), 0, 3)
            
            self.cluster_container = ClusterContainerWidget(
                self.neiview,
                stretchable=False,
                contents_accept_drops=False,
                row_size=4)
            self.lay.addWidget(self.cluster_container, 1, 0, 1, -1)
        def dragEnterEvent(self, event):
            self.cluster_container.dragEnterEvent(event)
        def dropEvent(self, event):
            self.cluster_container.dropEvent(event)
        def _get_clusters(self):
            return(self.cluster_container.get_clusters())
        @property 
        def spinbox_value(self):
            return(self.data["spinbox_value"])
        @spinbox_value.setter
        def spinbox_value(self, new_value):
            if isinstance(new_value, int):
                self.data["spinbox_value"] = new_value
                if self.config_spinbox.value() != new_value:
                    self.config_spinbox.setValue(new_value)
            else:
                raise ValueError(f"Spinbox value must be int, not {type(new_value)}")
        def cbk_config_spinbox_changed(self, value):
            self.spinbox_value = value
        def get_evaluator(self):
            clusters = self._get_clusters()
            valid_proteins = set()
            for cluster in clusters:
                for protein in cluster.proteins.values():
                    valid_proteins.add(protein)
            minimum_hits = self.spinbox_value
            if len(clusters) < 1:
                return(None)
            def evaluator(region):
                proteins = {ft.ref if ft.type == "cds" else None for ft in region.fts.values()}
                hits = 0
                for protein in proteins:
                    if protein in valid_proteins:
                        hits += 1
                return(hits >= minimum_hits)
            return(evaluator)
    class Rule_AtMost(Rule_AtLeast):
        message = "Must contain no more than "
        def __init__(self, hierarchy):
            super().__init__(hierarchy)
        def get_evaluator(self):
            clusters = self._get_clusters()
            invalid_proteins = set()
            for cluster in clusters:
                for protein in cluster.proteins.values():
                    invalid_proteins.add(protein)
            maximum_hits = self.spinbox_value
            if len(clusters) < 1:
                return(None)
            def evaluator(region):
                proteins = {ft.ref if ft.type == "cds" else None for ft in region.fts.values()}
                hits = 0
                for protein in proteins:
                    if protein in invalid_proteins:
                        hits += 1
                return(hits < maximum_hits)
            return(evaluator)
    
    options = {
        "Has all": Rule_All,
        "Has none": Rule_None,
        "Has at least N": Rule_AtLeast,
        "Has at most N": Rule_AtMost,
    }
 # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #

 #                            Other                                  #

 # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # #

_VALID_QTATTRS={"SolidLine", "DashLine", "DotLine", "DashDotLine", "DashDotDotLine",
                "SquareCap", "FlatCap", "RoundCap", "RoundJoin", "SolidPattern", 
                "Dense1Pattern", "Dense2Pattern", "Dense3Pattern", "Dense4Pattern", 
                "Dense5Pattern", "Dense6Pattern", "Dense7Pattern", "NoBrush", 
                "HorPattern", "VerPattern", "CrossPattern", "BDiagPattern", 
                "FDiagPattern", "DiagCrossPattern"}
_QTATTR_TO_STR = []
for item in _VALID_QTATTRS:
    _QTATTR_TO_STR.append((getattr(qtc.Qt, item), item))


def stringify(value):
    if isinstance(value, qtg.QColor):
        value = f"!color:{value.red()},{value.green()},{value.blue()},{value.alpha()}"
    elif any(value is item[0] for item in _QTATTR_TO_STR):
        for k,v in _QTATTR_TO_STR:
            if value is k:
                value = "!qtattr:" + v
                break
    elif isinstance(value, bool):
        # Bools always have to be checked before int
        #   because bool is a subclass of int in python
        #   for some stupid reason :)
        value = "!bool:"+str(value)
    elif isinstance(value, int):
        value = "!int:"+str(value)
    elif isinstance(value, float):
        value = "!float:"+str(value)
    elif isinstance(value, str):
        value = "!str:"+value
    elif value is None:
        value = "!none:"
    # Selection rules are only turned into a dict here, and are properly initialized
    #   later in dbdl.DatabaseManager.get_neiview after the rest of the neiview
    #   has been loaded.
    elif isinstance(value, RegionUI.Tag):
        value = ["!class:RegionUI.Tag", dbdl.stringify_data(value.data)]
    elif isinstance(value, SelectionManager.Rule_None):
        value = ["!class:SelectionManager.Rule_None", dbdl.stringify_data(value._to_dict())]
    elif isinstance(value, SelectionManager.Rule_All):
        value = ["!class:SelectionManager.Rule_All", dbdl.stringify_data(value._to_dict())]
    elif isinstance(value, SelectionManager.Rule_AtMost):
        value = ["!class:SelectionManager.Rule_AtMost", dbdl.stringify_data(value._to_dict())]
    elif isinstance(value, SelectionManager.Rule_AtLeast):
        value = ["!class:SelectionManager.Rule_AtLeast", dbdl.stringify_data(value._to_dict())]
    elif isinstance(value, nx.Graph):
        value = ["!class:nx.Graph", nx.node_link_data(value)]
    elif isinstance(value, qtg.QFont):
        value = ["!class:qtg.QFont", value.toString()]
    else:
        raise ValueError(f"Undefined data type. {type(value)}, {value}",)
    return(value)

def unstringify(value):
    if   isinstance(value, str) and value.startswith("!color:"):
        value = value.removeprefix("!color:").split(",")
        value = qtg.QColor(int(value[0]), int(value[1]),
                           int(value[2]), int(value[3]))
    elif isinstance(value, str) and value.startswith("!qtattr:"):
        value = getattr(qtc.Qt, value.removeprefix("!qtattr:"))
    elif isinstance(value, str) and value.startswith("!int:"):
        value = int(value.removeprefix("!int:"))
    elif isinstance(value, str) and value.startswith("!float:"):
        value = float(value.removeprefix("!float:"))
    elif isinstance(value, str) and value.startswith("!str:"):
        value = value.removeprefix("!str:")
    elif isinstance(value, str) and value.startswith("!bool:"):
        value = value.removeprefix("!bool:")
        if value == "True":
            value = True
        elif value == "False":
            value = False
        else:
            raise ValueError(f"Invalid bool value in stringified dict. ({value})")
    elif isinstance(value, str) and value.startswith("!none:"):
        value = None
    elif isinstance(value, list) and value[0] == "!class:RegionUI.Tag":
        new = RegionUI.Tag(None)
        new.update_data(dbdl.unstringify_data(value[1]))
        value = new
    elif isinstance(value, list) and value[0] == "!class:SelectionManager.Rule_All":
        value = [value[0], dbdl.unstringify_data(value[1])]
    elif isinstance(value, list) and value[0] == "!class:SelectionManager.Rule_None":
        value = [value[0], dbdl.unstringify_data(value[1])]
    elif isinstance(value, list) and value[0] == "!class:SelectionManager.Rule_AtLeast":
        value = [value[0], dbdl.unstringify_data(value[1])]
    elif isinstance(value, list) and value[0] == "!class:SelectionManager.Rule_AtMost":
        value = [value[0], dbdl.unstringify_data(value[1])]
    elif isinstance(value, list) and value[0] == "!class:nx.Graph":
        value = nx.node_link_graph(value[1])
    elif isinstance(value, list) and value[0] == "!class:qtg.QFont":
        font = qtg.QFont()
        font.fromString(value[1])
        value = font
    else:
        raise ValueError(f"Unknown datatype in value: {value}")
    return(value)

def stringify_dict(unstringy):
    stringy = dict()
    for key in unstringy:
        value = unstringy[key]
        value = stringify(value)
        stringy[key] = value
    return(stringy)

def unstringify_dict(stringy, optional=[]):
    unstringy = dict()
    for key in stringy:
        value = stringy[key]
        value = unstringify(value)
        unstringy[key] = value
    
    final = {key: None for key in optional}
    final.update(unstringy)
    return(final)




def launch_app():
    app = AppRoot()
    app.launch()
