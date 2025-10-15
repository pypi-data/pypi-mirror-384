import calendar
import io
import math
import os
import random
import re
import zipfile
from datetime import date, datetime, timedelta
from decimal import ROUND_HALF_UP, Decimal
from pathlib import Path

import numpy as np
import pandas as pd
from dateutil.relativedelta import relativedelta
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.core.paginator import EmptyPage, PageNotAnInteger, Paginator
from django.db.models import (
    Count,
    DecimalField,
    ExpressionWrapper,
    F,
    Q,
    Sum,
    Value,
)
from django.db.models.functions import Coalesce
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse_lazy
from django.utils import translation
from django.utils.encoding import escape_uri_path
from django.utils.translation import gettext as _
from django.views.decorators.http import require_POST
from django.views.generic import (
    CreateView,
    DeleteView,
    DetailView,
    ListView,
    UpdateView,
)
from fnschool import count_chinese_characters
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ..forms import (
    CategoryForm,
    ConsumptionForm,
    IngredientForm,
    PurchasedIngredientsWorkBookForm,
)
from ..models import Category, Consumption, Ingredient, MealType
from ..views import decimal_prec


def get_CNY_TEXT(amount):
    units = {
        "0": "\u96f6",  # ling2
        "1": "\u58f9",  # yi1
        "2": "\u8d30",  # er4
        "3": "\u53c1",  # san1
        "4": "\u8086",  # si4
        "5": "\u4f0d",  # wu3
        "6": "\u9646",  # liu4
        "7": "\u67d2",  # qi1
        "8": "\u634c",  # ba1
        "9": "\u7396",  # jiu3
    }

    levels = [
        "",
        "\u62fe",  # shi2
        "\u4f70",  # bai3
        "\u4edf",  # qian1
        "\u4e07",  # wan4
        "\u4ebf",  # yi4
        "\u5143",  # yuan2
        "\u89d2",  # jiao3
        "\u5206",  # fen1
        "\u6574",  # zheng3
    ]

    is_negative = False
    if amount < 0:
        is_negative = True
        amount = abs(amount)
    if amount == 0:
        return "\u96f6\u5143\u6574"  # ling2 yuan2 zheng3.

    amount = Decimal(str(amount)).quantize(
        Decimal("0.00"), rounding=ROUND_HALF_UP
    )
    amount_str = str(amount)

    integer_part = None
    decimal_part = None
    if "." in amount_str:
        integer_part, decimal_part = amount_str.split(".")
    else:
        integer_part = amount_str
        decimal_part = "00"

    result = []
    integer_part = integer_part.zfill(16)

    groups = [
        integer_part[-16:-12],
        integer_part[-12:-8],
        integer_part[-8:-4],
        integer_part[-4:],
    ]

    group_names = [
        "\u4e07",  # wan4
        "\u4ebf",  # yi4
        "\u4e07",  # wan4
        "\u5143",  # yuan2
    ]

    for i, group in enumerate(groups):
        group = group.lstrip("0")
        if not group:
            continue

        for j, digit in enumerate(group):
            if digit == "0":
                if result and result[-1] != "\u96f6":  # \\u96f6 is ling2 .
                    result.append("\u96f6")  # \\u96f6 is ling2 .
            else:
                result.append(units[digit])

                if len(group) - j - 1 > 0:
                    result.append(levels[len(group) - j - 1])

        if group_names[i]:
            result.append(group_names[i])

    if decimal_part != "00":
        if decimal_part[0] != "0":
            result.append(units[decimal_part[0]])
            result.append("\u89d2")  # \\u89d2 is jiao3 .

        if decimal_part[1] != "0":
            result.append(units[decimal_part[1]])
            result.append("\u5206")  # \\u5206 is fen1 .
    else:
        result.append("\u6574")  # \\u6574 is zheng3 .

    output = "".join(result)

    output = re.sub("\u96f6+", "\u96f6", output)
    output = re.sub("\u96f6([\u4e07\u4ebf])", r"\1", output)
    output = re.sub("\u96f6\u5143", "\u5143", output)
    output = re.sub("\u96f6\u89d2\u96f6\u5206", "", output)
    output = re.sub("\u96f6\u5206", "", output)

    if output.startswith("\u58f9\u62fe"):
        output = output.replace("\u58f9\u62fe", "\u62fe", 1)

    if is_negative:
        output = "\u8d1f" + output

    return output


def is_zh_CN():
    lang = translation.get_language()
    return lang.lower() in ["zh-cn", "zh-hans"]


def set_column_width_in_inches(worksheet, column, inches):
    char_width = inches * 96 / 7

    if isinstance(column, int):
        col_letter = get_column_letter(column)
    else:
        col_letter = column
    worksheet.column_dimensions[col_letter].width = char_width


def set_row_height_in_inches(worksheet, row, inches):
    points = inches * 72
    worksheet.row_dimensions[row].height = points


class CanteenWorkBook:
    def __init__(self, request, month, meal_type):
        self.wb = Workbook()
        self.wb[self.wb.sheetnames[0]].sheet_state = "hidden"
        self.cover_sheet = self.wb.create_sheet(title=_("Sheet Cover"))
        self.storage_sheet = self.wb.create_sheet(title=_("Sheet Storage"))
        self.storage_list_sheet = self.wb.create_sheet(
            title=_("Sheet Storage List")
        )
        self.non_storage_sheet = self.wb.create_sheet(
            title=_("Sheet Non-Storage")
        )
        self.non_storage_list_sheet = self.wb.create_sheet(
            title=_("Sheet Non-Storage List")
        )
        self.consumption_sheet = self.wb.create_sheet(
            title=_("Sheet Consumption")
        )
        self.consumption_list_sheet = self.wb.create_sheet(
            title=_("Sheet Consumption List")
        )
        self.surplus_sheet = self.wb.create_sheet(title=_("Sheet Surplus"))
        self.center_alignment = Alignment(
            horizontal="center", vertical="center"
        )
        self.thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        self.font_12 = Font(size=12)
        self.font_12_bold = Font(size=12, bold=True)
        self.font_14 = Font(size=14)
        self.font_16 = Font(size=16)
        self.font_16_bold = Font(size=16, bold=True)
        self.font_20_bold = Font(size=20, bold=True)

        self.request = request
        self.user = self.request.user
        self.meal_type = meal_type
        self.year = int(month.split("-")[0])
        self.month = int(month.split("-")[1])
        self.date_start = datetime(self.year, self.month, 1).date()
        self.date_end = datetime(
            self.year, self.month, calendar.monthrange(self.year, self.month)[1]
        ).date()
        self.is_zh_CN = is_zh_CN()

        self._is_school = None

    @property
    def is_school(self):
        if self._is_school:
            return self._is_school
        is_school = (
            (
                any(
                    [
                        name in self.user.affiliation
                        for name in [
                            "\u5e7c\u513f\u56ed",
                            "\u5c0f\u5b66",
                            "\u4e2d\u5b66",
                            "\u5927\u5b66",
                        ]
                    ]
                )
            )
            if self.is_zh_CN
            else False
        )
        self._is_school = is_school
        return self._is_school

    def fill_in_non_storage_sheet(self):
        sheet = self.non_storage_sheet
        user = self.user
        title_cell = sheet.cell(1, 1)
        title_cell.value = _(
            "Table of {superior_department} Canteen Non-Storaged Ingredients Statistics"
        ).format(
            superior_department=user.superior_department,
        )
        title_cell.font = self.font_16_bold
        title_cell.alignment = self.center_alignment
        for col_num, width in [
            [1, 2.23],
            [2, 3.14],
            [3, 3.14],
        ]:
            set_column_width_in_inches(sheet, col_num, width)
        sheet.merge_cells("A1:C1")

        sub_title_cell = sheet.cell(2, 1)
        sheet.merge_cells("A2:C2")
        sub_title_cell.font = self.font_12
        sub_title_cell.alignment = self.center_alignment
        sub_title_cell.value = _(
            "Affiliation: {affiliation}        Monetary Unit:         {year}.{month:0>2}.{day:0>2}"
        ).format(
            affiliation=user.affiliation,
            year=self.year,
            month=self.month,
            day=self.date_end.day,
        )

        header_row_num = 3
        header_category_cell = sheet.cell(header_row_num, 1)
        header_category_cell.value = _("Ingredient Categories (storage sheet)")

        header_total_price_cell = sheet.cell(header_row_num, 2)
        header_total_price_cell.value = _(
            "Ingredient Total Prices (storage sheet)"
        )

        header_note_cell = sheet.cell(header_row_num, 3)
        header_note_cell.value = _("Procurement Note")

        for cell in [
            header_category_cell,
            header_total_price_cell,
            header_note_cell,
        ]:
            cell.font = self.font_12
            cell.alignment = self.center_alignment
            cell.border = self.thin_border

        categories = Category.objects.filter(
            Q(user=user) & Q(is_disabled=False)
        ).all()

        set_row_height_in_inches(sheet, 1, 0.38)
        set_row_height_in_inches(sheet, 2, 0.22)
        set_row_height_in_inches(sheet, 3, 0.32)

        for index, category in enumerate(categories):

            row_num = header_row_num + 1 + index
            set_row_height_in_inches(sheet, row_num, 0.32)

            category_cell = sheet.cell(row_num, 1)
            category_cell.value = category.name

            ingredients = Ingredient.objects.filter(
                Q(user=user)
                & Q(category=category)
                & Q(storage_date__gte=self.date_start)
                & Q(storage_date__lte=self.date_end)
                & Q(meal_type=self.meal_type)
                & Q(is_disabled=False)
                & Q(is_ignorable=True)
            ).all()
            total_price_cell = sheet.cell(row_num, 2)
            total_price_cell.value = sum([i.total_price for i in ingredients])

            note_cell = sheet.cell(row_num, 3)

            for cell in [category_cell, total_price_cell, note_cell]:
                cell.font = self.font_12
                cell.alignment = self.center_alignment
                cell.border = self.thin_border

        ingredients = Ingredient.objects.filter(
            Q(user=user)
            & Q(storage_date__gte=self.date_start)
            & Q(storage_date__lte=self.date_end)
            & Q(meal_type=self.meal_type)
            & Q(is_disabled=False)
            & Q(is_ignorable=True)
        ).all()

        summary_row_num = len(categories) + header_row_num + 1
        summary_total_price = sum([i.total_price for i in ingredients])
        summary_total_price_cell = sheet.cell(summary_row_num, 1)
        summary_total_price_cell.border = self.thin_border
        summary_total_price_cell.value = (
            _(
                "Total Price Text: {total_price_text}        {total_price}"
            ).format(
                total_price_text=get_CNY_TEXT(summary_total_price),
                total_price=summary_total_price,
            )
            if self.is_zh_CN
            else _(
                "Total Price Text: {total_price_text}        {total_price}"
            ).format(
                total_price_text=str(summary_total_price),
                total_price=summary_total_price,
            )
        )
        sheet.merge_cells(f"A{summary_row_num}:C{summary_row_num}")
        set_row_height_in_inches(sheet, summary_row_num, 0.32)

        handler_row_num = summary_row_num + 1
        handler_cell = sheet.cell(handler_row_num, 1)
        handler_cell.border = self.thin_border
        handler_cell.value = _("Handler:")
        sheet.merge_cells(f"A{handler_row_num}:C{handler_row_num}")
        set_row_height_in_inches(sheet, handler_row_num, 0.32)

        reviewer_row_num = handler_row_num + 1
        reviewer_cell = sheet.cell(reviewer_row_num, 1)
        reviewer_cell.border = self.thin_border
        reviewer_cell.value = _("Reviewer:")
        sheet.merge_cells(f"A{reviewer_row_num}:C{reviewer_row_num}")
        set_row_height_in_inches(sheet, reviewer_row_num, 0.32)

        supervisor_row_num = reviewer_row_num + 1
        supervisor_cell = sheet.cell(supervisor_row_num, 1)
        supervisor_cell.border = self.thin_border
        supervisor_cell.value = (
            _("Principal's Signature:")
            if self.is_school
            else _("Supervisor's Signature:")
        )
        sheet.merge_cells(f"A{supervisor_row_num}:C{supervisor_row_num}")
        set_row_height_in_inches(sheet, supervisor_row_num, 0.32)

        note_row_num = supervisor_row_num + 1
        note_cell = sheet.cell(note_row_num, 1)
        note_cell.border = self.thin_border
        note_cell.value = (
            _(
                "Note: This form is a summary of all monthly food and "
                + "material inventory receipts from the cafeteria. After "
                + "verification, it will be signed and stamped with "
                + "the school seal by the principal as reimbursement "
                + "evidence."
            )
            if self.is_school
            else _(
                "Note: This form is a summary of all monthly food and "
                + "material inventory receipts from the cafeteria. "
                + "After verification, it will be signed and stamped "
                + "with the affiliation seal by the supervisor as "
                + "reimbursement evidence."
            )
        )
        sheet.merge_cells(f"A{note_row_num}:C{note_row_num}")
        set_row_height_in_inches(sheet, note_row_num, 0.27)

    def fill_in_consumption_sheet(self):
        sheet = self.consumption_sheet
        user = self.user
        title_cell = sheet.cell(1, 1)
        title_cell.value = _(
            "Table of {superior_department} Canteen Consumed Ingredients Statistics"
        ).format(
            superior_department=user.superior_department,
        )
        title_cell.font = self.font_16_bold
        title_cell.alignment = self.center_alignment
        for col_num, width in [
            [1, 2.23],
            [2, 3.14],
            [3, 3.14],
        ]:
            set_column_width_in_inches(sheet, col_num, width)
        sheet.merge_cells("A1:C1")

        sub_title_cell = sheet.cell(2, 1)
        sheet.merge_cells("A2:C2")
        sub_title_cell.font = self.font_12
        sub_title_cell.alignment = self.center_alignment
        sub_title_cell.value = _(
            "Affiliation: {affiliation}        Monetary Unit:         {year}.{month:0>2}.{day:0>2}"
        ).format(
            affiliation=user.affiliation,
            year=self.year,
            month=self.month,
            day=self.date_end.day,
        )

        header_row_num = 3
        header_category_cell = sheet.cell(header_row_num, 1)
        header_category_cell.value = _(
            "Ingredient Categories (Consumption Sheet)"
        )

        header_total_price_cell = sheet.cell(header_row_num, 2)
        header_total_price_cell.value = _(
            "Ingredient Total Prices (Consumption Sheet)"
        )

        header_note_cell = sheet.cell(header_row_num, 3)
        header_note_cell.value = _("Procurement Note (Consumption Sheet)")

        for cell in [
            header_category_cell,
            header_total_price_cell,
            header_note_cell,
        ]:
            cell.font = self.font_12
            cell.alignment = self.center_alignment
            cell.border = self.thin_border

        categories = Category.objects.filter(
            Q(user=user) & Q(is_disabled=False)
        ).all()

        set_row_height_in_inches(sheet, 1, 0.38)
        set_row_height_in_inches(sheet, 2, 0.22)
        set_row_height_in_inches(sheet, 3, 0.32)

        for index, category in enumerate(categories):

            row_num = header_row_num + 1 + index
            set_row_height_in_inches(sheet, row_num, 0.32)

            category_cell = sheet.cell(row_num, 1)
            category_cell.value = category.name

            ingredients = Ingredient.objects.filter(
                Q(user=user)
                & Q(category=category)
                & Q(
                    consumptions__date_of_using__range=(
                        self.date_start,
                        self.date_end,
                    )
                )
                & Q(meal_type=self.meal_type)
                & Q(is_disabled=False)
                & Q(is_ignorable=False)
            ).distinct()
            total_price_cell = sheet.cell(row_num, 2)
            total_price_consumed = Decimal("0.0")
            for i in ingredients:
                consumptions = i.consumptions.filter(is_disabled=False).all()
                total_price_consumed += sum(
                    [c.amount_used * i.unit_price for c in consumptions]
                )
            total_price_cell.value = total_price_consumed

            note_cell = sheet.cell(row_num, 3)

            for cell in [category_cell, total_price_cell, note_cell]:
                cell.font = self.font_12
                cell.alignment = self.center_alignment
                cell.border = self.thin_border

        ingredients = Ingredient.objects.filter(
            Q(user=user)
            & Q(
                consumptions__date_of_using__range=(
                    self.date_start,
                    self.date_end,
                )
            )
            & Q(meal_type=self.meal_type)
            & Q(category__is_disabled=False)
            & Q(is_disabled=False)
            & Q(is_ignorable=False)
        ).all()

        summary_row_num = len(categories) + header_row_num + 1
        summary_total_price = Decimal("0.0")
        for i in ingredients:
            consumptions = i.consumptions.filter(Q(is_disabled=False)).all()
            summary_total_price += sum(
                [c.amount_used * i.unit_price for c in consumptions]
            )
        summary_total_price_cell = sheet.cell(summary_row_num, 1)
        total_price_cell.border = self.thin_border
        summary_total_price_cell.value = (
            _(
                "Total Price Text: {total_price_text}        {total_price}"
            ).format(
                total_price_text=get_CNY_TEXT(summary_total_price),
                total_price=summary_total_price,
            )
            if self.is_zh_CN
            else _(
                "Total Price Text: {total_price_text}        {total_price}"
            ).format(
                total_price_text=str(summary_total_price),
                total_price=summary_total_price,
            )
        )
        sheet.merge_cells(f"A{summary_row_num}:C{summary_row_num}")
        set_row_height_in_inches(sheet, summary_row_num, 0.32)

        handler_row_num = summary_row_num + 1
        handler_cell = sheet.cell(handler_row_num, 1)
        handler_cell.border = self.thin_border
        handler_cell.value = _("Handler:")
        sheet.merge_cells(f"A{handler_row_num}:C{handler_row_num}")
        set_row_height_in_inches(sheet, handler_row_num, 0.32)

        reviewer_row_num = handler_row_num + 1
        reviewer_cell = sheet.cell(reviewer_row_num, 1)
        reviewer_cell.border = self.thin_border
        reviewer_cell.value = _("Reviewer:")
        sheet.merge_cells(f"A{reviewer_row_num}:C{reviewer_row_num}")
        set_row_height_in_inches(sheet, reviewer_row_num, 0.32)

        supervisor_row_num = reviewer_row_num + 1
        supervisor_cell = sheet.cell(supervisor_row_num, 1)
        supervisor_cell.border = self.thin_border
        supervisor_cell.value = (
            _("Principal's Signature:")
            if self.is_school
            else _("Supervisor's Signature:")
        )
        sheet.merge_cells(f"A{supervisor_row_num}:C{supervisor_row_num}")
        set_row_height_in_inches(sheet, supervisor_row_num, 0.32)

        note_row_num = supervisor_row_num + 1
        note_cell = sheet.cell(note_row_num, 1)
        note_cell.border = self.thin_border
        note_cell.value = (
            _(
                "Note: This form is a summary of all monthly food and "
                + "material consumption receipts from the cafeteria. After "
                + "verification, it will be signed and stamped with "
                + "the school seal by the principal as reimbursement "
                + "evidence."
            )
            if self.is_school
            else _(
                "Note: This form is a summary of all monthly food and "
                + "material consumption receipts from the cafeteria. "
                + "After verification, it will be signed and stamped "
                + "with the affiliation seal by the supervisor as "
                + "reimbursement evidence."
            )
        )
        sheet.merge_cells(f"A{note_row_num}:C{note_row_num}")
        set_row_height_in_inches(sheet, note_row_num, 0.27)

    def fill_in_storage_sheet(self):
        sheet = self.storage_sheet
        user = self.user
        title_cell = sheet.cell(1, 1)
        title_cell.value = _(
            "Table of {superior_department} Canteen Storaged Ingredients Statistics"
        ).format(
            superior_department=user.superior_department,
        )
        title_cell.font = self.font_16_bold
        title_cell.alignment = self.center_alignment
        for col_num, width in [
            [1, 2.23],
            [2, 3.14],
            [3, 3.14],
        ]:
            set_column_width_in_inches(sheet, col_num, width)
        sheet.merge_cells("A1:C1")

        sub_title_cell = sheet.cell(2, 1)
        sheet.merge_cells("A2:C2")
        sub_title_cell.font = self.font_12
        sub_title_cell.alignment = self.center_alignment
        sub_title_cell.value = _(
            "Affiliation: {affiliation}        Monetary Unit:         {year}.{month:0>2}.{day:0>2}"
        ).format(
            affiliation=user.affiliation,
            year=self.year,
            month=self.month,
            day=self.date_end.day,
        )

        header_row_num = 3
        header_category_cell = sheet.cell(header_row_num, 1)
        header_category_cell.value = _("Ingredient Categories (storage sheet)")

        header_total_price_cell = sheet.cell(header_row_num, 2)
        header_total_price_cell.value = _(
            "Ingredient Total Prices (storage sheet)"
        )

        header_note_cell = sheet.cell(header_row_num, 3)
        header_note_cell.value = _("Procurement Note")

        for cell in [
            header_category_cell,
            header_total_price_cell,
            header_note_cell,
        ]:
            cell.font = self.font_12
            cell.alignment = self.center_alignment
            cell.border = self.thin_border

        categories = Category.objects.filter(
            Q(user=user) & Q(is_disabled=False)
        ).all()

        set_row_height_in_inches(sheet, 1, 0.38)
        set_row_height_in_inches(sheet, 2, 0.22)
        set_row_height_in_inches(sheet, 3, 0.32)

        for index, category in enumerate(categories):

            row_num = header_row_num + 1 + index
            set_row_height_in_inches(sheet, row_num, 0.32)

            category_cell = sheet.cell(row_num, 1)
            category_cell.value = category.name

            ingredients = Ingredient.objects.filter(
                Q(user=user)
                & Q(category=category)
                & Q(storage_date__gte=self.date_start)
                & Q(storage_date__lte=self.date_end)
                & Q(meal_type=self.meal_type)
                & Q(is_disabled=False)
                & Q(is_ignorable=False)
            ).all()
            total_price_cell = sheet.cell(row_num, 2)
            total_price_cell.value = sum([i.total_price for i in ingredients])

            note_cell = sheet.cell(row_num, 3)

            for cell in [category_cell, total_price_cell, note_cell]:
                cell.font = self.font_12
                cell.alignment = self.center_alignment
                cell.border = self.thin_border

        ingredients = Ingredient.objects.filter(
            Q(user=user)
            & Q(storage_date__gte=self.date_start)
            & Q(storage_date__lte=self.date_end)
            & Q(meal_type=self.meal_type)
            & Q(is_disabled=False)
            & Q(is_ignorable=False)
        ).all()

        summary_row_num = len(categories) + header_row_num + 1
        summary_total_price = sum([i.total_price for i in ingredients])
        summary_total_price_cell = sheet.cell(summary_row_num, 1)
        total_price_cell.border = self.thin_border
        summary_total_price_cell.value = (
            _(
                "Total Price Text: {total_price_text}        {total_price}"
            ).format(
                total_price_text=get_CNY_TEXT(summary_total_price),
                total_price=summary_total_price,
            )
            if self.is_zh_CN
            else _(
                "Total Price Text: {total_price_text}        {total_price}"
            ).format(
                total_price_text=str(summary_total_price),
                total_price=summary_total_price,
            )
        )
        sheet.merge_cells(f"A{summary_row_num}:C{summary_row_num}")
        set_row_height_in_inches(sheet, summary_row_num, 0.32)

        handler_row_num = summary_row_num + 1
        handler_cell = sheet.cell(handler_row_num, 1)
        handler_cell.border = self.thin_border
        handler_cell.value = _("Handler:")
        sheet.merge_cells(f"A{handler_row_num}:C{handler_row_num}")
        set_row_height_in_inches(sheet, handler_row_num, 0.32)

        reviewer_row_num = handler_row_num + 1
        reviewer_cell = sheet.cell(reviewer_row_num, 1)
        reviewer_cell.border = self.thin_border
        reviewer_cell.value = _("Reviewer:")
        sheet.merge_cells(f"A{reviewer_row_num}:C{reviewer_row_num}")
        set_row_height_in_inches(sheet, reviewer_row_num, 0.32)

        supervisor_row_num = reviewer_row_num + 1
        supervisor_cell = sheet.cell(supervisor_row_num, 1)
        supervisor_cell.border = self.thin_border
        supervisor_cell.value = (
            _("Principal's Signature:")
            if self.is_school
            else _("Supervisor's Signature:")
        )
        sheet.merge_cells(f"A{supervisor_row_num}:C{supervisor_row_num}")
        set_row_height_in_inches(sheet, supervisor_row_num, 0.32)

        note_row_num = supervisor_row_num + 1
        note_cell = sheet.cell(note_row_num, 1)
        note_cell.border = self.thin_border
        note_cell.value = (
            _(
                "Note: This form is a summary of all monthly food and "
                + "material inventory receipts from the cafeteria. After "
                + "verification, it will be signed and stamped with "
                + "the school seal by the principal as reimbursement "
                + "evidence."
            )
            if self.is_school
            else _(
                "Note: This form is a summary of all monthly food and "
                + "material inventory receipts from the cafeteria. "
                + "After verification, it will be signed and stamped "
                + "with the affiliation seal by the supervisor as "
                + "reimbursement evidence."
            )
        )
        sheet.merge_cells(f"A{note_row_num}:C{note_row_num}")
        set_row_height_in_inches(sheet, note_row_num, 0.27)

    def fill_in_consumption_list_sheet(self):
        sheet = self.consumption_list_sheet
        user = self.user
        consumption_rows_count = 21
        categories = Category.objects.filter(
            Q(user=user) & Q(is_disabled=False)
        ).all()
        consumptions = Consumption.objects.filter(
            Q(is_disabled=False)
            & Q(ingredient__meal_type=self.meal_type)
            & Q(ingredient__user=user)
            & Q(date_of_using__gte=self.date_start)
            & Q(date_of_using__lte=self.date_end)
        ).all()
        consumption_row_height = 0.18
        consumption_rows_height = consumption_rows_count * 0.18

        consumption_dates = list(set([c.date_of_using for c in consumptions]))
        consumption_dates = sorted(consumption_dates)

        formed_consumptions = []
        for consumption_date in consumption_dates:
            dated_consumptions = [
                c for c in consumptions if c.date_of_using == consumption_date
            ]
            dated_consumptions = sorted(
                dated_consumptions,
                key=lambda i: (i.ingredient.category.name, i.ingredient.name),
            )

            dated_consumption_categories = list(
                set([c.ingredient.category for c in dated_consumptions])
            )
            empty_categories = [
                c for c in categories if not c in dated_consumption_categories
            ]
            same_date_count = math.ceil(
                len(dated_consumptions)
                / (consumption_rows_count - len(empty_categories))
            )
            step = math.floor(len(dated_consumptions) / same_date_count)
            sub_consumption_num = 1
            for index in range(0, len(dated_consumptions), step):
                split_dated_consumptions = dated_consumptions[
                    index : index + step
                ]
                split_dated_consumption = split_dated_consumptions[0]
                split_dated_consumption_categories = list(
                    set(
                        [
                            c.ingredient.category
                            for c in split_dated_consumptions
                        ]
                    )
                )
                split_empty_categories = [
                    c
                    for c in split_dated_consumption_categories
                    if not c in categories
                ]
                split_empty_categories = split_empty_categories + [
                    random.choice(categories)
                    for i in range(
                        consumption_rows_count
                        - len(split_dated_consumptions)
                        - len(split_empty_categories)
                    )
                ]

                fake_ingredients = [
                    Ingredient(
                        user=user,
                        name="",
                        storage_date=self.date_start,
                        meal_type=split_dated_consumption.ingredient.meal_type,
                        category=c,
                        quantity=0.0,
                        quantity_unit_name="",
                        total_price=0.0,
                        is_ignorable=False,
                    )
                    for c in split_empty_categories
                ]

                for fake_ingredient in fake_ingredients:
                    split_dated_consumptions.append(
                        Consumption(
                            ingredient=fake_ingredient,
                            date_of_using=consumption_date,
                            amount_used=Decimal("0"),
                            is_disabled=False,
                        )
                    )

                split_dated_consumptions = sorted(
                    split_dated_consumptions,
                    key=lambda c: (
                        c.ingredient.category.priority,
                        c.ingredient.category.name,
                    ),
                )

                consumption_date_index = sub_consumption_num
                formed_consumptions.append(
                    [
                        consumption_date,
                        consumption_date_index,
                        split_dated_consumptions,
                    ]
                )

                sub_consumption_num += 1

        consumption_num = 0
        for index, (
            consumption_date,
            consumption_date_index,
            dated_consumptions,
        ) in enumerate(formed_consumptions):
            row_num = (consumption_rows_count + 6) * index + 1

            title_cell_row_num = row_num
            title_cell = sheet.cell(title_cell_row_num, 1)
            title_cell.value = _("Consumption List Title (Consumption List Sheet)")
            title_cell.alignment = self.center_alignment
            title_cell.font = self.font_20_bold
            sheet.merge_cells(f"A{title_cell_row_num}:H{title_cell_row_num}")

            sub_title_affiliation_cell_row_num = title_cell_row_num + 1
            sub_title_affiliation_cell = sheet.cell(
                sub_title_affiliation_cell_row_num, 1
            )
            sub_title_affiliation_cell.font = self.font_12
            sub_title_affiliation_cell.value = (
                _("Affiliation Name: {affiliation_name}")
                if self.is_school
                else _("Principal Name: {affiliation_name}")
            ).format(affiliation_name=user.affiliation)

            sub_title_date_and_unit_cell_row_num = title_cell_row_num + 1
            sub_title_date_and_unit_cell = sheet.cell(
                sub_title_date_and_unit_cell_row_num, 4
            )
            sub_title_date_and_unit_cell.font = self.font_12
            sub_title_date_and_unit_cell.value = (
                _("{day:0>2} {month:0>2} {year}  Quantity Unit Name: CNY")
            ).format(year=self.year, month=self.month, day=consumption_date.day)

            if consumption_date_index < 2:
                consumption_num += 1

            prev_consumption_date = (
                formed_consumptions[index - 1][0]
                if 0 <= index - 1 < len(formed_consumptions)
                else None
            )
            next_consumption_date = (
                formed_consumptions[index + 1][0]
                if 0 < index + 1 < (len(formed_consumptions))
                else None
            )
            sub_title_num_cell_row_num = title_cell_row_num + 1
            sub_title_num_cell = sheet.cell(sub_title_num_cell_row_num, 7)
            sub_title_num_cell.font = self.font_12
            sub_title_num_cell.value = _(
                "Storage No. {consumption_num}"
            ).format(
                consumption_num=(
                    f"C{self.month:0>2}{consumption_num:0>2}"
                    if self.is_zh_CN
                    else f"C{self.month:0>2}{consumption_num:0>2}"
                )
            ) + (
                _("(Sub Storage No. {sub_consumption_num})").format(
                    sub_consumption_num=consumption_date_index
                )
                if (
                    (
                        next_consumption_date
                        and next_consumption_date == consumption_date
                    )
                    or (
                        prev_consumption_date
                        and prev_consumption_date == consumption_date
                    )
                )
                else ""
            )

            sheet.merge_cells(
                f"A{sub_title_affiliation_cell_row_num}:B{sub_title_affiliation_cell_row_num}"
            )
            sheet.merge_cells(
                f"D{sub_title_date_and_unit_cell_row_num}:F{sub_title_date_and_unit_cell_row_num}"
            )
            sheet.merge_cells(
                f"G{sub_title_num_cell_row_num}:H{sub_title_num_cell_row_num}"
            )

            font_12_cells = []
            header_row_num = title_cell_row_num + 2
            category_header_cell = sheet.cell(header_row_num, 1)
            category_header_cell.value = _("Category (Consumption List Sheet)")

            ingredient_name_header_cell = sheet.cell(header_row_num, 2)
            ingredient_name_header_cell.value = _(
                "Ingredient Name (Consumption List Sheet)"
            )

            quantity_unit_name_header_cell = sheet.cell(header_row_num, 3)
            quantity_unit_name_header_cell.value = _(
                "Quantity Unit Name (Consumption List Sheet)"
            )

            quantity_header_cell = sheet.cell(header_row_num, 4)
            quantity_header_cell.value = _("Quantity (Consumption List Sheet)")

            unit_price_header_cell = sheet.cell(header_row_num, 5)
            unit_price_header_cell.value = _(
                "Unit Price (Consumption List Sheet)"
            )

            total_price_header_cell = sheet.cell(header_row_num, 6)
            total_price_header_cell.value = _(
                "Total Price (Consumption List Sheet)"
            )

            ingredients_total_price_header_cell = sheet.cell(header_row_num, 7)
            ingredients_total_price_header_cell.value = _(
                "Ingredients Total Price (Consumption List Sheet)"
            )

            note_header_cell = sheet.cell(header_row_num, 8)
            note_header_cell.value = _("Note (Consumption List Sheet)")

            font_12_cells += [
                category_header_cell,
                ingredient_name_header_cell,
                quantity_unit_name_header_cell,
                quantity_header_cell,
                unit_price_header_cell,
                total_price_header_cell,
                ingredients_total_price_header_cell,
                note_header_cell,
            ]
            for cell in font_12_cells:
                cell.alignment = self.center_alignment
                cell.font = self.font_12
                cell.border = self.thin_border

            last_category = None

            first_consumption_row_num = header_row_num + 1
            for row_num in range(
                first_consumption_row_num,
                first_consumption_row_num + len(dated_consumptions) + 1,
            ):
                for col_num in range(1, 9):
                    cell = sheet.cell(row_num, col_num)
                    cell.font = self.font_12
                    cell.alignment = self.center_alignment
                    cell.border = self.thin_border

            for index, consumption in enumerate(dated_consumptions):
                consumption_row_num = header_row_num + 1 + index

                if not consumption.ingredient.category == last_category:
                    sheet.cell(
                        consumption_row_num,
                        1,
                        consumption.ingredient.category.name,
                    )
                    sheet.cell(
                        consumption_row_num,
                        7,
                        sum(
                            [
                                c.ingredient.unit_price * c.amount_used
                                for c in dated_consumptions
                                if c.ingredient.category
                                == consumption.ingredient.category
                            ]
                        )
                        or "",
                    )
                    consumptions_same_category_len = len(
                        [
                            c
                            for c in dated_consumptions
                            if c.ingredient.category
                            == consumption.ingredient.category
                        ]
                    )
                    sheet.merge_cells(
                        f"A{consumption_row_num}:A{consumptions_same_category_len+consumption_row_num-1}"
                    )
                    sheet.merge_cells(
                        f"G{consumption_row_num}:G{consumptions_same_category_len+consumption_row_num-1}"
                    )
                    last_category = consumption.ingredient.category

                sheet.cell(consumption_row_num, 2, consumption.ingredient.name)
                sheet.cell(
                    consumption_row_num,
                    3,
                    consumption.ingredient.quantity_unit_name,
                )
                sheet.cell(
                    consumption_row_num,
                    4,
                    (
                        consumption.ingredient.quantity
                        if consumption.ingredient.quantity
                        else ""
                    ),
                )
                sheet.cell(
                    consumption_row_num,
                    5,
                    (
                        consumption.ingredient.unit_price
                        if consumption.ingredient.unit_price
                        else ""
                    ),
                )
                sheet.cell(
                    consumption_row_num,
                    6,
                    (
                        consumption.ingredient.unit_price
                        * consumption.amount_used
                        if consumption.ingredient.unit_price
                        and consumption.amount_used
                        else ""
                    ),
                )
                set_row_height_in_inches(
                    sheet, consumption_row_num, consumption_row_height
                )

            summary_total_price = sum(
                [
                    c.ingredient.unit_price * c.amount_used
                    for c in dated_consumptions
                ]
            )
            summary_row_num = header_row_num + len(dated_consumptions) + 1
            sheet.cell(summary_row_num, 1, _("Summary (Storage List Sheet)"))
            sheet.cell(summary_row_num, 6, summary_total_price)
            sheet.cell(summary_row_num, 7, summary_total_price)

            summary_row_height = consumption_row_height
            set_row_height_in_inches(sheet, summary_row_num, summary_row_height)

            signature_row_num = summary_row_num + 1
            signature_cell = sheet.cell(signature_row_num, 1)
            signature_cell.value = _(
                "   Reviewer:        Handler:{handler} 　    Weigher:      Warehouseman: 　"
            ).format(handler=user.username)
            signature_cell.font = self.font_14
            signature_cell.alignment = self.center_alignment
            sheet.merge_cells(f"A{signature_row_num}:H{signature_row_num}")
            set_row_height_in_inches(sheet, signature_row_num, 0.22)

        for col_num, col_width in [
            [1, 1.13],
            [2, 1.98],
            [3, 0.85],
            [4, 0.85],
            [5, 0.88],
            [6, 1.28],
            [7, 1.19],
            [8, 0.78],
        ]:
            set_column_width_in_inches(sheet, col_num, col_width)

    def fill_in_storage_list_sheet(self):
        sheet = self.storage_list_sheet
        user = self.user
        ingredient_rows_count = 21
        ingredients = Ingredient.objects.filter(
            Q(user=user)
            & Q(is_disabled=False)
            & Q(is_ignorable=False)
            & Q(storage_date__gte=self.date_start)
            & Q(storage_date__lte=self.date_end)
            & Q(meal_type=self.meal_type)
        ).all()
        categories = Category.objects.filter(
            Q(user=user) & Q(is_disabled=False)
        ).all()
        ingredient_row_height = 0.18
        ingredient_rows_height = ingredient_rows_count * 0.18
        storage_dates = sorted(list(set([i.storage_date for i in ingredients])))

        storaged_ingredients = []
        for storage_date in storage_dates:
            dated_ingredients = [
                i for i in ingredients if i.storage_date == storage_date
            ]
            dated_ingredients = sorted(
                dated_ingredients, key=lambda i: (i.category.name, i.name)
            )
            dated_ingredient_categories = list(
                set([i.category for i in ingredients])
            )
            empty_categories = [
                c for c in categories if not c in dated_ingredient_categories
            ]
            same_date_count = math.ceil(
                len(dated_ingredients)
                / (ingredient_rows_count - len(empty_categories))
            )
            step = math.floor(len(dated_ingredients) / same_date_count)
            sub_storage_num = 1
            for index in range(0, len(dated_ingredients), step):
                split_dated_ingredients = dated_ingredients[
                    index : index + step
                ]
                split_dated_ingredient = split_dated_ingredients[0]
                split_dated_ingredient_categories = list(
                    set([i.category for i in split_dated_ingredients])
                )
                split_empty_categories = [
                    c
                    for c in split_dated_ingredient_categories
                    if not c in categories
                ]
                split_empty_categories = split_empty_categories + [
                    random.choice(categories)
                    for i in range(
                        ingredient_rows_count
                        - len(split_dated_ingredients)
                        - len(split_empty_categories)
                    )
                ]

                split_dated_ingredients += [
                    Ingredient(
                        user=user,
                        name="",
                        storage_date=storage_date,
                        meal_type=split_dated_ingredient.meal_type,
                        category=c,
                        quantity=0.0,
                        quantity_unit_name="",
                        total_price=0.0,
                        is_ignorable=False,
                    )
                    for c in split_empty_categories
                ]

                split_dated_ingredients = sorted(
                    split_dated_ingredients,
                    key=lambda i: (i.category.priority, i.category.name),
                )

                storage_date_index = sub_storage_num
                storaged_ingredients.append(
                    [storage_date, storage_date_index, split_dated_ingredients]
                )

                sub_storage_num += 1

        storage_num = 0
        for index, (
            storage_date,
            storage_date_index,
            dated_ingredients,
        ) in enumerate(storaged_ingredients):
            row_num = (ingredient_rows_count + 6) * index + 1

            title_cell_row_num = row_num
            title_cell = sheet.cell(title_cell_row_num, 1)
            title_cell.value = _("Storage List (Storage Sheet)")
            title_cell.alignment = self.center_alignment
            title_cell.font = self.font_20_bold
            sheet.merge_cells(f"A{title_cell_row_num}:H{title_cell_row_num}")

            sub_title_affiliation_cell_row_num = title_cell_row_num + 1
            sub_title_affiliation_cell = sheet.cell(
                sub_title_affiliation_cell_row_num, 1
            )
            sub_title_affiliation_cell.font = self.font_12
            sub_title_affiliation_cell.value = (
                _("Affiliation Name: {affiliation_name}")
                if self.is_school
                else _("Principal Name: {affiliation_name}")
            ).format(affiliation_name=user.affiliation)

            sub_title_date_and_unit_cell_row_num = title_cell_row_num + 1
            sub_title_date_and_unit_cell = sheet.cell(
                sub_title_date_and_unit_cell_row_num, 4
            )
            sub_title_date_and_unit_cell.font = self.font_12
            sub_title_date_and_unit_cell.value = (
                _("{day:0>2} {month:0>2} {year}  Quantity Unit Name: CNY")
            ).format(year=self.year, month=self.month, day=storage_date.day)

            if storage_date_index < 2:
                storage_num += 1

            prev_storage_date = (
                storaged_ingredients[index - 1][0]
                if 0 <= index - 1 < len(storaged_ingredients)
                else None
            )
            next_storage_date = (
                storaged_ingredients[index + 1][0]
                if 0 < index + 1 < (len(storaged_ingredients) - 1)
                else None
            )
            sub_title_num_cell_row_num = title_cell_row_num + 1
            sub_title_num_cell = sheet.cell(sub_title_num_cell_row_num, 7)
            sub_title_num_cell.font = self.font_12
            sub_title_num_cell.value = _("Storage No. {storage_num}").format(
                storage_num=(
                    f"R{self.month:0>2}{storage_num:0>2}"
                    if self.is_zh_CN
                    else f"S{self.month:0>2}{storage_num:0>2}"
                )
            ) + (
                _("(Sub Storage No. {sub_storage_num})").format(
                    sub_storage_num=storage_date_index
                )
                if (
                    (next_storage_date and next_storage_date == storage_date)
                    or (prev_storage_date and prev_storage_date == storage_date)
                )
                else ""
            )

            sheet.merge_cells(
                f"A{sub_title_affiliation_cell_row_num}:B{sub_title_affiliation_cell_row_num}"
            )
            sheet.merge_cells(
                f"D{sub_title_date_and_unit_cell_row_num}:F{sub_title_date_and_unit_cell_row_num}"
            )
            sheet.merge_cells(
                f"G{sub_title_num_cell_row_num}:H{sub_title_num_cell_row_num}"
            )

            font_12_cells = []
            header_row_num = title_cell_row_num + 2
            category_header_cell = sheet.cell(header_row_num, 1)
            category_header_cell.value = _("Category (Storage List Sheet)")

            ingredient_name_header_cell = sheet.cell(header_row_num, 2)
            ingredient_name_header_cell.value = _(
                "Ingredient Name (Storage List Sheet)"
            )

            quantity_unit_name_header_cell = sheet.cell(header_row_num, 3)
            quantity_unit_name_header_cell.value = _(
                "Quantity Unit Name (Storage List Sheet)"
            )

            quantity_header_cell = sheet.cell(header_row_num, 4)
            quantity_header_cell.value = _("Quantity (Storage List Sheet)")

            unit_price_header_cell = sheet.cell(header_row_num, 5)
            unit_price_header_cell.value = _("Unit Price (Storage List Sheet)")

            total_price_header_cell = sheet.cell(header_row_num, 6)
            total_price_header_cell.value = _(
                "Total Price (Storage List Sheet)"
            )

            ingredients_total_price_header_cell = sheet.cell(header_row_num, 7)
            ingredients_total_price_header_cell.value = _(
                "Ingredients Total Price (Storage List Sheet)"
            )

            note_header_cell = sheet.cell(header_row_num, 8)
            note_header_cell.value = _("Note (Storage List Sheet)")

            font_12_cells += [
                category_header_cell,
                ingredient_name_header_cell,
                quantity_unit_name_header_cell,
                quantity_header_cell,
                unit_price_header_cell,
                total_price_header_cell,
                ingredients_total_price_header_cell,
                note_header_cell,
            ]
            for cell in font_12_cells:
                cell.alignment = self.center_alignment
                cell.font = self.font_12
                cell.border = self.thin_border

            last_category = None

            first_ingredient_row_num = header_row_num + 1
            for row_num in range(
                first_ingredient_row_num,
                first_ingredient_row_num + len(dated_ingredients) + 1,
            ):
                for col_num in range(1, 9):
                    cell = sheet.cell(row_num, col_num)
                    cell.font = self.font_12
                    cell.alignment = self.center_alignment
                    cell.border = self.thin_border

            for index, ingredient in enumerate(dated_ingredients):
                ingredient_row_num = header_row_num + 1 + index

                if not ingredient.category == last_category:
                    sheet.cell(ingredient_row_num, 1, ingredient.category.name)
                    sheet.cell(
                        ingredient_row_num,
                        7,
                        sum(
                            [
                                float(i.total_price)
                                for i in dated_ingredients
                                if i.category == ingredient.category
                            ]
                        )
                        or "",
                    )
                    ingredients_same_category_len = len(
                        [
                            i
                            for i in dated_ingredients
                            if i.category == ingredient.category
                        ]
                    )
                    sheet.merge_cells(
                        f"A{ingredient_row_num}:A{ingredients_same_category_len+ingredient_row_num-1}"
                    )
                    sheet.merge_cells(
                        f"G{ingredient_row_num}:G{ingredients_same_category_len+ingredient_row_num-1}"
                    )
                    last_category = ingredient.category

                sheet.cell(ingredient_row_num, 2, ingredient.name)
                sheet.cell(ingredient_row_num, 3, ingredient.quantity_unit_name)
                sheet.cell(
                    ingredient_row_num,
                    4,
                    f"{ingredient.quantity:.2f}" if ingredient.quantity else "",
                )
                sheet.cell(
                    ingredient_row_num,
                    5,
                    (
                        f"{ingredient.unit_price:.2f}"
                        if ingredient.unit_price
                        else ""
                    ),
                )
                sheet.cell(
                    ingredient_row_num,
                    6,
                    (
                        f"{ingredient.total_price:.2f}"
                        if ingredient.total_price
                        else ""
                    ),
                )
                set_row_height_in_inches(
                    sheet, ingredient_row_num, ingredient_row_height
                )

            summary_total_price = sum(
                [float(i.total_price) for i in dated_ingredients]
            )
            summary_row_num = header_row_num + len(dated_ingredients) + 1
            sheet.cell(summary_row_num, 1, _("Summary (Storage List Sheet)"))
            sheet.cell(summary_row_num, 6, summary_total_price)
            sheet.cell(summary_row_num, 7, summary_total_price)

            summary_row_height = ingredient_row_height
            set_row_height_in_inches(sheet, summary_row_num, summary_row_height)

            signature_row_num = summary_row_num + 1
            signature_cell = sheet.cell(signature_row_num, 1)
            signature_cell.value = _(
                "   Reviewer:        Handler:{handler} 　    Weigher:      Warehouseman: 　"
            ).format(handler=user.username)
            signature_cell.font = self.font_14
            signature_cell.alignment = self.center_alignment
            sheet.merge_cells(f"A{signature_row_num}:H{signature_row_num}")
            set_row_height_in_inches(sheet, signature_row_num, 0.22)

        for col_num, col_width in [
            [1, 1.13],
            [2, 1.98],
            [3, 0.85],
            [4, 0.85],
            [5, 0.88],
            [6, 1.28],
            [7, 1.19],
            [8, 0.78],
        ]:
            set_column_width_in_inches(sheet, col_num, col_width)

    def fill_in_cover_sheet(self):
        sheet = self.cover_sheet
        user = self.user
        title_cell = sheet.cell(1, 1)
        title_cell.value = _(
            "Table of {affiliation} Canteen Ingredients Procurement Statistics in {month:0>2} {year}"
        ).format(
            affiliation=user.affiliation,
            year=self.year,
            month=self.month,
        )
        title_cell.font = self.font_16_bold
        title_cell.alignment = self.center_alignment
        for col_num, width in [
            [1, 1.96],
            [2, 2.26],
            [3, 4.44],
        ]:
            set_column_width_in_inches(sheet, col_num, width)
        sheet.merge_cells("A1:C1")

        header_row_num = 2
        header_category_cell = sheet.cell(header_row_num, 1)
        header_category_cell.value = _("Ingredient Categories (cover sheet)")

        header_total_price_cell = sheet.cell(header_row_num, 2)
        header_total_price_cell.value = _("Ingredient Total Prices")

        header_note_cell = sheet.cell(header_row_num, 3)
        header_note_cell.value = _("Procurement Note")

        for cell in [
            header_category_cell,
            header_total_price_cell,
            header_note_cell,
        ]:
            cell.alignment = self.center_alignment
            cell.border = self.thin_border
            cell.font = self.font_12_bold

        categories = Category.objects.filter(
            Q(user=user) & Q(is_disabled=False)
        ).all()

        set_row_height_in_inches(sheet, 1, 0.60)
        set_row_height_in_inches(sheet, 2, 0.44)

        for index, category in enumerate(categories):

            row_num = header_row_num + 1 + index
            set_row_height_in_inches(sheet, row_num, 0.44)

            category_cell = sheet.cell(row_num, 1)
            category_cell.value = category.name

            ingredients = Ingredient.objects.filter(
                Q(user=user)
                & Q(category=category)
                & Q(storage_date__gte=self.date_start)
                & Q(storage_date__lte=self.date_end)
                & Q(meal_type=self.meal_type)
                & Q(is_disabled=False)
            ).all()
            total_price_cell = sheet.cell(row_num, 2)
            total_price_cell.value = sum([i.total_price for i in ingredients])

            note_cell = sheet.cell(row_num, 3)
            note_cell.value = _(
                "Total price of storaged ingredients is {0}, total price of non-storaged ingredients is {1}."
            ).format(
                sum([i.total_price for i in ingredients if not i.is_ignorable]),
                sum([i.total_price for i in ingredients if i.is_ignorable]),
            )

            for cell in [category_cell, total_price_cell, note_cell]:
                cell.font = self.font_12
                cell.alignment = self.center_alignment
                cell.border = self.thin_border

        ingredients = Ingredient.objects.filter(
            Q(user=user)
            & Q(storage_date__gte=self.date_start)
            & Q(storage_date__lte=self.date_end)
            & Q(meal_type=self.meal_type)
            & Q(is_disabled=False)
        ).all()

        summary_row_num = len(categories) + header_row_num + 1
        summary_index_cell = sheet.cell(summary_row_num, 1)
        summary_index_cell.value = _("Procurement Summary")

        summary_total_price_cell = sheet.cell(summary_row_num, 2)
        summary_total_price_cell.value = sum(
            [i.total_price for i in ingredients]
        )

        summary_note_cell = sheet.cell(summary_row_num, 3)
        summary_note_cell.value = _(
            "Total price of storaged ingredients is {0}, total price of non-storaged ingredients is {1}."
        ).format(
            sum([i.total_price for i in ingredients if not i.is_ignorable]),
            sum([i.total_price for i in ingredients if i.is_ignorable]),
        )

        for cell in [
            summary_index_cell,
            summary_total_price_cell,
            summary_note_cell,
        ]:
            cell.font = self.font_12_bold
            cell.alignment = self.center_alignment
            cell.border = self.thin_border

        set_row_height_in_inches(sheet, summary_row_num, 0.44)

    def fill_in_surplus_sheet(self):
        sheet = self.surplus_sheet
        user = self.user
        ingredient_rows_count = 17

        ingredients = Ingredient.objects.filter(
            Q(user=user)
            & Q(storage_date__lte=self.date_end)
            & Q(meal_type=self.meal_type)
            & Q(is_disabled=False)
            & Q(is_ignorable=False)
        ).all()

        consumptions = []
        for ingredient in ingredients:
            consumptions += [
                c
                for c in ingredient.consumptions.filter(
                    Q(date_of_using__lte=self.date_end) & Q(is_disabled=False)
                ).all()
            ]

        inventory_days = []
        dates_of_using = sorted(
            list(set([c.date_of_using for c in consumptions]))
        )
        for i, date_of_using in enumerate(dates_of_using):
            if i + 1 < len(dates_of_using):
                if (date_of_using + timedelta(days=1)) < dates_of_using[i + 1]:
                    inventory_days.append(date_of_using)
            else:
                inventory_days.append(date_of_using)

        if len(inventory_days) < 1:
            inventory_days.insert(-1, self.date_end)

        inventory_days.insert(
            0, (self.date_start.replace(day=1) - timedelta(days=1))
        )
        formed_ingredients = []
        for inventory_day in inventory_days:
            inventory_day_ingredients = []
            for ingredient in ingredients:
                if (
                    ingredient.storage_date > inventory_day
                    and inventory_day != inventory_days[-1]
                ):
                    continue
                remaining_quantity = ingredient.quantity - sum(
                    [
                        c.amount_used
                        for c in ingredient.consumptions.filter(
                            Q(is_disabled=False)
                            & Q(date_of_using__lte=inventory_day)
                        ).all()
                    ]
                )
                if remaining_quantity > Decimal("0.0"):
                    inventory_day_ingredients.append(ingredient)
            form_count = len(inventory_day_ingredients) / ingredient_rows_count
            surplus_ingredients_len = (
                len(inventory_day_ingredients) % ingredient_rows_count
            )
            fake_ingredients_len = (
                ingredient_rows_count - surplus_ingredients_len
            )

            inventory_day_ingredients = sorted(
                inventory_day_ingredients, key=lambda i: i.category.name
            )

            if len(inventory_day_ingredients) < 1:
                continue
            s_ingredient0 = inventory_day_ingredients[0]
            inventory_day_ingredients += [
                Ingredient(
                    user=user,
                    storage_date=self.date_start,
                    name="",
                    meal_type=self.meal_type,
                    category=s_ingredient0.category,
                    quantity=Decimal("0"),
                    quantity_unit_name="",
                    total_price=Decimal("0.0"),
                    is_ignorable=False,
                    is_disabled=False,
                )
                for i in range(fake_ingredients_len)
            ]
            for index in range(
                0, len(inventory_day_ingredients), ingredient_rows_count
            ):
                split_ingredients = inventory_day_ingredients[
                    index : index + ingredient_rows_count
                ]

                formed_ingredients.append(
                    [inventory_day, index, split_ingredients]
                )

        for index, (
            inventory_day,
            inventory_day_index,
            ingredients,
        ) in enumerate(formed_ingredients):

            title_row_num = (ingredient_rows_count + 8) * index + 1
            title_cell = sheet.cell(title_row_num, 1)
            title_cell.value = _("Table of Surplus Ingredients")
            title_cell.font = self.font_20_bold
            title_cell.alignment = self.center_alignment

            set_row_height_in_inches(sheet, title_row_num, 0.31)
            sheet.merge_cells(f"A{title_row_num}:I{title_row_num}")

            sub_title_row_num = title_row_num + 1
            sub_title_affiliation_date_cell = sheet.cell(sub_title_row_num, 1)
            sub_title_affiliation_date_cell.value = (
                _(
                    "Principal Name: {affiliation}                {year}.{month:0>2}.{day:0>2} (Sub-title of Surplus Sheet)"
                )
                if self.is_school
                else _(
                    "Affiliation Name: {affiliation}                {year}.{month:0>2}.{day:0>2} (Sub-title of Surplus Sheet)"
                )
            ).format(
                affiliation=user.affiliation,
                year=inventory_day.year,
                month=inventory_day.month,
                day=inventory_day.day,
            )

            sub_title_affiliation_date_cell.font = self.font_12
            sub_title_affiliation_date_cell.alignment = self.center_alignment
            sheet.merge_cells(f"A{sub_title_row_num}:I{sub_title_row_num}")

            set_row_height_in_inches(sheet, sub_title_row_num, 0.20)

            __style_row_end = sub_title_row_num + ingredient_rows_count + 5 + 1
            for row in range(sub_title_row_num, __style_row_end):
                set_row_height_in_inches(sheet, row, 0.20)
                for col in range(1, 10):
                    cell = sheet.cell(row, col)
                    cell.font = self.font_12
                    cell.alignment = self.center_alignment
                    if row < (__style_row_end - 1):
                        cell.border = self.thin_border

            header0_row_num = sub_title_row_num + 1

            header0_ingredient_name_cell = sheet.cell(header0_row_num, 1)
            header0_ingredient_name_cell.value = _(
                "Ingredient Name (Surplus Sheet)"
            )
            sheet.merge_cells(f"A{header0_row_num}:A{header0_row_num+1}")
            header0_quantity_unit_name_cell = sheet.cell(header0_row_num, 2)
            header0_quantity_unit_name_cell.value = _(
                "Ingredient Quantity Unit Name (Surplus Sheet)"
            )
            sheet.merge_cells(f"B{header0_row_num}:B{header0_row_num+1}")
            header0_recorded_cell = sheet.cell(header0_row_num, 3)
            header0_recorded_cell.value = _("Recorded (Surplus Sheet)")
            sheet.merge_cells(f"C{header0_row_num}:D{header0_row_num}")
            header0_actual_cell = sheet.cell(header0_row_num, 5)
            header0_actual_cell.value = _("Actual (Surplus Sheet)")
            sheet.merge_cells(f"E{header0_row_num}:F{header0_row_num}")
            header0_difference_cell = sheet.cell(header0_row_num, 7)
            header0_difference_cell.value = _("Difference (Surplus Sheet)")
            sheet.merge_cells(f"G{header0_row_num}:H{header0_row_num}")
            header0_reason_cell = sheet.cell(header0_row_num, 9)
            header0_reason_cell.value = _("Reason (Surplus Sheet)")
            sheet.merge_cells(f"I{header0_row_num}:I{header0_row_num+1}")

            header1_row_num = header0_row_num + 1
            header1_recorded_quantity_cell = sheet.cell(header1_row_num, 3)
            header1_recorded_quantity_cell.value = _(
                "Recorded Quantity (Surplus Sheet)"
            )
            header1_recorded_total_price_cell = sheet.cell(header1_row_num, 4)
            header1_recorded_total_price_cell.value = _(
                "Recorded Total Price (Surplus Sheet)"
            )
            header1_actual_quantity_cell = sheet.cell(header1_row_num, 5)
            header1_actual_quantity_cell.value = _(
                "Actual Quantity (Surplus Sheet)"
            )
            header1_actual_total_price_cell = sheet.cell(header1_row_num, 6)
            header1_actual_total_price_cell.value = _(
                "Actual Total Price (Surplus Sheet)"
            )
            header1_difference_quantity_cell = sheet.cell(header1_row_num, 7)
            header1_difference_quantity_cell.value = _(
                "Difference Quantity (Surplus Sheet)"
            )
            header1_difference_total_price_cell = sheet.cell(header1_row_num, 8)
            header1_difference_total_price_cell.value = _(
                "Difference Total Price (Surplus Sheet)"
            )

            summary_total_price = Decimal("0.0")
            formed_ingredients_index = index
            for index, ingredient in enumerate(ingredients):
                ingredient_row_num = header1_row_num + index + 1
                ingredient_quantity = (
                    ingredient.quantity
                    - sum(
                        [
                            c.amount_used
                            for c in ingredient.consumptions.filter(
                                Q(date_of_using__lte=inventory_day)
                                & Q(is_disabled=False)
                            ).all()
                        ]
                    )
                    if ingredient.id
                    else Decimal("0.0")
                )
                ingredient_total_price = (
                    ingredient_quantity * ingredient.unit_price
                )
                summary_total_price += ingredient_total_price

                ingredient_name_cell = sheet.cell(ingredient_row_num, 1)
                ingredient_name_cell.value = ingredient.name
                if ingredient.name:
                    ingredient_name_cell.comment = Comment(
                        _(
                            "{meal_type} ({category}, Checked/Storaged at {storage_date})"
                        ).format(
                            meal_type=(
                                (
                                    ingredient.meal_type.abbreviation
                                    or ingredient.meal_type.name
                                )
                                if ingredient.meal_type
                                else ""
                            ),
                            category=ingredient.category,
                            storage_date=ingredient.storage_date,
                        ),
                        user.username,
                    )
                sheet.cell(ingredient_row_num, 2, ingredient.quantity_unit_name)
                sheet.cell(ingredient_row_num, 3, ingredient_quantity or "")
                sheet.cell(ingredient_row_num, 4, ingredient_total_price or "")
                sheet.cell(ingredient_row_num, 5, ingredient_quantity or "")
                sheet.cell(ingredient_row_num, 6, ingredient_total_price or "")
                sheet.cell(ingredient_row_num, 7, "")
                sheet.cell(ingredient_row_num, 8, "")
                sheet.cell(ingredient_row_num, 9, "")

            summary_row_num = header1_row_num + ingredient_rows_count + 1

            prev_inventory_day = (
                formed_ingredients[formed_ingredients_index - 1][0]
                if 0 <= formed_ingredients_index - 1 < len(formed_ingredients)
                else None
            )
            next_inventory_day = (
                formed_ingredients[formed_ingredients_index + 1][0]
                if 0 < formed_ingredients_index + 1 < len(formed_ingredients)
                else None
            )
            summary_col1_value = ""

            if next_inventory_day and next_inventory_day == inventory_day:
                summary_col1_value = _(
                    "Sub0-summary Total Price (Surplus Sheet)"
                )
            else:
                summary_col1_value = _("Summary Total Price (Surplus Sheet)")
                summary_total_price = Decimal("0.0")
                ingredients_list = [
                    __ingredients
                    for __inventory_day, __inventory_day_index, __ingredients in formed_ingredients
                    if __inventory_day == inventory_day
                ]
                for ingredients in ingredients_list:
                    for ingredient in ingredients:
                        ingredient_quantity = ingredient.quantity - (
                            sum(
                                [
                                    c.amount_used
                                    for c in ingredient.consumptions.filter(
                                        Q(is_disabled=False)
                                        & Q(date_of_using__lte=inventory_day)
                                    ).all()
                                ]
                            )
                            if ingredient.id
                            else Decimal("0.0")
                        )
                        summary_total_price += (
                            ingredient_quantity * ingredient.unit_price
                        )

            sheet.cell(summary_row_num, 1, summary_col1_value)

            sheet.cell(summary_row_num, 4, summary_total_price)
            sheet.cell(summary_row_num, 6, summary_total_price)

            bottom_mote_row_num = summary_row_num + 1
            bottom_mote_cell = sheet.cell(bottom_mote_row_num, 1)
            bottom_mote_cell.value = _("Note: Inventory once a week.")
            sheet.merge_cells(f"A{bottom_mote_row_num}:I{bottom_mote_row_num}")
            signature_row_num = bottom_mote_row_num + 1
            signature_cell = sheet.cell(
                signature_row_num,
                1,
                _(
                    "   Reviewer:        Handler:{handler_name}    Weigher:        Warehouseman: 　     "
                ).format(handler_name=user.username),
            )
            sheet.merge_cells(f"A{signature_row_num}:I{signature_row_num}")

        for col, width in [
            [1, 2.08],
            [2, 0.49],
            [3, 0.75],
            [4, 1.14],
            [5, 0.89],
            [6, 1.15],
            [7, 0.74],
            [8, 0.74],
            [9, 0.56],
        ]:
            set_column_width_in_inches(sheet, col, width)

    def fill_in_non_storage_list_sheet(self):
        sheet = self.non_storage_list_sheet
        user = self.user
        ingredient_rows_count = 11

        ingredients = Ingredient.objects.filter(
            Q(user=user)
            & Q(storage_date__gte=self.date_start)
            & Q(storage_date__lte=self.date_end)
            & Q(meal_type=self.meal_type)
            & Q(is_disabled=False)
            & Q(is_ignorable=True)
        ).all()

        categories = list(set([i.category for i in ingredients]))

        category_ingredients = []
        for category in categories:
            _ingredients = [i for i in ingredients if i.category == category]
            split_count = math.ceil(len(_ingredients) / ingredient_rows_count)
            for i in range(0, len(_ingredients), ingredient_rows_count):
                _split_ingredients = _ingredients[i : i + ingredient_rows_count]
                _split_ingredient0 = _split_ingredients[0]
                if len(_split_ingredients) < ingredient_rows_count:
                    _split_ingredients += [
                        Ingredient(
                            user=user,
                            storage_date=None,
                            name="",
                            meal_type=_split_ingredient0.meal_type,
                            category=_split_ingredient0.category,
                            quantity=Decimal("0"),
                            quantity_unit_name=None,
                            total_price=Decimal("0"),
                            is_ignorable=_split_ingredient0.is_ignorable,
                        )
                        for i in range(
                            ingredient_rows_count - len(_split_ingredients)
                        )
                    ]
                category_ingredients.append([category, _split_ingredients])

        for index, (category, c_ingredients) in enumerate(category_ingredients):
            row_num = (ingredient_rows_count + 5) * index + 1
            title_row_num = row_num
            title_cell = sheet.cell(title_row_num, 1)
            title_cell.value = _(
                "Table of Non-storaged Ingredients ({category})"
            ).format(category=category)
            title_cell.font = self.font_20_bold
            title_cell.alignment = self.center_alignment

            set_row_height_in_inches(sheet, title_row_num, 0.42)
            sheet.merge_cells(f"A{title_row_num}:G{title_row_num}")

            sub_title_row_num = title_row_num + 1
            sub_title_affiliation_cell = sheet.cell(sub_title_row_num, 1)
            sub_title_affiliation_cell.value = (
                _("Principal Name: {affiliation}")
                if self.is_school
                else _("Affiliation Name: {affiliation}")
            ).format(affiliation=user.affiliation)

            sub_title_date_cell = sheet.cell(sub_title_row_num, 4)
            sub_title_date_cell.value = _(
                "{year}.{month:0>2}.{day:0>2} (Non-storage list sheet)"
            ).format(year=self.year, month=self.month, day=self.date_end.day)

            for cell in [sub_title_affiliation_cell, sub_title_date_cell]:
                cell.font = self.font_14
                cell.alignment = self.center_alignment

            sheet.merge_cells(f"A{sub_title_row_num}:C{sub_title_row_num}")
            sheet.merge_cells(f"D{sub_title_row_num}:G{sub_title_row_num}")

            set_row_height_in_inches(sheet, sub_title_row_num, 0.33)

            header_row_num = sub_title_row_num + 1
            for col, value in [
                [1, _("Procurement Date (Non-storage list sheet)")],
                [2, _("Ingredient Name (Non-storage list sheet)")],
                [3, _("Unit Name of Quantity (Non-storage list sheet)")],
                [4, _("Quantity (Non-storage list sheet)")],
                [5, _("Unit Price (Non-storage list sheet)")],
                [6, _("Total Price (Non-storage list sheet)")],
                [7, _("Note (Non-storage list sheet)")],
            ]:
                cell = sheet.cell(header_row_num, col)
                cell.value = value
                cell.font = self.font_16
                cell.alignment = self.center_alignment
                cell.border = self.thin_border

            set_row_height_in_inches(sheet, header_row_num, 0.30)

            for i_index, ingredient in enumerate(c_ingredients):
                ingredient_row_num = header_row_num + 1 + i_index

                storage_date_cell = sheet.cell(ingredient_row_num, 1)
                storage_date_cell.value = (
                    _(
                        "{year}.{month:0>2}.{day:0>2} (Column of Non-storage list sheet)"
                    ).format(
                        year=ingredient.storage_date.year,
                        month=ingredient.storage_date.month,
                        day=ingredient.storage_date.day,
                    )
                    if ingredient.storage_date
                    else ""
                )
                name_cell = sheet.cell(ingredient_row_num, 2)
                name_cell.value = ingredient.name
                quantity_unit_name_cell = sheet.cell(ingredient_row_num, 3)
                quantity_unit_name_cell.value = ingredient.quantity_unit_name
                quantity_cell = sheet.cell(ingredient_row_num, 4)
                quantity_cell.value = (
                    ingredient.quantity if ingredient.quantity else ""
                )
                unit_price_cell = sheet.cell(ingredient_row_num, 5)
                unit_price_cell.value = (
                    ingredient.unit_price if ingredient.unit_price else ""
                )
                total_price_cell = sheet.cell(ingredient_row_num, 6)
                total_price_cell.value = (
                    ingredient.total_price if ingredient.total_price else ""
                )
                note_cell = sheet.cell(ingredient_row_num, 7)
                note_cell.value = ""

                for col in range(1, 8):
                    cell = sheet.cell(ingredient_row_num, col)
                    cell.font = self.font_12
                    cell.alignment = self.center_alignment
                    cell.border = self.thin_border

                set_row_height_in_inches(sheet, ingredient_row_num, 0.30)

            summary_row_num = header_row_num + ingredient_rows_count + 1

            next_category, ___ = (
                category_ingredients[index + 1]
                if (index + 1) < len(category_ingredients)
                else (None, None)
            )
            summary_note_cell = sheet.cell(summary_row_num, 2)
            summary_total_price_cell = sheet.cell(summary_row_num, 6)

            c_total_price = Decimal("0.0")
            if (not next_category) or (next_category != category):
                summary_note_cell.value = _("Summary (Non-storage list sheet)")
                c_ingredients_list = [
                    _c_ingredients
                    for _category, _c_ingredients in category_ingredients
                    if _category == category
                ]
                c_total_price = Decimal("0.0")
                for _c_ingredients in c_ingredients_list:
                    c_total_price += sum(
                        [i.total_price for i in _c_ingredients]
                    )
                summary_total_price_cell.value = (
                    f"{c_total_price:.{decimal_prec}f}"
                )

            else:
                summary_note_cell.value = _(
                    "Sub-summary (Non-storage list sheet)"
                )
                c_total_price += sum([i.total_price for i in c_ingredients])
                summary_total_price_cell.value = (
                    f"{c_total_price:.{decimal_prec}f}"
                )

            for col in range(1, 8):
                cell = sheet.cell(summary_row_num, col)
                cell.font = self.font_14
                cell.alignment = self.center_alignment
                cell.border = self.thin_border

            set_row_height_in_inches(sheet, summary_row_num, 0.30)

        for col, width in [
            [1, 1.17],
            [2, 1.67],
            [3, 1.30],
            [4, 0.86],
            [5, 0.95],
            [6, 1.28],
            [7, 1.04],
        ]:
            set_column_width_in_inches(sheet, col, width)

    def fill_in_food_sheets(self):
        user = self.request.user
        date_start = self.date_start
        date_end = self.date_end
        meal_type = self.meal_type

        ingredients = Ingredient.objects.filter(
            Q(is_disabled=False)
            & Q(is_ignorable=False)
            & Q(user=user)
            & Q(meal_type=meal_type)
        ).all()

    def fill_in(self):
        self.fill_in_cover_sheet()
        self.fill_in_storage_sheet()
        self.fill_in_storage_list_sheet()
        self.fill_in_non_storage_sheet()
        self.fill_in_non_storage_list_sheet()
        self.fill_in_consumption_sheet()
        self.fill_in_consumption_list_sheet()
        self.fill_in_surplus_sheet()
        # self.fill_in_food_sheets()

        return self.wb


def get_workbook_zip(request, month):
    meal_types = MealType.objects.annotate(
        ingredients_count=Count("ingredients")
    ).filter(
        Q(user=request.user) & Q(is_disabled=False) & Q(ingredients_count__gt=0)
    )

    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
        for meal_type in meal_types:
            filename = (
                _(
                    "Canteen {meal_type} Daybook WorkBook ({month}) of {affiliation}"
                ).format(
                    meal_type=(
                        (meal_type.abbreviation or meal_type.name)
                        if meal_type
                        else ""
                    ),
                    month=month.replace("-", ""),
                    affiliation=request.user.affiliation,
                )
                + ".xlsx"
            )

            wb = CanteenWorkBook(request, month, meal_type).fill_in()
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)

            zip_file.writestr(filename, excel_buffer.getvalue())

    zip_buffer.seek(0)
    return zip_buffer
